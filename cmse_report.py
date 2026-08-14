#!/usr/bin/env python
r"""MMSEA_MSPi_CMSE Dashboard - MMSEA response-file loading + MSPI file loading.

Modeled on the "Paid Dates - Medical" static-HTML dashboards and on the
MMSEA-2026 tab of
  \\trgfile1\Shared\DIG\Data Business Delivery Team\Delivery Schedule\2026\ClientTracker.2026.xlsx

Formerly two separate dashboards (CMSEReport.html + MSPIReport.html), merged
2026-08-14.  Six tabs:
  1. Calendar  - Client / File Type / Frequency / Handling + Jan..Dec grid, one
                 cell per month holding the ADO ticket number(s) of the loads
                 that landed that month ("E" = expected, carried over from the
                 tracker; pink = tracker outreach/late date).
  2. Loads     - every cmse_new..SourceLog row for the in-scope SourceIds,
                 SourceLogId desc, expandable to the per-file ImportStaging
                 breakdown (EntryName / DNDispositionCode / StagingStatus /
                 record counts).
  3. MMSEA_Report - the layout of MMSEA_Report_20250605.xlsx: one row per
                 SourceLog x StagingStatus, in the spreadsheet's column order.
  4. MSPi Monthly - the MSPI dashboard's client/contract x Jan..Dec matrix of
                 files loaded, ported unchanged.
  5. MSPi Report - the layout of "MSPI Report_20250605.xlsx", which is an export
                 of MSP.Report.MSPIMonthlyReport (a table that stopped being
                 refreshed at TapeID 3949 / 2026-05-01).  Computed live so it
                 stays current; verified column-for-column against that table.
  6. Reference - the File Type / Display Length table that sits to the right of
                 the calendar on the Excel tab, plus the SourceId->SourceName
                 key, the Client/ClientId key and the legend.

Data:
  TRGRepSQL3 / cmse_new    Source, SourceLog, ImportStaging, Client, ClientCodeLookup
  TRGINTP3 / MSP           etl.Tape + history.DET / history.PRM / history.MSP
  ADO / TFS                MMSEA work items, matched to a load by its PCN
                           (SourceLog.ProductionControlId) appearing in the
                           work-item description.

Everything expensive (the ImportStaging rollup, the per-PCN ADO lookups) is
cached in cmse_report_cache.json and only re-queried for new / recent rows, so
the 5am weekday refresh is cheap.  `--full` rebuilds from scratch.

Manual run:  python C:\Users\tls2\.claude\projects\H--\cmse_report.py
"""

import json
import os
import re
import shutil
import subprocess
import sys
import time
from collections import defaultdict, OrderedDict
from datetime import datetime, timedelta

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, r"H:\\")
try:
    from report_export import EXPORT_CSS, EXPORT_JS, EXPORT_UI
except ImportError:  # export dropdown is optional
    EXPORT_CSS = EXPORT_JS = EXPORT_UI = ""

SERVER = "TRGRepSQL3"
DATABASE = "cmse_new"
SOURCE_IDS = [4, 5, 6, 12, 13, 14, 15, 18, 19, 20]
WINDOW_START = "2025-01-01"

# ---- MSPI (merged in from the standalone MSPI File Report, 2026-08-14) ------
MSPI_SERVER = "TRGINTP3"
MSPI_DATABASE = "MSP"
# 310 = DET file, 300 = PRM file.  TableID 320 rows are "MSP Archive [nnnn]"
# stubs with no ProdCtrlNo or file, and are excluded - which is also exactly the
# row set of Report.MSPIMonthlyReport.
MSPI_TYPE_BY_TABLEID = {"310": "DET", "300": "PRM"}
# Every one of the 3,789 rows in Report.MSPIMonthlyReport is "Raw"; the table has
# never held any other value, and nothing on etl.Tape distinguishes the two.
MSPI_RAW_OR_TRANSFORM = "Raw"

ADO_BASE = "https://devops.ado.rawlingslou.prod/TFS2012/Rawlings"
ADO_WEB = ADO_BASE + "/_workitems/edit/"

TRACKER = (r"\\trgfile1\Shared\DIG\Data Business Delivery Team\Delivery Schedule"
           r"\2026\ClientTracker.2026.xlsx")
TRACKER_SHEET = "MMSEA - 2026"
TRACKER_YEAR = 2026
LATE_FILL = "FFFFC7CE"          # pink "Late - Outreach Date" fill on the Excel tab

CACHE = os.path.join(HERE, "cmse_report_cache.json")

REPORT_NAME = "MMSEA_MSPi_CMSE"

OUTPUT_PATHS = [
    r"\\trgfile1\Shared\DIG\Data Business Delivery Team\Delivery Schedule"
    r"\Daily Status Reports\%s.html" % REPORT_NAME,
    os.path.join(HERE, "%s.html" % REPORT_NAME),
    r"C:\Users\tls2\OneDrive - Machinify\Documents\Reports\%s.html" % REPORT_NAME,
]

# The two dashboards this one replaces.  Their old file names are still written,
# as copies of the merged report, so existing links and bookmarks keep working
# and never go stale.
LEGACY_NAMES = ["CMSEReport.html", "MSPIReport.html"]

# SourceId -> (tracker row label, file-format spec label)
SOURCE_TYPE = {
    4:  ("HEW",  "HEW"),
    5:  ("MSP",  "MSP"),
    6:  ("NMSP", "Non-MSP"),
    12: ("NMSP", "Aetna NMSP"),
    13: ("NMSP", "Aetna NMSP"),
    14: ("MSP",  "Aetna MSP"),
    15: ("MSP",  "Aetna MSP"),
    18: ("HEW",  "Aetna HEW"),
    19: ("NMSP", "Aetna NMSP"),
    20: ("MSP",  "Aetna MSP"),
}

# File Type / Display Length table from the right of the Excel calendar (S1:X8)
FILE_SPECS = [
    ("HEW",        302,  300,  "Yes - %HO%",    "38",   "29"),
    ("MSP",        802,  800,  "Yes - %MSPR%",  "51",   "175"),
    ("Non-MSP",    502,  500,  "Yes",           "215",  "5"),
    ("Aetna HEW",  1203, 1201, "None",          "825",  "-"),
    ("Aetna MSP",  2050, 2048, "None",          "1244", "-"),
    ("Aetna NMSP", 1879, 1877, "None",          "1515", "-"),
]

# tracker client label -> cmse_new Client.ClientName
TRACKER_ALIAS = {
    "aetna": "Aetna",
    "bcbsfl": "BCBSFL",
    "bcbsks": "BCBSKS",
    "bcbsnc": "BCBS NC",
    "bcbssc": "BCBS SC",
    "bsca": "BSCA_FACETS",
    "carefirst": "CareFirst",
    "cigna": "Cigna",
    "emblem": "Emblem",
    "excellus": "Excellus",
    "harvardpilgrim": "HARVARDP",
    "healthnet": "HealthNet",
    "hne": "HealthNewEngland",
    "jhhc": "JohnsHopkins",
    "kaiser": "Kaiser",
    "kaiser_wa": "Kaiser_WA",
    "medica": "Medica",
    "mmoh": "Medical Mutual of Ohio",
    "oscar": "Oscar",
    "premera": "Premera Blue Cross",
    "tufts": "TUFTS",
}

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# ImportStaging.StagingStatus
STAGING_STATUS = [
    ("1",  "Record has been acquired from a source and staged in CMSE"),
    ("2",  "Record is currently in the process of being validated"),
    ("3",  "Record has been validated \u2014 keep checking. May have to enter ticket"),
    ("6",  "Record having passed validation, is in the process of being imported "
           "into the CMSE patient tables"),
    ("7",  "Record has been imported into the CMSE patient tables"),
    ("11", "Record failed validation for a reason of invalid SSN"),
    ("12", "Record failed validation for a reason of invalid DOB"),
    ("13", "Record failed validation for a reason of invalid last name"),
    ("14", "Record failed for invalid effective dates"),
]

# Tools linked from the dashboard toolbar: (label, UNC path, protocol url or "")
#
# A browser will not execute an .exe from a file:// link, so FileDate goes
# through the per-user "filedate:" URL protocol registered by
# Register-FileDate-Protocol.reg (which is published next to the report).
# File Transformer is a ClickOnce .application, which the shell can open
# directly from a file:// url.
TOOLS = [
    ("File Date Change",
     r"\\trgfile1\Shared\DIG\Data Business Delivery Team\Data Delivery Documentation\FileDate.exe",
     "filedate:open"),
    ("File Transformer",
     r"\\trgfile1\Operations\Software\_Source\ToolBox\File Transformer\File Transformer.application",
     "filetransformer:open"),
]

# Published next to the report so other users can self-register the protocols by
# double-clicking the .reg.  The .cmd is the launcher the filetransformer:
# handler points at, so it has to sit in that same folder.
REG_FILE = "Register-CMSE-Tools.reg"
SIDECARS = [REG_FILE, "LaunchFileTransformer.vbs"]

# "R or T" from the MMSEA_Report spreadsheet, keyed on (ClientId, SourceId).
#
# The value is constant per feed - all 782 loads in MMSEA_Report_20250605.xlsx
# agree, with no (client, file type) pair carrying both letters - but nothing in
# cmse_new drives it, so the lookup is transcribed from that spreadsheet.  Feeds
# that started after it was taken (HARVARDP, Kaiser_WA, BSCA_FACETS, Excellus)
# aren't in it and render blank until someone supplies their letter.
RT_BY_FEED = {
    (2,   4): "R",   # Emblem / Hew Response File
    (2,   5): "R",   # Emblem / MSP Response File
    (3,   4): "R",   # United / Hew Response File
    (3,   5): "T",   # United / MSP Response File
    (5,  19): "T",   # Aetna / Aetna Traditional AIS NonMSP File
    (5,  20): "T",   # Aetna / Aetna Traditional AIS MSP
    (6,   4): "R",   # CareFirst / Hew Response File
    (6,   5): "T",   # CareFirst / MSP Response File
    (6,   6): "T",   # CareFirst / Non MSP Response File
    (12,  4): "R",   # HealthNet / Hew Response File
    (12,  5): "R",   # HealthNet / MSP Response File
    (16,  4): "R",   # TUFTS / Hew Response File
    (16,  5): "T",   # TUFTS / MSP Response File
    (17,  4): "T",   # Medical Mutual of Ohio / Hew Response File
    (17,  5): "R",   # Medical Mutual of Ohio / MSP Response File
    (31,  4): "R",   # BCBS NC / Hew Response File
    (31,  5): "R",   # BCBS NC / MSP Response File
    (36,  5): "R",   # BCBS SC / MSP Response File
    (37,  5): "R",   # Premera Blue Cross / MSP Response File
    (37,  6): "T",   # Premera Blue Cross / Non MSP Response File
    (41,  4): "T",   # BCBSKS / Hew Response File
    (41,  5): "T",   # BCBSKS / MSP Response File
    (42,  4): "R",   # Cigna / Hew Response File
    (42,  5): "R",   # Cigna / MSP Response File
    (42,  6): "T",   # Cigna / Non MSP Response File
    (43,  4): "R",   # Kaiser / Hew Response File
    (43,  5): "R",   # Kaiser / MSP Response File
    (44,  5): "T",   # Cambia / MSP Response File
    (44,  6): "T",   # Cambia / Non MSP Response File
    (45,  4): "T",   # Oscar / Hew Response File
    (45,  5): "T",   # Oscar / MSP Response File
    (46,  5): "R",   # BCBSFL / MSP Response File
    (46,  6): "R",   # BCBSFL / Non MSP Response File
    (48,  5): "R",   # GEHA / MSP Response File
    (48,  6): "R",   # GEHA / Non MSP Response File
    (61,  5): "T",   # JohnsHopkins / MSP Response File
    (62,  4): "T",   # Medica / Hew Response File
    (62,  5): "T",   # Medica / MSP Response File
    (63,  5): "T",   # BCBSVT / MSP Response File
    (102, 5): "T",   # HealthNewEngland / MSP Response File
}

# Columns of the MMSEA_Report spreadsheet that nothing in cmse_new reproduces.
# Verified 2026-08-14 against MMSEA_Report_20250605.xlsx: two loads with
# near-identical ImportStaging profiles (HNE MSP 32225 and GEHA MSP 31824) carry
# 0%/0% and 100%/100%, so no record-level predicate can produce them - they come
# from outside this database.  Rendered blank rather than guessed.
UNSOURCED_COLUMNS = ["E %", "MC %", "# of Invs"]

# Loads whose ticket the PCN search can't find (the PCN isn't written in any
# work item description).  Keyed on the ProductionControlId; a key that is the
# bare base number also covers its suffixed variants ("9537705" -> "9537705_01").
PCN_ADO_OVERRIDE = {
    "9537705": 931813,   # Kaiser HEW 2026-02-17/18 (base + _01.._04)
    "991597":  941594,   # HealthNet HEW 2026-03-31
    "984687":  940095,   # Excellus MSP 2026-03-25
    "984688":  940095,
    "984690":  940095,
    "984691":  940095,
    "984692":  940095,
}


# --------------------------------------------------------------------------- #
# SQL
# --------------------------------------------------------------------------- #

def sql(query, timeout=900, server=None, database=None, sep="|"):
    """Run a query via sqlcmd, return a list of column lists (no header).

    `sep` is the column delimiter; the MSPI queries use a tab because their
    file-path column can plausibly contain other punctuation.
    """
    p = subprocess.run(
        ["sqlcmd", "-S", server or SERVER, "-d", database or DATABASE, "-E",
         "-W", "-w", "65535", "-s", sep, "-h", "-1",
         "-Q", "SET NOCOUNT ON; " + query],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
        timeout=timeout)
    if p.returncode != 0:
        raise RuntimeError("sqlcmd failed: %s" % ((p.stderr or p.stdout or "")[:500]))
    out = []
    for line in (p.stdout or "").splitlines():
        line = line.rstrip("\r")
        if not line or line.startswith("(") or set(line) <= set("-" + sep):
            continue
        # char columns come back space-padded; NULL prints as the literal "NULL"
        out.append([("" if c.strip() == "NULL" else c.strip()) for c in line.split(sep)])
    return out


def fetch_sources():
    rows = sql("SELECT SourceId, SourceName, ISNULL(SourceDescription,'') "
               "FROM dbo.Source ORDER BY SourceId")
    return [{"id": int(r[0]), "name": r[1], "desc": r[2]} for r in rows if r[0].isdigit()]


def fetch_clients():
    rows = sql("""
        SELECT C.ClientId, C.ClientName,
               STUFF((SELECT ', ' + X.SMARTClientCode
                        FROM dbo.ClientCodeLookup X (NOLOCK)
                       WHERE X.ClientId = C.ClientId
                    ORDER BY X.SMARTClientCode FOR XML PATH('')), 1, 2, '')
          FROM dbo.Client C (NOLOCK)
         WHERE C.ClientId IN (SELECT DISTINCT ClientId FROM dbo.SourceLog (NOLOCK)
                               WHERE SourceId IN (%s) AND ImportStartDate >= '%s')
      ORDER BY C.ClientName
    """ % (",".join(map(str, SOURCE_IDS)), WINDOW_START))
    return [{"id": int(r[0]), "name": r[1], "codes": (r[2] if len(r) > 2 else "")}
            for r in rows if r[0].isdigit()]


def fetch_loads():
    rows = sql("""
        SELECT L.SourceLogId, L.SourceId, ISNULL(L.ClientId,0), ISNULL(C.ClientName,''),
               L.EntryName,
               CONVERT(varchar(19), L.ImportStartDate, 120),
               ISNULL(CONVERT(varchar(19), L.ImportCompleteDate, 120), ''),
               ISNULL(L.RecordCount,0), ISNULL(L.ImportSuccessCount,0),
               ISNULL(L.ImportFailedcount,0),
               ISNULL(CONVERT(varchar(3), L.MIRProcessed), ''),
               ISNULL(CONVERT(varchar(19), L.MIRProcessedDate, 120), ''),
               ISNULL(L.ProductionControlId,''),
               ISNULL(L.EntitlementAgeCount,0), ISNULL(L.EntitlementDisabilityCount,0),
               ISNULL(L.EntitlementEsrdCount,0)
          FROM dbo.SourceLog L (NOLOCK)
          LEFT JOIN dbo.Client C (NOLOCK) ON C.ClientId = L.ClientId
         WHERE L.SourceId IN (%s) AND L.ImportStartDate >= '%s'
      ORDER BY L.SourceLogId DESC
    """ % (",".join(map(str, SOURCE_IDS)), WINDOW_START))
    loads = []
    for r in rows:
        if not r[0].isdigit():
            continue
        loads.append({
            "sl": int(r[0]), "src": int(r[1]), "cid": int(r[2]), "client": r[3],
            "entry": r[4], "start": r[5], "done": r[6],
            "rec": int(r[7]), "ok": int(r[8]), "bad": int(r[9]),
            "mir": r[10], "mird": r[11],
            "pcn": r[12],
            "age": int(r[13]), "dis": int(r[14]), "esrd": int(r[15]),
        })
    return loads


def fetch_staging(source_log_ids):
    """ImportStaging rollup for the given SourceLogIds -> {sl: [[dn, status, n], ...]}."""
    out = defaultdict(list)
    ids = sorted(source_log_ids)
    for i in range(0, len(ids), 400):
        chunk = ids[i:i + 400]
        rows = sql("""
            SELECT I.SourceLogId, ISNULL(I.DNDispositionCode,''),
                   ISNULL(CONVERT(varchar(12), I.StagingStatus),''), COUNT(*)
              FROM dbo.ImportStaging I (NOLOCK)
             WHERE I.SourceLogId IN (%s)
          GROUP BY I.SourceLogId, I.DNDispositionCode, I.StagingStatus
          ORDER BY I.SourceLogId, I.DNDispositionCode, I.StagingStatus
        """ % ",".join(map(str, chunk)), timeout=1800)
        for r in rows:
            if r[0].isdigit():
                out[int(r[0])].append([r[1], r[2], int(r[3])])
    return out


def fetch_unique_members(source_log_ids):
    """The spreadsheet's "U M Count" -> {sl: n}.

    Verified against MMSEA_Report_20250605.xlsx: it is the number of distinct
    SSN + last name + first initial + DOB combinations in ImportStaging.  Not
    distinct SSN (26733 has 139, the sheet says 141) and not SSN + DOB alone
    (140).  Dropping DOB from the key agrees on small files but silently
    over-collapses big ones - it was wrong on 50 of the 86 loads the dashboard
    and the spreadsheet share.  Middle initial and HIC number are NOT part of
    the key.
    """
    out = {}
    ids = sorted(source_log_ids)
    for i in range(0, len(ids), 200):
        chunk = ids[i:i + 200]
        rows = sql("""
            SELECT I.SourceLogId,
                   COUNT(DISTINCT ISNULL(I.PatientSSN,'') + '|'
                                + ISNULL(I.PatientLastName,'') + '|'
                                + ISNULL(I.PatientFirstInitial,'') + '|'
                                + ISNULL(CONVERT(varchar(10), I.PatientDOB, 112),''))
              FROM dbo.ImportStaging I (NOLOCK)
             WHERE I.SourceLogId IN (%s)
          GROUP BY I.SourceLogId
        """ % ",".join(map(str, chunk)), timeout=1800)
        for r in rows:
            if r[0].isdigit():
                out[int(r[0])] = int(r[1])
    return out


# --------------------------------------------------------------------------- #
# MSPI  (TRGINTP3 / MSP)
# --------------------------------------------------------------------------- #

# A contract dot-segment: optional routing 'R', a letter, then 3-5 digits
# (anchored so date/time segments like D260622 / T0858590 don't match).
MSPI_CONTRACT_SEG_RE = re.compile(r"^R?[A-Z]\d{3,5}$")
# Drop a single leading routing 'R' only when a real contract (letter+digit)
# follows it: RH3890 -> H3890, RR6694 -> R6694, but R6694 stays R6694.
MSPI_LEAD_R_RE = re.compile(r"^R(?=[A-Z]\d)")
# Data-date token in the leaf name, e.g. D260623 -> 2026-06-23 (Date Extracted).
MSPI_DATE_TOKEN_RE = re.compile(r"^D(\d{2})(\d{2})(\d{2})$")
# Aetna MAO_DATA / MAODATA files carry no contract token -> label them 'MAO'.
MSPI_MAO_DATA_RE = re.compile(r"^MAO[ _]?DATA", re.I)


def mspi_leaf(path):
    return re.split(r"[\\/]", path or "")[-1]


def mspi_path_client(path):
    """Folder after ...\\MSP\\ in the path.

    This is the Monthly tab's grouping key.  It is NOT the MSPi Report tab's
    Client - that one comes from the history tables and is blank for the 112
    tapes that loaded no records (see build()).
    """
    segs = [s for s in re.split(r"[\\/]", path or "") if s]
    for i, seg in enumerate(segs[:-1]):
        if seg.upper() == "MSP" and i + 1 < len(segs):
            return segs[i + 1]
    return ""


def mspi_contract(path):
    """Contract token in the leaf file name, leading routing 'R' dropped.

    e.g. H0342 -> H0342, EFTO.RH3890... -> H3890, P.RR6694... -> R6694,
    H5580_MSPCOBMA... -> H5580.  Aetna MAO_DATA files have no contract token
    and are labelled 'MAO'.
    """
    leaf = mspi_leaf(path)
    if MSPI_MAO_DATA_RE.match(leaf):
        return "MAO"
    for seg in re.split(r"[._ ]", leaf.upper()):
        if MSPI_CONTRACT_SEG_RE.match(seg):
            return MSPI_LEAD_R_RE.sub("", seg)
    return ""


def mspi_extracted(path):
    """Data date parsed from the D-token in the leaf name (Date Extracted)."""
    for seg in mspi_leaf(path).upper().split("."):
        m = MSPI_DATE_TOKEN_RE.match(seg)
        if m:
            return "20%s-%s-%s" % (m.group(1), m.group(2), m.group(3))
    return ""


def fetch_mspi_tapes():
    """One row per MSPI file: etl.Tape, TableID 300 (PRM) / 310 (DET)."""
    rows = sql("""
        SELECT t.TapeID, t.TableID, ISNULL(CONVERT(varchar(20), t.ProdCtrlNo),''),
               t.FileName, ISNULL(CONVERT(varchar(20), t.FileSize),''),
               ISNULL(CONVERT(varchar(10), t.FileCreateDate, 120),''),
               ISNULL(CONVERT(varchar(19), t.FileLoadDate, 120),'')
          FROM MSP.etl.Tape t
         WHERE t.TableID IN (300, 310)
      ORDER BY t.TapeID DESC
    """, server=MSPI_SERVER, database=MSPI_DATABASE, sep="\t")
    out = []
    for r in rows:
        if not r[0].isdigit():
            continue
        out.append({
            "tape": int(r[0]), "tableid": r[1], "prod": r[2], "path": r[3],
            "size": int(r[4]) if r[4].isdigit() else 0,
            "created": r[5], "loaded": r[6],
        })
    return out


def fetch_mspi_stats(min_tape):
    """Per-TapeID record + unique-HICN counts -> {tape: [det,deth,prm,prmh,msp,client]}.

    The TapeID floor matters: an unbounded GROUP BY over history.DET +
    history.PRM + history.MSP (36M rows) is ~75s, while the same query pruned to
    recent tapes is well under a second.
    """
    rows = sql("""
        WITH d AS (SELECT TapeID, COUNT_BIG(*) c, COUNT(DISTINCT HICN) h, MAX(Client) client
                     FROM MSP.history.DET WHERE TapeID >= %(t)d GROUP BY TapeID),
             p AS (SELECT TapeID, COUNT_BIG(*) c, COUNT(DISTINCT HICN) h, MAX(Client) client
                     FROM MSP.history.PRM WHERE TapeID >= %(t)d GROUP BY TapeID),
             m AS (SELECT TapeID, COUNT_BIG(*) c, MAX(Client) client
                     FROM MSP.history.MSP WHERE TapeID >= %(t)d GROUP BY TapeID)
        SELECT t.TapeID, ISNULL(d.c,0), ISNULL(d.h,0), ISNULL(p.c,0), ISNULL(p.h,0),
               ISNULL(m.c,0), ISNULL(COALESCE(d.client, p.client, m.client),'')
          FROM MSP.etl.Tape t
          LEFT JOIN d ON d.TapeID = t.TapeID
          LEFT JOIN p ON p.TapeID = t.TapeID
          LEFT JOIN m ON m.TapeID = t.TapeID
         WHERE t.TableID IN (300, 310) AND t.TapeID >= %(t)d
    """ % {"t": min_tape}, timeout=1800, server=MSPI_SERVER,
           database=MSPI_DATABASE, sep="\t")
    out = {}
    for r in rows:
        if r[0].isdigit():
            out[int(r[0])] = [int(r[1]), int(r[2]), int(r[3]), int(r[4]), int(r[5]),
                              r[6]]
    return out


# --------------------------------------------------------------------------- #
# ADO
# --------------------------------------------------------------------------- #

def _curl(args, timeout=180):
    p = subprocess.run(["curl.exe", "-s", "--negotiate", "-u", ":"] + args,
                       capture_output=True, text=True, encoding="utf-8",
                       errors="replace", timeout=timeout)
    if not p.stdout:
        raise RuntimeError("empty response from ADO")
    return json.loads(p.stdout.lstrip("\ufeff"))


def pcn_override(pcn):
    """Hand-supplied work item for a PCN, honouring base-number keys."""
    if pcn in PCN_ADO_OVERRIDE:
        return PCN_ADO_OVERRIDE[pcn]
    base = re.match(r"\d+", pcn or "")
    if base:
        return PCN_ADO_OVERRIDE.get(base.group(0))
    return None


def _ado_desc_search(term):
    q = ("SELECT [System.Id] FROM WorkItems WHERE [System.TeamProject]='Rawlings' "
         "AND [System.Description] CONTAINS '%s'" % term.replace("'", "''"))
    try:
        d = _curl([ADO_BASE + "/_apis/wit/wiql?api-version=5.0",
                   "-H", "Content-Type: application/json",
                   "-d", json.dumps({"query": q})])
    except Exception:
        return []
    return sorted(w["id"] for w in d.get("workItems", []))


def ado_find_pcn(pcn):
    """Work item ids whose description mentions this PCN, lowest (original) first.

    WIQL CONTAINS is word-based, so a suffixed PCN is one token: searching
    "1079229" does NOT find a description that says "1079229_1".  Try the
    ProductionControlId exactly as stored first, then fall back to its leading
    digits (which is what most tickets write).
    """
    terms = [pcn]
    base = re.match(r"\d+", pcn)
    if base and base.group(0) != pcn:
        terms.append(base.group(0))
    for term in terms:
        hits = _ado_desc_search(term)
        if hits:
            return hits
    return []


def ado_details(ids):
    """{id: {title, state, tags}} for the given work item ids."""
    out = {}
    ids = sorted(set(ids))
    for i in range(0, len(ids), 180):
        chunk = ids[i:i + 180]
        try:
            d = _curl([ADO_BASE + "/_apis/wit/workitems?ids=" + ",".join(map(str, chunk)) +
                       "&fields=System.Id,System.Title,System.State,System.Tags"
                       "&api-version=5.0&errorPolicy=omit"], timeout=300)
        except Exception as e:
            print("[warn] ADO detail batch failed: %s" % e)
            continue
        for w in d.get("value", []) or []:
            if not w:
                continue
            f = w.get("fields", {})
            out[w["id"]] = {"title": f.get("System.Title", ""),
                            "state": f.get("System.State", ""),
                            "tags": f.get("System.Tags", "") or ""}
    return out


# --------------------------------------------------------------------------- #
# Tracker (Excel)
# --------------------------------------------------------------------------- #

def _norm(s):
    return re.sub(r"[^a-z0-9_]", "", (s or "").strip().lower())


def read_tracker():
    """-> ({(client, filetype): {"freq","hand","cells":{month:{"v","late"}}}}, note)."""
    try:
        import openpyxl
    except ImportError:
        return {}, "openpyxl not installed - tracker overlay skipped"
    try:
        wb = openpyxl.load_workbook(TRACKER, data_only=True, read_only=False)
        ws = wb[TRACKER_SHEET]
    except Exception as e:
        return {}, "tracker unavailable (%s)" % type(e).__name__

    out = OrderedDict()
    client = ftype = ""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=17):
        a = row[0].value
        b = row[1].value
        if a is not None and str(a).strip():
            if str(a).strip().lower().startswith("e ="):   # legend row
                break
            client = str(a).strip()
            ftype = ""
        if b is not None and str(b).strip():
            ftype = str(b).strip()
        if not client:
            continue
        label = TRACKER_ALIAS.get(_norm(client), client)
        key = (label, ftype or "—")
        ent = out.setdefault(key, {"freq": "", "hand": "", "cells": {}})
        if row[2].value and not ent["freq"]:
            ent["freq"] = str(row[2].value).strip()
        if row[3].value and not ent["hand"]:
            ent["hand"] = str(row[3].value).strip()
        for m in range(12):                       # cols F..Q -> idx 5..16
            c = row[5 + m]
            v = c.value
            if v is None or str(v).strip() == "":
                continue
            if isinstance(v, datetime):
                txt = v.strftime("%m/%d/%Y")
            elif isinstance(v, float) and v == int(v):
                txt = str(int(v))
            else:
                txt = str(v).strip()
            fill = c.fill
            rgb = ""
            try:
                if fill is not None and fill.patternType:
                    rgb = str(fill.fgColor.rgb or "")
            except Exception:
                rgb = ""
            slot = ent["cells"].setdefault(m + 1, {"v": [], "late": False})
            slot["v"].append(txt)
            if rgb.upper() == LATE_FILL:
                slot["late"] = True
    for ent in out.values():
        for slot in ent["cells"].values():
            slot["v"] = " / ".join(slot["v"])
    return out, ""


# --------------------------------------------------------------------------- #
# Cache
# --------------------------------------------------------------------------- #

def load_cache(full):
    if full or not os.path.exists(CACHE):
        return {"staging": {}, "um": {}, "pcn": {}, "pcn_checked": {}, "wi": {},
                "mspi": {}}
    try:
        with open(CACHE, "r", encoding="utf-8") as f:
            c = json.load(f)
    except Exception:
        return {"staging": {}, "um": {}, "pcn": {}, "pcn_checked": {}, "wi": {},
                "mspi": {}}
    for k in ("staging", "um", "pcn", "pcn_checked", "wi", "mspi"):
        c.setdefault(k, {})
    return c


def save_cache(cache):
    tmp = CACHE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(cache, f, separators=(",", ":"))
    os.replace(tmp, CACHE)


# --------------------------------------------------------------------------- #
# Build
# --------------------------------------------------------------------------- #

def build_mspi(cache, now, full=False):
    """The two MSPi tabs' row set: one dict per DET/PRM file."""
    print("[info] %s / %s" % (MSPI_SERVER, MSPI_DATABASE))
    tapes = fetch_mspi_tapes()
    print("[info] %d MSPI files" % len(tapes))

    recent_cut = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    need = [t["tape"] for t in tapes
            if str(t["tape"]) not in cache["mspi"] or t["loaded"][:10] >= recent_cut]
    if need:
        floor = min(need)
        print("[info] MSPI record counts for %d tapes (TapeID >= %d)..."
              % (len(need), floor))
        t0 = time.time()
        got = fetch_mspi_stats(floor)
        for tp in need:
            cache["mspi"][str(tp)] = got.get(tp, [0, 0, 0, 0, 0, ""])
        print("[info]   %.0fs" % (time.time() - t0))

    rows = []
    for t in tapes:
        det, deth, prm, prmh, msp, hclient = cache["mspi"].get(
            str(t["tape"]), [0, 0, 0, 0, 0, ""])
        rows.append({
            "tape": t["tape"],
            "type": MSPI_TYPE_BY_TABLEID.get(t["tableid"], t["tableid"]),
            "prod": t["prod"],                      # = the report's VolSerNum
            # Monthly groups on the path folder (every file has one); the MSPi
            # Report tab shows `rclient`, the history-table Client, which the
            # source report leaves blank for tapes that loaded no records.
            "client": mspi_path_client(t["path"]) or hclient,
            "rclient": hclient,
            "contract": mspi_contract(t["path"]),
            "file": mspi_leaf(t["path"]),
            "path": t["path"],
            "extracted": mspi_extracted(t["path"]),
            "created": t["created"],
            "loaded": t["loaded"],
            "size": t["size"],
            "det": det, "deth": deth, "prm": prm, "prmh": prmh, "msp": msp,
        })
    return rows


def build(full=False):
    cache = load_cache(full)
    now = datetime.now()

    print("[info] %s / %s" % (SERVER, DATABASE))
    sources = fetch_sources()
    clients = fetch_clients()
    loads = fetch_loads()
    print("[info] %d loads, %d clients" % (len(loads), len(clients)))

    # -- ImportStaging: everything uncached, plus anything loaded in the last 30
    #    days (a recent load can still be re-summarized).
    recent_cut = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    want = {l["sl"] for l in loads
            if str(l["sl"]) not in cache["staging"] or l["start"][:10] >= recent_cut}
    before = {str(sl): cache["staging"].get(str(sl)) for sl in want}
    if want:
        print("[info] ImportStaging rollup for %d SourceLogIds..." % len(want))
        t0 = time.time()
        got = fetch_staging(want)
        for sl in want:
            cache["staging"][str(sl)] = got.get(sl, [])
        print("[info]   %.0fs" % (time.time() - t0))

    # -- "U M Count" for the MMSEA_Report tab.  COUNT(DISTINCT ...) over a
    #    multi-million-row Aetna file is minutes, not seconds, so unlike the
    #    rollup this is only redone when the load is new or its ImportStaging
    #    rollup actually changed - re-querying the whole 30-day window every
    #    morning would add several minutes to the 5am run for nothing.
    want_um = {l["sl"] for l in loads
               if str(l["sl"]) not in cache["um"]
               or (l["sl"] in want
                   and before[str(l["sl"])] != cache["staging"][str(l["sl"])])}
    if want_um:
        print("[info] unique-member count for %d SourceLogIds..." % len(want_um))
        t0 = time.time()
        got = fetch_unique_members(want_um)
        for sl in want_um:
            cache["um"][str(sl)] = got.get(sl, 0)
        print("[info]   %.0fs" % (time.time() - t0))

    # -- ADO: resolve each PCN once.  Unresolved PCNs are retried while the load
    #    is still young (a ticket may be written after the file lands).
    todo = []
    for l in loads:
        # Key on the ProductionControlId exactly as stored - it is sometimes
        # suffixed ("1079229_1", "801765-2") and ado_find_pcn() handles the
        # fallback to the bare leading digits itself.
        pcn = l["pcn"].strip()
        l["pcnk"] = pcn
        l["wiover"] = pcn_override(pcn)
        if not pcn or l["wiover"]:
            continue
        if pcn in cache["pcn"]:
            continue
        last = cache["pcn_checked"].get(pcn)
        if last and l["start"][:10] < (now - timedelta(days=45)).strftime("%Y-%m-%d"):
            continue                       # old and still ticket-less: stop asking
        todo.append(pcn)
    todo = sorted(set(todo))
    if todo:
        print("[info] ADO lookup for %d PCNs..." % len(todo))
        t0 = time.time()
        for n, pcn in enumerate(todo, 1):
            hits = ado_find_pcn(pcn)
            if hits:
                cache["pcn"][pcn] = hits
            cache["pcn_checked"][pcn] = now.strftime("%Y-%m-%d")
            if n % 50 == 0:
                print("[info]   %d/%d" % (n, len(todo)))
        print("[info]   %.0fs" % (time.time() - t0))

    # -- work item titles/states for everything we matched
    wanted_wi = set()
    for l in loads:
        for wi in cache["pcn"].get(l.get("pcnk") or "", []):
            wanted_wi.add(wi)
        if l.get("wiover"):
            wanted_wi.add(l["wiover"])
    missing_wi = [w for w in wanted_wi if str(w) not in cache["wi"]]
    if missing_wi:
        print("[info] ADO details for %d work items..." % len(missing_wi))
        for wi, det in ado_details(missing_wi).items():
            cache["wi"][str(wi)] = det
    # refresh state for anything not yet Closed
    stale = [int(w) for w in cache["wi"]
             if int(w) in wanted_wi and cache["wi"][w].get("state") not in ("Closed", "Removed")]
    if stale:
        for wi, det in ado_details(stale).items():
            cache["wi"][str(wi)] = det

    mspi = build_mspi(cache, now, full=full)

    save_cache(cache)

    # -- attach ADO + staging to each load, prefer an MMSEA-tagged ticket
    for l in loads:
        hits = [l["wiover"]] if l.get("wiover") else cache["pcn"].get(l.get("pcnk") or "", [])
        pick = None
        for wi in hits:
            det = cache["wi"].get(str(wi), {})
            blob = (det.get("tags", "") + " " + det.get("title", "")).lower()
            if "mmsea" in blob:
                pick = wi
                break
        if pick is None and hits:
            pick = hits[0]
        l["wi"] = pick
        det = cache["wi"].get(str(pick), {}) if pick else {}
        l["wit"] = det.get("title", "")
        l["wis"] = det.get("state", "")
        l["stg"] = [[(dn or "").strip(), (st or "").strip(), n]
                    for dn, st, n in cache["staging"].get(str(l["sl"]), [])]
        l["ft"] = SOURCE_TYPE.get(l["src"], ("?", "?"))[0]
        l["spec"] = SOURCE_TYPE.get(l["src"], ("?", "?"))[1]
        l["file"] = l["entry"].rsplit("\\", 1)[-1]
        l["um"] = cache["um"].get(str(l["sl"]), 0)
        l["rt"] = RT_BY_FEED.get((l["cid"], l["src"]), "")

    tracker, tracker_note = read_tracker()
    if tracker_note:
        print("[warn] %s" % tracker_note)

    # -- calendar rows: union of what CMSE has loaded and what the tracker lists
    cal = OrderedDict()

    def cal_row(client, ftype, cid):
        key = (client, ftype)
        if key not in cal:
            t = tracker.get(key, {})
            cal[key] = {"client": client, "cid": cid, "ft": ftype,
                        "freq": t.get("freq", ""), "hand": t.get("hand", ""),
                        "m": {}, "tr": {str(k): v for k, v in t.get("cells", {}).items()}}
        elif cid and not cal[key]["cid"]:
            cal[key]["cid"] = cid
        return cal[key]

    name_by_id = {c["id"]: c["name"] for c in clients}
    for l in loads:
        client = l["client"] or name_by_id.get(l["cid"], "(unknown %d)" % l["cid"])
        row = cal_row(client, l["ft"], l["cid"])
        ym = l["start"][:7]
        row["m"].setdefault(ym, []).append(
            {"sl": l["sl"], "wi": l["wi"], "pcn": l["pcn"], "d": l["start"][8:10],
             "rec": l["rec"], "file": l["file"], "src": l["src"]})
    for (client, ftype) in tracker:
        cal_row(client, ftype, 0)

    id_by_name = {c["name"]: c["id"] for c in clients}
    for row in cal.values():
        if not row["cid"]:
            row["cid"] = id_by_name.get(row["client"], 0)

    years = sorted({l["start"][:4] for l in loads} | {str(TRACKER_YEAR)})
    cal_rows = sorted(cal.values(), key=lambda r: (r["client"].lower(), r["ft"]))

    # StagingStatus inventory for the reference tab: every documented status,
    # plus anything observed that isn't documented yet.
    st_ct = defaultdict(int)
    for l in loads:
        for dn, st, n in l["stg"]:
            st_ct[st or "(blank)"] += n
    defined = dict(STAGING_STATUS)
    st_rows = [[code, desc, st_ct.get(code, 0)] for code, desc in STAGING_STATUS]
    st_rows += [[code, "", n] for code, n in sorted(st_ct.items())
                if code not in defined]

    return {
        "generated": now.strftime("%Y-%m-%d %H:%M"),
        "name": REPORT_NAME,
        "mspi": mspi,
        "mspiRawOrTransform": MSPI_RAW_OR_TRANSFORM,
        "windowStart": WINDOW_START,
        "years": years,
        "trackerYear": str(TRACKER_YEAR),
        "trackerNote": tracker_note,
        "cal": cal_rows,
        "loads": loads,
        "sources": sources,
        "scope": SOURCE_IDS,
        "clients": clients,
        "specs": [list(s) for s in FILE_SPECS],
        "srcType": {str(k): list(v) for k, v in SOURCE_TYPE.items()},
        "st": st_rows,
        "stDef": dict(STAGING_STATUS),
        "unsourced": UNSOURCED_COLUMNS,
        "tools": [list(t) for t in TOOLS],
        "regFile": REG_FILE,
        "regDir": os.path.dirname(OUTPUT_PATHS[0]),
        "adoBase": ADO_WEB,
    }


# --------------------------------------------------------------------------- #
# HTML
# --------------------------------------------------------------------------- #

HTML_TEMPLATE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>MMSEA_MSPi_CMSE</title>
<style>
  :root {
    --bg:#f4f6f9; --card:#fff; --border:#e3e7ec; --text:#24292f; --muted:#656d76;
    --accent:#2c5f8a; --accent-dark:#1f3d5c; --band:#f7f9fb; --hover:#eef4fb;
    --ok:#e8f5e9; --ok-line:#2e7d32; --late:#ffc7ce; --late-text:#9c0006;
    --exp:#f1f3f5; --exp-text:#6b7280;
  }
  * { box-sizing:border-box; }
  body { margin:0; font-family:"Segoe UI",-apple-system,BlinkMacSystemFont,sans-serif;
         background:var(--bg); color:var(--text); font-size:13px; }
  header { background:var(--accent-dark); color:#39ff14; padding:14px 24px; }
  header h1 { margin:0; font-size:18px; font-weight:600; }
  header .meta { font-size:12px; opacity:.85; margin-top:4px; color:#dbe6f0; }
  main { padding:14px 24px 40px; }
  .controls { display:flex; flex-wrap:wrap; gap:12px; align-items:center; margin-bottom:12px; }
  .seg { display:inline-flex; border:1px solid var(--border); border-radius:6px;
         overflow:hidden; background:#fff; }
  .seg button { border:0; background:#fff; color:var(--accent); padding:6px 14px;
                cursor:pointer; font-size:13px; border-right:1px solid var(--border); }
  .seg button:last-child { border-right:0; }
  .seg button.active { background:var(--accent); color:#fff; }
  select, input[type=text] { border:1px solid var(--border); border-radius:6px;
                             padding:6px 9px; font-size:13px; background:#fff;
                             color:var(--text); }
  input#search { min-width:230px; }
  .grow { flex:1; }
  button.export { background:var(--accent); color:#fff; border:0; border-radius:6px;
                  padding:6px 14px; cursor:pointer; font-size:13px; }
  button.export:hover { background:var(--accent-dark); }
  button.link { background:none; border:0; color:var(--accent); cursor:pointer;
                font-size:13px; text-decoration:underline; padding:0; }
  .kpis { display:flex; flex-wrap:wrap; gap:10px; margin-bottom:12px; }
  .kpi { background:var(--card); border:1px solid var(--border); border-radius:8px;
         padding:8px 14px; min-width:112px; }
  .kpi .n { font-size:19px; font-weight:600; font-variant-numeric:tabular-nums; }
  .kpi .l { font-size:11px; color:var(--muted); text-transform:uppercase;
            letter-spacing:.4px; }
  .wrap { background:var(--card); border:1px solid var(--border); border-radius:8px;
          overflow:auto; display:block; max-width:100%; max-height:72vh; }
  table { border-collapse:collapse; width:auto; }
  th, td { padding:4px 9px; white-space:nowrap; text-align:left;
           border-bottom:1px solid var(--border); }
  thead th { position:sticky; top:0; background:#fff; z-index:3; font-weight:600;
             color:var(--muted); border-bottom:2px solid var(--border); font-size:11.5px;
             text-transform:uppercase; letter-spacing:.3px; }
  thead th.sortable { cursor:pointer; }
  thead th.sortable:hover { color:var(--accent); }
  tbody tr:nth-child(even) td { background:var(--band); }
  tbody tr:hover td { background:var(--hover); }
  td.num, th.num { text-align:right; font-variant-numeric:tabular-nums; }
  td.mid, th.mid { text-align:center; }
  td.bad { color:#c62828; }
  a { color:var(--accent); }
  #loads td.entry { max-width:44ch; overflow:hidden; text-overflow:ellipsis; }
  #loads td.mir { cursor:help; border-bottom:1px dotted var(--muted); }
  #loads th.hint { cursor:help; }
  /* MMSEA_Report */
  #mmsea td.entry { max-width:46ch; overflow:hidden; text-overflow:ellipsis; }
  #mmsea td.na { color:#b9bfc7; cursor:help; }
  #mmsea th.na { color:#b9bfc7; cursor:help; }
  #mmsea td.sts { cursor:help; }
  /* MSPi Report */
  #msr td.entry { max-width:60ch; overflow:hidden; text-overflow:ellipsis; }
  #msr td.zero { color:#b6bec8; cursor:help; }
  #msr td.noclient { color:#b6bec8; font-style:italic; cursor:help; }
  /* MSPi filter bars (each MSPi tab carries its own - the toolbar selects at
     the top of the page are MMSEA-scoped) */
  .subbar { display:flex; flex-wrap:wrap; gap:10px; align-items:flex-end;
            background:var(--card); border:1px solid var(--border);
            border-radius:8px; padding:9px 12px; margin-bottom:12px; }
  .subbar .field { display:flex; flex-direction:column; gap:3px; }
  .subbar label { font-size:10.5px; color:var(--muted); text-transform:uppercase;
                  letter-spacing:.4px; font-weight:700; }
  .subbar .hint { font-size:12px; color:var(--muted); align-self:center; }
  .subbar button { background:#fff; color:var(--accent); border:1px solid var(--accent);
                   border-radius:6px; padding:6px 12px; cursor:pointer; font-size:13px; }
  .subbar button:hover { background:var(--hover); }
  .pager { display:flex; align-items:center; gap:12px; margin-top:10px;
           justify-content:flex-end; font-size:12px; color:var(--muted); }
  .pager button { background:#fff; color:var(--accent); border:1px solid var(--accent);
                  border-radius:6px; padding:4px 10px; cursor:pointer; font-size:12px; }
  /* Monthly Summary - a KPI row of stat tiles, not a chart.  All tiles wear the
     same ink: they are four different measures, not four series, so colouring
     them apart would be decorative. */
  .sum-card { background:var(--card); border:1px solid var(--border);
              border-radius:10px; padding:20px 24px 22px; max-width:940px; }
  .sum-h { font-size:23px; font-weight:600; color:var(--accent-dark); margin:0; }
  .sum-sub { font-size:12.5px; color:var(--muted); margin:3px 0 0; }
  .sum-group { margin-top:20px; }
  .sum-group > h3 { margin:0 0 9px; font-size:11px; font-weight:700;
                    color:var(--muted); text-transform:uppercase;
                    letter-spacing:.6px; display:flex; align-items:baseline; gap:8px; }
  .sum-group > h3 span.tot { font-size:12px; color:var(--accent);
                             text-transform:none; letter-spacing:0; font-weight:600; }
  .tiles { display:flex; flex-wrap:wrap; gap:12px; }
  .tile { border:1px solid var(--border); border-radius:9px; padding:12px 16px 13px;
          min-width:168px; background:var(--band); }
  /* proportional figures: tabular-nums makes a standalone display number look
     loose (every digit as wide as a 0) - keep tabular for table columns only */
  .tile .v { font-size:31px; font-weight:600; line-height:1.06;
             color:var(--accent-dark); font-variant-numeric:proportional-nums; }
  .tile .l { font-size:12.5px; color:var(--text); margin-top:5px; }
  .tile .d { font-size:11.5px; color:var(--muted); margin-top:4px; }
  .sum-prose { margin-top:22px; padding-top:16px; border-top:1px solid var(--border); }
  .sum-prose h4 { margin:0 0 8px; font-size:13.5px; font-weight:600; }
  .sum-prose ul { margin:0; padding-left:20px; line-height:1.85; font-size:13.5px; }
  .sum-prose .none { color:var(--muted); font-style:italic; }
  .sum-copy { margin-top:12px; }
  /* MSPi Monthly matrix */
  .matrix-wrap { overflow:auto; max-height:72vh; border:1px solid var(--border);
                 border-radius:8px; background:var(--card); }
  table.matrix { border-collapse:separate; border-spacing:0; width:auto; }
  table.matrix th, table.matrix td { border-bottom:1px solid #eef1f5; white-space:nowrap; }
  table.matrix thead th { position:sticky; top:0; z-index:3;
    background:linear-gradient(#ffffff,#f2f6fb); color:var(--accent-dark);
    padding:9px 10px; font-size:11px; text-align:center; font-weight:700;
    text-transform:uppercase; letter-spacing:.6px;
    border-bottom:2px solid var(--accent); }
  table.matrix thead th.rowhdr { z-index:4; text-align:left; }
  table.matrix th.rowhdr, table.matrix td.rowhdr { padding:6px 14px; font-size:12.5px;
    position:sticky; overflow:hidden; text-overflow:ellipsis; }
  table.matrix td.rowhdr { background:#fff; color:var(--text); z-index:2; }
  table.matrix th.c-client, table.matrix td.c-client { left:0;
    width:190px; min-width:190px; max-width:190px; border-right:1px solid #eef1f5; }
  table.matrix th.c-contract, table.matrix td.c-contract { left:190px;
    width:120px; min-width:120px; max-width:120px;
    font-variant-numeric:tabular-nums; color:var(--muted);
    border-right:2px solid var(--border); }
  table.matrix td.day { text-align:center; padding:5px 6px; min-width:62px; }
  table.matrix td.day.hit { cursor:pointer; }
  .mk { display:inline-flex; align-items:center; justify-content:center;
        min-width:22px; height:22px; border-radius:6px; padding:0 8px;
        background:#e3f4ea; color:#1c7a44; font-weight:700; font-size:12px;
        box-shadow:inset 0 0 0 1px rgba(28,122,68,.28);
        transition:transform .08s ease; font-variant-numeric:tabular-nums; }
  table.matrix td.day.hit:hover .mk { background:#cdecd8; transform:scale(1.12); }
  tr.client-row td { background:#f5f9fd; border-bottom:1px solid #dce7f3; }
  tr.client-row td.rowhdr { cursor:pointer; background:#eef4fb; }
  tr.client-row:hover td { background:#e6effb; }
  tr.client-row:hover td.rowhdr { background:#dfeafb; }
  tr.client-row td.c-client { font-weight:700; color:var(--accent-dark);
                              box-shadow:inset 3px 0 0 var(--accent); }
  tr.client-row .mk { background:var(--accent); color:#fff; box-shadow:none; }
  tr.client-row td.day.hit:hover .mk { background:var(--accent-dark); }
  .tri { display:inline-block; width:12px; margin-right:5px; color:var(--accent);
         font-size:10px; }
  tr.contract-row:hover td { background:#f4f8fd; }
  tr.contract-row:hover td.rowhdr { background:#eef4fb; }
  /* NB: c-contract is position:sticky, itself the containing block for this
     absolute marker - do NOT add position:relative (it cancels the sticky
     frozen column and throws the alignment out). */
  tr.contract-row td.c-contract { padding-left:24px; }
  tr.contract-row td.c-contract::before { content:""; position:absolute; left:13px;
    top:50%; width:5px; height:5px; margin-top:-3px; border-radius:50%;
    background:#c2ccd8; }
  #tooltip { position:fixed; z-index:9999; pointer-events:none; background:#1f2a37;
    color:#fff; border-radius:6px; padding:8px 10px; font-size:12px; line-height:1.45;
    box-shadow:0 4px 16px rgba(0,0,0,.3); max-width:460px; }
  #tooltip .tt-file + .tt-file { margin-top:6px; border-top:1px solid #3a4757;
                                 padding-top:6px; }
  #tooltip .tt-name { font-family:Consolas,"Courier New",monospace; word-break:break-all; }
  #tooltip .tt-label { color:#9fb4cc; }
  /* toolbar tool links */
  .tools { display:flex; gap:14px; align-items:center; margin:0 0 12px;
           font-size:12px; color:var(--muted); }
  .tools a { color:var(--accent); text-decoration:none; font-weight:600; }
  .tools a:hover { text-decoration:underline; }
  .tools .cp { background:none; border:0; cursor:pointer; color:var(--muted);
               font-size:12px; padding:0 0 0 3px; }
  .tools .cp:hover { color:var(--accent); }
  .tools .sep { color:var(--border); }
  /* calendar */
  #cal td.mo { text-align:center; min-width:52px; padding:3px 6px;
               font-variant-numeric:tabular-nums; }
  #cal td.mo.has { background:var(--ok) !important; color:var(--ok-line);
                   font-weight:700; }
  #cal td.mo.trk { color:var(--muted); font-weight:700; }
  #cal td.mo.late { background:var(--late) !important; color:var(--late-text);
                    font-weight:600; }
  #cal td.mo.exp { color:var(--exp-text); }
  #cal td.cli { position:sticky; left:0; background:#fff; z-index:2; font-weight:600; }
  #cal tbody tr:nth-child(even) td.cli { background:var(--band); }
  #cal tbody tr:hover td.cli { background:var(--hover); }
  .pill { display:inline-block; padding:1px 7px; border-radius:9px; font-size:11px;
          border:1px solid var(--border); background:#fff; }
  .pill.closed { background:#e8f5e9; border-color:#c8e6c9; color:#1b5e20; }
  .pill.open { background:#fff4e5; border-color:#ffe0b2; color:#8a5300; }
  .pill.none { background:#fdecea; border-color:#f8c9c4; color:#8a1c13; }
  tr.det td { background:#fbfcfd !important; padding:0; }
  tr.det table { margin:6px 0 10px 28px; width:auto; }
  tr.det th, tr.det td { border-bottom:1px solid #eef1f4; font-size:12px;
                         padding:2px 10px; }
  tr.det thead th { position:static; text-transform:none; letter-spacing:0; }
  tr.det td.def { white-space:normal; max-width:52ch; color:var(--muted);
                  line-height:1.4; }
  tr.load td.exp { cursor:pointer; color:var(--accent); user-select:none; width:18px; }
  .legend { display:flex; flex-wrap:wrap; gap:16px; align-items:center;
            margin:10px 0 0; font-size:12px; color:var(--muted); }
  .sw { display:inline-block; width:13px; height:13px; border-radius:3px;
        border:1px solid var(--border); vertical-align:-2px; margin-right:5px; }
  .cards { display:flex; flex-wrap:wrap; gap:16px; align-items:flex-start; }
  .cards .col { display:flex; flex-direction:column; gap:16px; align-items:flex-start; }
  #stkey td.def { white-space:normal; max-width:46ch; line-height:1.45;
                  color:var(--muted); }
  .card { background:var(--card); border:1px solid var(--border); border-radius:8px;
          overflow:hidden; }
  .card h2 { margin:0; padding:9px 14px; font-size:13px; background:#f7f9fb;
             border-bottom:1px solid var(--border); }
  .card .body { padding:0; max-height:60vh; overflow:auto; }
  .card.wide { flex:1 1 560px; max-width:820px; }
  #clikey td:last-child { white-space:normal; line-height:1.5; color:var(--muted);
                          font-size:12px; }
  .card p.note { margin:0; padding:8px 14px; font-size:12px; color:var(--muted);
                 border-top:1px solid var(--border); }
  tr.dim td { color:var(--muted); }
  .empty { padding:26px; text-align:center; color:var(--muted); }
  [hidden] { display:none !important; }
__EXPORT_CSS__
</style>
</head>
<body>
<header>
  <h1>MMSEA_MSPi_CMSE &mdash; MMSEA Response Files &amp; MSPI File Loading</h1>
  <div class="meta" id="meta"></div>
</header>
<main>
  <div class="controls">
    <div class="seg" id="tabs">
      <button data-tab="loads" class="active">Loads</button>
      <button data-tab="cal">Calendar</button>
      <button data-tab="mmsea">MMSEA_Report</button>
      <button data-tab="msmo">MSPi Monthly</button>
      <button data-tab="msr">MSPi Report</button>
      <button data-tab="sum">Monthly Summary</button>
      <button data-tab="ref">File Types &amp; Keys</button>
    </div>
    <select id="year" class="cal-only"></select>
    <select id="client"></select>
    <select id="source"></select>
    <select id="ticket" class="loads-only">
      <option value="">All loads</option>
      <option value="y">Has ADO ticket</option>
      <option value="n">No ADO ticket</option>
    </select>
    <input id="search" type="text" placeholder="Find client, file, PCN, ticket&hellip;">
    <button class="link" id="clear">Reset</button>
    <span class="grow"></span>
    __EXPORT_UI__
  </div>

  <div class="tools" id="tools"></div>
  <div class="kpis" id="kpis"></div>

  <section id="view-cal" hidden>
    <div class="wrap"><table id="cal"><thead id="cal-head"></thead><tbody id="cal-body"></tbody></table></div>
    <div class="legend">
      <span><span class="sw" style="background:var(--ok)"></span><b>X</b> = loaded to CMSE</span>
      <span><b style="color:var(--muted)">X</b> = tracker shows loaded, no CMSE load</span>
      <span><span class="sw" style="background:var(--exp)"></span>E = expected</span>
      <span><span class="sw" style="background:var(--late)"></span>Not loaded &ndash; late / outreach
        <span style="opacity:.75">(hover for the tracker date)</span></span>
      <span id="cal-note"></span>
    </div>
  </section>

  <section id="view-loads">
    <div class="wrap"><table id="loads"><thead id="loads-head"></thead><tbody id="loads-body"></tbody></table></div>
    <div class="legend"><span>Click a row to expand its ImportStaging breakdown
      (EntryName / DNDispositionCode / StagingStatus / records).</span>
      <button class="link" id="expand-all">Expand all</button></div>
  </section>

  <section id="view-mmsea" hidden>
    <div class="wrap"><table id="mmsea"><thead id="mmsea-head"></thead><tbody id="mmsea-body"></tbody></table></div>
    <div class="legend">
      <span>Layout of <b>MMSEA_Report_20250605.xlsx</b> &mdash; one row per load
        per StagingStatus.</span>
      <span id="mmsea-note"></span>
    </div>
  </section>

  <section id="view-msmo" hidden>
    <div class="subbar">
      <div class="field"><label for="mo-year">Year</label><select id="mo-year"></select></div>
      <div class="field"><label for="mo-client">Client</label>
        <select id="mo-client"><option value="">All clients</option></select></div>
      <div class="field"><label for="mo-contract">Contract</label>
        <select id="mo-contract"><option value="">All contracts</option></select></div>
      <button id="mo-expand">Expand all</button>
      <button id="mo-collapse">Collapse all</button>
      <span class="hint">Cell = files loaded that month &middot; click a client to
        expand its contracts &middot; hover a <b>count</b> for date extracted,
        file name &amp; record count.</span>
    </div>
    <div class="matrix-wrap">
      <table class="matrix" id="matrix"><thead id="matrix-head"></thead>
        <tbody id="matrix-body"></tbody></table>
    </div>
  </section>

  <section id="view-msr" hidden>
    <div class="subbar">
      <div class="field"><label for="msr-client">Client</label>
        <select id="msr-client"><option value="">All clients</option></select></div>
      <div class="field"><label for="msr-type">File Type</label>
        <select id="msr-type"><option value="">All types</option>
          <option value="DET">DET</option><option value="PRM">PRM</option></select></div>
      <div class="field"><label for="msr-contract">Contract</label>
        <select id="msr-contract"><option value="">All contracts</option></select></div>
      <div class="field"><label for="msr-file">File Name</label>
        <input type="text" id="msr-file" placeholder="contains&hellip;"></div>
      <div class="field"><label for="msr-from">Created From</label>
        <input type="date" id="msr-from"></div>
      <div class="field"><label for="msr-to">Created To</label>
        <input type="date" id="msr-to"></div>
      <button id="msr-reset">Reset</button>
    </div>
    <div class="wrap"><table id="msr"><thead id="msr-head"></thead>
      <tbody id="msr-body"></tbody></table></div>
    <div class="pager">
      <span id="msr-pageinfo"></span>
      <button id="msr-prev">&laquo; Prev</button>
      <button id="msr-next">Next &raquo;</button>
    </div>
    <div class="legend">
      <span>Layout of <b>MSPI Report_20250605.xlsx</b> (an export of
        <code>MSP.Report.MSPIMonthlyReport</code>), computed live from
        <code>etl.Tape</code> + <code>history.*</code> so it stays current.</span>
      <span id="msr-note"></span>
    </div>
  </section>

  <section id="view-sum" hidden>
    <div class="subbar">
      <div class="field"><label for="sum-month">Month</label>
        <select id="sum-month"></select></div>
      <button id="sum-prev-m">&laquo; Previous</button>
      <button id="sum-next-m">Next &raquo;</button>
      <span class="hint">Files received and loaded in the selected month.</span>
    </div>
    <div class="sum-card" id="sum-card"></div>
  </section>

  <section id="view-ref" hidden>
    <div class="cards">
      <div class="col">
        <div class="card"><h2>File Type &mdash; layout</h2><div class="body">
          <table id="specs"></table></div>
          <p class="note">From the MMSEA&nbsp;-&nbsp;2026 tab of ClientTracker.2026.xlsx.</p></div>
        <div class="card"><h2>StagingStatus</h2><div class="body">
          <table id="stkey"></table></div>
          <p class="note">Record counts across every load in window.</p></div>
      </div>
      <div class="col">
        <div class="card"><h2>SourceId key</h2><div class="body">
          <table id="srckey"></table></div>
          <p class="note">Bold rows are the SourceIds this dashboard reports on.</p></div>
      </div>
      <div class="col">
        <div class="card wide"><h2>Client key</h2><div class="body">
          <table id="clikey"></table></div>
          <p class="note">ClientId / ClientName from cmse_new..Client, joined to
            ClientCodeLookup for the SMART client codes.</p></div>
      </div>
    </div>
  </section>
</main>
<div id="tooltip" style="display:none"></div>

<script type="application/json" id="data">__DATA_JSON__</script>
<script>__EXPORT_JS__</script>
<script>
(function () {
  const D = JSON.parse(document.getElementById('data').textContent);
  const MN = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
  const SRC = {}; D.sources.forEach(s => SRC[s.id] = s.name);
  const $ = id => document.getElementById(id);
  const esc = s => (s == null ? '' : String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;')
                                              .replace(/>/g,'&gt;').replace(/"/g,'&quot;'));
  const nf = n => (n == null ? '' : Number(n).toLocaleString('en-US'));

  const latestYear = D.years[D.years.length - 1];
  let S = { tab:'loads', year:latestYear, client:'', source:'', ticket:'', q:'',
            sortK:'sl', sortD:-1, msortK:'', msortD:1, open:new Set() };

  // ---- tool links ---------------------------------------------------------
  // A UNC path becomes file://host/share/... and every segment must be
  // percent-encoded - unencoded spaces are why "File Transformer" never opened.
  const fileUrl = p => 'file:' + p.replace(/\\/g, '/').split('/')
                                  .map(encodeURIComponent).join('/');
  $('tools').innerHTML = '<span>Tools:</span>' + D.tools.map(([label, path, proto], i) => {
    // a registered URL protocol launches the app; file:// is the fallback
    const href = proto || fileUrl(path);
    const tip = esc(path) + (proto
      ? '\n\nOpens via the ' + proto.split(':')[0] + ': protocol. Nothing happens? '
        + 'Run ' + D.regFile + ' once (link at right), then restart the browser.'
      : '\n\nOpens in its own app. If the browser blocks it, use the copy button '
        + 'and paste the path into the Windows Run box (Win+R).');
    return (i ? '<span class="sep">|</span>' : '') +
      `<span><a href="${href}"${proto ? '' : ' target="_blank" rel="noopener"'} ` +
      `title="${tip}">${esc(label)}</a>` +
      `<button class="cp" data-path="${esc(path)}" title="Copy path">⧉</button></span>`;
  }).join('')
    + `<span class="sep">|</span><span style="opacity:.8">first time on this PC? run ` +
      `<a href="${fileUrl(D.regDir + '\\' + D.regFile)}">${esc(D.regFile)}</a> once ` +
      `&middot; or copy a path with ⧉ and use Win+R</span>`;
  $('tools').addEventListener('click', e => {
    const b = e.target.closest('button.cp'); if (!b) return;
    const p = b.dataset.path;
    if (navigator.clipboard && navigator.clipboard.writeText) {
      navigator.clipboard.writeText(p);
    } else {                                  // file:// pages often lack the async API
      const ta = document.createElement('textarea');
      ta.value = p; document.body.appendChild(ta); ta.select();
      try { document.execCommand('copy'); } catch (_) {}
      document.body.removeChild(ta);
    }
    const old = b.textContent; b.textContent = '✓';
    setTimeout(() => { b.textContent = old; }, 900);
  });

  // ---- filter widgets -----------------------------------------------------
  $('year').innerHTML = D.years.map(y => `<option${y===S.year?' selected':''}>${y}</option>`).join('');
  const clientNames = [...new Set(D.cal.map(r => r.client).concat(D.loads.map(l => l.client)))]
                        .filter(Boolean).sort((a,b) => a.localeCompare(b));
  $('client').innerHTML = '<option value="">All clients</option>' +
    clientNames.map(c => `<option>${esc(c)}</option>`).join('');
  $('source').innerHTML = '<option value="">All source types</option>' +
    D.scope.map(id => `<option value="${id}">${id} &middot; ${esc(SRC[id]||'')}</option>`).join('');

  $('meta').textContent = 'Generated ' + D.generated
    + ' \u00b7 MMSEA: TRGRepSQL3 / CMSE_New, SourceId ' + D.scope.join(', ') + ', '
    + D.loads.length.toLocaleString('en-US') + ' loads from ' + (D.years[0] || '')
    + ' forward \u00b7 MSPi: TRGINTP3 / MSP, '
    + (D.mspi || []).length.toLocaleString('en-US') + ' files';
  $('cal-note').textContent = D.trackerNote ? ('\u26a0 ' + D.trackerNote) : '';

  // ---- helpers ------------------------------------------------------------
  const hay = l => [l.client, l.file, l.entry, l.pcn, l.wi, l.wit, l.sl, SRC[l.src], l.ft]
                     .join(' ').toLowerCase();

  // the ticket dropdown is a Loads-tab control; MMSEA_Report passes useTicket=false
  function filteredLoads(useTicket) {
    const q = S.q.toLowerCase();
    const t = useTicket === false ? '' : S.ticket;
    return D.loads.filter(l =>
      (!S.client || l.client === S.client) &&
      (!S.source || String(l.src) === S.source) &&
      (t !== 'y' || l.wi) && (t !== 'n' || !l.wi) &&
      (!q || hay(l).includes(q)));
  }

  function filteredCal() {
    const q = S.q.toLowerCase();
    return D.cal.filter(r => {
      if (S.client && r.client !== S.client) return false;
      if (S.source && SRC[S.source] && D.srcType[S.source] && D.srcType[S.source][0] !== r.ft) return false;
      if (!q) return true;
      if ((r.client + ' ' + r.ft + ' ' + r.freq + ' ' + r.hand).toLowerCase().includes(q)) return true;
      return Object.keys(r.m).some(ym => ym.startsWith(S.year) &&
        r.m[ym].some(x => (x.file + ' ' + x.pcn + ' ' + (x.wi||'')).toLowerCase().includes(q)));
    });
  }

  // ---- calendar -----------------------------------------------------------
  function renderCal() {
    const rows = filteredCal();
    $('cal-head').innerHTML = '<tr><th>Client</th><th class="mid">Client Id</th>' +
      '<th>File Type</th><th>Frequency</th><th>Handling</th>' +
      MN.map(m => `<th class="mid">${m}</th>`).join('') + '</tr>';

    if (!rows.length) { $('cal-body').innerHTML =
      '<tr><td colspan="17" class="empty">No rows match.</td></tr>'; return; }

    const isTrackerYear = S.year === D.trackerYear;
    const html = rows.map(r => {
      const cells = MN.map((_, i) => {
        const ym = S.year + '-' + String(i+1).padStart(2,'0');
        const loads = r.m[ym] || [];
        if (loads.length) {
          const t = `${loads.length} file${loads.length>1?'s':''} loaded ${MN[i]} ${S.year}\n`
                  + loads.map(x => `\u00b7 ${x.file} (${MN[i]} ${x.d}, ${nf(x.rec)} records)`)
                         .join('\n');
          return `<td class="mo has" title="${esc(t)}">X</td>`;
        }
        const tr = isTrackerYear ? r.tr[String(i+1)] : null;
        if (tr) {
          // the tracker writes a bare ticket number to mean "loaded to CMSE"
          if (!tr.late && /^\d+$/.test(tr.v))
            return `<td class="mo trk" title="Tracker shows loaded (${esc(tr.v)}); no CMSE load this month">X</td>`;
          // not loaded: a red cell keeps its colour but shows no date and no "E"
          if (tr.late)
            return `<td class="mo late" title="Not loaded &mdash; tracker outreach/late: ${esc(tr.v)}"></td>`;
          return `<td class="mo exp">${esc(tr.v)}</td>`;
        }
        return '<td class="mo"></td>';
      }).join('');
      return `<tr><td class="cli">${esc(r.client)}</td><td class="mid">${r.cid||''}</td>` +
             `<td>${esc(r.ft)}</td><td>${esc(r.freq)}</td><td>${esc(r.hand)}</td>${cells}</tr>`;
    }).join('');
    $('cal-body').innerHTML = html;
  }

  // ---- loads --------------------------------------------------------------
  const LCOLS = [
    ['', '', null],
    ['sl', 'SourceLogId', 'num'],
    ['src', 'SourceId', 'mid hint'],
    ['ft', 'File Type', null],
    ['client', 'Client Name', null],
    ['cid', 'Client Id', 'mid'],
    ['file', 'Entry Name', 'entry'],
    ['start', 'Import Start', null],
    ['done', 'Import Complete', null],
    ['rec', 'Records', 'num'],
    ['ok', 'Success', 'num'],
    ['bad', 'Failed', 'num'],
    ['pcn', 'PCN', 'mid'],
    ['wi', 'ADO Ticket', 'mid'],
  ];

  function sortLoads(rows) {
    const k = S.sortK, d = S.sortD;
    const val = l => k === 'srcname' ? (SRC[l.src] || '') : (l[k] == null ? '' : l[k]);
    return rows.slice().sort((a, b) => {
      const x = val(a), y = val(b);
      if (typeof x === 'number' && typeof y === 'number') return (x - y) * d;
      return String(x).localeCompare(String(y), undefined, {numeric:true}) * d;
    });
  }

  function renderLoads() {
    const rows = sortLoads(filteredLoads());
    $('loads-head').innerHTML = '<tr>' + LCOLS.map(([k, lbl, cls]) =>
      k ? `<th class="${cls||''} sortable" data-k="${k}"` +
          (k === 'src' ? ' title="Hover a SourceId cell for its SourceName"' : '') +
          `>${lbl}${S.sortK===k?(S.sortD>0?' \u25b2':' \u25bc'):''}</th>`
        : '<th></th>').join('') + '</tr>';

    if (!rows.length) { $('loads-body').innerHTML =
      `<tr><td colspan="${LCOLS.length}" class="empty">No loads match.</td></tr>`; return; }

    const out = [];
    for (const l of rows) {
      const open = S.open.has(l.sl);
      const wiCell = l.wi
        ? `<a href="${D.adoBase}${l.wi}" target="_blank" title="${esc(l.wit)}">${l.wi}</a>`
        : '<span class="pill none">none</span>';
      // Aetna (ClientId 5) carries the MIR flags; only that client gets the tooltip
      const isAetna = l.client === 'Aetna' && l.cid === 5;
      const mirTip = isAetna
        ? `MIRProcessed: ${l.mir === '1' ? 'Yes' : l.mir === '0' ? 'No' : '\u2014'}`
          + `\nMIRProcessedDate: ${l.mird || '\u2014'}`
        : '';
      const cliCell = isAetna
        ? `<td class="mir" title="${esc(mirTip)}">${esc(l.client)}</td>`
        : `<td>${esc(l.client)}</td>`;
      out.push(`<tr class="load" data-sl="${l.sl}">` +
        `<td class="exp">${open?'\u2212':'+'}</td>` +
        `<td class="num">${l.sl}</td>` +
        `<td class="mid" title="${esc(SRC[l.src]||'')}">${l.src}</td>` +
        `<td>${esc(l.ft)}</td>` + cliCell +
        `<td class="mid">${l.cid||''}</td>` +
        `<td class="entry" title="${esc(l.entry)}">${esc(l.file)}</td>` +
        `<td>${esc(l.start)}</td><td>${esc(l.done)}</td>` +
        `<td class="num">${nf(l.rec)}</td><td class="num">${nf(l.ok)}</td>` +
        `<td class="num${l.bad ? ' bad' : ''}">${nf(l.bad)}</td>` +
        `<td class="mid">${esc(l.pcn)}</td>` +
        `<td class="mid">${wiCell}</td></tr>`);
      if (open) {
        const tot = l.stg.reduce((s, r) => s + r[2], 0);
        const body = l.stg.length
          ? l.stg.map(([dn, sts, n]) =>
              `<tr><td>${esc(l.file)}</td><td class="mid">${esc(dn)||'\u2014'}</td>` +
              `<td class="mid">${esc(sts)||'\u2014'}</td>` +
              `<td class="def">${esc(D.stDef[sts]||'')}</td>` +
              `<td class="num">${nf(n)}</td></tr>`).join('')
            + `<tr class="dim"><td colspan="4"><b>Total</b></td><td class="num"><b>${nf(tot)}</b></td></tr>`
          : '<tr class="dim"><td colspan="5">No ImportStaging rows for this SourceLogId.</td></tr>';
        out.push(`<tr class="det"><td colspan="${LCOLS.length}"><table><thead><tr>` +
          '<th>EntryName</th><th class="mid">DNDispositionCode</th>' +
          '<th class="mid">StagingStatus</th><th>StagingStatus definition</th>' +
          '<th class="num">Records</th>' +
          `</tr></thead><tbody>${body}</tbody></table></td></tr>`);
      }
    }
    $('loads-body').innerHTML = out.join('');
  }

  // ---- MMSEA_Report -------------------------------------------------------
  // Column order is the MMSEA_Report_20250605.xlsx column order.  E %, MC % and
  // # of Invs have no counterpart anywhere in cmse_new (see UNSOURCED_COLUMNS
  // in the generator) so they render as an em-dash rather than a guess.
  const MCOLS = [
    ['client',  'Client',               null],
    ['ftname',  'File Type',            null],
    ['rt',      'R or T',               'mid'],
    ['idate',   'Import Date',          null],
    ['epct',    'E %',                  'num na'],
    ['mcpct',   'MC %',                 'num na'],
    ['rec',     'Import Count',         'num'],
    ['ok',      'Success',              'num'],
    ['bad',     'Failed',               'num'],
    ['age',     'Age',                  'num'],
    ['dis',     'Dis',                  'num'],
    ['esrd',    'ESRD',                 'num'],
    ['um',      'U M Count',            'num'],
    ['invs',    '# of Invs',            'num na'],
    ['st',      'Staging Status',       'mid sts'],
    ['stn',     'Staging Status Count', 'num'],
    ['sl',      'SourceLogID',          'num'],
    ['file',    'File Name',            'entry'],
  ];
  const MNA = new Set(D.unsourced || []);
  const NA_TIP = 'Not available from cmse_new — this column comes from '
               + 'outside the CMSE database and is left blank.';

  function mmseaRows() {
    const out = [];
    for (const l of filteredLoads(false)) {
      const agg = {};
      (l.stg || []).forEach(([dn, st, n]) => {
        const k = (st || '').trim();
        agg[k] = (agg[k] || 0) + n;
      });
      const base = { client:l.client, ftname:SRC[l.src] || '', rt:l.rt || '',
                     idate:(l.start || '').slice(0, 10), rec:l.rec, ok:l.ok, bad:l.bad,
                     age:l.age, dis:l.dis, esrd:l.esrd, um:l.um, sl:l.sl,
                     file:l.file, entry:l.entry, src:l.src, ft:l.ft };
      const stats = Object.keys(agg).sort((a, b) => Number(a) - Number(b));
      if (!stats.length) out.push(Object.assign({ st:'', stn:null }, base));
      else stats.forEach(s => out.push(Object.assign({ st:s, stn:agg[s] }, base)));
    }
    // spreadsheet order: client, file type, newest import first
    const k = S.msortK, d = S.msortD;
    if (k) {
      out.sort((a, b) => {
        const x = a[k] == null ? '' : a[k], y = b[k] == null ? '' : b[k];
        if (typeof x === 'number' && typeof y === 'number') return (x - y) * d;
        return String(x).localeCompare(String(y), undefined, {numeric:true}) * d;
      });
    } else {
      out.sort((a, b) =>
        a.client.localeCompare(b.client) || a.ftname.localeCompare(b.ftname) ||
        b.idate.localeCompare(a.idate) || b.sl - a.sl ||
        String(a.st).localeCompare(String(b.st), undefined, {numeric:true}));
    }
    return out;
  }

  function renderMmsea() {
    const rows = mmseaRows();
    $('mmsea-head').innerHTML = '<tr>' + MCOLS.map(([k, lbl, cls]) =>
      `<th class="${cls || ''} sortable" data-k="${k}"` +
      (MNA.has(lbl) ? ` title="${esc(NA_TIP)}"` : '') +
      `>${esc(lbl)}${S.msortK === k ? (S.msortD > 0 ? ' ▲' : ' ▼') : ''}</th>`)
      .join('') + '</tr>';

    if (!rows.length) { $('mmsea-body').innerHTML =
      `<tr><td colspan="${MCOLS.length}" class="empty">No rows match.</td></tr>`; return; }

    $('mmsea-body').innerHTML = rows.map(r => {
      const stTip = r.st ? (D.stDef[r.st] || 'undocumented StagingStatus') : '';
      return '<tr>' +
        `<td>${esc(r.client)}</td>` +
        `<td>${esc(r.ftname)}</td>` +
        `<td class="mid">${esc(r.rt)}</td>` +
        `<td>${esc(r.idate)}</td>` +
        `<td class="num na" title="${esc(NA_TIP)}">—</td>` +
        `<td class="num na" title="${esc(NA_TIP)}">—</td>` +
        `<td class="num">${nf(r.rec)}</td>` +
        `<td class="num">${nf(r.ok)}</td>` +
        `<td class="num${r.bad ? ' bad' : ''}">${nf(r.bad)}</td>` +
        `<td class="num">${nf(r.age)}</td>` +
        `<td class="num">${nf(r.dis)}</td>` +
        `<td class="num">${nf(r.esrd)}</td>` +
        `<td class="num">${nf(r.um)}</td>` +
        `<td class="num na" title="${esc(NA_TIP)}">—</td>` +
        `<td class="mid sts" title="${esc(stTip)}">${esc(r.st) || '—'}</td>` +
        `<td class="num">${r.stn == null ? '—' : nf(r.stn)}</td>` +
        `<td class="num">${r.sl}</td>` +
        `<td class="entry" title="${esc(r.entry)}">${esc(r.file)}</td></tr>`;
    }).join('');

    $('mmsea-note').textContent = MNA.size
      ? '⚠ ' + [...MNA].join(', ') + ' are not held in cmse_new and stay blank.'
      : '';
  }

  // ---- MSPi shared --------------------------------------------------------
  const MS = D.mspi || [];
  const MSPAGE = 100;
  // A file's own record count: DET files load history.DET, PRM files load
  // history.PRM.  (The MSPi Report tab's "Load count" column is deliberately
  // DET-only - see renderMsr.)
  const msRecords = r => (r.type === 'DET' ? r.det : r.prm) || 0;

  const fillOpts = (sel, values, keep) => {
    const first = sel.options[0] ? sel.options[0].outerHTML : '';
    sel.innerHTML = first + values.map(v => `<option>${esc(v)}</option>`).join('');
    if (keep && values.indexOf(keep) !== -1) sel.value = keep;
  };

  // ---- MSPi Monthly -------------------------------------------------------
  const MOS = { year:'', client:'', contract:'', expanded:new Set() };
  const msYears = [...new Set(MS.map(r => r.loaded).filter(Boolean)
                                .map(s => s.slice(0, 4)))].sort().reverse();
  $('mo-year').innerHTML = msYears.map(y => `<option>${y}</option>`).join('');
  MOS.year = msYears[0] || '';
  if (MOS.year) $('mo-year').value = MOS.year;
  fillOpts($('mo-client'), [...new Set(MS.map(r => r.client).filter(Boolean))].sort());

  function moContracts() {          // contract list cascades off the client
    const src = MOS.client ? MS.filter(r => r.client === MOS.client) : MS;
    fillOpts($('mo-contract'),
             [...new Set(src.map(r => r.contract).filter(Boolean))].sort(),
             MOS.contract);
    if ($('mo-contract').value !== MOS.contract) MOS.contract = $('mo-contract').value;
  }
  moContracts();

  // All-time universe of client -> [contracts] so every pair is listed each
  // year whether or not it loaded (a blank cell = nothing loaded that month).
  const MO_PAIRS = {};
  MS.forEach(r => { (MO_PAIRS[r.client] = MO_PAIRS[r.client] || new Set()).add(r.contract); });
  Object.keys(MO_PAIRS).forEach(k => {
    MO_PAIRS[k] = [...MO_PAIRS[k]].sort((a, b) => a.localeCompare(b));
  });

  const tip = $('tooltip');
  let tipMap = {}, tipId = 0;
  function showTip(evt, id) {
    const files = tipMap[id];
    if (!files) return;
    tip.innerHTML = files.map(f =>
      '<div class="tt-file"><div class="tt-name">' + esc(f.type) + ' · ' +
      esc(f.file) + '</div>' +
      '<div><span class="tt-label">Date Extracted:</span> ' + (esc(f.extracted) || '—') + '</div>' +
      '<div><span class="tt-label">Date Loaded:</span> ' + esc((f.loaded || '').slice(0, 10)) + '</div>' +
      '<div><span class="tt-label">File Count:</span> ' + nf(msRecords(f)) + ' records</div></div>'
    ).join('');
    tip.style.display = 'block';
    moveTip(evt);
  }
  function moveTip(evt) {
    const pad = 14, w = tip.offsetWidth, h = tip.offsetHeight;
    let x = evt.clientX + pad, y = evt.clientY + pad;
    if (x + w > window.innerWidth) x = evt.clientX - w - pad;
    if (y + h > window.innerHeight) y = evt.clientY - h - pad;
    tip.style.left = Math.max(4, x) + 'px';
    tip.style.top = Math.max(4, y) + 'px';
  }
  const hideTip = () => { tip.style.display = 'none'; };

  function monthCells(map) {
    let out = '';
    for (let mo = 1; mo <= 12; mo++) {
      const files = map[mo];
      if (files && files.length) {
        const id = 'x' + (tipId++);
        tipMap[id] = files;
        out += `<td class="day hit" data-tip="${id}"><span class="mk">${nf(files.length)}</span></td>`;
      } else out += '<td class="day"></td>';
    }
    return out;
  }

  function renderMonthly() {
    const yr = MOS.year, head = $('matrix-head'), body = $('matrix-body');
    tipMap = {}; tipId = 0;
    head.innerHTML = '<tr><th class="rowhdr c-client">Client</th>' +
      '<th class="rowhdr c-contract">Contract</th>' +
      MN.map(m => `<th>${m}</th>`).join('') + '</tr>';
    if (!yr) { body.innerHTML = '<tr><td class="empty" colspan="14">No data.</td></tr>'; return; }

    const byClient = {};
    for (const r of MS) {
      if (!r.loaded || r.loaded.slice(0, 4) !== yr) continue;
      const c = byClient[r.client] || (byClient[r.client] = { contracts:{}, agg:{} });
      const ct = c.contracts[r.contract] || (c.contracts[r.contract] = {});
      const mo = Number(r.loaded.slice(5, 7));
      (ct[mo] = ct[mo] || []).push(r);
      (c.agg[mo] = c.agg[mo] || []).push(r);
    }

    let names = Object.keys(MO_PAIRS);
    if (MOS.client) names = names.filter(n => n === MOS.client);
    if (MOS.contract) names = names.filter(n => MO_PAIRS[n].indexOf(MOS.contract) !== -1);
    names.sort((a, b) => a.localeCompare(b));
    if (!names.length) {
      body.innerHTML = '<tr><td class="empty" colspan="14">No clients match.</td></tr>';
      return;
    }

    const forceOpen = !!MOS.contract;   // a contract filter opens its client
    const out = [];
    for (const name of names) {
      const c = byClient[name] || { contracts:{}, agg:{} };
      const open = forceOpen || MOS.expanded.has(name);
      out.push(`<tr class="client-row" data-client="${esc(name)}">` +
        `<td class="rowhdr c-client"><span class="tri">${open ? '▼' : '▶'}</span>${esc(name)}</td>` +
        '<td class="rowhdr c-contract"></td>' + monthCells(c.agg) + '</tr>');
      if (!open) continue;
      let cts = MO_PAIRS[name].slice();
      if (MOS.contract) cts = cts.filter(ct => ct === MOS.contract);
      for (const ct of cts) {
        out.push('<tr class="contract-row"><td class="rowhdr c-client"></td>' +
          `<td class="rowhdr c-contract">${esc(ct) || '<span style="color:var(--muted)">(none)</span>'}</td>` +
          monthCells(c.contracts[ct] || {}) + '</tr>');
      }
    }
    body.innerHTML = out.join('');
  }

  // ---- MSPi Report --------------------------------------------------------
  // Column order is the MSPI Report_20250605.xlsx column order.
  const MRCOLS = [
    ['rclient', 'Client',             null],
    ['type',    'FileType',           'mid'],
    ['file',    'FileName',           'entry'],
    ['prod',    'VolSerNum',          'num'],
    ['load',    'Load count',         'num'],
    ['hicn',    'Unique HICN Count',  'num'],
    ['rt',      'Raw Or Transform',   'mid'],
    ['created', 'Create Date',        null],
  ];
  const MR = { client:'', type:'', contract:'', file:'', from:'', to:'',
               sortK:'', sortD:1, page:0 };
  fillOpts($('msr-client'), [...new Set(MS.map(r => r.rclient).filter(Boolean))].sort());
  fillOpts($('msr-contract'), [...new Set(MS.map(r => r.contract).filter(Boolean))].sort());

  function msrFiltered() {
    const q = MR.file.trim().toLowerCase();
    return MS.filter(r =>
      (!MR.client || r.rclient === MR.client) &&
      (!MR.type || r.type === MR.type) &&
      (!MR.contract || r.contract === MR.contract) &&
      (!q || r.file.toLowerCase().includes(q)) &&
      (!MR.from || (r.created && r.created >= MR.from)) &&
      (!MR.to || (r.created && r.created <= MR.to)));
  }

  // "Load count" / "Unique HICN Count" are history.DET counts keyed on the
  // file's own TapeID, so a PRM file is always 0 — that is what the source
  // report does, and the real history.PRM count is in the cell's tooltip.
  const mrVal = (r, k) => k === 'load' ? r.det
                        : k === 'hicn' ? r.deth
                        : k === 'rt'   ? D.mspiRawOrTransform
                        : k === 'prod' ? Number(r.prod || 0)
                        : (r[k] == null ? '' : r[k]);

  function msrSorted() {
    const rows = msrFiltered();
    if (!MR.sortK) return rows.sort((a, b) => b.tape - a.tape);   // newest first
    const k = MR.sortK, d = MR.sortD;
    return rows.sort((a, b) => {
      const x = mrVal(a, k), y = mrVal(b, k);
      if (typeof x === 'number' && typeof y === 'number') return (x - y) * d || (b.tape - a.tape);
      return String(x).localeCompare(String(y), undefined, {numeric:true}) * d
             || (b.tape - a.tape);
    });
  }

  function renderMsr() {
    const rows = msrSorted();
    const pages = Math.max(1, Math.ceil(rows.length / MSPAGE));
    if (MR.page >= pages) MR.page = pages - 1;
    if (MR.page < 0) MR.page = 0;
    const start = MR.page * MSPAGE;
    const slice = rows.slice(start, start + MSPAGE);

    $('msr-head').innerHTML = '<tr>' + MRCOLS.map(([k, lbl, cls]) =>
      `<th class="${cls || ''} sortable" data-k="${k}">${esc(lbl)}` +
      `${MR.sortK === k ? (MR.sortD > 0 ? ' ▲' : ' ▼') : ''}</th>`).join('') + '</tr>';

    if (!slice.length) {
      $('msr-body').innerHTML =
        `<tr><td colspan="${MRCOLS.length}" class="empty">No files match.</td></tr>`;
      $('msr-pageinfo').textContent = '0 files';
      return;
    }

    $('msr-body').innerHTML = slice.map(r => {
      const fileTip = r.path + '\nTapeID ' + r.tape +
        '\nLoaded ' + (r.loaded || '—') +
        '\nDate extracted ' + (r.extracted || '—') +
        '\nFile size ' + nf(r.size) + ' bytes';
      // the zero a PRM row always shows: say what the file actually holds
      const zeroTip = 'The source report counts history.DET only, so a PRM file '
        + 'is always 0.\nThis file holds ' + nf(r.prm) + ' history.PRM records ('
        + nf(r.prmh) + ' unique HICNs) and ' + nf(r.msp) + ' history.MSP records.';
      const cell = (v, tipWhenZero) => v
        ? `<td class="num">${nf(v)}</td>`
        : `<td class="num zero" title="${esc(tipWhenZero)}">0</td>`;
      const noRecs = !r.det && !r.prm && !r.msp;
      const clientCell = r.rclient
        ? `<td>${esc(r.rclient)}</td>`
        : `<td class="noclient" title="${esc('This tape loaded no records, so the '
            + 'source report has no client for it. Path folder: ' + (r.client || '—'))}">`
          + `${esc(r.client) || '—'}</td>`;
      return '<tr>' + clientCell +
        `<td class="mid">${esc(r.type)}</td>` +
        `<td class="entry" title="${esc(fileTip)}">${esc(r.file)}</td>` +
        `<td class="num">${esc(r.prod)}</td>` +
        cell(r.det, r.type === 'PRM' ? zeroTip
                    : (noRecs ? 'This tape loaded no records.' : zeroTip)) +
        cell(r.deth, r.type === 'PRM' ? zeroTip
                     : (noRecs ? 'This tape loaded no records.' : zeroTip)) +
        `<td class="mid">${esc(D.mspiRawOrTransform)}</td>` +
        `<td>${esc(r.created)}</td></tr>`;
    }).join('');

    $('msr-pageinfo').textContent = nf(start + 1) + '–' + nf(start + slice.length)
      + ' of ' + nf(rows.length) + ' files';
    $('msr-note').textContent = '⚠ PRM rows show 0 — the source report counts '
      + 'history.DET only. Hover a 0 for the real PRM/MSP counts.';
  }

  // ---- Monthly Summary ----------------------------------------------------
  // A handful of headline numbers -> a KPI row of stat tiles, not a chart.
  const MONTH_FULL = ['January','February','March','April','May','June','July',
                      'August','September','October','November','December'];
  const monthLabel = ym => MONTH_FULL[Number(ym.slice(5, 7)) - 1] + ' ' + ym.slice(0, 4);

  // MMSEA file types, in the order the summary lists them.  D.srcType maps a
  // SourceId to [tracker label, spec label]; the tracker label is what l.ft holds.
  const SUM_FT = [['MSP', 'MSP'], ['NMSP', 'Non-MSP'], ['HEW', 'HEW']];

  const sumMonths = [...new Set(
    D.loads.map(l => (l.start || '').slice(0, 7))
      .concat(MS.map(r => (r.loaded || '').slice(0, 7)))
  )].filter(Boolean).sort().reverse();

  $('sum-month').innerHTML = sumMonths
    .map(m => `<option value="${m}">${esc(monthLabel(m))}</option>`).join('');
  const SUM = { month: sumMonths[0] || '' };
  if (SUM.month) $('sum-month').value = SUM.month;

  function sumFor(ym) {
    const loads = D.loads.filter(l => (l.start || '').slice(0, 7) === ym);
    const files = MS.filter(r => (r.loaded || '').slice(0, 7) === ym);
    const ft = {};
    SUM_FT.forEach(([k]) => {
      const rows = loads.filter(l => l.ft === k);
      ft[k] = { n: rows.length, clients: new Set(rows.map(l => l.client)).size };
    });
    return {
      ft,
      mmsea: SUM_FT.reduce((s, [k]) => s + ft[k].n, 0),
      // One "set" = the DET + PRM pair of a single delivery.  Pair on
      // VolSerNum + file name, NOT client + file name: 42 older tapes sit at
      // ...\MSP\DET\ and ...\MSP\PRM\ with no client folder, so their two halves
      // carry different clients and would each count as a set of their own.
      // prod|file agrees with max(DET, PRM) in all 113 months; client|file was
      // wrong in 15 of them.
      sets: new Set(files.map(r => r.prod + '|' + r.file)).size,
      det: files.filter(r => r.type === 'DET').length,
      prm: files.filter(r => r.type === 'PRM').length,
      mspiClients: new Set(files.map(r => r.client)).size,
    };
  }

  // "+3 vs April" / "no change vs April" - deliberately in muted ink with no
  // red/green: more files loaded is neither good nor bad, so a direction colour
  // would assert a judgement the data doesn't support.
  function deltaLine(cur, ym, pick) {
    const i = sumMonths.indexOf(ym);
    const prev = i >= 0 && i + 1 < sumMonths.length ? sumMonths[i + 1] : null;
    if (!prev) return '';
    const was = pick(sumFor(prev));
    const d = cur - was;
    const arrow = d > 0 ? '▲ +' : d < 0 ? '▼ −' : '';
    return (d === 0 ? 'no change' : arrow + nf(Math.abs(d)))
           + ' vs ' + MONTH_FULL[Number(prev.slice(5, 7)) - 1];
  }

  // MSPI reaches back to 2017 but the MMSEA half of this dashboard only loads
  // from WINDOW_START, so for earlier months the MMSEA counts would all be a
  // misleading zero - say "not covered" instead of reporting nothing happened.
  const MMSEA_FROM = (D.windowStart || '').slice(0, 7);
  const mmseaCovered = ym => !MMSEA_FROM || ym >= MMSEA_FROM;
  const partialMonth = ym => ym === D.generated.slice(0, 7);

  function sumText(ym, s) {
    const lines = [];
    if (mmseaCovered(ym)) {
      SUM_FT.forEach(([k, lbl]) => {
        lines.push(nf(s.ft[k].n) + ' ' + lbl + ' file' + (s.ft[k].n === 1 ? '' : 's')
                   + ' ' + (s.ft[k].n === 1 ? 'was' : 'were') + ' received and loaded');
      });
    } else {
      lines.push('MMSEA response files are not covered before '
                 + monthLabel(MMSEA_FROM + '-01'));
    }
    lines.push(nf(s.sets) + ' set' + (s.sets === 1 ? '' : 's') + ' of MSPI files '
               + (s.sets === 1 ? 'was' : 'were') + ' received and loaded');
    return { head: 'Summary for the month of ' + monthLabel(ym) + ':', lines };
  }

  function renderSum() {
    const ym = SUM.month;
    if (!ym) { $('sum-card').innerHTML = '<div class="empty">No data.</div>'; return; }
    const s = sumFor(ym);
    const t = sumText(ym, s);

    const tile = (v, label, sub, delta) =>
      `<div class="tile"><div class="v">${nf(v)}</div><div class="l">${esc(label)}</div>` +
      `<div class="d">${esc(sub)}${delta ? ' &middot; ' + esc(delta) : ''}</div></div>`;

    const partial = partialMonth(ym);
    $('sum-card').innerHTML =
      `<h2 class="sum-h">${esc(monthLabel(ym))}` +
        (partial ? ' <span class="sum-sub" style="font-size:12.5px">(month to date)</span>' : '') +
      '</h2>' +
      '<p class="sum-sub">Files received and loaded &middot; MMSEA response files '
        + 'from cmse_new, MSPI files from MSP'
        + (partial ? ' &middot; the month is still in progress, so the comparison '
                     + 'with last month is against a full month' : '') + '</p>' +

      '<div class="sum-group"><h3>MMSEA response files' +
        (mmseaCovered(ym) ? `<span class="tot">${nf(s.mmsea)} total</span>` : '') +
        '</h3>' +
      (mmseaCovered(ym)
        ? '<div class="tiles">' + SUM_FT.map(([k, lbl]) => tile(s.ft[k].n, lbl + ' files',
            s.ft[k].clients + (s.ft[k].clients === 1 ? ' client' : ' clients'),
            deltaLine(s.ft[k].n, ym, x => x.ft[k].n))).join('') + '</div>'
        : '<p class="sum-sub">Not covered &mdash; this dashboard loads MMSEA '
          + 'response files from ' + esc(monthLabel(MMSEA_FROM + '-01'))
          + ' forward. The MSPI side goes back further.</p>') +
      '</div>' +

      '<div class="sum-group"><h3>MSPI' +
        `<span class="tot">${nf(s.det + s.prm)} files</span></h3><div class="tiles">` +
      tile(s.sets, 'MSPI file sets',
           nf(s.det) + ' DET + ' + nf(s.prm) + ' PRM · '
             + s.mspiClients + (s.mspiClients === 1 ? ' client' : ' clients'),
           deltaLine(s.sets, ym, x => x.sets)) +
      '</div></div>' +

      '<div class="sum-prose"><h4>' + esc(t.head) + '</h4><ul>' +
      t.lines.map(l => `<li${/^0 /.test(l) ? ' class="none"' : ''}>${esc(l)}</li>`).join('') +
      '</ul><div class="sum-copy"><button class="link" id="sum-copy-btn">' +
      'Copy this summary</button></div></div>';

    $('sum-copy-btn').addEventListener('click', () => {
      const txt = t.head + '\n\n' + t.lines.join('\n');
      if (navigator.clipboard && navigator.clipboard.writeText) {
        navigator.clipboard.writeText(txt);
      } else {                                 // file:// pages often lack the async API
        const ta = document.createElement('textarea');
        ta.value = txt; document.body.appendChild(ta); ta.select();
        try { document.execCommand('copy'); } catch (_) {}
        document.body.removeChild(ta);
      }
      const b = $('sum-copy-btn');
      b.textContent = 'Copied ✓';
      setTimeout(() => { b.textContent = 'Copy this summary'; }, 1200);
    });
  }

  // ---- reference ----------------------------------------------------------
  function renderRef() {
    $('specs').innerHTML = '<thead><tr><th>File Type</th><th class="num">Display Length</th>' +
      '<th class="num">File Length</th><th>Header &amp; Trailer</th>' +
      '<th class="num">Entitlement Reason Position</th><th class="num">SSN Location</th></tr></thead><tbody>' +
      D.specs.map(s => `<tr><td><b>${esc(s[0])}</b></td><td class="num">${s[1]}</td>` +
        `<td class="num">${s[2]}</td><td>${esc(s[3])}</td>` +
        `<td class="num">${esc(s[4])}</td><td class="num">${esc(s[5])}</td></tr>`).join('') +
      '</tbody>';

    $('srckey').innerHTML = '<thead><tr><th class="num">SourceId</th><th>SourceName</th>' +
      '<th>File Type</th><th>Description</th></tr></thead><tbody>' +
      D.sources.map(s => {
        const inScope = D.scope.includes(s.id);
        const t = D.srcType[s.id];
        return `<tr class="${inScope?'':'dim'}"><td class="num">${inScope?'<b>'+s.id+'</b>':s.id}</td>` +
               `<td>${inScope?'<b>'+esc(s.name)+'</b>':esc(s.name)}</td>` +
               `<td>${t?esc(t[1]):''}</td><td>${esc(s.desc)}</td></tr>`;
      }).join('') + '</tbody>';

    $('clikey').innerHTML = '<thead><tr><th class="num">Client Id</th><th>Client Name</th>' +
      '<th>SMART Client Codes</th></tr></thead><tbody>' +
      D.clients.map(c => `<tr><td class="num">${c.id}</td><td><b>${esc(c.name)}</b></td>` +
        `<td>${esc(c.codes||'')}</td></tr>`).join('') + '</tbody>';

    $('stkey').innerHTML = '<thead><tr><th class="num">Status</th><th>Definition</th>' +
      '<th class="num">Records</th></tr></thead><tbody>' +
      D.st.map(([code, desc, n]) =>
        `<tr class="${n?'':'dim'}"><td class="num"><b>${esc(code)}</b></td>` +
        `<td class="def">${esc(desc) || '<i>undocumented</i>'}</td>` +
        `<td class="num">${n ? nf(n) : '—'}</td></tr>`).join('') + '</tbody>';
  }

  // ---- KPIs ---------------------------------------------------------------
  function renderKpis() {
    let cards;
    if (S.tab === 'cal') {
      const rows = filteredCal();
      let loaded = 0, cells = 0;
      rows.forEach(r => MN.forEach((_, i) => {
        const ym = S.year + '-' + String(i+1).padStart(2,'0');
        if (r.m[ym]) { loaded++; cells += r.m[ym].length; }
      }));
      cards = [['Rows', rows.length], ['Months with a load', loaded],
               ['Files loaded ' + S.year, cells],
               ['Clients', new Set(rows.map(r => r.client)).size]];
    } else if (S.tab === 'loads') {
      const rows = filteredLoads();
      cards = [['Loads', rows.length],
               ['Records', rows.reduce((s, l) => s + l.rec, 0)],
               ['Failed', rows.reduce((s, l) => s + l.bad, 0)],
               ['With ADO ticket', rows.filter(l => l.wi).length],
               ['No ADO ticket', rows.filter(l => !l.wi).length],
               ['Clients', new Set(rows.map(l => l.client)).size]];
    } else if (S.tab === 'mmsea') {
      const loads = filteredLoads(false);
      cards = [['Loads', loads.length],
               ['Rows', mmseaRows().length],
               ['Import count', loads.reduce((s, l) => s + l.rec, 0)],
               ['Unique members', loads.reduce((s, l) => s + (l.um || 0), 0)],
               ['Failed', loads.reduce((s, l) => s + l.bad, 0)],
               ['Clients', new Set(loads.map(l => l.client)).size]];
    } else if (S.tab === 'msr') {
      const rows = msrFiltered();
      cards = [['Files', rows.length],
               ['Clients', new Set(rows.map(r => r.rclient || r.client)).size],
               ['DET records', rows.reduce((s, r) => s + r.det, 0)],
               ['PRM records', rows.reduce((s, r) => s + r.prm, 0)],
               ['Unique HICNs (DET)', rows.reduce((s, r) => s + r.deth, 0)],
               ['Latest create', rows.reduce((m, r) => r.created > m ? r.created : m, '') || '—']];
    } else if (S.tab === 'msmo') {
      const yr = MOS.year;
      const rows = MS.filter(r => r.loaded && r.loaded.slice(0, 4) === yr &&
        (!MOS.client || r.client === MOS.client) &&
        (!MOS.contract || r.contract === MOS.contract));
      cards = [['Files ' + yr, rows.length],
               ['Clients', new Set(rows.map(r => r.client)).size],
               ['Contracts', new Set(rows.map(r => r.contract).filter(Boolean)).size],
               ['Records', rows.reduce((s, r) => s + msRecords(r), 0)]];
    } else {
      cards = [['File types', D.specs.length], ['Source types in scope', D.scope.length],
               ['Clients', D.clients.length]];
    }
    // a card value may be a string (e.g. a date) - only format the numbers
    $('kpis').innerHTML = cards.map(([l, n]) =>
      `<div class="kpi"><div class="n">${typeof n === 'string' ? esc(n) : nf(n)}</div>` +
      `<div class="l">${l}</div></div>`).join('');
  }

  // ---- render -------------------------------------------------------------
  function render() {
    ['cal', 'loads', 'mmsea', 'msmo', 'msr', 'sum', 'ref']
      .forEach(t => $('view-' + t).hidden = S.tab !== t);
    // the toolbar selects are MMSEA-scoped; the MSPi tabs carry their own bar
    const mmseaTab = S.tab === 'cal' || S.tab === 'loads' || S.tab === 'mmsea';
    document.querySelectorAll('.cal-only').forEach(e => e.hidden = S.tab !== 'cal');
    document.querySelectorAll('.loads-only').forEach(e => e.hidden = S.tab !== 'loads');
    $('source').hidden = !(S.tab === 'loads' || S.tab === 'mmsea');
    $('client').hidden = !mmseaTab;
    $('search').hidden = !mmseaTab;
    $('clear').hidden = !mmseaTab;
    hideTip();
    // the Monthly Summary tab is itself a KPI row - a second one above it would
    // just be two competing stat strips
    $('kpis').hidden = S.tab === 'sum';
    if (S.tab !== 'sum') renderKpis();
    if (S.tab === 'cal') renderCal();
    else if (S.tab === 'loads') renderLoads();
    else if (S.tab === 'mmsea') renderMmsea();
    else if (S.tab === 'msmo') renderMonthly();
    else if (S.tab === 'msr') renderMsr();
    else if (S.tab === 'sum') renderSum();
    else renderRef();
  }

  // ---- events -------------------------------------------------------------
  $('tabs').addEventListener('click', e => {
    const b = e.target.closest('button'); if (!b) return;
    S.tab = b.dataset.tab;
    [...$('tabs').children].forEach(x => x.classList.toggle('active', x === b));
    render();
  });
  $('year').addEventListener('change', e => { S.year = e.target.value; render(); });
  $('client').addEventListener('change', e => { S.client = e.target.value; render(); });
  $('source').addEventListener('change', e => { S.source = e.target.value; render(); });
  $('ticket').addEventListener('change', e => { S.ticket = e.target.value; render(); });
  let t; $('search').addEventListener('input', e => {
    clearTimeout(t); const v = e.target.value;
    t = setTimeout(() => { S.q = v; render(); }, 160);
  });
  $('clear').addEventListener('click', () => {
    S.client = S.source = S.ticket = S.q = ''; S.year = latestYear; S.open.clear();
    S.msortK = ''; S.msortD = 1;
    $('client').value = ''; $('source').value = ''; $('ticket').value = '';
    $('search').value = ''; $('year').value = latestYear; render();
  });
  $('loads-head').addEventListener('click', e => {
    const th = e.target.closest('th[data-k]'); if (!th) return;
    const k = th.dataset.k;
    if (S.sortK === k) S.sortD = -S.sortD; else { S.sortK = k; S.sortD = 1; }
    renderLoads();
  });
  $('mmsea-head').addEventListener('click', e => {
    const th = e.target.closest('th[data-k]'); if (!th) return;
    const k = th.dataset.k;
    if (S.msortK === k) {
      if (S.msortD < 0) { S.msortK = ''; S.msortD = 1; }   // third click = sheet order
      else S.msortD = -1;
    } else { S.msortK = k; S.msortD = 1; }
    renderMmsea();
  });
  $('loads-body').addEventListener('click', e => {
    const tr = e.target.closest('tr.load'); if (!tr) return;
    if (e.target.closest('a')) return;
    const sl = Number(tr.dataset.sl);
    if (S.open.has(sl)) S.open.delete(sl); else S.open.add(sl);
    renderLoads();
  });
  // ---- Monthly Summary events ---------------------------------------------
  $('sum-month').addEventListener('change', e => { SUM.month = e.target.value; renderSum(); });
  const stepMonth = d => {                 // sumMonths is newest-first
    const i = sumMonths.indexOf(SUM.month);
    const j = i + d;
    if (j < 0 || j >= sumMonths.length) return;
    SUM.month = sumMonths[j];
    $('sum-month').value = SUM.month;
    renderSum();
  };
  $('sum-prev-m').addEventListener('click', () => stepMonth(1));
  $('sum-next-m').addEventListener('click', () => stepMonth(-1));

  // ---- MSPi events --------------------------------------------------------
  $('mo-year').addEventListener('change', e => { MOS.year = e.target.value; render(); });
  $('mo-client').addEventListener('change', e => {
    MOS.client = e.target.value; moContracts(); render();
  });
  $('mo-contract').addEventListener('change', e => { MOS.contract = e.target.value; render(); });
  $('mo-expand').addEventListener('click', () => {
    Object.keys(MO_PAIRS).forEach(n => MOS.expanded.add(n)); renderMonthly();
  });
  $('mo-collapse').addEventListener('click', () => { MOS.expanded.clear(); renderMonthly(); });
  $('matrix-body').addEventListener('click', e => {
    const row = e.target.closest('tr.client-row'); if (!row) return;
    const n = row.dataset.client;
    if (MOS.expanded.has(n)) MOS.expanded.delete(n); else MOS.expanded.add(n);
    renderMonthly();
  });
  $('matrix-body').addEventListener('mouseover', e => {
    const td = e.target.closest('td.hit');
    if (td && td.dataset.tip) showTip(e, td.dataset.tip);
  });
  $('matrix-body').addEventListener('mousemove', e => {
    if (tip.style.display === 'block') moveTip(e);
  });
  $('matrix-body').addEventListener('mouseout', e => {
    const to = e.relatedTarget;
    if (!to || !to.closest || !to.closest('td.hit')) hideTip();
  });

  [['msr-client', 'client'], ['msr-type', 'type'], ['msr-contract', 'contract'],
   ['msr-file', 'file'], ['msr-from', 'from'], ['msr-to', 'to']]
    .forEach(([id, key]) => $(id).addEventListener('input', () => {
      MR[key] = $(id).value; MR.page = 0; render();
    }));
  $('msr-reset').addEventListener('click', () => {
    ['msr-client', 'msr-type', 'msr-contract', 'msr-file', 'msr-from', 'msr-to']
      .forEach(id => $(id).value = '');
    Object.assign(MR, { client:'', type:'', contract:'', file:'', from:'', to:'',
                        sortK:'', sortD:1, page:0 });
    render();
  });
  $('msr-head').addEventListener('click', e => {
    const th = e.target.closest('th[data-k]'); if (!th) return;
    const k = th.dataset.k;
    if (MR.sortK === k) {
      if (MR.sortD < 0) { MR.sortK = ''; MR.sortD = 1; }   // third click = TapeID desc
      else MR.sortD = -1;
    } else { MR.sortK = k; MR.sortD = 1; }
    MR.page = 0; renderMsr();
  });
  $('msr-prev').addEventListener('click', () => { MR.page--; renderMsr(); });
  $('msr-next').addEventListener('click', () => { MR.page++; renderMsr(); });

  $('expand-all').addEventListener('click', () => {
    const rows = filteredLoads();
    if (rows.every(l => S.open.has(l.sl))) S.open.clear();
    else rows.forEach(l => S.open.add(l.sl));
    $('expand-all').textContent = S.open.size ? 'Collapse all' : 'Expand all';
    renderLoads();
  });

  // ---- export -------------------------------------------------------------
  function buildExport() {
    if (S.tab === 'cal') {
      const rows = filteredCal();
      const isT = S.year === D.trackerYear;
      return {
        name: 'CMSE_Calendar_' + S.year,
        title: D.name + ' \u2014 MMSEA Calendar ' + S.year,
        subtitle: (S.client || 'All clients') + ' \u00b7 generated ' + D.generated,
        headers: ['Client', 'Client Id', 'File Type', 'Frequency', 'Handling', ...MN],
        rows: rows.map(r => [r.client, r.cid || '', r.ft, r.freq, r.hand,
          ...MN.map((_, i) => {
            const ym = S.year + '-' + String(i+1).padStart(2,'0');
            if (r.m[ym]) return 'X';
            const tr = isT ? r.tr[String(i+1)] : null;
            if (!tr || tr.late) return '';
            return /^\d+$/.test(tr.v) ? 'X' : tr.v;
          })]),
        note: 'TRGRepSQL3 / CMSE_New \u00b7 X = loaded to CMSE, E = expected',
        rowsPerSlide: 12, fontSz: 800,
      };
    }
    if (S.tab === 'loads') {
      const rows = sortLoads(filteredLoads());
      return {
        name: 'CMSE_Loads',
        title: D.name + ' \u2014 MMSEA SourceLog',
        subtitle: (S.client || 'All clients') + ' \u00b7 ' + rows.length +
                  ' loads \u00b7 generated ' + D.generated,
        headers: ['SourceLogId','SourceId','Source Name','File Type','Client Name',
                  'Client Id','EntryName','Import Start','Import Complete',
                  'Records','Success','Failed','PCN','ADO Ticket'],
        rows: rows.map(l => [l.sl, l.src, SRC[l.src]||'', l.ft, l.client, l.cid,
          l.entry, l.start, l.done, l.rec, l.ok, l.bad, l.pcn, l.wi||'']),
        note: 'cmse_new..SourceLog, SourceId ' + D.scope.join(', '),
        rowsPerSlide: 12, fontSz: 700,
      };
    }
    if (S.tab === 'sum') {
      const s = sumFor(SUM.month);
      const t = sumText(SUM.month, s);
      return {
        name: 'Monthly_Summary_' + SUM.month,
        title: D.name + ' — Monthly Summary',
        subtitle: t.head.replace(/:$/, '') + ' · generated ' + D.generated,
        headers: ['Measure', 'Files', 'Detail'],
        rows: [
          ...(mmseaCovered(SUM.month)
              ? SUM_FT.map(([k, lbl]) => [lbl + ' files', s.ft[k].n,
                                          s.ft[k].clients + ' clients'])
              : [['MMSEA response files', '', 'not covered before '
                  + monthLabel(MMSEA_FROM + '-01')]]),
          ['MSPI file sets', s.sets, s.det + ' DET + ' + s.prm + ' PRM, '
                                     + s.mspiClients + ' clients'],
        ],
        note: t.lines.join(' · '),
        rowsPerSlide: 8, fontSz: 1200,
      };
    }
    if (S.tab === 'msr') {
      const rows = msrSorted();
      return {
        name: 'MSPi_Report',
        title: D.name + ' — MSPi Report',
        subtitle: (MR.client || 'All clients') + ' · ' + rows.length +
                  ' files · generated ' + D.generated,
        headers: MRCOLS.map(c => c[1]),
        rows: rows.map(r => [r.rclient, r.type, r.file, r.prod, r.det, r.deth,
                             D.mspiRawOrTransform, r.created]),
        note: 'TRGINTP3 / MSP · etl.Tape + history.DET · Load count and Unique '
              + 'HICN Count are history.DET only, so PRM rows are 0',
        rowsPerSlide: 12, fontSz: 700,
      };
    }
    if (S.tab === 'msmo') {
      const yr = MOS.year;
      const byClient = {};
      for (const r of MS) {
        if (!r.loaded || r.loaded.slice(0, 4) !== yr) continue;
        if (MOS.client && r.client !== MOS.client) continue;
        if (MOS.contract && r.contract !== MOS.contract) continue;
        const c = byClient[r.client] || (byClient[r.client] = {});
        const ct = c[r.contract] || (c[r.contract] = {});
        const mo = Number(r.loaded.slice(5, 7));
        ct[mo] = (ct[mo] || 0) + 1;
      }
      const out = [];
      Object.keys(byClient).sort().forEach(cl => {
        Object.keys(byClient[cl]).sort().forEach(ct => {
          out.push([cl, ct || '(none)',
            ...MN.map((_, i) => byClient[cl][ct][i + 1] || '')]);
        });
      });
      return {
        name: 'MSPi_Monthly_' + yr,
        title: D.name + ' — MSPi Monthly ' + yr,
        subtitle: (MOS.client || 'All clients') + ' · generated ' + D.generated,
        headers: ['Client', 'Contract', ...MN],
        rows: out,
        note: 'TRGINTP3 / MSP · cell = files loaded that month',
        rowsPerSlide: 12, fontSz: 800,
      };
    }
    if (S.tab === 'mmsea') {
      const rows = mmseaRows();
      return {
        name: 'MMSEA_Report',
        title: D.name + ' — MMSEA Report',
        subtitle: (S.client || 'All clients') + ' · ' + rows.length +
                  ' rows · generated ' + D.generated,
        headers: MCOLS.map(c => c[1]),
        rows: rows.map(r => [r.client, r.ftname, r.rt, r.idate, '', '',
          r.rec, r.ok, r.bad, r.age, r.dis, r.esrd, r.um, '',
          r.st, r.stn == null ? '' : r.stn, r.sl, r.file]),
        note: 'cmse_new..SourceLog + ImportStaging · '
              + (D.unsourced || []).join(', ') + ' are not held in cmse_new',
        rowsPerSlide: 12, fontSz: 600,
      };
    }
    return {
      name: 'CMSE_FileTypes',
      title: D.name + ' \u2014 File Types',
      subtitle: 'Generated ' + D.generated,
      headers: ['File Type','Display Length','File Length','Header & Trailer',
                'Entitlement Reason Position','SSN Location'],
      rows: D.specs.map(s => s.slice()),
      note: 'From ClientTracker.2026.xlsx, MMSEA - 2026 tab',
    };
  }
  if (window.RptExport) RptExport.wire('btn-export', 'exp-menu', buildExport);

  render();
})();
</script>
</body>
</html>
"""


def generate_html(data):
    return (HTML_TEMPLATE
            .replace("__EXPORT_CSS__", EXPORT_CSS)
            .replace("__EXPORT_UI__", EXPORT_UI)
            .replace("__EXPORT_JS__", EXPORT_JS)
            .replace("__DATA_JSON__", json.dumps(data, separators=(",", ":"))))


def main():
    full = "--full" in sys.argv
    data = build(full=full)
    html = generate_html(data)

    primary = OUTPUT_PATHS[0]
    written = None
    for path in OUTPUT_PATHS:
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            if written and os.path.exists(written):
                shutil.copyfile(written, path)
            else:
                with open(path, "w", encoding="utf-8") as f:
                    f.write(html)
                written = path
            print("[done] %s" % path)
        except (PermissionError, OSError) as e:
            print("[warn] couldn't write %s: %s" % (path, e))
    if not written:
        raise SystemExit("[error] no output written")
    if written != primary:
        print("[warn] primary path %s was not written" % primary)

    # keep the two retired dashboards' file names alive as copies, so nobody is
    # left staring at a stale CMSEReport.html / MSPIReport.html
    for path in OUTPUT_PATHS:
        for name in LEGACY_NAMES:
            dst = os.path.join(os.path.dirname(path), name)
            try:
                shutil.copyfile(written, dst)
            except (PermissionError, OSError) as e:
                print("[warn] couldn't refresh %s: %s" % (dst, e))

    # publish the protocol registration + ClickOnce launcher next to the report
    for name in SIDECARS:
        src = os.path.join(HERE, name)
        if not os.path.exists(src):
            print("[warn] %s missing - the tool links may not launch" % src)
            continue
        for path in OUTPUT_PATHS:
            dst = os.path.join(os.path.dirname(path), name)
            if os.path.abspath(dst) == os.path.abspath(src):
                continue
            try:
                shutil.copyfile(src, dst)
            except (PermissionError, OSError) as e:
                print("[warn] couldn't publish %s: %s" % (dst, e))


if __name__ == "__main__":
    main()
