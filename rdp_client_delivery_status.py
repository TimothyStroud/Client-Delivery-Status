"""
RAMP Client Delivery Status Report — calendar view.

Data sources (all combined per scheduled cell):
  - SQL [DHTStats].[DHT].[TableList] on TRGUTIL10  — canonical certification date
  - RAMP /api/Ramp/Snap/SnapQueueStatus            — snap completion (for snap-only clients)
  - RAMP /api/Ramp/Queue/List                      — load-job completion
  - RAMP /api/Ramp/Job/List                        — to detect Inactive
  - ADO WIQL                                       — tickets tagged 'Delivery Ticket'

Each (client, scheduled-day) cell in the calendar resolves to:
  - Date  (MM/DD)  if the client certified that day in DHT
  - "Snap"          if the client snapped that day but does not certify (snap-only)
  - "L"             if a load/snap is currently in progress for that client today
  - blank           otherwise

Client name suffix conventions (matching the All Clients tab key):
  - (s)  SLA Client
  - (p)  Rx Client Post Snap
  - (n)  Not Delivered (special)
  - M -  Monthly client prefix (placed dynamically on day ticket fired / snap completed)
"""
import calendar
import glob as globmod
import json
import os
import re
import subprocess
import tempfile
from collections import defaultdict
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ============================================================
#                    static configuration
# ============================================================
ADO_BASE   = "https://devops.ado.rawlingslou.prod/TFS2012/Rawlings"
ADO_LINK   = "https://devops.ado.rawlingslou.prod/TFS2012/AppDev/_workitems/edit/{}"
RAMP_BASE  = "http://ramp"
SQL_SERVER = "TRGUTIL10"
SQL_DB     = "DHTStats"
OUTPUT_DIR = r"\\trgfile1\Shared\DIG\Data Business Delivery Team\Delivery Schedule\Daily Status Reports"
# 2026 monthly schedule files maintained by the team. Their 'All Clients' tab
# is the source of truth for past months (Jan–Apr 2026); the live DHT/RAMP
# data we fetch has only 3 months of cert history and rolling queue/snap, so
# we overlay those files for closed months instead of re-deriving cells.
EXPECTED_DATES_DIR = r"\\trgfile1\Shared\DIG\Data Business Delivery Team\Delivery Schedule\2026"
EXPECTED_DATES_FILES = {
    1: "202601_ExpectedClientDates_JAN.xlsx",
    2: "202602_ExpectedClientDates_FEB.xlsx",
    3: "202603_ExpectedClientDates_MAR.xlsx",
    4: "202604_ExpectedClientDates_APR.xlsx",
}
# Marker strings the team uses in those files that should pink-shade the cell.
ALL_CLIENTS_ALERT_MARKERS = {"no data", "load failure", "missing files",
                             "deployment", "snap hold", "field changes",
                             "outreach", "empty", "disabled", "inactive",
                             "cleanup"}
# Clients whose empty/True date cells in the snapshot files should render as
# ✓ (per user 2026-06-03 — Feb–Apr files use boolean True as a verify-load
# placeholder for non-cert days, but the team reads them as "loaded/snapped").
ALL_CLIENTS_FILL_CHECKMARK = {
    "AetnaHRP", "AetnaRCE", "AetnaRx", "NCStateAetna",
    "CVSPBMRx", "ESIPBMRx", "PrimePBMRx", "MedImpactPBMRx",
}
# Project-folder copy (canonical filename — overwritten each run).
LOCAL_COPY_DIR = r"C:\Users\tls2\.claude\projects\H--"
# OneDrive copy with fixed filename so a single Notion link stays valid run-to-run.
ONEDRIVE_COPY_PATH = r"C:\Users\tls2\OneDrive - Machinify\Documents\Reports\ClientDeliveryStatus.xlsx"

# Recognise three title formats: Snap and Mine | Load and Snap | Kaiser - SNAP/MINE
TITLE_RE = re.compile(
    r"^\s*(Snap and Mine|Load and Snap|Kaiser\s*-\s*SNAP/MINE)\s*-\s*([^-\s][^-]*?)\s*(?:-|$)",
    re.I,
)

# ADO ticket client identifier → additional normalized strings to match against
# RAMP JobName (substring) / FeedName / ClientName (equality).
CLIENT_ALIASES = {
    "SamaritanHealth":      ["samaritan"],
    "HealthNetCA":          ["healthnet"],
    # JohnsHopkins RAMP feed is named "JHHC Medical" — without these aliases,
    # 'JHHC Medical 0110 Load' wouldn't match find_matching_jobs / snap_idx.
    # Added 2026-05-19 per user: "JohnsHopkins 'JHHC Medical 0110 Load' should be an 'L'".
    "JohnsHopkins":         ["johnshopkins", "jhhc", "jhhcmedical"],
    # JHHCPassfile (monthly): the 'JHHC Passfile Email' job → snap_idx key
    # "jhhcpassfileemail" (no digit code in the name). Distinct from
    # JohnsHopkins ('JHHC Medical 0110 Load'). Per user 2026-06-08.
    "JHHCPassfile":         ["jhhcpassfileemail"],
    # ElevanceMMMRx (daily): RAMP jobs are 'ElevanceMMMRx Masterload 0110 Load' /
    # '0120 Snap'. build_snap_index strips the trailing "load" off "Masterload",
    # so the snap-index key is "elevancemmmrxmaster" — alias it so snap_on_day's
    # strict-equality match fires for ✓. Per user 2026-06-16: daily ✓ on load+snap.
    "ElevanceMMMRx":        ["elevancemmmrxmaster"],
    # BCBSNorthCarolinaFEP (monthly, ~10th): the monthly delivery loads through
    # 'BCBSNC FEP CareFirst 0100 Stage' → '0110 Load' (snap-index key
    # "bcbsncfepcarefirst"). The bare "bcbsncfep" alias only reaches the
    # 'BCBSNC FEP Daily' feed, which stages+loads EVERY night and is NOT the
    # monthly delivery — see LOAD_NAME_REQUIRED. Per user 2026-08-03.
    "BCBSNorthCarolinaFEP": ["bcbsncfep", "bcbsncfepcarefirst"],
    # NCStateAetna's daily load runs through 'Aetna RCE 310 ETL Load' (Feed
    # "RCE Medical" → key "aetnarce" after stripping). Include aetnarce alias
    # so NCStateAetna gets ✓ from those same daily completions.
    "NCStateAetna":         ["ncstateaetna", "aetnarce"],
    "WPRxDMGCOBMining":     ["wellpointedwardrxdmgcobmining", "wpwedmgcobmining"],
    "HumanaRx":             ["humanarx"],
    # OptumPBMRx: RAMP load is "Optum 0110 PBM Load" → digit-collapsed
    # "optumpbmload"; alias "optumpbm" gives the substring match for both
    # is_loading_today and snap_idx lookups.
    "OptumPBMRx":           ["optumpbm"],
    # GEHA: the actual RAMP load is "GEHA UMR 0110 Load" → prefix "gehaumr".
    # Plain "geha" alone wouldn't strict-match "gehaumr" in snap_idx.
    "GEHA":                 ["geha", "gehaumr"],
    "EmblemFacets":         ["emblemfacets"],
    "AetnaQNXT":            ["aetnaqnxt"],
    "AetnaQNXTRx":          ["aetnaqnxtrx"],
    "ElixirRx":             ["elixirrx"],
    "Tufts_PublicPlan":     ["tuftspublicplan"],
    # MMOH MHS jobs ('MMOH MHS 0100 Stage' / '… 0120 Snap') derive the snap_idx
    # prefix "mmohmhs" — alias it so the stage-file ✓ snap match reaches them.
    "MedicalMutualMHS":     ["medicalmutualmhs", "mmohmhs"],
    # WellpointEdwardRx — RAMP jobs are named "Wellpoint RX Claims …" /
    # "Wellpoint RX Claims HealthSun …" so we need the prefix-derived
    # snap_idx keys to be reachable via strict-equality aliases. Per user
    # 2026-05-20: "'Wellpoint RX Claims HealthSun 0130 Load' is running,
    # so there should be an 'L' on WellpointEdwardRx."
    "WellpointEdwardRx":    ["wellpointedwardrx", "wellpointrxclaims",
                             "wellpointrxclaimshealthsun"],
    # WellpointRxElig (monthly 'M - WellpointRx Elig'): the Wellpoint RX Elig
    # pipeline (0100 Stage -> 0110 Load -> 0120 Snap -> 0130 Mine). Both the
    # Load and Snap snap_idx entries derive the prefix key "wellpointrxelig"
    # (distinguished by kind). Per user 2026-06-12: L off '0110 Load', ✓ off
    # '0120 Snap'. Distinct from WellpointEdwardRx (Claims) jobs.
    "WellpointRxElig":      ["wellpointrxelig"],
    # HarvardPilgrim — RAMP job is "HarvardPilgrim Claims 0110 Load" so
    # the snap_idx prefix key is "harvardpilgrimclaims" (the trailing-
    # word strip doesn't catch "Claims"). Per user 2026-05-20:
    # "'HarvardPilgrim Claims 0110 Load' ran last night, so there should
    # be an 'L' until certification."
    "HarvardPilgrim":       ["harvardpilgrim", "harvardpilgrimclaims"],
    # UPMC — RAMP job "UPMC Masterload 0110 Load". The trailing-keyword
    # strip incorrectly takes "load" off "Masterload" leaving snap_idx
    # key "upmcmaster". Per user 2026-05-20: "'UPMC Masterload 0110 Load'
    # finished, so it should show an 'L' until certification."
    "UPMC":                 ["upmc", "upmcmaster", "upmcmasterload"],
    # BCBSAR is a MEDICAL client — its only jobs are 'BCBSAR Medical 0100/0110'.
    # The bare "bcbsar" alias was dropped 2026-06-12 because it substring-matched
    # the sibling BCBSARRx jobs ('BCBSARRx COBC/MasterLoad …'), wrongly pulling
    # the BCBSARRx COBC load failure onto BCBSAR. Matched via "bcbsarmedical"
    # only; see CLIENT_PRIMARY_KEY_OVERRIDE. (DHT cert key "bcbsar" is still
    # yielded by _keys_for_client's base normalize, so cert lookup is unaffected.)
    "BCBSAR":               ["bcbsarmedical"],
    "BCBSARRx":             ["bcbsarrx"],
    "MedStar":              ["medstar"],
    # HealthNewEngland shows up as "HNE Medical" on the RAMP Dashboard.
    # The bare "hne" alias was REMOVED 2026-08-05: find_matching_jobs does a
    # SUBSTRING match, and "hne" is a substring of "healthnet…" (healt-HNE-t),
    # so every HealthNetCA job matched HealthNewEngland too — a Failed
    # 'HealthNet 0110 Claims Load' painted "Load Failure" on HNE's Thursday
    # cell (per user: "HealthNewEngland is not in a Load failure status").
    # "hnemedical" still matches 'HNE Medical 0100 Stage/0110 Load' (the real
    # delivery jobs) and the base key "healthnewengland" still serves the DHT
    # cert lookup. Same class of fix as Oscar/Medica in CLIENT_PRIMARY_KEY_OVERRIDE.
    "HealthNewEngland":     ["healthnewengland", "hnemedical"],
    # AetnaRx queue variants: each distinct JobName prefix becomes its own
    # snap_idx key under JobName-only indexing.
    # "aetnarxclaims" covers the snap step "AetnaRx Claims 0130 Start Snap"
    # (plural Claims) — distinct from the load step "AetnaRX Claim 0120 Load".
    "AetnaRx":              ["aetnarx", "aetnarxo20", "aetnarxclaim", "aetnarxclaims",
                             "aetnarxnewelig", "aetnarxmining",
                             "aetnarxcobcmasterload", "aetnarxcobc",
                             "aetnarxcaqh", "aetnarxtrr", "aetnarxihp"],
    # CenteneFidelis / CenteneFidelisRx job prefixes
    "CenteneFidelis":       ["centenefidelis", "centenefidelismedical"],
    "CenteneFidelisRx":     ["centenefidelisrx", "centenefidelisrxmasterload"],
    # WellCare / WellCareRx job prefixes
    # 'wellcare' as a substring also lives inside 'wellcarerx', so we use
    # CLIENT_PRIMARY_KEY_OVERRIDE to swap the auto-derived primary key for
    # 'wellcaremedical'. The substring match then no longer crosses into
    # WellCareRx territory.
    "WellCare":             ["wellcaremedical"],
    # WellCareRx: per user 2026-05-18, ancillary jobs (COBC, ABII) are NOT
    # load indicators — exclude their snap_idx keys from matching so that a
    # successful COBC load this week does not trip the "loaded this week" L.
    # 2026-08-06: added "wellcarerxmaster" — the REAL snap_idx key. build_snap_index
    # slices the JobName before the first digit ("WellCareRx Masterload ") then strips
    # a trailing step word with `(load|stage|snap|...)$` — which has no word boundary,
    # so it eats the "load" inside "Master*load*" → key "wellcarerxmaster". The
    # "wellcarerxmasterload" alias below therefore NEVER matched under
    # _src_matches_client's strict equality, so snap_in_week always returned None and
    # the Friday cell went blank between load and cert (every week was being patched
    # by hand via MANUAL_OVERRIDES). Same latent bug shape as UPMC, which already
    # carries both "upmcmaster" and "upmcmasterload".
    "WellCareRx":           ["wellcarerx", "wellcarerxmaster",
                             "wellcarerxmasterload"],
    # Oscar / OscarRx job prefixes. NOTE: the bare "oscar" alias was REMOVED
    # 2026-07-23 — it substring-matched "oscarrx..." in find_matching_jobs and
    # put a false "L" on Oscar Medical from a Ready 'Oscar RX 0110 Load'. Only
    # "oscarmedical" (which does NOT appear inside "oscarrx...") is kept; the DHT
    # cert key "oscar" is still yielded by _keys_for_client's base. See the
    # matching CLIENT_PRIMARY_KEY_OVERRIDE["Oscar"] entry.
    "Oscar":                ["oscarmedical"],
    "OscarRx":              ["oscarrx", "oscarrxmasterload", "oscarrxabii"],
    # Centene / CenteneRx job prefixes
    "Centene":              ["centene", "centenemedical"],
    "CenteneRx":            ["centenerx", "centenerxhnt", "centenerxhntelig"],
    # AetnaHRP / AetnaRCE job prefixes
    "AetnaHRP":             ["aetnahrp"],
    "AetnaRCE":             ["aetnarce"],
    "AetnaQNXTRx":          ["aetnaqnxtrx", "aetnaqnxtrxmasterload"],
    "AetnaQNXT":            ["aetnaqnxt", "aetnaqnxtmasterload", "aetnaqnxtmspi", "aetnaqnxtcaqh"],
    # NCStateAetna alias to aetnarce (uses same ETL load)
    "NCStateAetna":         ["ncstateaetna", "aetnarce", "ncstateaetnamasterload"],
    # ESIPBMRx 0120 Start Snap → "esipbmrx"
    "ESIPBMRx":             ["esipbmrx"],
    # CareSource variants
    "CareSource":           ["caresource"],
    "CareSourceRx":         ["caresourcerx"],
    # Kaiser WA: ONLY exact "kaiserwa" — keep separate from KaiserWARx etc.
    "Kaiser_WA":            ["kaiserwa"],
    "Kaiser_WARx":          ["kaiserwarx"],
    # KaiserPrePayCOB
    "KaiserPrePayCOB":      ["kaiserprepaycob", "kaiserpareoprepay"],
    # Kaiser SC/NC Pareo — RAMP feeds are "Kaiser Pareo SC" and
    # "Kaiser Pareo NC&TPMG" so the normalized form has "pareo" before
    # the state code. Without these aliases find_matching_jobs misses them.
    # Per user 2026-05-19: "KaiserSCPareo & KaiserNCPareo should have an 'L'".
    "KaiserSCPareo":        ["kaiserscpareo", "kaiserpareosc"],
    "KaiserNCPareo":        ["kaiserncpareo", "kaiserpareonc"],
    # Kaiser_MASTapestry — RAMP renamed feed to "Kaiser Pareo MAS" per user
    # 2026-06-03. Keeps default `kaisermastapestry` match for historical
    # 'Kaiser MAS Tapestry' JobNames; adds `kaiserpareomas` for the new form.
    "Kaiser_MASTapestry":   ["kaisermastapestry", "kaiserpareomas"],
    # MMOH (WC) monthly: the report's "MMOH (WC)" row tracks the Workers' Comp
    # load 'MMO 0110 WC Load' (Stage->Load, no snap/cert). build_snap_index
    # emits the feed+sub-feed key "mmowc" for it (the bare-prefix key "mmo" is
    # too short to index). Per user 2026-06-08 (correction): this row is the WC
    # load, not the MMOHRx monthly claim. Distinct from MedicalMutualOH
    # ('MMOH Claims 0110 Load') and the weekly MMOHRx ('MMOHRx Weekly Claim').
    "MMOH":                 ["mmowc"],
    # MedicalMutualOH (monthly, cert-only): its load is 'MMOH Claims 0110 Load'
    # → snap_idx key "mmohclaims" (per user 2026-06-08). The auto-derived
    # primary key "medicalmutualoh" never matches the RAMP "MMOH ..." JobNames.
    "MedicalMutualOH":      ["mmohclaims"],
    # MMOHRx weekly Tue: only 'MMOHRx Weekly Claim 0110 Load' counts. COBC
    # alias dropped 2026-05-18 so MMOHRx COBC Successful loads don't trip L.
    "MMOHRx":               ["mmohrx", "mmohrxweeklyclaim"],
    # MMOHRxMonthly (monthly 'M - MMOHRx'): the MMOHRx Monthly Claim pipeline
    # (Stage->Load 0110->Snap 0120). Both the Load and Snap snap_idx entries
    # derive key "mmohrxmonthlyclaim" (distinguished by kind). Per user
    # 2026-06-09: L off '0110 Load', ✓ off '0120 Snap'. Distinct from the
    # weekly MMOHRx (Weekly Claim) and from MMOH (WC).
    "MMOHRxMonthly":        ["mmohrxmonthlyclaim"],
    # Cigna variants
    "CignaFacets":          ["cignafacets"],
    # CignaRx: only 'Cigna RX 0110 Load' counts (key 'cignarx'). COBC and
    # Daily PassFile have their own snap_idx keys and shouldn't trigger L
    # for CignaRx — per user 2026-05-18.
    "CignaRx":              ["cignarx"],
    "CignaPower":           ["cignapower"],
    "CignaProClaims":       ["cignaproclaims"],
    # Premera / PremeraMedAdv*
    "Premera":              ["premera"],
    "PremeraMedAdvVIS":     ["premeramedadvvis"],
    "PremeraMedAdvRx":      ["premeramedadvrx"],
    # Tufts variants
    "TuftsMedPref":         ["tuftsmedpref"],
    "Tufts_Audit_CIT":      ["tuftsauditcit"],
    "WebTPA":               ["webtpa"],
    "BCBSPuertoRico":       ["bcbspuertorico"],
    "NCState":              ["ncstate"],
    "MedImpactPBMRx":       ["medimpactpbmrx"],
    "CenteneRx":            ["centenerx"],
    # BCBSFLEligibilityLoad: the RAMP job is "BCBSFL Eligibility ..." → key
    # "bcbsfleligibility" after stripping load/stage/digits.
    "BCBSFLEligibilityLoad": ["bcbsfleligibility"],
    # Kaiser monthly aliases
    "Kaiser_GE":            ["kaiserge"],
    "Kaiser_AmbCO":         ["kaiserambulanceco", "kaiserambco"],
    "Kaiser_AmbGA":         ["kaiserambulancega", "kaiserambga"],
    "Kaiser_AmbHI":         ["kaiserambulancehi", "kaiserambhi"],
    "Kaiser_AmbNW":         ["kaiserambulancenw", "kaiserambnw"],
    "Kaiser_AmbN":          ["kaiserambulancenc", "kaiserambn"],
    "Kaiser_AmbS":          ["kaiserambulancesc", "kaiserambs"],
    "Kaiser_AmbM":          ["kaiserambulancemas", "kaiserambulancema", "kaiserambm"],
}

# --------- canonical client lists (from the manual ExpectedClientDates sheet) ---------
# Daily clients (load + snap every weekday; cert cadence varies).
DAILY_CLIENTS = ["AetnaHRP", "AetnaRCE", "AetnaRx", "ElevanceMMMRx", "NCStateAetna"]
KAISER_PREPAY_CLIENT = "KaiserPrePayCOB"

# Fixed display order for the Daily section (per user 2026-06-16) — NOT
# alphabetical. Every daily client shows on every weekday (blank or "-" when it
# didn't load). Matched by label prefix so client suffixes like "(s)"/"(p)"
# still resolve. "Kaiser Submission" sits between KaiserPrePayCOB and
# NCStateAetna. (Aetna MSPI moved to Monthly Ad Hoc 2026-06-25.)
DAILY_ORDER = [
    "AetnaHRP",
    "AetnaRCE",
    "AetnaRx",
    "ElevanceMMMRx",
    "KaiserPrePayCOB",
    "Kaiser Submission",
    "NCStateAetna",
]


def _daily_order_key(label):
    """Sort key mapping a daily row label to its fixed-order index."""
    for idx, name in enumerate(DAILY_ORDER):
        if label.startswith(name):
            return idx
    return len(DAILY_ORDER)

# Weekly clients keyed by canonical name -> list of weekday names they deliver on.
WEEKLY_CLIENTS = {
    # === MONDAY ===
    "BCBSKSMedAdv":          ["Monday"],
    "Cambia":                ["Monday"],
    "CignaPower":            ["Monday"],
    "CignaProClaims":        ["Monday"],
    "CVSPBMRx":              ["Monday"],
    "EverNorthRx":           ["Monday"],
    "GEHA":                  ["Monday"],
    "HealthNetCA":           ["Monday"],
    "MedicaDean":            ["Monday"],
    "Tufts_Audit_CIT":       ["Monday"],
    "TuftsMedPref":          ["Monday"],
    "TuftsRx":               ["Monday"],
    # === TUESDAY ===
    "BCBSAR":                ["Tuesday"],
    "BCBSARRx":              ["Tuesday"],
    "BCBSFL":                ["Tuesday"],
    "Centene":               ["Tuesday"],
    "CenteneQualChoice":     ["Tuesday"],
    "CignaFacets":           ["Tuesday"],
    "CignaRx":               ["Tuesday"],
    "HMSA":                  ["Tuesday"],
    "HMSA_Rx":               ["Tuesday"],
    "JohnsHopkins":          ["Tuesday"],
    "MedStar":               ["Tuesday"],
    "MMOHRx":                ["Tuesday"],
    "Wellmark":              ["Tuesday"],
    # === WEDNESDAY ===
    "CareSource":            ["Wednesday"],
    "CenteneFidelis":        ["Wednesday"],
    "CenteneFidelisRx":      ["Wednesday"],
    "EmblemRx":              ["Wednesday"],
    "ExcellusRx":            ["Wednesday"],
    "HarvardPilgrim":        ["Wednesday"],
    "Medica":                ["Wednesday"],
    "Oscar":                 ["Wednesday"],
    "WellpointEdwardRx":     ["Wednesday"],
    # === THURSDAY ===
    "HealthNewEngland":      ["Thursday"],
    "Kaiser_CO":             ["Thursday"],
    "Kaiser_GA":             ["Thursday"],
    "Kaiser_HI":             ["Thursday"],
    "Kaiser_MASTapestry":    ["Thursday"],
    "CareSourceRx":          ["Thursday"],
    "Kaiser_NW":             ["Thursday"],
    "KaiserNCPareo":         ["Thursday"],
    "KaiserSCPareo":         ["Thursday"],
    "Premera":               ["Thursday"],
    "PrimePBMRx":            ["Thursday"],
    "UPMC":                  ["Thursday"],   # moved from Tuesday per user 2026-07-02
    # === FRIDAY ===
    "CenteneRx":             ["Friday"],
    "OscarRx":               ["Friday"],
    "WebTPA":                ["Friday"],
    "WellCare":              ["Friday"],
    "WellCareRx":            ["Friday"],
    # Snap-only Monday slots kept from earlier user instructions:
    # ESIPBMRx is MONTHLY (handled by MONTHLY_CLIENTS) — not weekly.
    # OptumPBMRx is monthly (loaded once per month via TRGETL3 tape RAW1/2/3) —
    # placed dynamically by determine_monthly() on the actual tape-load date.
}

# Clients that should always show as Inactive (pink shade) regardless of
# RAMP/DHT detection. User-confirmed list.
# HealthNetCA added 2026-05-18: 'HealthNet 0100 Claims Stage' disabled in RAMP.
# Kaiser_AmbM removed 2026-05-19 (no longer inactive). Snap re-enabled
# 2026-06-08 — now treated as a normal monthly Kaiser_Amb cert feed (see
# MONTHLY_CERT_ONLY_CLIENTS / SNAP_KIND_ONLY_CLIENTS).
# Oscar (weekly Wed) was Inactive 7/1-7/8; reactivated 2026-07-15 (backfill
# certified 7/15 covering 7/1 & 7/8). 2026-07-17: re-added to FORCED_INACTIVE
# (all future dates Inactive). 2026-07-23: REACTIVATED per user — Oscar Medical is
# back in RAMP (Stage/Load jobs running again; last cert 7/15) so it must track
# live loads/cert/L via the normal STAGE_FILE_CELL path. Removed from
# FORCED_INACTIVE. The 7/1/7/8/7/15 cells keep their MANUAL_OVERRIDES (dates) as a
# historical anchor; future Wed cells now render real state.
# ESIPBMRx (monthly, tape/snap-driven) added 2026-07-15 per user: mark Inactive
# until loading resumes ("can go to 'L' once loading starts again").
# Tufts_Audit_CIT (weekly Mon) added 2026-07-17 per user: mark Inactive.
# Tufts_PublicPlan REACTIVATED 2026-07-27 per user — "Active again. All missing
# months through this month will be certified on the next ticket (today)." Removed
# from FORCED_INACTIVE so its DHT cert dates render on every month tab as the
# catch-up ticket certifies. Also added to AUTO_INACTIVE_EXCLUDE below (same as
# CareFirstRx) so, if its 0100/0110 RAMP jobs are still disabled, the auto-inactive
# sweep doesn't short-circuit determine_monthly() before the cert lookup. Its
# STAGE_FILE_CELL_CLIENTS wiring provides the ✓ fallback when a month has no cert.
# TuftsRx REACTIVATED 2026-07-28 per user — "no longer Inactive; will be certified
# today for all past Monthly deliveries; checkmarks can be added to the weekly
# cells." Removed from FORCED_INACTIVE + added to AUTO_INACTIVE_EXCLUDE (same as
# Tufts_PublicPlan/CareFirstRx). Weekly Monday cells now resolve via the
# STAGE_FILE_CELL_CLIENTS ✓ fallback (all claims files 2/16–7/27 loaded PS=50).
# The monthly (10th) cert catch-up was still "Email sent, Ready for Certification
# review" in DHT at reactivation (not yet Certified) and, once flipped, all rows
# carry today's CertTimestamp — so the auto cert lookup would only place July.
# May/June/July monthly cells are therefore pinned to the 7/28 cert date via
# MONTHLY_MONTH_MARKER_OVERRIDES (a real per-month DHT cert still auto-wins).
# 2026-07-28 per user: "the only clients still Inactive are HealthNetCA &
# MedicalMutualMHS." ESIPBMRx + Tufts_Audit_CIT REACTIVATED (removed here + added
# to AUTO_INACTIVE_EXCLUDE so the RAMP auto-sweep can't re-flag them) — they now
# render real cert/snap history. MedicalMutualMHS ADDED (monthly stage-file client
# not yet delivering) → shows "Inactive".
# 2026-08-14 per user: HealthNetCA REACTIVATED — "loaded backfill of 3/20 to 3/27
# data, which will be certified today. Please remove 'Inactive' from HealthNetCA
# and put (3/20-3/27) next to name." Removed from FORCED_INACTIVE (its Monday
# cells now render real load/cert state) + added to AUTO_INACTIVE_EXCLUDE so the
# RAMP auto-sweep can't re-flag it while the 0100/0110 jobs are still disabled.
# The backfill window is shown via CLIENT_DISPLAY_NAME → "HealthNetCA (3/20-3/27)".
FORCED_INACTIVE = {"MedicalMutualMHS"}

# Date-gated inactivation: {client: cutoff_date}. Cells on/after the cutoff show
# "Inactive"; cells BEFORE the cutoff render their normal history (cert dates /
# ✓ / etc.). Unlike FORCED_INACTIVE (which is always-on for the whole client and
# only preserves past-day markers before *today*), this pins the exact date a
# client stopped delivering while keeping the earlier real record intact.
# TuftsMedPref & Tufts_Audit_CIT (both weekly Mon) → Inactive 7/6/26 forward per
# user 2026-07-29 (deliveries stopped as of that Monday; earlier weeks keep their
# history). Takes precedence over the cert lookup in resolve_marker.
FORCED_INACTIVE_FROM = {
    "TuftsMedPref":    date(2026, 7, 6),
    # Tufts_Audit_CIT REMOVED 2026-08-20 per user: it is loading again, catching
    # up the weeks of 7/6 through 8/17 (all seven Mondays pinned "L" in
    # MANUAL_OVERRIDES; they will all certify on the same date). The 7/6 cutoff
    # would have overridden those cells with "Inactive".
    # Oscar (weekly Wed) → Inactive 7/22/26 forward per user 2026-07-29. Earlier
    # cells keep their history (incl. the 7/1/7/8/7/15 → 7/15 cert overrides).
    "Oscar":           date(2026, 7, 22),
    # TuftsRx → Inactive 8/17/26 forward per user 2026-08-19: deliveries stop for
    # an implementation of HarvardPilgrim data. TuftsRx has BOTH a weekly Monday
    # row and a MONTHLY row, so the cutoff is honored in resolve_marker (weekly
    # cells 8/17 onward) and in determine_monthly (Sept 2026 forward; the August
    # monthly cell anchors to the 10th, before the cutoff, so it keeps its real
    # cert/L state). Everything through Mon 8/10 keeps its history.
    "TuftsRx":         date(2026, 8, 17),
}

# Clients whose load is running but snap step is disabled in RAMP — show
# marker "Snap" with pink shading on the expected delivery day. Mechanism kept
# wired for future use; Kaiser_AmbM removed 2026-06-08 (snap re-enabled).
SNAP_DISABLED_CLIENTS = set()

# For certain clients, the "L" (currently loading) indicator should only fire
# when a Ready/Running job's JobName contains one of the listed substrings.
# Ancillary jobs (COBC, IHP, ABII, etc.) running do NOT mean the client is
# loading. Per user 2026-05-18: "A WellCareRx COBC job ran, but that is not
# the indicator for Loading. We only want to look for MasterLoad or Claims Load."
LOAD_NAME_REQUIRED = {
    # AetnaRx: only the main "Claim 0110 Split Load" / "Claim 0120 Load" /
    # "Claim 0130 Start Snap" / "MasterLoad" steps count. Ancillary
    # "Claim 0150 RTA Load", "0132 ETL4 O20 Load", "IHP", "COBC" etc.
    # running Ready does NOT count as the client currently loading.
    # "claim 0130" added 2026-05-20: when the 0130 Start Snap completes
    # in early morning (attribution date = today, after the 0120 Load
    # finishes overnight on yesterday's date), today's cell gets ✓ from
    # the snap step. Per user: "AetnaRx should be a checkmark for today."
    "AetnaRx":           ("claim 0110", "claim 0120", "claim 0130", "masterload"),
    # BCBSFL (weekly Tue): only 'BCBSFL 0110 Claims Load' counts — not CMS
    # Referral Load, Claims Stage, or Claims Start Snap.
    "BCBSFL":            ("claims load",),
    # TuftsRx: only the 'Tufts RX Claims *' jobs are the delivery. Added
    # 2026-08-04 per user — a running 'Tufts RX Elig 0110 Load' is an
    # eligibility file, not the claims delivery, and was painting "L".
    "TuftsRx":           ("claim",),
    # MMOH (WC) monthly: only 'MMO 0110 WC Load' is the load indicator (per
    # user 2026-06-08 correction — this row is the Workers' Comp load). The
    # "wc load" keyword matches it but not the WC Stage step or any other
    # MMOH* job (MMOH Claims / GEN / ICD10 / MMOHRx...).
    "MMOH":              ("wc load",),
    # MedicalMutualOH (monthly, cert-only): only 'MMOH Claims 0110 Load' counts
    # (plural "Claims" — distinct from the MMOHRx monthly/weekly singular).
    "MedicalMutualOH":   ("claims 0110 load",),
    # MMOHRx weekly Tue: only 'MMOHRx Weekly Claim 0110 Load' counts —
    # filter excludes the Monthly Claim Stage and Weekly Claim Stage.
    "MMOHRx":            ("weekly claim 0110 load",),
    # MMOHRxMonthly: the 'MMOHRx Monthly Claim' Load (0110) and Snap (0120)
    # steps. "monthly claim" matches both (excludes Weekly Claim / COBC); the
    # Stage step is dropped by is_loading_today's stage filter, and ✓ comes
    # from the Snap step via SNAP_KIND_ONLY_CLIENTS.
    "MMOHRxMonthly":     ("monthly claim",),
    # CignaRx (weekly Tue): only 'Cigna RX 0110 Load' counts. COBC Load,
    # Daily PassFile, and other ancillary jobs share the "cignarx" matching
    # prefix and would otherwise trip the L indicator. Per user 2026-05-19:
    # "'Cigna RX 0110 Load' is not running, so CignaRx should not have an 'L'."
    # Pattern "rx 0110 load" matches "Cigna RX 0110 Load" but not
    # "Cigna RX COBC 0110 Load" (the "COBC" between "RX" and "0110" breaks
    # the substring).
    "CignaRx":           ("rx 0110 load",),
    # ElevanceMMMRx (daily): only the Masterload load + snap steps signal
    # delivery. Excludes the ancillary 'ElevanceMMMRx COBC 0110 Load' (and the
    # 0100 Stage / 0130 Post Snap) from both is_loading_today and the ✓ lookup.
    # Per user 2026-06-16.
    "ElevanceMMMRx":     ("masterload 0110 load", "masterload 0120 snap"),
    # WellCareRx: narrowed to the Masterload Load step 2026-06-08 (was
    # "masterload","claim") so 'WellCareRx Masterload 0100 Stage' no longer
    # trips L via the snap-index activity path. Delivery = Masterload 0110 Load.
    # 2026-08-06: added the 0120 Snap step (same shape as ElevanceMMMRx above) so
    # is_loading_today keeps the cell at "L" while the snap is still running — a
    # cert client is supposed to hold L through BOTH steps until certification,
    # but the load-only keyword cleared it the moment 0110 finished. Still excludes
    # 0100 Stage / 0130 Post Snap / 0140 MINE Snap and the COBC + ABII jobs.
    "WellCareRx":        ("masterload 0110 load", "masterload 0120 snap"),
    # BCBSSC (monthly, ~20th): delivery = 'BCBSSC 0110 Load'. build_snap_index
    # indexes Stage jobs with kind="load", so a Ready/Resolved 'BCBSSC 0100
    # Stage' was satisfying load_this_month and painting "L" on the expected day
    # while nothing had actually loaded. Per user 2026-08-03: "BCBSSC is not
    # loading, just staging." The full prefix "bcbssc 0110 load" also excludes
    # the sibling 'BCBSSC RX 0110 Load' (BCBSSCRx is a separate client that the
    # "bcbssc" substring key otherwise matches).
    "BCBSSC":            ("bcbssc 0110 load",),
    # BCBSNorthCarolinaFEP (monthly, ~10th): delivery = 'BCBSNC FEP CareFirst
    # 0110 Load' (5/10, 6/10, 7/10 → certs 5/11, 6/10, 7/10). The
    # 'BCBSNC FEP Daily 0100 Stage/0110 Load' pair is a separate DAILY feed that
    # succeeds every night, so load_this_month was always true and the cell sat
    # at "L" until the monthly cert landed. Per user 2026-08-03:
    # "BCBSNorthCarolinaFEP is also not Loading." ('BCBSNC FEP MHS 0110 Load'
    # is dormant — no runs since May; add "mhs 0110 load" here if it resumes.)
    "BCBSNorthCarolinaFEP": ("carefirst 0110 load",),
    # BCBSARRx: delivery = 'BCBSARRx MasterLoad 0110 Load'. Narrowed 2026-06-12
    # so the ancillary 'BCBSARRx COBC 0110 Load' failure no longer trips a
    # "Load Failure" (nor L) for the claims cycle. Per user: the failure is COBC,
    # not claims.
    "BCBSARRx":          ("masterload 0110 load",),
    # OscarRx: main load is 'Oscar RX 0110 Load' (no "claim"/"masterload"
    # in the name). Added "rx 0110 load" 2026-05-20 per user:
    # "Oscar Rx 0110 Load is running and was not picked up."
    "OscarRx":           ("masterload", "claim", "rx 0110 load"),
    # CenteneRx: narrowed to the Claims Load step 2026-06-08 — bare "claim"
    # matched 'Centene RX 0130 Claims Stage'. Delivery = '0140 Claims Load'.
    "CenteneRx":         ("claims load",),
    # CenteneFidelisRx: only 'Centene Fidelis Rx 0130 Claims Load' (and the
    # MasterLoad variant) should drive L. Narrowed from "claim" to "claims
    # load" 2026-06-08 — the bare "claim" matched 'Centene Fidelis Rx 0120
    # Claims Stage', which (as a completed entry in the snap index) tripped the
    # cert-client "activity -> L" path. The is_loading_today stage filter
    # already excluded a Ready/Running stage, but the snap-index activity path
    # had no stage guard, so the narrower keyword is the fix.
    "CenteneFidelisRx":  ("masterload", "claims load"),
    # CenteneFidelis (Medical) 2026-08-19 per user: "CenteneFidelis Medical is
    # not Loading, it is currently Staging." Same fix as the Rx sibling — the
    # completed 'Centene Fidelis Medical 0110 Eligibility Load' sat in the snap
    # index as a "load" kind entry, so the cert-client "activity this week -> L"
    # path painted an L while only 'Medical 0120 Claims Stage' was Ready.
    # Delivery = '0130 Claims Load' (or a MasterLoad variant).
    "CenteneFidelis":    ("masterload", "claims load"),
    # AetnaQNXT / AetnaQNXTRx: narrowed to the Masterload Load step 2026-06-08
    # so 'Masterload 0100 Stage' no longer trips L via the activity path.
    "AetnaQNXTRx":       ("masterload 0110 load",),
    "AetnaQNXT":         ("masterload 0110 load",),
    # WellpointEdwardRx: narrowed to the two Claims Load steps 2026-06-08 so
    # 'Claims 0100 Stage' / 'Claims HealthSun 0120 Stage' no longer trip L.
    "WellpointEdwardRx": ("claims 0110 load", "claims healthsun 0130 load"),
    # WellpointRxElig: only the Elig pipeline counts. "rx elig" matches BOTH
    # 'Wellpoint RX Elig 0110 Load' (→ L) and 'Wellpoint RX Elig 0120 Snap'
    # (→ ✓ via the snap-index path), but excludes the Claims/COBC jobs.
    "WellpointRxElig":   ("rx elig",),
    # HAPRx (per user 2026-06-04): only the main Claims load counts —
    # `HAPRx 0110 Load` and any future `HAPRx Masterload 0110 ...`. The
    # substring `haprx 0110 load` matches the main load but NOT the COBC /
    # TPLCov variants (`HAPRx COBC 0110 Load`, `HAPRx TPLCov 0110 Load`)
    # because the modifier breaks the substring — same trick as CignaRx.
    "HAPRx":             ("haprx 0110 load", "haprx masterload 0110"),
    # ElixirRx and PremeraMedAdvRx (per user 2026-06-04): only Claim Load
    # jobs trigger L; snap / mine / SFTP jobs must not. Patterns match
    # `Premera MedAdv Rx Claims 0110 Load` and any future Elixir Claims
    # 0110 Load variant.
    "ElixirRx":          ("claim 0110 load", "claims 0110 load", "masterload"),
    "PremeraMedAdvRx":   ("claim 0110 load", "claims 0110 load", "masterload"),
    # BCBSNC (per user 2026-06-05): only the main `BCBSNC Claims 0110 Load`
    # counts. Excludes CAQH, COBC, Claris Health, MSPI, Rx, Daily Passfile —
    # all of which have their own 0110 Load lines.
    "BCBSNC":            ("bcbsnc claims 0110 load",),
    # AetnaRCE / NCStateAetna (both DAILY, driven by the shared 'Aetna RCE 310
    # ETL Load'). Added 2026-08-13: EVERY 'Aetna RCE *' job collapses to the same
    # snap-index key "aetnarce", and build_snap_index stores non-snap steps with
    # kind="load" — so a Successful 'Aetna RCE 300 ETL Stage' (or a Resolved
    # 'Aetna RCE 400 Daily Snap' card, or 'Aetna RCE 210 CM9 Load') painted a
    # false "✓" on the day's cell even though the ETL Load had not started. Per
    # user 2026-08-13: "the 8/13/26 load for AetnaRCE & NCStateAetna have not
    # started yet." The delivery signal for both clients is the 310 ETL Load
    # only; NCStateAetna also keeps its own weekly delivery snap.
    "AetnaRCE":          ("310 etl load",),
    "NCStateAetna":      ("310 etl load", "ncstateaetna 0110 snap"),
    # NCState (MONTHLY, per user 2026-08-18: "NCState did not Load, the Stage job
    # was kicked off, but no files were found"). Same bug class as BCBSSC:
    # build_snap_index indexes 'NC State 0100 Stage' with kind="load", so the
    # Resolved 3-second Stage card (QueueId 1425026, 8/17 12:46) satisfied
    # load_this_month and determine_monthly step 4b painted "L" on the expected
    # (8/14) cell while nothing had loaded. Restricting to the real load job means
    # the cell reads "No Data" until 'NC State 0110 Load' actually runs — and
    # flips to "L" on its own the moment it does. NCState's ✓ is unaffected: it
    # comes from the RAMP snap-endpoint (kind="snap"), which _load_name_allowed
    # bypasses for SNAP_ONLY_CLIENTS. The full 'nc state' prefix also keeps the
    # NCStateRx sibling jobs out (see CLIENT_PRIMARY_KEY_OVERRIDE).
    "NCState":           ("nc state 0110 load",),
    # WebTPA (weekly Fri) 2026-08-24 per user: "WebTPA is not loading for 8/28
    # delivery. The 8/21/26 is Staging, but waiting for missing files from the
    # client." Same bug class as CenteneFidelis / NCState: build_snap_index
    # indexes 'WebTPA 0100 Stage' with kind="load", so the Resolved stage card
    # (QueueId 1428143, 8/24 09:32 — with 1429761 still Ready) satisfied
    # snap_in_week and the cert-client "activity this week -> L" path painted an
    # "L" on the 8/28 cell while nothing had loaded. Delivery = 'WebTPA 0110
    # Load'; the cell flips back to "L" on its own when that job runs. Also
    # excludes the 'WebTPA MFT Logfile' cards.
    "WebTPA":            ("webtpa 0110 load",),
    # Cambia (weekly Mon) 2026-08-24 per user: "Cambia is not Loading, just
    # Staging." Same bug class — 'Cambia 0120 PassFile Load' (Successful 8/23)
    # and 'cambia LogFile' land in snap_idx as kind="load" and painted an "L"
    # while only 'Cambia 0100 Claims Stage' was Ready. Delivery = the Claims Load.
    "Cambia":            ("claims load",),
}

# Soft, self-clearing cell labels — {(client, day): label}. Unlike
# MANUAL_OVERRIDES (which always wins), a soft label renders ONLY while the
# auto-resolved marker is still empty ("" / "No Data"). The moment real activity
# appears — "L" when the load starts, "✓" when it lands, or a cert date — the
# live marker takes over and the label disappears with no manual cleanup.
#
# 2026-08-13 per user: the 8/13 AetnaRCE / NCStateAetna load had not started —
# both are behind, held up by the monthly 'Aetna 0110 Subro Load' (Ready since
# 06:18). 'Aetna RCE 300 ETL Stage' finished 08:38 but the 310 ETL Load has not
# begun, so the cells would otherwise be blank.
SOFT_OVERRIDES = {
    ("AetnaRCE",     date(2026, 8, 13)): "Delayed",
    ("NCStateAetna", date(2026, 8, 13)): "Delayed",
    # 2026-08-19 per user: CenteneFidelis Medical's Wednesday cell is not
    # loading yet — 'Centene Fidelis Medical 0120 Claims Stage' is still Ready.
    # Self-clears the moment the Claims Load lands / the week certifies.
    # 2026-08-24 per user: Cambia (weekly Mon) is staging, not loading —
    # 'Cambia 0100 Claims Stage' Ready since 8/22, no Claims Load yet.
    ("Cambia", date(2026, 8, 24)): "Staging",
    # 2026-08-24 per user: "CenteneFidelis for 8/26/26 is loading Elig and will
    # move to Claims next." LOAD_NAME_REQUIRED restricts CenteneFidelis to the
    # Claims Load, so the running Eligibility Load can't paint L on its own.
    ("CenteneFidelis", date(2026, 8, 26)): "L",
    # 2026-08-24 per user: EDW_WGS (monthly, expected the 20th) is in a load
    # failure — 'Wellpoint 0100 EDW Pull EDW_WGS' Failed 8/21 20:48 -> 8/23 08:56.
    # has_recent_failure only matches LOAD-named jobs, and the EDW feeds use the
    # verb "Pull", so the failure never surfaced on its own.
    ("EDW_WGS", date(2026, 8, 20)): "Load Failure",
    # (2026-08-24 WebTPA 8/21 "Staging" removed 2026-08-25: the missing client
    # files arrived, 'WebTPA 0110 Load' ran 8/25 11:02 -> 11:29 and DHT certified
    # 8/25 11:57 — the 8/21 cell now carries that cert date via MANUAL_OVERRIDES.)
}

# FILE-GATED overrides — {(client, day): (marker, directory, filename_glob)}.
# The marker is pinned on the cell for as long as the gate file is ABSENT from
# `directory`; the moment a file matching `filename_glob` appears there the entry
# stops applying and the cell resolves normally. Self-clearing, so unlike a
# MANUAL_OVERRIDES pin it never needs to be deleted by hand.
#
# Use this whenever the user says "keep showing X until <file> loads" — the
# client's landing folders already record load progress (e.g. the AetnaHRP claims
# pipeline moves each extract from Claim\ToLoad → Claim\Load → Claim\Loaded, or
# → Claim\BadFile on a reject), so "did it load?" is a directory test.
#
# 2026-08-18 per user: "AetnaHRP for 8/17 is also a Load Failure and should
# remain as such until the \\ETL2\Clients\AetnaHRP\Data\Claim\
# VENDOR.CB-CLAIMS-EXTRACT.260816 file loads." The 8/17 'Aetna 0110 HRP Load'
# (QueueId 1425253) FAILED 18:34→22:36, the 0100 Stage retry failed 22:39 and the
# 'Aetna HRP Delivery Ticket' job failed 8/18 07:51 — but the 8/17 Monday cell was
# reading "✓" because its Sat/Sun lookback found the 8/16 05:18 'Aetna 0120 HRP
# Snap' (which belongs to the PREVIOUS file, the .260815 extract now sitting in
# Claim\Loaded). Gate on the .260816 extract reaching Claim\Loaded.
FILE_GATED_OVERRIDES = {
    ("AetnaHRP", date(2026, 8, 17)): (
        "Load Failure",
        r"\\ETL2\Clients\AetnaHRP\Data\Claim\Loaded",
        "VENDOR.CB-CLAIMS-EXTRACT.260816*",
    ),
}

# Cache so one report run does a single directory probe per (dir, pattern) —
# these are UNC paths and every month tab calls place() for the same cell.
_FILE_GATE_CACHE = {}


def file_gate_satisfied(directory, pattern):
    """True once a file matching `pattern` exists in `directory`.

    An unreachable/missing directory counts as NOT satisfied, so a network
    hiccup leaves the pinned marker in place rather than silently clearing it
    (the safe direction — the cell keeps reporting the known-bad state).
    """
    key = (directory, pattern)
    if key not in _FILE_GATE_CACHE:
        try:
            hit = bool(globmod.glob(os.path.join(directory, pattern)))
        except OSError:
            hit = False
        if not hit and not os.path.isdir(directory):
            print(f"[warn] file gate: directory unreachable — {directory}")
        _FILE_GATE_CACHE[key] = hit
    return _FILE_GATE_CACHE[key]

# Manual cell overrides — (client, scheduled_date) → marker. Marker can be:
#   - a date object (rendered as MM/DD/YY)
#   - one of the marker strings: "✓", "L", "No Data", "Load Failure",
#     "Inactive", "Deployment", "" (blank)
# Use sparingly — only for one-off corrections that the data sources can't
# express on their own (e.g. retroactively assigning a cert date to a Friday
# cell, or marking a known deployment-blocked Wednesday).
MANUAL_OVERRIDES = {
    ("AetnaHRP",  date(2026, 5, 1)): "✓",
    ("WebTPA",    date(2026, 5, 1)): "No Data",
    ("CenteneRx", date(2026, 5, 1)): date(2026, 5, 5),
    ("CenteneRx", date(2026, 5, 8)): date(2026, 5, 5),
    ("Medica",    date(2026, 5, 6)): "Deployment",
    # 2026-05-20: CignaFacets 5/12 Tue cycle certified 5/19 (Mon, outside the
    # default Mon-Fri 5/11-5/15 backward window). Per user: "missing past
    # dates … CignaFacets on 5/12/26." Pin the late cert explicitly.
    ("CignaFacets", date(2026, 5, 12)): date(2026, 5, 19),
    # 2026-05-21: Premera certified 5/20 14:32 (Wed, one day before its Thu
    # scheduled cycle). Forward cert window 5/21-5/27 misses it; pin the
    # cert date explicitly. Per user: "Premera Load was certified on 5/20."
    ("Premera",      date(2026, 5, 21)): date(2026, 5, 20),
    # 2026-05-28: ExcellusRx 5/20 finished loading and certified today;
    # pin 5/28 as the cert date for the 5/20 cycle cell.
    ("ExcellusRx",   date(2026, 5, 20)): date(2026, 5, 28),
    # 2026-05-29: ExcellusRx 5/27 cycle loaded and certified today.
    ("ExcellusRx",   date(2026, 5, 27)): date(2026, 5, 29),
    # 2026-05-28: CenteneRx 5/22 cycle certified today alongside the 5/29
    # load job. Pin 5/28 cert date for the 5/22 cell. (5/29 cell will pick
    # up the cert naturally via DHT detection.)
    ("CenteneRx",    date(2026, 5, 22)): date(2026, 5, 28),
    # 2026-05-28: KaiserSCPareo certified today (Thu). Pin the cert date
    # explicitly so the cell shows the cert regardless of when DHT indexes.
    ("KaiserSCPareo", date(2026, 5, 28)): date(2026, 5, 28),
    # 2026-05-28: Premera Commercial 0110 Load finished + certified today.
    ("Premera",       date(2026, 5, 28)): date(2026, 5, 28),
    # 2026-06-01: AetnaRx 6/1 was a dupe of 5/31 files; flipped to ✓ on
    # 2026-06-05 once the associated data landed.
    ("AetnaRx",       date(2026, 6, 1)): "✓",
    # 2026-06-03: AetnaRx Snap failed — staff will manually correct. Force ✓
    # for today; remove this override once the next day's run picks up the
    # corrected snap completion naturally.
    ("AetnaRx",       date(2026, 6, 3)): "✓",
    # Kaiser_MASTapestry / KaiserSCPareo blank-until-cert overrides cleared
    # 2026-06-04 — both certified, auto-detection now surfaces the date.
    # (2026-06-04 TuftsRx Mon 6/1 → "Inactive" override removed 2026-07-28 when
    # TuftsRx was reactivated — the 6/1 claims file loaded, so the weekly cell now
    # earns its ✓ via the STAGE_FILE_CELL_CLIENTS fallback like the other Mondays.)
    # 2026-06-08: BCBSFL weekly delivery skipped this week (one-off). Show
    # "Skip" on the 6/9/26 Tuesday cell only. Remove after this week.
    ("BCBSFL",        date(2026, 6, 9)): "Skip",
    # 2026-06-09: CignaFacets certified last week per user (TapeID 3593,
    # [dbo].[vwMiningCache_Full], CertTimestamp 6/2). DHT.TableList row is still
    # at "Email sent, Ready for Certification review" (not "Certified"), so the
    # auto cert lookup misses it and the 6/2 cell flagged with "!". Pin the
    # 6/2/26 cert date. Remove once DHT flips the 3593 row to Certified (then
    # cert_in_week surfaces it automatically).
    ("CignaFacets",   date(2026, 6, 2)): date(2026, 6, 2),
    # 2026-06-16: CignaFacets 6/9 cycle (TapeID 3598) was Certified and shown as
    # 6/9/26, but DHT reverted it to "Email sent, Ready for Certification review"
    # (a known client issue), so the live cert lookup misses it and the cell went
    # to a pink "!". Pin the cert date back. Also seeds the sticky-cert cache so
    # the regression can't recur. Per user: "Once a client gets certified, do not
    # change the cell to an '!'." Remove once DHT flips 3598 back to Certified.
    ("CignaFacets",   date(2026, 6, 9)): date(2026, 6, 9),
    # 2026-07-28: BCBSARRx (weekly Tue) certified all four May weeks (5/5, 5/12,
    # 5/20, 5/26) but the 5/5 cert carries an ANOMALOUS StatTimestamp of 5/22 in
    # DHT, so cert_in_week piled it onto the 5/18/5/25 weeks and left 5/12 with a
    # pink "!" and 5/19 showing the wrong 5/26 date. Pin both Tuesday cells to
    # their real cert dates. (June's 6/8-week "!" is a GENUINE no-cert gap — DHT
    # jumps 6/1→6/16 — so it is left flagged.)
    ("BCBSARRx",      date(2026, 5, 12)): date(2026, 5, 12),
    ("BCBSARRx",      date(2026, 5, 19)): date(2026, 5, 20),
    # WellCare / WellCareRx delivery tracking (updated 2026-06-24):
    #  - WellCare Medical: 6/12 cell certified 6/17; 6/19 cell certified 6/18.
    #  - WellCareRx loaded 6/12+6/19 together and CERTIFIED 6/24 (per user; DHT
    #    6/24 cert covers StatTimestamp weeks 6/19 + 6/24). Per user, pin 6/24 on
    #    the 6/12, 6/19 AND 6/26 cells — the combined load's 6/12 week has no own
    #    StatTimestamp so 6/12 needs the explicit pin; 6/19 & 6/26 pinned too.
    ("WellCare",      date(2026, 6, 12)): date(2026, 6, 17),
    ("WellCare",      date(2026, 6, 19)): date(2026, 6, 18),
    ("WellCareRx",    date(2026, 6, 12)): date(2026, 6, 24),
    ("WellCareRx",    date(2026, 6, 19)): date(2026, 6, 24),
    ("WellCareRx",    date(2026, 6, 26)): date(2026, 6, 24),
    # 2026-08-03: the normal weekly WellCareRx deliveries for 7/24 and 7/31 both
    # CERTIFIED today (DHT WellCareRx cert 8/3 14:57, StatTimestamps 7/25 + 7/30
    # — one cert covering both weeks). Per user: "the WellCareRx certified today
    # goes to the 7/24 & 7/31 dates." Pin the 8/3 cert date on both Friday cells
    # — the 7/24 cell's Mon-Fri window (7/20-7/24) can't reach the Sat 7/25
    # StatTimestamp on its own. Supersedes the 7/31 "L" pins. (The Ad Hoc reload
    # certified 7/22 still sits on its own labeled row below.)
    ("WellCareRx",    date(2026, 7, 24)): date(2026, 8, 3),
    ("WellCareRx",    date(2026, 7, 31)): date(2026, 8, 3),
    # 2026-06-24: CenteneRx/OscarRx 6/22 certs land on the 6/19 cell
    # automatically via the StatTimestamp system. The 6/26 blank overrides were
    # REMOVED 2026-06-24 — OscarRx & CenteneRx are now LOADING their 6/26
    # delivery and must show "L" (the blanks were hiding it).
    # 2026-06-29: Centene Medical (weekly Tue) was rendering "Load Failure" on
    # its 6/30 cell, but per user it was NOT failed — it had Snapped and was
    # ready for certification. Pinned to "L" while snapped-not-yet-certified.
    # REMOVED 2026-06-30: DHT cert landed 6/29 09:33, so the 6/30 cell now
    # surfaces the cert date automatically (week of 6/29 contains the cert).
    # 2026-07-02: Premera certified 7/1 14:55 (Wed, one day before its Thu 7/2
    # cell). Forward window for 7/2 [7/2-7/8] misses it, and the prior Thu 6/25
    # window [6/25-7/1] would otherwise grab the 7/1 cert and overwrite 6/25's
    # own 6/25 cert. Pin each cell to its own cert date. Per user: "Premera was
    # certified on 7/1/26."
    ("Premera",      date(2026, 6, 25)): date(2026, 6, 25),
    ("Premera",      date(2026, 7, 2)):  date(2026, 7, 1),
    # 2026-07-06: CenteneRx 7/3 (Fri) loaded and certifies today (7/6). The cert
    # lands on Monday 7/6 — OUTSIDE the 7/3 cell's Mon-Fri window (6/29-7/3) — so
    # the auto backward-window lookup can't reach it (CenteneRx has no forward
    # CERT_DIRECTION). Pin the 7/6 cert date on the 7/3 cell. Per user 2026-07-06.
    ("CenteneRx",    date(2026, 7, 3)):  date(2026, 7, 6),
    # 2026-07-07: Aetna RCE 310 ETL Load (JobId 2257) Failed at 00:59; pinned
    # "Load Failure" on the 7/7 daily cells for AetnaRCE + NCStateAetna (same
    # Aetna RCE ETL). 2026-07-08: per user the 7/7 load has since completed and
    # Snapped, so the overrides were removed — the 7/7 cells now resolve to ✓
    # via auto snap detection.
    # 2026-07-08: HarvardPilgrim (weekly Wed) — the current load processed the
    # 6/30 data file, which is the 7/1 delivery running late (NOT a 7/8 one).
    # Per user: "the HarvardPilgrim load was for file 6/30, which will be the 7/1
    # load." Pin "L" on 7/1 (loaded, awaiting cert) to clear the pink "!", and
    # blank the phantom current-week L on 7/8 (one weekly delivery, not two).
    # Replace 7/1 with the cert date once it certifies.
    # 2026-07-09: 'HarvardPilgrim Claims 0110 Load' is now running for the 7/8
    # weekly delivery — flip the 7/8 cell from blank to "L". Per user.
    # 2026-07-10: both deliveries loaded AND certified (per user). Two Claims
    # 0110 Loads ran: Load A 7/7 13:35->7/8 00:02 (the late 6/30 file = 7/1
    # delivery) certified 7/8 10:46 (StatTimestamp 7/7); Load B 7/9->7/9 18:45
    # (the 7/8 weekly delivery) certified 7/10 09:03 (StatTimestamp 7/9). Both
    # StatTimestamps fall in the 7/6 week so CERT_WEEK_IDX collapses them onto
    # one cell — pin each cert date explicitly to keep the two deliveries split.
    ("HarvardPilgrim", date(2026, 7, 1)): date(2026, 7, 8),
    ("HarvardPilgrim", date(2026, 7, 8)): date(2026, 7, 10),
    # 2026-07-08: HMSA_Rx (weekly Tue) received an EMPTY file this week — a known
    # occasional occurrence per user. Mark the 7/7 Tuesday cell "Empty" (pink,
    # matching the team's ALL_CLIENTS_ALERT_MARKERS convention). Remove next week.
    ("HMSA_Rx",       date(2026, 7, 7)): "Empty",
    # 2026-07-29: HMSA_Rx (weekly Tue) 7/28 file is EMPTY again (per user). Mark
    # the 7/28 Tuesday cell "Empty" with NO pink — alert_state returns False for
    # "Empty". Remove next week.
    ("HMSA_Rx",       date(2026, 7, 28)): "Empty",
    # 2026-07-31: AetnaRCE & NCStateAetna (both DAILY, driven by the shared
    # 'Aetna RCE 310 ETL Load'). The 7/28 re-send SUCCEEDED (RAMP QueueId 1411504,
    # 7/28 13:38 -> Successful 16:33) → keep "✓" on 7/28 (both clients).
    # 2026-08-02: the resumed loads all landed and snapped, so the 7/29 / 7/30 /
    # 7/31 "L" pins were REMOVED — those cells now auto-resolve to "✓" from the
    # RAMP queue (7/29 08:47->11:05 Resolved, 7/30 09:10->09:53 Successful +
    # 10:34->13:42 Resolved, 7/31 10:40->8/1 01:25 Successful; "Resolved" counts
    # as success). Certification for the 7/31 delivery date runs 8/3 — see
    # CERT_CELL_REMAP, which lands that cert on the 7/31 cell instead of 8/3.
    ("AetnaRCE",      date(2026, 7, 28)): "✓",
    ("NCStateAetna",  date(2026, 7, 28)): "✓",
    # 2026-07-30: CVSPBMRx (weekly Monday) — per user the 7/27 delivery is complete;
    # pin "✓" and lock it in on the 7/27 cell. The weekly RAW_MEMBR_ELIG_20260725
    # file isn't in tape yet (auto cvspbm_delivered can't fire), so this manual ✓
    # both shows the checkmark and locks the cell against regression.
    ("CVSPBMRx",      date(2026, 7, 27)): "✓",
    # 2026-08-20 per user: Tufts_Audit_CIT (weekly Mon) "is currently loading for
    # the weeks of 7/6 through 8/17 … mark all Monday's as Loading and they will
    # also get the same certification date." Removed from FORCED_INACTIVE_FROM
    # (was Inactive 7/6 forward) and every Monday in that catch-up range pinned
    # to "L". When the catch-up certifies, replace all seven values with that one
    # cert date.
    # 2026-08-25 per user: "Tufts_Audit_CIT for 7/6 to 8/17 should have 8/20 cert
    # date" — the catch-up certified, so all seven Mondays now carry 08/20/26.
    # (DHT logged CertTimestamp 8/21 11:39 with StatTimestamp 8/20 15:54; the
    # user's date is the delivery/cert date for the whole catch-up range. The
    # auto path can only ever place ONE cell per cert week, hence the pins.)
    # This week's Mon 8/24 cell is NOT pinned — it picks up its own 8/24 cert
    # (CertTimestamp 8/24 15:18, StatTimestamp 8/24) automatically.
    ("Tufts_Audit_CIT", date(2026, 7, 6)):  date(2026, 8, 20),
    ("Tufts_Audit_CIT", date(2026, 7, 13)): date(2026, 8, 20),
    ("Tufts_Audit_CIT", date(2026, 7, 20)): date(2026, 8, 20),
    ("Tufts_Audit_CIT", date(2026, 7, 27)): date(2026, 8, 20),
    ("Tufts_Audit_CIT", date(2026, 8, 3)):  date(2026, 8, 20),
    ("Tufts_Audit_CIT", date(2026, 8, 10)): date(2026, 8, 20),
    ("Tufts_Audit_CIT", date(2026, 8, 17)): date(2026, 8, 20),
    # 2026-08-20 per user: "The 8/17 for CVSPBMRx should be '!' and the Ad Hoc
    # for CVSPBMRx should be 'Load Failure'." The carry-over Ad Hoc card
    # (QueueId 1413143, 'CVS PBM RX 0110 Load', started 7/30 11:19) finally
    # FAILED 8/18 19:37. That ended the in-flight carry-over, so (a) the Ad Hoc
    # row disappeared and (b) has_recent_failure repainted the weekly 8/17 cell
    # "Load Failure" — but the failure is the backfill's, not the weekly cycle's.
    # Flag 8/17 as a plain miss ("!"), and carry the failure on its own Ad Hoc
    # row (see ADDITIONAL_ENTRIES, 8/18). Replace with ✓/cert once 8/17 lands.
    # 2026-08-24 per user: "CVSPBMRx for 8/17 has loaded and Snapped 8/23" —
    # 'CVS PBM RX 0110 Load' 8/21 16:01 -> 8/23 21:44 Successful, then
    # '0120 Start Snap' 8/23 21:44 -> 22:13 Successful. Was pinned "!" while the
    # Ad Hoc backfill failure was outstanding; flip to a delivered checkmark.
    ("CVSPBMRx",      date(2026, 8, 17)): "✓",
    # 2026-08-24 per user: "CenteneFidelis for 8/19/26 change to Missing in
    # Pink" — the 8/19 Wednesday delivery never arrived (was labelled "Staging"
    # on 8/19). "Missing" is a pink problem-state marker (see alert_state).
    ("CenteneFidelis", date(2026, 8, 19)): "Missing",
    # 2026-07-09: Oscar (weekly Wed) — was Inactive 7/1 & 7/8. 2026-07-15: per
    # user a backfill certified today (7/15) covering BOTH the 7/1 and 7/8
    # deliveries. 2026-07-17: per user the 7/15 cell certified 7/15 (show the
    # date); all future Oscar Medical dates are Inactive (Oscar re-added to FORCED_INACTIVE).
    ("Oscar",         date(2026, 7, 1)):  date(2026, 7, 15),
    ("Oscar",         date(2026, 7, 8)):  date(2026, 7, 15),
    ("Oscar",         date(2026, 7, 15)): date(2026, 7, 15),
    # 2026-07-09: AetnaHRP (daily) was in a Load Failure state — the 7/5, 7/6,
    # 7/7 data had not loaded, and the 7/8 certification did NOT include those
    # files. 2026-07-15: per user the backlog loaded and certified; the last-week
    # Load-Failure cells (7/6, 7/7, 7/9) now show ✓ (corrected from a 7/15 cert
    # date per user). The 7/8 cell keeps its own 7/8 cert.
    ("AetnaHRP",      date(2026, 7, 6)): "✓",
    ("AetnaHRP",      date(2026, 7, 7)): "✓",
    ("AetnaHRP",      date(2026, 7, 8)): date(2026, 7, 8),
    ("AetnaHRP",      date(2026, 7, 9)): "✓",
    # 2026-08-19 per user: mark AetnaHRP's 8/17 cell with a checkmark.
    ("AetnaHRP",      date(2026, 8, 17)): "✓",
    # 2026-07-29: TuftsMedPref (weekly Mon) → Inactive 7/6/26 forward per user
    # (deliveries stopped). Handled by FORCED_INACTIVE_FROM now; the old 7/13
    # cert / 7/6 "No Data" pins were removed so resolve_marker returns "Inactive".
    # 2026-07-13: BCBSAR (Medical, weekly Tue) — today's cert (7/13) covers BOTH
    # last week AND this week. Per user, mark last week's Tue 7/7 and this week's
    # Tue 7/14 cells with the 7/13 cert date.
    ("BCBSAR",        date(2026, 7, 7)):  date(2026, 7, 13),
    ("BCBSAR",        date(2026, 7, 14)): date(2026, 7, 13),
    # 2026-07-13: BCBSARRx (weekly Tue) — certified today (7/13). Per user, pin
    # last week's Tue 7/7 cell with the 7/13 cert date.
    ("BCBSARRx",      date(2026, 7, 7)):  date(2026, 7, 13),
    # (2026-07-17 Tufts_Audit_CIT 7/13 → "Inactive" override removed 2026-07-28:
    # Tufts_Audit_CIT reactivated per user — "the only clients still Inactive are
    # HealthNetCA & MedicalMutualMHS" — so this cell resolves normally again.)
    # 2026-07-17: WebTPA (weekly Fri) certified today (7/17); per user this cert
    # covers BOTH the 7/10 and 7/17 weeks. Pin the 7/17 cert date on each Friday
    # cell (the single cert would otherwise land on only one week).
    ("WebTPA",        date(2026, 7, 10)): date(2026, 7, 17),
    ("WebTPA",        date(2026, 7, 17)): date(2026, 7, 17),
    # 2026-08-25 per user: "WEBTPA for 8/21 should be 8/25, not the 8/28 cell."
    # The client's missing files finally arrived, 'WebTPA 0110 Load' ran 8/25
    # 11:02 -> 11:29 Successful and DHT certified 8/25 11:57 — but that cert's
    # StatTimestamp is 8/25, so stat_week_monday puts it in the 8/24 week and
    # cert_in_week would stamp it on THIS week's Friday (8/28) cell. It belongs
    # to the late 8/21 delivery. Pin it here, and see CELL_ACTIVITY_AFTER for the
    # 8/28 cell (which must ignore the 8/25 cert/load and wait for its own).
    ("WebTPA",        date(2026, 8, 21)): date(2026, 8, 25),
    # 2026-07-22: Kaiser_NW (Kaiser Pareo NW, weekly Thu) was rendering
    # "Load Failure" on its 7/23 cell — but per user it is CLEAR. The failure
    # is a spurious trailing card in RAMP: 'Kaiser Pareo NW 0110 Load' (JobId
    # 2160) logged a 4-second Failed (QueueId 1406222, 7/21 13:19:43) created
    # right AFTER two Resolved recovery cards (7/20 14:35, 7/21 13:17). Because
    # has_recent_failure isn't day-specific it pinned onto the only current-week
    # Kaiser_NW cell (Thu 7/23). Last full load certified 7/16; the 7/17 batch was
    # "Ready for Certification review". 2026-07-23: DHT certified Kaiser_NW today
    # (CertID 1297227-9, CertTimestamp 2026-07-23 07:29:51). Per user "Kaiser_NW
    # for today can be marked as certified" → show the 7/23 cert date on the Thu
    # cell. (The cert's StatTimestamp is 7/17, so cert_in_week attributes it to the
    # prior week's band; this override surfaces it on the current Thu 7/23 cell to
    # match the sibling Kaiser feeds. Remove once next week's cycle certifies.)
    ("Kaiser_NW",     date(2026, 7, 23)): date(2026, 7, 23),
    # 2026-08-10 per user: "the UPMC certification today will be for the 8/6
    # delivery." UPMC (weekly Thu) loaded 8/8 20:44-21:24 (QueueId 1419512) and
    # DHT sat at "Email sent, Ready for Certification review" (StatTimestamp
    # 8/8 21:13), leaving the 8/6 cell a pink "!". Pin today's cert date on 8/6.
    # (Once DHT flips to Certified this pin is redundant but harmless: the cert's
    # StatTimestamp 8/8 is a Saturday, so stat_week_monday -> 8/3 and cert_in_week
    # attributes it to this same Thu 8/6 cell. No CERT_CELL_REMAP needed.) The
    # separate load that started today 8/10 00:52 (QueueId 1420054, still Ready)
    # is the NEXT cycle and paints "L" on Thu 8/13 — unaffected.
    ("UPMC",          date(2026, 8, 6)):  date(2026, 8, 10),
    # 2026-07-29: HealthNetCA (weekly Mon) was FORCED_INACTIVE while its resumed
    # loads were under review, so no cell-by-cell "L"-suppression pins were needed.
    # 2026-08-14: HealthNetCA left FORCED_INACTIVE (3/20-3/27 backfill certified
    # today) — its Monday cells now resolve normally from cert/load activity, and
    # per user the dormant weeks that used to read "Inactive" should show a pink
    # "!" instead (no pins here). The claims date range for each backfill load is
    # attached to the label per-cell — see HEALTHNETCA_* below.
    #
    # 2026-08-24 per user: "HealthNetCA for 4/3-4/10 certified on 8/20/26,
    # 4/10-4/24 certified on 8/24/26." Four backfill weeks certified inside two
    # calendar weeks, so the automatic one-cert-per-week attribution can't place
    # them (cert_in_week keys off stat_week_monday(StatTimestamp), and every one of
    # these certs has a StatTimestamp in the 8/17 week — 8/19, 8/20, 8/20 23:24,
    # 8/21 — so CERT_WEEK_IDX[healthnetca][8/17] collapses to the LATEST cert,
    # 8/24, and the 8/24 cell gets nothing). Pin each cell to the cert that
    # belongs to it, by TapeID:
    #   3/27-4/3  TapeID 16097-16103 → cert 8/19 → 8/17 cell (label already pinned)
    #   4/3-4/10  TapeID 16104-16111 → cert 8/20 → its own labeled ADDITIONAL_ENTRIES
    #                                               row on 8/20 (only one Monday
    #                                               cell exists in that week)
    #   4/10-4/17 TapeID 16113-16119 ┐
    #   4/17-4/24 TapeID 16120-16127 ┘ both → cert 8/24 → 8/24 cell as "(4/10-4/24)"
    ("HealthNetCA",   date(2026, 8, 17)): date(2026, 8, 19),
    ("HealthNetCA",   date(2026, 8, 24)): date(2026, 8, 24),
}

# --- Cert-to-cell reattribution ---------------------------------------------
# {(client, cert_date): target_cell_day} — a DHT certification that RAN on
# `cert_date` actually covers the delivery on `target_cell_day`. cert_on_day()
# both (a) pulls the cert forward/backward onto the target cell and (b) hides it
# from the cell it would otherwise have landed on, so one cert can't paint two
# cells.
#
# Why: for the daily Aetna clients cert_on_day only widens a FRIDAY cell to pick
# up Sat/Sun certs — a cert that runs on the following MONDAY is out of reach and
# would instead stamp its date on the Monday cell (which belongs to the weekend
# ETL load and should read "✓").
#
# 2026-08-02 per user: "AetnaRCE & NCStateAetna will be certified on 8/3, which
# will be for the 7/31 delivery date." Until the cert lands, 7/31 shows "✓"
# (loaded + snapped) and 8/3 shows "✓" from the Sat/Sun ETL loads; once DHT
# certifies on 8/3 the 7/31 cell flips to 08/03/26 and 8/3 stays "✓". Safe to
# leave in place after the fact — it only ever matches that one cert date.
CERT_CELL_REMAP = {
    ("AetnaRCE",     date(2026, 8, 3)): date(2026, 7, 31),
    ("NCStateAetna", date(2026, 8, 3)): date(2026, 7, 31),
}

# --- Per-cell "ignore earlier activity" gate --------------------------------
# {(client, cell_day): cutoff_date} — for THIS cell only, any cert / load / snap
# activity dated on or before `cutoff_date` belongs to an EARLIER delivery, so it
# is ignored and the cell stays blank until something happens AFTER the cutoff.
#
# Why this and not a blank MANUAL_OVERRIDES pin: it is self-clearing. A pin of ""
# would also hide the cell's own future cert / "L" / "Load Failure" until someone
# deletes it by hand; the cutoff only masks the known-stale activity.
#
# Use it whenever a LATE delivery's load+cert land inside the NEXT delivery's
# week — the late cert gets pinned onto its own (earlier) cell via
# MANUAL_OVERRIDES, and the next cell needs to not double-report the same run.
#
# 2026-08-25 per user: "WEBTPA for 8/21 should be 8/25, not the 8/28 cell."
# WebTPA (weekly Fri) delivered 8/21 four days late — 'WebTPA 0110 Load' 8/25
# 11:02 -> 11:29 Successful, DHT cert 8/25 11:57 (StatTimestamp 8/25, i.e. the
# 8/24 week). Without the cutoff the 8/28 cell shows 08/25/26 from cert_in_week,
# or an "L" from the weekly cert-client "activity this week" fallback.
CELL_ACTIVITY_AFTER = {
    ("WebTPA", date(2026, 8, 28)): date(2026, 8, 25),
}

# --- Monthly cert-to-month reattribution ------------------------------------
# {(client, cert_date): (year, month)} — the monthly analogue of CERT_CELL_REMAP:
# a DHT certification that RAN on `cert_date` actually covers the delivery for a
# DIFFERENT month. latest_cert_in_month() both (a) counts the cert for the target
# month's tab (rendered on that month's expected day, since the cert date itself
# falls outside the month) and (b) HIDES it from the month it actually ran in, so
# one cert can't fill two month tabs.
#
# Why: a monthly delivery that arrives late certifies in the following calendar
# month. Without the remap, step 1 of determine_monthly attributes the cert to
# the month it ran in — leaving the real delivery month at "No Data" and painting
# a cert date on a month that hasn't delivered yet.
#
# 2026-08-05 per user: "BCBSSCRx will be certified today for the July delivery."
# The July file loaded 8/5 00:22, snapped 8/5 02:31 and certified 8/5 09:55 → the
# July tab shows 08/05/26 on its expected (18th-19th) cell and August stays open
# for its own delivery (held at "No Data" by MONTHLY_MONTH_MARKER_OVERRIDES so the
# July delivery's 8/5 load can't paint a premature "L" there).
MONTHLY_CERT_MONTH_REMAP = {
    ("BCBSSCRx", date(2026, 8, 5)): (2026, 7),
}

# --- Sticky certifications --------------------------------------------------
# Once a (client, scheduled-day) cell has rendered a real cert date, remember it
# so a later DHT status reversion (a row flipping from "Certified" back to e.g.
# "Email sent, Ready for Certification review") can't regress the cell to a
# blank / "No Data" pink "!". Per user 2026-06-16: "Once a client gets certified,
# do not change the cell to an '!'." Live certs always win (a fresh cert date
# overwrites the remembered one); the cache only fills in when the live lookup
# would otherwise leave the cell empty.
#
# LOCK-IN POLICY (per user 2026-07-28: "once past clients have been certified/
# Snapped and updated, lock in that information unless told otherwise"): the cache
# now remembers BOTH cert dates AND "✓" snap marks — the STRONGEST state a cell
# has ever shown (a cert date outranks a ✓). When a later run regresses a cell to
# blank/"No Data" (because the source data has aged out of the fetch window), the
# remembered marker is restored. Manual overrides are the "told otherwise" escape:
# they are never overwritten by a restore. Two key namespaces in the same JSON:
#   "client|YYYY-MM-DD"     → day-keyed (weekly/daily/kaiser/NYShip — stable day)
#   "M|client|YYYY-MM"      → month-keyed "day~marker" (monthly clients, whose
#                             placement day moves to the cert/snap date, so a
#                             day key can't be matched on the regressed run)
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CERT_STICKY_PATH = os.path.join(_SCRIPT_DIR, "cert_sticky_cache.json")
STICKY_CERTS = {}


def load_sticky_certs():
    global STICKY_CERTS
    try:
        with open(CERT_STICKY_PATH, "r", encoding="utf-8") as f:
            STICKY_CERTS = json.load(f)
    except (OSError, ValueError):
        STICKY_CERTS = {}
    return STICKY_CERTS


def save_sticky_certs():
    try:
        with open(CERT_STICKY_PATH, "w", encoding="utf-8") as f:
            json.dump(STICKY_CERTS, f, indent=0, sort_keys=True)
    except OSError as e:
        print(f"[warn] couldn't write sticky cert cache: {e}")


def _marker_rank(m):
    """Strength of a positive delivery marker: cert date (2) > "✓" snap (1) >
    everything else (0). Used so a remembered cert date is never clobbered by a
    later ✓, and a ✓ is never clobbered by a regression."""
    if isinstance(m, date):
        return 2
    if m == "✓":
        return 1
    return 0


def _stored_rank(s):
    """Rank of a stored (string) sticky value: "✓" → 1, any other non-empty
    string is treated as an ISO cert date → 2, empty/missing → 0."""
    if s == "✓":
        return 1
    if isinstance(s, str) and s:
        return 2
    return 0


def apply_sticky_cert(client, day, marker, alert, from_manual):
    """Lock in a certified/snapped cell (per user 2026-07-28). Remember the
    STRONGEST positive marker a cell has shown (cert date > "✓" snap); if a later
    run regresses the cell to blank/"No Data" because the source data aged out of
    the fetch window, restore the remembered marker (and clear the "!" alert).

    Manual overrides (`from_manual`) are the "told otherwise" escape: an
    intentional manual blank/No Data is left alone and never restored over. A
    live marker that is itself positive always flows through (so a ✓→cert-date
    upgrade shows), and only bumps the stored value UP in rank (never down).
    """
    key = f"{client}|{day.isoformat()}"
    # CELL_ACTIVITY_AFTER cells: a marker remembered from activity on/before the
    # cutoff belongs to an EARLIER delivery, so it must not be restored here —
    # otherwise the cache silently undoes the gate on every later run (this is
    # exactly what happened to WebTPA's 8/28 cell, cached from the run that saw
    # the 8/25 cert before the gate existed). A genuine post-cutoff marker is
    # re-stored immediately below via the live_rank path.
    cutoff = CELL_ACTIVITY_AFTER.get((client, day))
    if cutoff is not None:
        stored = STICKY_CERTS.get(key)
        if stored:
            try:
                stale = date.fromisoformat(stored) <= cutoff
            except ValueError:
                stale = True          # a "✓" earned by the pre-cutoff activity
            if stale:
                STICKY_CERTS.pop(key, None)
    live_rank = _marker_rank(marker)
    if live_rank > 0:
        if live_rank >= _stored_rank(STICKY_CERTS.get(key)):
            STICKY_CERTS[key] = marker.isoformat() if isinstance(marker, date) else marker
        return marker, alert
    if from_manual:
        return marker, alert
    if marker in ("", "No Data"):
        stored = STICKY_CERTS.get(key)
        if stored == "✓":
            return "✓", False
        if isinstance(stored, str) and stored:
            try:
                return date.fromisoformat(stored), False
            except ValueError:
                pass
    return marker, alert


def apply_sticky_monthly(client, day, marker, year, month, today, told_otherwise):
    """Month-keyed lock-in for MONTHLY clients, whose placement day moves to the
    cert/snap date — so the day-keyed cache can't be matched on the run where the
    cell regresses (which lands on the expected day instead). Remembers the
    strongest (day, marker) seen for (client, year, month) and restores it when a
    later run in the SAME-or-past month regresses to No Data/Inactive.

    `told_otherwise` = a manual monthly override / FORCED_INACTIVE governs this
    cell → don't restore over it. Returns (day, marker), possibly restored.
    """
    mkey = f"M|{client}|{year:04d}-{month:02d}"
    if _marker_rank(marker) > 0:
        cur = f"{day.isoformat()}~{marker.isoformat() if isinstance(marker, date) else marker}"
        prev = STICKY_CERTS.get(mkey, "")
        prev_marker = prev.split("~", 1)[1] if "~" in prev else ""
        if _marker_rank_str(marker) >= _stored_rank(prev_marker):
            STICKY_CERTS[mkey] = cur
        return day, marker
    if told_otherwise:
        return day, marker
    # Only restore for the current or a past month (never fabricate a future one).
    if (year, month) <= (today.year, today.month) and mkey in STICKY_CERTS:
        try:
            sd, sm = STICKY_CERTS[mkey].split("~", 1)
            rd = date.fromisoformat(sd)
            rm = "✓" if sm == "✓" else date.fromisoformat(sm)
            return rd, rm
        except (ValueError, KeyError):
            pass
    return day, marker


def _marker_rank_str(m):
    """_marker_rank for a value that may be a date or the string "✓"."""
    return 2 if isinstance(m, date) else (1 if m == "✓" else 0)


# ADO ticket IDs to hyperlink onto specific Load-Failure cells. Keyed by
# (client, day) — same convention as MANUAL_OVERRIDES. When a cell renders
# "Load Failure" AND has an entry here, the marker text becomes a clickable
# link to the TFS work item. Per user 2026-05-20: "For Load failures that
# have an ADO, like 954657 for Centene Medical, added as a link to the
# 'Load Failure' comment."
LOAD_FAILURE_ADO_LINKS = {
    ("ExcellusRx", date(2026, 5, 20)): 955578,  # 'Excellus - Rx - ExcellusRx 0110 Load'
    ("AetnaHRP",   date(2026, 5, 22)): 956353,  # 'Aetna 0110 HRP Load' failure
    # 2026-05-26: current-week Load-Failure links per user.
    ("CignaFacets", date(2026, 5, 26)): 956575,  # 'ProdSupp - Cigna - Audit/Subro - Cigna Facets 0110 Load Failure'
    ("ExcellusRx",  date(2026, 5, 27)): 955578,  # same active ExcellusRx 0110 Load story
    # Centene removed 2026-05-21 — 0110 Claims Load restarted (Ready 14:46),
    # is_loading_today now auto-returns "L".
    # MMOHRx removed 2026-05-21 — Weekly Claim 0110 Load finished.
}

# Monthly clients whose placement (day AND/OR marker) should be forced,
# overriding determine_monthly's auto-detection. Keyed by client. Used for
# one-off corrections: e.g. EDW feeds certified late but anchored to their
# expected day, or AetnaQNXT mid-cycle visible on a specific day.
# Value: (placement_date, marker). marker may be a date object (cert-style)
# or a string ("L", "Load Failure", "No Data", etc.).
MONTHLY_PLACEMENT_OVERRIDES = {
    # 2026-05-21: EDW feeds certified this morning at 07:54-07:56 but user
    # wants them anchored to 5/20 (their expected delivery day). Marker
    # shows the actual 5/21 cert date.
    "EDW_ASE":    (date(2026, 5, 20), date(2026, 5, 21)),
    "EDW_C_FAC":  (date(2026, 5, 20), date(2026, 5, 21)),
    "EDW_C_NAS":  (date(2026, 5, 20), date(2026, 5, 21)),
    "EDW_Empire": (date(2026, 5, 20), date(2026, 5, 21)),
    # EDW_WGS certified 2026-05-22 per user. Anchored to 5/20 (expected
    # delivery day); marker shows the actual 5/22 cert date.
    "EDW_WGS":    (date(2026, 5, 20), date(2026, 5, 22)),
    # 2026-05-21: AetnaQNXT — Masterload started 5/18; anchored to 5/19.
    # "AUTO" marker: cell shows L while pending, then auto-updates to the
    # actual cert date once it lands — but placement stays on 5/19 even
    # though the cert may arrive on a different day (per user 2026-05-22:
    # "At the next update, AetnaQNXT will have been certified. Please
    # leave it on the 5/19 date.").
    "AetnaQNXT":  (date(2026, 5, 19), "AUTO"),
    # Kaiser_AmbM override removed 2026-06-08 — snap re-enabled; it now follows
    # the standard Kaiser_Amb cert-only placement (Thursday anchor).
    # BCBSFL Elig moved to MONTHLY_MONTH_MARKER_OVERRIDES 2026-07-21 (the old
    # 2026-06-29 "Load Failure" pin is gone — June's file reloaded and delivered
    # late on 7/6, now shown as a "BCBSFL Elig (June)" row in ADDITIONAL_ENTRIES).
    # 2026-07-15: HumanaRx (monthly, snap/load-as-delivery) was auto-rendering ✓,
    # but per user it is still in Staging and has NOT loaded. 2026-07-17: per user
    # HumanaRx Stage is in RAMP, not loading — change to "No Data".
    # 2026-08-18: kept only to stop July auto-resolving to a false ✓ off the
    # lingering Ready 'HumanaRx 0100 Stage' card. The row itself no longer renders
    # — HumanaRx is in MONTHLY_ONLY_WHEN_ACTIVE, so a "No Data" month is dropped
    # from the calendar entirely instead of showing a placeholder.
    "HumanaRx": (date(2026, 7, 15), "No Data"),
    # 2026-08-26 per user: "Remove NCState from 8/14 and add to 8/26/26 as
    # 'Discontinued'." The August delivery never arrived (only the empty
    # 'NC State 0100 Stage' card ran on 8/17 — see LOAD_NAME_REQUIRED["NCState"]),
    # so the auto path parked a pink "No Data" on the expected 8/14 cell. This
    # override moves the single August row off 8/14 onto 8/26 and labels it
    # "Discontinued" (pink, like "Inactive" — the client has stopped delivering).
    # Scoped to August 2026 only (step 0a checks ov_day's year/month), so
    # May/June/July keep their real ✓ history; Sept 2026 forward is dropped
    # entirely via MONTHLY_RETIRED_FROM["NCState"].
    "NCState":    (date(2026, 8, 26), "Discontinued"),
    # 2026-08-27 per user: "The Molina certification is not ready to run yet, so
    # add 'Molina (Implementation)' to the 8/27/26 Monthly section with 8/27/26
    # as the cert date." That pinned Implementation row lives in
    # ADDITIONAL_ENTRIES (it needs the "(Implementation)" label, which the
    # standing row can't carry without renaming Molina everywhere). This entry
    # exists only to SUPPRESS the standing August row so August can't show two
    # Molina cells: Molina is in MONTHLY_ONLY_WHEN_ACTIVE, so a "No Data" month
    # is dropped from the calendar entirely (same trick as HumanaRx above).
    # Scoped to Aug 2026 only (step 0a checks ov_day's year/month), so if the
    # cert processes in September the standing row surfaces there normally.
    # Remove this (and the ADDITIONAL_ENTRIES row) when the user confirms Molina
    # is in full production and the weekly cadence gets wired.
    "Molina":     (date(2026, 8, 27), "No Data"),
}

# Monthly clients that appear ONLY when something actually happened — no
# placeholder row on a fixed expected day. A month whose marker resolves to
# "No Data" / blank / "Inactive" is dropped from the calendar; "L", "✓", a cert
# date and "Load Failure" all still render, on whatever day the activity landed.
#
# 2026-08-18 per user: "Change HumanaRx only show up when Loading starts, and not
# as a set date on the calendar." HumanaRx has no MONTHLY_EXPECTED_DAY_RANGE entry,
# so determine_monthly was falling back to the generic 15th (rendered 8/17) and
# parking a pink "No Data" there every month. Same spirit as the OptumPBMRx
# per-RAW rows and ADHOC_MONTHLY_SNAP_CLIENTS: only-when-present.
#
# 2026-08-26 per user: "Add 'Molina' to the dashboard once the certification
# processes. This is a new implementation that appears to be Weekly, but add to
# the Monthly section for the Implementation delivery on ADO 975084."
# Molina (ADO 975084 'Snap and Mine - Molina/ConnectiCare - PCN 1071309 to
# 1088600', go-live 08/28/26, loaded on ETL4 outside RAMP) has NO 0100/0110 RAMP
# jobs yet — only 'Molina MFT Logfile', which build_snap_index skips — so the DHT
# cert is the ONLY signal. As of this change all 34 Molina PCNs in
# DHTStats.DHT.TableList are 'Ready for Stats' (no StatTimestamp, no
# CertTimestamp). Only-when-active means the row simply appears, on its own cert
# date, the moment the certification processes; nothing shows before then (no
# pink "No Data" placeholder for a client that has no delivery history).
MONTHLY_ONLY_WHEN_ACTIVE = {"HumanaRx", "Molina"}

# Per-(client, year, month) marker override for monthly clients. Unlike
# MONTHLY_PLACEMENT_OVERRIDES (one entry per client), this keys on the specific
# month so a single client can be forced across SEVERAL months. The marker is
# placed on the client's normal expected day; a real DHT cert for that month
# always wins (checked first in determine_monthly), so the marker only shows
# until the cert lands. Value = marker string ("L", "No Data", "✓", ...) OR a
# date(...) which renders as a pinned cert date (MM/DD/YY, no pink) — used to
# stamp a same-day catch-up cert onto months the auto lookup can't attribute.
MONTHLY_MONTH_MARKER_OVERRIDES = {
    # 2026-07-16: ElixirRx data for May, June, and July has loaded and will be
    # certified tomorrow (per user). Show "L" (loaded, awaiting cert) on each
    # month's expected day until the DHT cert lands — the cert auto-wins once it
    # does. Remove these three entries after ElixirRx certifies.
    ("ElixirRx", 2026, 5): "L",
    ("ElixirRx", 2026, 6): "L",
    ("ElixirRx", 2026, 7): "L",
    # 2026-08-24 per user: "ElixirRx should have the '!'". May/June/July all
    # certified together on 7/16 (DHT ElixirRx CertTimestamp 2026-07-16 18:08 —
    # July's cell shows that date via step 1) but AUGUST never delivered, and
    # MONTHLY_BLANK_UNTIL_CERT was leaving the August cell silently blank instead
    # of flagging the miss. "!" renders an EMPTY pink-shaded cell (converted in
    # place()), so the month reads as missed rather than not-yet-due. A real
    # August DHT cert auto-wins (step 1 runs before this override), so this
    # self-clears the moment August certifies.
    ("ElixirRx", 2026, 8): "!",
    # 2026-07-17: Kaiser_WARx (monthly cert-only) — per user, show "L" (loading,
    # awaiting cert) on its expected day, NOT the auto "No Data". A DHT cert this
    # month auto-wins (checked before this override). The Friday pink escalation
    # is separately suppressed for Kaiser_WARx in alert_state. Remove once it
    # certifies.
    ("Kaiser_WARx", 2026, 7): "L",
    # 2026-08-05: BCBSSCRx's July delivery finally loaded (8/5 00:22), snapped
    # (8/5 02:31) and CERTIFIED 8/5 09:55 — per user "BCBSSCRx will be certified
    # today for the July delivery." The July tab now surfaces that cert via
    # MONTHLY_CERT_MONTH_REMAP (which also hides it from August), so the old
    # ("BCBSSCRx", 2026, 7) "No Data" hold is gone. August is held at "No Data"
    # until its OWN delivery loads — otherwise the July delivery's 8/5 load would
    # satisfy load_this_month and paint a premature "L" on August's 18th/19th
    # cell. A real August DHT cert still auto-wins (checked before this override).
    # Remove once August's own file loads.
    ("BCBSSCRx", 2026, 8): "No Data",
    # 2026-07-21: BCBSFL Elig — June's file delivered LATE on 7/6 (shown as a
    # separate "BCBSFL Elig (June)" row). That stray 7/6 snap was pulling July's
    # auto-placement onto 7/6; July's 25th slot (renders Fri 7/24) was held at
    # "No Data" until July's own file arrived.
    # 2026-08-05 per user: "The BCBSFL Elig on 7/24 loaded successfully on
    # 8/1/26." 'BCBSFL 0110 Eligibility Load' ran 7/29 14:03 -> 8/1 04:25
    # (Successful), so July DID deliver — flip the hold to "✓". Keeping it as a
    # pinned marker (rather than deleting the entry) holds the row on the 7/24
    # expected cell the user is tracking: build_snap_index attributes the load by
    # START date, so the auto path would otherwise move the ✓ onto a 7/29 cell.
    ("BCBSFLEligibilityLoad", 2026, 7): "✓",
    # 2026-07-28: TuftsRx reactivated — all past monthly deliveries certified today
    # (per user). The DHT catch-up cert was still "Email sent, Ready for
    # Certification review" at reactivation and, once flipped, every row carries
    # today's CertTimestamp, so the auto cert lookup would only place July. Pin the
    # 7/28 cert DATE on May/June/July's expected (10th) cell — a date value renders
    # as the cert date (MM/DD/YY) with no pink, exactly like a real cert. A real
    # per-month DHT cert still auto-wins (checked before this override). Remove
    # these once DHT attributes distinct per-month certs (unlikely — same-day
    # catch-up), or once these months roll off the report.
    ("TuftsRx", 2026, 5): date(2026, 7, 28),
    ("TuftsRx", 2026, 6): date(2026, 7, 28),
    ("TuftsRx", 2026, 7): date(2026, 7, 28),
}

# Extra rows injected into the calendar after standard placement runs. Use for
# one-off catch-up entries that don't fit the regular weekly/monthly cadence.
# Tuple: (section, day, label, marker, alert, highlight)
#   section ∈ {"daily", "weekly", "monthly", "kaiser"}
ADDITIONAL_ENTRIES = [
    # Medica catch-up for 5/1/26 claims — certified 2026-05-18 09:13:42
    # (DHT). Display cert date in the Mon cell.
    ("weekly", date(2026, 5, 18), "Medica (5/1/26)", date(2026, 5, 18), False, None),
    # EverNorthRx backsweep files 21, 22, 23 certified 2026-05-22 per user.
    ("weekly", date(2026, 5, 22), "EverNorthRx (21,22,23 BS)", date(2026, 5, 22), False, None),
    # 2026-07-21: BCBSFL Elig June cycle delivered late on 7/6 (per user — the
    # 7/6 Elig activity was June's, not July's). Show it as a labeled ✓ row so
    # the late delivery is recorded; July's own row sits on 7/24 (No Data).
    ("monthly", date(2026, 7, 6), "BCBSFL Elig (June)", "✓", False, None),
    # 2026-07-24: the current WellCareRx delivery is an Ad Hoc RELOAD, certified
    # 7/22 (DHT WellCareRx [mining].[RxClaim], StatTimestamp 7/22 10:13). Surface
    # it as its own labeled row showing the cert date; the normal weekly Friday
    # 7/24 cell is separately pinned to "No Data" (weekly not yet loading).
    ("weekly", date(2026, 7, 24), "WellCareRx (Ad Hoc)", date(2026, 7, 22), False, None),
    # 2026-08-05/08-06: a `HealthNetCA (Ad Hoc 3/20-3/27)` row lived here for the
    # ad hoc re-load of the 3/20–3/27 data (JobId 1812; failed 8/5, restarted
    # 8/6 07:10). REMOVED 2026-08-10 per user: "The HealthNetCA (3/20-3/27) can
    # be removed. The file was backed out and on hold for now." HealthNetCA stays
    # FORCED_INACTIVE, so the regular Monday cells continue to read "Inactive"
    # and no HealthNetCA activity surfaces anywhere. Re-add a labeled row here if
    # the re-load is picked back up.
    # 2026-08-18: a `HealthNetCA (3/27-4/3)` -> "Load Failure" row sat on 8/14.
    # REMOVED 2026-08-19 per user: that week reloaded and CERTIFIED (DHT
    # HealthNetCA CertTimestamp 2026-08-19 14:57), and belongs on the 8/17 cell.
    # No labeled row is needed — the regular weekly Monday row picks the cert up
    # automatically (8/19 falls in the Mon-Fri week of 8/17) and
    # HEALTHNETCA_RANGE_LABEL_OVERRIDES tags that cell "(3/27-4/3)".
    # 2026-08-20: the "CVSPBMRx (Ad Hoc)" → "Load Failure" row (8/18, when
    # QueueId 1413143 died after running 7/30→8/18) is NOT pinned here — the
    # auto Ad Hoc row emits it on the failure date now (see cvspbm_adhoc_failed),
    # which self-clears when the backfill is re-run.
    # 2026-08-24 per user: "HealthNetCA for 4/3-4/10 certified on 8/20/26."
    # THREE backfill weeks certified inside the 8/17 week (3/27-4/3 on 8/19,
    # 4/3-4/10 on 8/20, 4/10-4/24 on 8/24) and there is only one Monday cell per
    # week, so this middle one gets its own labeled row on its cert day (Thu 8/20).
    # The 8/17 Monday cell stays (3/27-4/3)/8/19 and the 8/24 Monday cell carries
    # (4/10-4/24)/8/24 — both pinned in MANUAL_OVERRIDES.
    ("weekly", date(2026, 8, 20), "HealthNetCA (4/3-4/10)",
     date(2026, 8, 20), False, None),
    # 2026-08-27 per user: "The Molina certification is not ready to run yet, so
    # go ahead and add 'Molina (Implementation)' to the 8/27/26 Monthly section
    # with 8/27/26 as the cert date. I'll provide an update once this client is
    # in full production." Molina (ADO 975084, go-live 08/28/26) has no RAMP
    # pipeline and its 34 PCNs in DHT are still 'Ready for Stats', so there is no
    # live signal to resolve from — the row is pinned by hand. A date marker
    # renders as the cert date (MM/DD/YY, no pink). The standing monthly row is
    # suppressed for August via MONTHLY_PLACEMENT_OVERRIDES["Molina"] so this is
    # the only Molina cell on the August tab.
    # Both Implementation rows highlighted yellow per user 2026-08-27.
    ("monthly", date(2026, 8, 27), "Molina (Implementation)",
     date(2026, 8, 27), False, "yellow"),
    # 2026-08-27 per user: add 'ModaRx (Implementation)' to today's Monthly
    # section; the standing 'ModaRx' monthly row starts on the 5th from
    # September forward (MONTHLY_STARTS_FROM). RAMP feed 905 is built out
    # ('ModaRx 0100 Stage' / '0110 Load' / '0120 Snap' — the Snap job is still
    # Enabled=0) but nothing has certified yet, so the row is pinned by hand.
    # Per user 2026-08-27 the Date shows 8/27/26 (same treatment as Molina)
    # rather than the "Implementation" marker.
    ("monthly", date(2026, 8, 27), "ModaRx (Implementation)",
     date(2026, 8, 27), False, "yellow"),
]

# CignaRx EOM/SOM cycle — at the start of each month a second CignaRx cycle
# closes out the prior month's tail; user 2026-06-03: "typically marked as an
# exception after certification." A `CignaRx (EOM/SOM)` row is injected on
# the first Tuesday of every month (matching regular CignaRx Tuesday). Marker
# is auto-detected from cert/load activity in a window straddling the month
# boundary; override here when the cert is recorded as an exception.
# Key: (year, month) of the SOM side. Value can be:
#   - a date / marker string ("✓", "L", "No Data", "Load Failure", "") — the
#     marker; the row stays on its default first-Tuesday placement, and
#   - a (placement_date, marker) TUPLE — forces BOTH the calendar column the row
#     lands on AND the marker (used when the EOM/SOM cycle should sit on a
#     non-Tuesday day, e.g. its data/exception date).
CIGNARX_EOM_SOM_OVERRIDES = {
    # 2026-07-06: the July EOM/SOM cycle certified 7/2 (StatTimestamp/data date
    # 7/1, mining rows recorded as "Exception"). Per user, place it on 7/1/26
    # instead of the default first-Tuesday (7/7). Show 7/1/26 in the cell.
    (2026, 7): (date(2026, 7, 1), date(2026, 7, 1)),
}

# Per-client cert window direction (default = backward / same Mon-Fri week).
# "forward" = look forward 7 days from scheduled day (used when a cert that
# lands after the scheduled day belongs to that scheduled day's cycle, like
# Premera where a Mon 5/11 cert completes the previous Thursday 5/7 cycle).
CERT_DIRECTION = {
    "Premera": "forward",
}

# Monthly clients that must remain anchored to their expected day even after
# a snap/load completes — only a DHT cert moves them. Per-user spec:
# "BCBSKS & BCBSKSMedAdv Monthly clients should always be on the 15th".
# 2026-05-15: Kaiser_Amb* feeds (CO/GA/HI/N/NW/S) added — user wants them
# anchored to the Thursday cert day even while loading. Kaiser_AmbM added
# 2026-06-08 (snap re-enabled) — now certifies with the rest of the feeds.
MONTHLY_CERT_ONLY_CLIENTS = {
    "BCBSKS", "BCBSKSMedAdv", "BCBSSCRx", "CareFirstRx",
    "Kaiser_AmbCO", "Kaiser_AmbGA", "Kaiser_AmbHI",
    "Kaiser_AmbM",
    "Kaiser_AmbN", "Kaiser_AmbNW", "Kaiser_AmbS",
    # Kaiser_WA: per user 2026-05-18, load completion alone is not delivery —
    # the cell should stay L on the expected day until the cycle truly
    # completes. Previously in LOAD_AS_DELIVERY, which auto-✓'d on a Successful
    # load even when the actual data was empty/incomplete.
    "Kaiser_WA",
    # Christus: per user 2026-06-01, stays L until DHT cert lands — snap
    # completion alone should not flip the cell to ✓.
    "Christus",
    # BCBSNorthCarolinaFEP: per user 2026-06-02, no ✓ from snap activity —
    # stays L on the expected Friday until DHT cert lands.
    "BCBSNorthCarolinaFEP",
    # BSCA_Facets: per user 2026-06-03, snap activity alone must not surface
    # ✓ — stays L (or No Data) until cert lands.
    "BSCA_Facets",
    # HealthSpring_FWA: per user 2026-06-03, same rule — loaded today but
    # cert will arrive separately; no snap-only ✓.
    "HealthSpring_FWA",
    # ElixirRx and PremeraMedAdvRx per user 2026-06-04: only Claim Load
    # job and DHT certification should move them off No Data / L. A snap
    # alone must not surface a ✓ for either client.
    "ElixirRx", "PremeraMedAdvRx",
    # BCBSVT and BSCA_Medicare per user 2026-06-08: leave as "L" until
    # certification — snap activity alone must not flip the cell to ✓.
    "BCBSVT", "BSCA_Medicare",
    # HAP_Medical and HAPRx per user 2026-06-08: same — stay "L" until cert.
    "HAP_Medical", "HAPRx",
    # Per user 2026-06-08: the following monthly clients should never show a
    # checkmark — stay "L" until the cert date lands (snap activity alone must
    # not flip to ✓). New 'MasterLoad 0110 Load' implementations also default
    # to this cert-style behavior (see the auto-discovery block in main()).
    "AetnaQNXT", "AetnaQNXTRx",
    "BCBSNC", "BCBSNC_Rx", "BCBSPuertoRico", "BCBSSC",
    "CareFirstDC", "CareFirstFacets", "CareFirstNasco",
    "Chickering",
    "EDW_ASE", "EDW_C_FAC", "EDW_C_NAS", "EDW_Empire", "EDW_WGS",
    "EmblemFacets",
    "Kaiser_WARx",
    "MedicalMutualMHS", "MedicalMutualOH",
    # ModaRx — new implementation, per user 2026-08-27: track the monthly
    # claims load/snap/certification only. RAMP feed 905 runs ONE load job
    # ('ModaRx 0110 Load') for both the weekly Elig/Other-Insurance files and
    # the monthly claims file (the staged files under 'ModaRx 0100 Stage' are
    # what distinguish them), so a load completion alone must not surface a ✓ —
    # stay "L" until the DHT certification lands.
    "ModaRx",
    # Molina — new implementation (ADO 975084). Cert-only per user 2026-08-26
    # ("add Molina once the certification processes"): the DHT cert is the
    # delivery signal, so a Snap-and-Mine completion alone must not surface a ✓.
    "Molina",
    "NCStateRx",
    "PremeraMedAdvVIS",
    "SamaritanHealth",
    # TuftsRx per user 2026-08-04: the MONTHLY row is a certification milestone
    # and "should always be on the 10th of the month" — weekly claims/elig load
    # activity must not drag it onto today. (The weekly Monday cells still get
    # their own ✓ per delivered claims file via STAGE_FILE_CELL_CLIENTS.)
    "TuftsRx",
}

# Monthly clients that should show an empty Date cell (rather than "No Data")
# until the cert lands. Per-user: "M - BCBSKSMedAdv had data, so have it blank
# each month until the ticket, Snap, Certification process finish on/near 15th".
#
# ElixirRx added 2026-08-18 per user: "Change ElixirRx to not say No Data." It is
# already in MONTHLY_CERT_ONLY_CLIENTS, so this returns "" on its expected day
# (the Tuesday of the 10th-15th week) instead of the pink "No Data" it was showing
# from August forward. May/June/July still read "L" via their
# MONTHLY_MONTH_MARKER_OVERRIDES entries (step 1a runs before this branch), and a
# real DHT cert always wins (step 1).
MONTHLY_BLANK_UNTIL_CERT = {"BCBSKSMedAdv", "ElixirRx"}

# Monthly clients whose "No Data" should always be shaded regardless of the
# 7-day grace window. ElixirRx removed 2026-08-18 — it no longer renders "No
# Data" at all (see MONTHLY_BLANK_UNTIL_CERT), so the force-shade would only
# have pink-shaded a blank cell into a stray "!".
FORCE_SHADE_NO_DATA = {"MedicalMutualMHS"}

# Clients that should be rendered with bold label (no fill).
BOLD_LABEL = {"Aetna NMSP - MMSEA", "AetnaMMSEA"}
# Clients with yellow label fill (kept empty per user — Aetna NMSP changed to bold only).
YELLOW_HIGHLIGHT = set()

MONTHLY_CLIENTS = {
    "AetnaQNXT", "AetnaQNXTRx", "AetnaRx_LegacyDMG", "AetnaSubro",
    "BCBSFLEligibilityLoad", "BCBSKS", "BCBSKSMedAdv",
    "BCBSNC", "BCBSNC_Rx", "BCBSNorthCarolinaFEP",
    "BCBSPuertoRico", "BCBSSC", "BCBSSCRx", "BCBSVT",
    "BSCA_Facets", "BSCA_Medicare",
    "CareFirstDC", "CareFirstFacets", "CareFirstNasco", "CareFirstRx",
    "Chickering", "Christus",
    "EDW_ASE", "EDW_C_FAC", "EDW_C_NAS", "EDW_Empire", "EDW_WGS",
    "ElixirRx",
    "EmblemFacets",
    "HAP_Medical", "HAPRx", "HealthSpring_FWA", "HumanaRx",
    "JHHCPassfile",
    "Kaiser_AmbCO", "Kaiser_AmbGA", "Kaiser_AmbHI", "Kaiser_AmbM", "Kaiser_AmbN",
    "Kaiser_AmbNW", "Kaiser_AmbS",
    "Kaiser_GE",
    "Kaiser_WA", "Kaiser_WARx",
    "MedicalMutualMHS", "MedicalMutualOH", "MedImpactPBMRx",
    "MMOH", "MMOHRxMonthly",
    "ModaRx",                           # new implementation — see MONTHLY_STARTS_FROM
    "Molina",                           # new implementation (see below)
    "NCState", "NCStateRx",
    "WellpointRxElig",                  # monthly load->snap (L on load, ✓ on snap)
    "ESIPBMRx",                         # monthly snap-only (RAMP snap-driven)
    "OptumPBMRx",                       # monthly, tape-driven
    "PremeraMedAdvRx", "PremeraMedAdvVIS",
    "SamaritanHealth", "Tufts_PublicPlan", "TuftsRx",
}

# Monthly clients retired mid-year: drop from the report starting the given
# (year, month) — kept on earlier month tabs where they were still active.
# Per user 2026-07-21: EDW_ASE retired; the EDW feeds now pull under
# 'Wellpoint 0100 EDW Pull …' jobs and ASE is no longer delivered.
MONTHLY_RETIRED_FROM = {
    "EDW_ASE": (2026, 7),
    # NCState discontinued per user 2026-08-26 — its final row is the 8/26/26
    # "Discontinued" cell (MONTHLY_PLACEMENT_OVERRIDES); drop the client from
    # September 2026 forward. May–Aug 2026 tabs keep their history.
    "NCState": (2026, 9),
}

# Inverse of MONTHLY_RETIRED_FROM: new monthly clients that only start appearing
# from the given (year, month) forward — earlier month tabs stay clean.
# ModaRx added 2026-08-27 per user: "Add 'ModaRx' to the monthly section on the
# 5th starting in September forward." (Today's Implementation row is a one-off
# ADDITIONAL_ENTRIES entry on the August tab.)
MONTHLY_STARTS_FROM = {
    "ModaRx": (2026, 9),
}

# Ad-hoc MONTHLY snap-driven clients (per user 2026-06-25): appear ONCE per
# month in the MONTHLY section, dated to when a specific RAMP job FINISHES
# (Successful/Resolved) — no fixed expected day, never flagged missing. Used when
# a client can't be certified, so its cell is dated off a RAMP job instead of a
# DHT cert. Value = exact RAMP JobName. UnitedRx(P) can't certify, so it's dated
# off 'United 0130 RX Post Snap' (JobId 10218).
ADHOC_MONTHLY_SNAP_CLIENTS = {
    "UnitedRx(p)": "United 0130 RX Post Snap",
}

# --- HealthNetCA claims backfill ---------------------------------------------
# HealthNetCA (weekly Mon) is working through a backfill: each load covers an
# older week of claims, so the calendar cell needs the CLAIMS date range next to
# the name — e.g. `HealthNetCA (3/20-3/27)` on the 8/10 cell for the load that
# ran 8/13. Per user 2026-08-14: "Backfill loading for HealthNetCA will
# continue, so please add them to the report when they load and provide the date
# range of claims … The date range will show up in the HealthNet 0100 Claims
# Stage job."
#
# The range is the pair of dates in the claims filename that the RAMP
# 'HealthNet 0100 Claims Stage' job (JobId 1811) lists:
#   HNT_VENDOR_CLAIM_MEDICAL_<plan>_<YYYYMMDD>_<YYYYMMDD>.txt
# That RAMP job is currently Enabled=0 while the backfill is driven manually, so
# its ramp.FileLog rows stop at 5/9/26 and can't be used. TRGETL1.HealthNetCA
# dbo.tblTape carries the same filenames plus when each one actually loaded
# (FileDate; note FileLoaded is NULL for this client and there is no etl.Tape /
# ProcessStatus here — it's the legacy tape table), so that's the source.
# FileTypeID 6 = Claims (7 = Eligibility, 8 = Providers, 13 = Groups) — only the
# claims files carry a range, and only they should label the cell.
HEALTHNETCA_CLIENT      = "HealthNetCA"
HEALTHNETCA_TAPE_SERVER = "TRGETL1"
HEALTHNETCA_CLAIM_TYPE  = 6
# Only label loads from the backfill program forward. Before this the client was
# on its normal weekly cadence (the 5/6 + 5/9 loads), and tagging those cells
# with a range would rewrite settled history.
HEALTHNETCA_BACKFILL_FROM = date(2026, 7, 27)
HEALTHNETCA_CLAIM_RANGE_RE = re.compile(
    r"HNT_VENDOR_CLAIM.*?_(\d{8})_(\d{8})", re.IGNORECASE)
# Claims windows that reached tblTape but whose LOAD failed / was backed out.
# Excluded from the cell label so the week reads only what actually delivered.
# tblTape gets a row when the file is picked up, not when it successfully loads
# (there is no ProcessStatus on this legacy table), so a failed load is
# indistinguishable from a good one here and has to be listed by hand.
#
# 2026-08-18 per user: "HealthNetCA for 8/10 is only for (3/20-3/27), the
# 3/27-4/3 load failed." Both weeks landed in the 8/10 week (3/20-3/27 on 8/13,
# 3/27-4/3 on 8/14) and healthnetca_range_labels merged the contiguous pair into
# a single "(3/20-4/3)" span. The failed week is surfaced on its own labeled
# "HealthNetCA (3/27-4/3)" → "Load Failure" row in ADDITIONAL_ENTRIES.
#
# 2026-08-19: that 3/27-4/3 week reloaded and certified, but it is still kept
# here so it doesn't merge back into the 8/10 label — per user it belongs on the
# 8/17 cell, which HEALTHNETCA_RANGE_LABEL_OVERRIDES supplies. (The set is
# really "ranges excluded from their own load-week label".)
HEALTHNETCA_FAILED_RANGES = {
    (date(2026, 3, 27), date(2026, 4, 3)),
}

# {cell_monday: " (M/D-M/D)"} — force the backfill range shown on a given weekly
# cell, overriding what healthnetca_range_labels() derives from tblTape load
# dates. Needed when a week's claims files were picked up in one calendar week
# but the delivery is recorded on a later cell.
#
# 2026-08-19 per user: "remove the Load Failure for HealthNetCA (3/27-4/3) and
# mark it as certified and place on 8/17 cell." The files loaded 8/14 (the 8/10
# week) but the successful reload certified 8/19, so the 8/17 cell carries it.
#
# 2026-08-24 per user: "4/10-4/24 certified on 8/24/26." Those two chunks
# (4/10-4/17 loaded 8/20, 4/17-4/24 loaded 8/21) both landed in the 8/17 week but
# certified together on Mon 8/24, so the 8/24 cell carries the merged span. The
# 8/17 pin keeps that cell on the 3/27-4/3 week it certified for (8/19), and the
# 4/3-4/10 week (cert 8/20) rides its own ADDITIONAL_ENTRIES row.
HEALTHNETCA_RANGE_LABEL_OVERRIDES = {
    date(2026, 8, 17): " (3/27-4/3)",
    date(2026, 8, 24): " (4/10-4/24)",
}

# Override display name for a client (the label only; client_key stays the same).
CLIENT_DISPLAY_NAME = {
    "BCBSFLEligibilityLoad": "BCBSFL Elig",
    "MMOH":                  "MMOH (WC)",
    "MMOHRxMonthly":         "MMOHRx",
    "WellpointRxElig":       "WellpointRx Elig",
    "JHHCPassfile":          "JHHC Passfile",
    "Kaiser_AmbCO":          "KaiserAmbCO",
    "Kaiser_AmbGA":          "KaiserAmbGA",
    "Kaiser_AmbHI":          "KaiserAmbHI",
    "Kaiser_AmbM":           "Kaiser_AmbM",
    "Kaiser_AmbN":           "KaiserAmbN",
    "Kaiser_AmbNW":          "KaiserAmbNW",
    "Kaiser_AmbS":           "Kaiser_AmbS",
    "Kaiser_GE":             "KaiserGE",
}

# Expected delivery window per monthly client, as (start_day, end_day) of month.
# Used for both placement (end of range) and "late" detection (today > end + 7).
MONTHLY_EXPECTED_DAY_RANGE = {
    "Chickering":             (1, 1),
    "Christus":               (1, 1),
    "MedicalMutualMHS":       (1, 1),
    "NCStateRx":              (1, 1),
    "MedicalMutualOH":        (3, 8),
    # ModaRx: monthly CLAIMS delivery only, anchored to the 5th per user
    # 2026-08-27. No MONTHLY_PLACEMENT_WEEKDAY entry so it stays on the 5th
    # itself rather than spreading across that work-week.
    "ModaRx":                 (5, 5),
    "MedImpactPBMRx":         (5, 10),
    "AetnaQNXTRx":            (5, 10),
    "BCBSVT":                 (5, 10),
    "BSCA_Facets":            (5, 10),
    "BSCA_Medicare":          (5, 10),
    "HAP_Medical":            (5, 10),
    "HAPRx":                  (5, 10),
    "HealthSpring_FWA":       (5, 10),
    "MMOH":                   (5, 10),
    "NCState":                (5, 10),
    # WellpointRxElig loads ~11th, snaps ~12th each month (per RAMP history).
    "WellpointRxElig":        (10, 12),
    "PremeraMedAdvVIS":       (5, 10),
    "PremeraMedAdvRx":        (5, 10),
    "TuftsRx":                (10, 10),
    "AetnaQNXT":              (10, 15),
    "AetnaSubro":             (11, 11),
    "BCBSNC":                 (10, 15),
    "BCBSNorthCarolinaFEP":   (10, 15),
    "BCBSPuertoRico":         (10, 15),
    "ElixirRx":               (10, 15),
    "Kaiser_WA":              (10, 15),
    "Kaiser_WARx":            (10, 15),
    "SamaritanHealth":        (10, 15),
    # Tufts_PublicPlan: pinned to the 10th per user 2026-08-04 ("keep on the
    # 10th until the next load"). Was (10, 15) with a Friday spread, which in a
    # month where the 15th is a Saturday rolled the row all the way out to the
    # 21st. Range end = 10 and no MONTHLY_PLACEMENT_WEEKDAY entry ⇒ the row sits
    # on the 10th itself (next Monday if the 10th is a weekend).
    "Tufts_PublicPlan":       (10, 10),
    "BCBSKS":                 (15, 15),
    "BCBSKSMedAdv":           (15, 15),
    "BCBSNC_Rx":              (15, 15),
    "AetnaMMSEA":             (15, 15),
    "AetnaRx_LegacyDMG":      (16, 16),
    "BCBSSC":                 (15, 20),
    # CareFirst clients moved to 19 to spread out from the 5/20 cluster
    "CareFirstDC":            (15, 19),
    "CareFirstFacets":        (15, 19),
    "CareFirstNasco":         (15, 19),
    "CareFirstRx":            (15, 19),
    "EmblemFacets":           (15, 20),
    "EDW_ASE":                (20, 20),
    "EDW_C_FAC":              (20, 20),
    "EDW_C_NAS":              (20, 20),
    "EDW_Empire":             (20, 20),
    "EDW_WGS":                (20, 20),
    "BCBSFLEligibilityLoad":  (25, 25),
    # Molina — go-live 08/28/26 per ADO 975084. Anchors the Implementation
    # delivery on the 28th; the DHT cert date always wins placement, and
    # MONTHLY_ONLY_WHEN_ACTIVE suppresses the row until something lands.
    "Molina":                 (28, 28),
    "AetnaRx_LegacyDMG":      (16, 16),
    # BCBSSCRx delays one week — per user, this month's load belongs next week.
    "BCBSSCRx":               (18, 19),
    # Kaiser monthly clients
    "Kaiser_GE":              (15, 20),
    "Kaiser_AmbCO":           (21, 21),
    "Kaiser_AmbGA":           (21, 21),
    "Kaiser_AmbHI":           (21, 21),
    "Kaiser_AmbM":            (21, 21),
    "Kaiser_AmbN":            (21, 21),
    "Kaiser_AmbNW":           (21, 21),
    "Kaiser_AmbS":            (21, 21),
    # HumanaRx: no fixed expected day — see MONTHLY_ONLY_WHEN_ACTIVE (the row is
    # dropped entirely until a load starts).
}

# Per-(client, year, month) PLACEMENT-DAY override for monthly clients — moves the
# row to a different calendar day for one month without touching the marker logic.
# Overrides both the placeholder and expected_date in determine_monthly, so a
# cert-only client's "L" / "No Data" / cert date all render on the given day.
# Use this (not MONTHLY_PLACEMENT_OVERRIDES) when the DELIVERY DATE slipped but the
# marker should keep resolving from live data.
#
# 2026-08-18 per user: "The Kaiser Amb files will be delivered on 8/20. Please move
# them from 8/13." All seven Kaiser_Amb feeds anchor to the closest Thursday to the
# 15th (8/13 for August); this shifts August one week out to Thu 8/20. The markers
# stay live — the 8/18 00:46 Stage → 0110 Load runs keep them at "L" (Kaiser_Amb*
# are MONTHLY_CERT_ONLY + SNAP_KIND_ONLY) until each feed's DHT cert lands.
MONTHLY_PLACEMENT_DAY_OVERRIDES = {
    ("Kaiser_AmbCO", 2026, 8): date(2026, 8, 20),
    ("Kaiser_AmbGA", 2026, 8): date(2026, 8, 20),
    ("Kaiser_AmbHI", 2026, 8): date(2026, 8, 20),
    ("Kaiser_AmbM",  2026, 8): date(2026, 8, 20),
    ("Kaiser_AmbN",  2026, 8): date(2026, 8, 20),
    ("Kaiser_AmbNW", 2026, 8): date(2026, 8, 20),
    ("Kaiser_AmbS",  2026, 8): date(2026, 8, 20),
}

# Spread monthly clients across the Mon-Fri week of their anchor day.
# Weekday is 0=Mon, 1=Tue, 2=Wed, 3=Thu, 4=Fri. The override snaps
# expected_date to the same Mon-Fri work-week as the original anchor day.
# Clients NOT in this dict stay on their MONTHLY_EXPECTED_DAY_RANGE end day.
# Per user 2026-05-26: avoid stacking everyone on the 10th and 15th. Pinned
# clients (BCBSKS, BCBSKSMedAdv, Kaiser_WA, BCBSSCRx, CareFirst*) are left
# off this list — they keep their explicit anchor.
MONTHLY_PLACEMENT_WEEKDAY = {
    # Day-10 cluster — spread Mon-Fri of the work-week of the 10th.
    "AetnaQNXTRx":            0,  # Mon
    "BCBSVT":                 1,  # Tue
    "BSCA_Facets":            2,  # Wed
    "BSCA_Medicare":          3,  # Thu
    "HAP_Medical":            4,  # Fri
    "HAPRx":                  0,  # Mon
    "HealthSpring_FWA":       1,  # Tue
    "MMOH":                   2,  # Wed
    "MedImpactPBMRx":         3,  # Thu
    "NCState":                4,  # Fri
    "PremeraMedAdvRx":        0,  # Mon
    "PremeraMedAdvVIS":       1,  # Tue
    # TuftsRx removed 2026-08-04 per user: "the Monthly delivery ... should
    # always be on the 10th of the month". With no spread entry its (10, 10)
    # range anchors the monthly row to the 10th itself instead of the Wednesday
    # of the 10th's work-week. (Tufts_PublicPlan likewise has no entry.)
    # Day-15 cluster — spread Mon-Fri of the work-week of the 15th.
    # (AetnaMMSEA omitted — handled separately by nmsp_mmsea_date.)
    "AetnaQNXT":              1,  # Tue
    "BCBSNC":                 2,  # Wed
    "BCBSNC_Rx":              3,  # Thu
    "BCBSNorthCarolinaFEP":   4,  # Fri
    "BCBSPuertoRico":         0,  # Mon
    "ElixirRx":               1,  # Tue
    "Kaiser_WARx":            2,  # Wed
    "SamaritanHealth":        3,  # Thu
}

# Monthly clients whose anchor day should snap to the CLOSEST weekday
# (Sat → previous Fri, Sun → next Mon) instead of the default
# next_monday_if_weekend (Sat → next Mon). Per user 2026-05-26: BCBSFL Elig
# anchor 25th should land on the closest weekday in months where 25 is Sat/Sun.
CLOSEST_WEEKDAY_CLIENTS = {"BCBSFLEligibilityLoad"}

# Clients in implementation phase — render "Implementation" in the date cell
# on every scheduled day from start_date until the first DHT cert lands. Cells
# before start_date are suppressed entirely.
# Per user 2026-05-26: BCBSAR was a new Tuesday weekly implementation from June.
# BCBSAR removed 2026-06-08 — Implementation marker dropped; certifying this
# week, so it now behaves as a normal weekly Tuesday cert-style client.
# Mechanism kept wired for the next implementation client.
IMPLEMENTATION_CLIENTS = {}

# Weekly clients whose cells BEFORE a given date render empty (no row, no pink
# "!") — for clients with no real history before they went live. Keyed by
# client → cutoff date; any scheduled day strictly before the cutoff is skipped.
# Per user 2026-06-08: BCBSAR — blank all dates before this week (Mon 6/8/26).
BLANK_BEFORE = {
    "BCBSAR": date(2026, 6, 8),
    # ElevanceMMMRx — only show June 2026 forward (per user 2026-06-16); earlier
    # cells were implementation-phase noise.
    "ElevanceMMMRx": date(2026, 6, 1),
}

# Clients whose cells ON/AFTER a given date render empty (the standing row is
# dropped) — the mirror image of BLANK_BEFORE. Earlier history is preserved.
# Keyed by client → cutoff date; any scheduled day on/after the cutoff is skipped.
# Per user 2026-07-29: ElevanceMMMRx removed from the report starting 7/13/26 —
# it should only appear going forward as an AdHoc load (added manually via
# ADDITIONAL_ENTRIES when one actually arrives).
BLANK_FROM = {
    "ElevanceMMMRx": date(2026, 7, 13),
}

# Clients whose "is delivered" signal is exclusively from TRGETL3 tape loads.
# Lookups for these clients ignore RAMP snap entries entirely.
TAPE_ONLY_CLIENTS = {"OptumPBMRx", "ESIPBMRx", "MedImpactPBMRx"}

# Jobs that signal a delivery but whose name lacks the usual load/stage/snap/
# mine keyword, so build_snap_index would otherwise skip them. Matched on the
# lowercased, stripped JobName. Per user 2026-06-08: 'JHHC Passfile Email'
# completion is the JHHCPassfile monthly ✓.
EXTRA_INDEXED_JOBS = {"jhhc passfile email"}

# Snap destination filter — when a client uses a specific snap destination,
# only count snap entries matching that destination string.
# (MMOH's Pharmacy filter removed 2026-06-08 — MMOH is now the WC load, which
# delivers via 'MMO 0110 WC Load' completion, not a Pharmacy snap.)
SNAP_DESTINATION_FILTER = {
}

# Clients that show ✓ when snapped (not just blank when no cert).
# Combines:  daily clients (✓ on snap-only days), PBMRx clients, and the
# small set of "select" snap-deliverable clients the user named.
SNAP_ONLY_CLIENTS = {
    "OptumPBMRx",
    "ESIPBMRx", "CVSPBMRx",
    "MedImpactPBMRx", "PrimePBMRx",
    "AetnaSubro", "HumanaRx",
    "WPRxDMGCOBMining",
    "BCBSKSMedAdv", "TuftsRx",   # snap weekly, cert monthly
    "NCState",                   # blocked from DHT cert by Chimera; track via snap
}

# These clients ONLY get ✓ when an actual SNAP step completes — a load-step
# completion alone doesn't trigger ✓.
SNAP_KIND_ONLY_CLIENTS = {
    # PBMRx snap clients
    "ESIPBMRx", "MedImpactPBMRx", "PrimePBMRx", "CVSPBMRx",
    # Other snap-required monthly/weekly clients
    # (MMOH removed 2026-06-08 — now the WC load, ✓ on load completion not snap)
    "AetnaSubro", "TuftsRx", "NCState", "WPRxDMGCOBMining",
    # MMOHRxMonthly: ✓ off the 'MMOHRx Monthly Claim 0120 Snap' step (per user
    # 2026-06-09) — L during the 0110 Load, ✓ when the Snap completes.
    "MMOHRxMonthly",
    # WellpointRxElig: ✓ off the 'Wellpoint RX Elig 0120 Snap' step (per user
    # 2026-06-12) — L during the 0110 Load, ✓ at the Snap step.
    "WellpointRxElig",
    # Kaiser_GE needs snap-step completion (0120 Snap).
    "Kaiser_GE",
    # Kaiser ambulance feeds: per user 2026-05-15, must wait for an actual
    # snap step (Kaiser Ambulance NC/CO/GA/HI/MAS/NW/S 0120 Snap) — a load-step
    # completion alone leaves the cell in "L" (load done, snap pending).
    # Kaiser_AmbM added 2026-06-08 (snap re-enabled).
    "Kaiser_AmbCO", "Kaiser_AmbGA", "Kaiser_AmbHI", "Kaiser_AmbM",
    "Kaiser_AmbN", "Kaiser_AmbNW", "Kaiser_AmbS",
    # AetnaHRP added 2026-05-19 per user — snap step must complete; the load
    # alone is not delivery (cell stays "L" between load done and snap done).
    "AetnaHRP",
    # NOTE: Daily Aetna clients AetnaRCE, AetnaRx, NCStateAetna are still
    # NOT in this set — their ✓ fires on the respective Load job completion
    # (Aetna RCE 310 ETL Load / AetnaRX Claim 0120 Load), not the snap step.
}

# These clients get ✓ on LOAD completion (load = delivery for them).
# BCBSKSMedAdv weekly: ✓ after 'BCBSKS Med Adv 0110 Load' finishes.
# Daily Aetna clients (AetnaRCE, AetnaRx, NCStateAetna): ✓ on the LOAD step
# — the subsequent Start Snap step should NOT keep the cell in "L".
# AetnaHRP REMOVED 2026-05-19 per user: "AetnaHRP did not Snap yet from the
# 5/18/26 load. The 5/18/26 HRP should still be an 'L'." AetnaHRP now requires
# the snap step to complete — see SNAP_KIND_ONLY_CLIENTS.
LOAD_AS_DELIVERY_CLIENTS = {
    "OptumPBMRx", "HumanaRx", "BCBSKSMedAdv",
    "AetnaRCE", "AetnaRx", "NCStateAetna",
}

# Clients whose LOAD step uses a non-standard verb ("Pull" instead of "Load")
# in the JobName. Per user 2026-07-21: the EDW feeds load via
# 'Wellpoint 0100 EDW Pull <feed>' jobs — treat "pull" as a load token so
# is_loading_today surfaces "L" while the pull is running.
PULL_AS_LOAD_CLIENTS = {
    "EDW_ASE", "EDW_C_FAC", "EDW_C_NAS", "EDW_Empire", "EDW_WGS",
}

# Clients whose "L" is driven ONLY by the LOAD step — a running snap/mine step
# does NOT keep them at L. Unlike LOAD_AS_DELIVERY_CLIENTS, this set does NOT
# affect ✓ resolution (so these clients can still be SNAP_KIND_ONLY and only
# get ✓ from an actual snap completion). Used solely by is_loading_today.
# Per user 2026-06-12 for WellpointRxElig: L during '0110 Load', then ✓ at the
# '0120 Snap' — the post-snap '0130 Mine' running must NOT revert ✓ back to L.
L_ON_LOAD_ONLY_CLIENTS = {
    "WellpointRxElig",
}

# Clients we are NOT actively working (no certification expected), but whose
# load pipeline is running for implementation/testing. Cell behavior:
#   - currently loading → "L"
#   - load+snap finished → blank (NOT a ✓, NOT a cert date)
# New auto-discovered MasterLoad implementations that are NOT PBMRx default
# into this set; promote out of it once the client is being actively delivered.
# (ElevanceMMMRx promoted out 2026-06-16 — now a DAILY_CLIENTS client with ✓ on
# load+snap days; see DAILY_CLIENTS / CLIENT_ALIASES / LOAD_NAME_REQUIRED.)
IMPLEMENTATION_LOAD_ONLY_CLIENTS = set()

# Weekly cert clients whose "L" means ONLY "a load is running right now". The
# generic weekly-cert fallback keeps a cell at "L" for the whole week once ANY
# load/snap activity is indexed (the "CenteneRx & WellCareRx should have an L
# since they have not been certified" rule), but the Kaiser Pareo weekly feeds
# load early in the week and don't SNAP until Wednesday afternoon/evening, so a
# week-long "L" misreads as "still loading" for days after the load finished.
# Per user 2026-08-05: "Clear the Kaiser weekly 'L' once the load finishes."
# Cell lifecycle for these clients: "L" while a load is genuinely in flight →
# blank once it finishes → the Thursday cert date when DHT certifies. Load
# failures and the Friday-pink escalation are unaffected.
LOADING_L_ONLY_CLIENTS = {
    "Kaiser_CO", "Kaiser_GA", "Kaiser_HI", "Kaiser_MASTapestry",
    "Kaiser_NW", "KaiserNCPareo", "KaiserSCPareo",
}

# Override the auto-derived primary key for clients whose name is a substring
# of another client's name (causing spurious substring matches in
# find_matching_jobs). Per user 2026-05-18: WellCare jobs were detected from
# WellCareRx Ready entries because 'wellcare' ⊂ 'wellcarerx'.
CLIENT_PRIMARY_KEY_OVERRIDE = {
    "WellCare": "wellcaremedical",
    # AetnaQNXT's normalized key "aetnaqnxt" is a prefix-substring of
    # "aetnaqnxtrx", so find_matching_jobs would wrongly match AetnaQNXTRx's
    # jobs. Override to "aetnaqnxtmasterload" 2026-06-08 — matches AetnaQNXT's
    # Masterload jobs (its delivery; CAQH/MSPI are ancillary) but NOT any
    # AetnaQNXTRx job ("aetnaqnxt" + "rx" + "masterload" is not contiguous).
    # _keys_for_client still yields base "aetnaqnxt" for the DHT cert lookup.
    "AetnaQNXT": "aetnaqnxtmasterload",
    # BCBSAR (Medical) ⊂ BCBSARRx — without this, "bcbsar" substring-matched the
    # BCBSARRx jobs and pulled the BCBSARRx COBC load failure onto BCBSAR. Match
    # only the Medical jobs. Per user 2026-06-12. _keys_for_client still yields
    # base "bcbsar" for the DHT cert lookup.
    "BCBSAR": "bcbsarmedical",
    # "medica" is a substring of "medical", so find_matching_jobs("Medica")
    # falsely matched every "*Medical*" load (Centene Medical, WellCare Medical,
    # BCBSAR Medical, …) — a failed 'WellCare Medical 0110 Load' was showing up
    # as a "Load Failure" on Medica's cell (per user 2026-06-16: "Medica 0110
    # Load ran w/o error"). Override to "medica0", which matches Medica's own
    # step jobs ('Medica 0110 Load' → "medica0110load") but NOT "…medical0…".
    # _keys_for_client still yields base "medica" for DHT cert / snap-index
    # lookups (those use strict equality against the "medica" prefix key).
    "Medica": "medica0",
    # "ncstate" (NCState Medical's key) is a substring of "ncstaterx", so
    # find_matching_jobs("NCState") falsely matched NCStateRx's jobs — a
    # loading 'NC State Rx 0110 Load' put a spurious "L" on NCState while
    # NCState itself was not loading. Override to "ncstate0", which matches
    # NCState's own step jobs ('NC State 0110 Load' → "ncstate0110load") but
    # NOT "ncstaterx0110load" ("ncstate" + "rx", not "ncstate" + "0"). Same
    # trick as Medica. _keys_for_client still yields base "ncstate" for the
    # snap-index strict-equality lookups. Per user 2026-07-01.
    "NCState": "ncstate0",
    # "hmsa" (HMSA Medical's key) is a substring of "hmsarx", so
    # find_matching_jobs("HMSA") falsely matched HMSA_Rx's jobs — the failed
    # 'HMSA RX Claims 0110 Load' put a spurious "Load Failure" on HMSA Medical
    # while HMSA's own 'HMSA Claims 0110 Load' ran fine. Override to
    # "hmsaclaims", which matches HMSA Medical's own jobs
    # ('HMSA Claims 0110 Load' → "hmsaclaims0110load") but NOT
    # "hmsarxclaims0110load" ("hmsa" + "rx" + "claims", not contiguous). Same
    # trick as Medica/NCState. _keys_for_client still yields base "hmsa" for
    # DHT cert / snap-index lookups. Per user 2026-07-07.
    "HMSA": "hmsaclaims",
    # "oscar" (Oscar Medical's key) is a substring of "oscarrx", so
    # find_matching_jobs("Oscar") falsely matched OscarRx's jobs — a Ready
    # 'Oscar RX 0110 Load' put a spurious "L" on Oscar Medical. Override to
    # "oscarmedical", which matches Oscar Medical's own jobs
    # ('Oscar Medical 0110 Load' → "oscarmedical0110load") but NOT
    # "oscarrx0110load". Same trick as Medica/NCState/HMSA. _keys_for_client
    # still yields base "oscar" for the DHT cert lookup (cert DatabaseName is
    # "Oscar"). Surfaced 2026-07-23 when Oscar was reactivated from
    # FORCED_INACTIVE. The bare "oscar" alias in CLIENT_ALIASES was also removed.
    "Oscar": "oscarmedical",
}

# NYShip_Rx fires four times per month — on the 1st, 8th, 16th, 24th
# (or the next Monday if that date is a weekend).
NYSHIP_DAYS = [1, 8, 16, 24]
NYSHIP_LABEL = {1: "1st", 8: "8th", 16: "16th", 24: "24th"}

# One-off per-cell overrides for the NYShip_Rx rotation, keyed by
# (year, month, daynum). Value is a marker string ("L") or a date(...) cert
# date (renders MM/DD/YY). Bypasses resolve_marker for that specific cell.
# Per user 2026-06-08: the 1st & 8th loaded together this cycle — both marked
# "L", and both get the SAME certification date once it lands. One-off; clear
# these entries after this June 2026 cycle certifies.
NYSHIP_OVERRIDES = {
    # 1st & 8th loaded together this cycle and both certified 6/9/26 (DHT
    # NYSHIP_RX, status Certified). Both cells show the shared cert date. The
    # 1st cell needs the override because a 6/9 cert is outside its 6/1-6/5
    # auto-detection window. One-off; clear after the June 2026 cycle.
    (2026, 6, 1): date(2026, 6, 9),
    (2026, 6, 8): date(2026, 6, 9),
    # 2026-06-23: the 16th load ran 6/22 (for the 6/16 data). It has since
    # certified — DHT NYSHIP_RX cert 6/23 (StatTimestamp 6/22) — so the stale "L"
    # was swapped to the 6/23 cert date on 2026-07-28.
    (2026, 6, 16): date(2026, 6, 23),
    # 2026-07-28: the July 24th delivery has loaded and certified (per user).
    # DHT NYSHIP_RX cert 7/28 (StatTimestamp 7/27). The 7/24 cell sits in the
    # week of Mon 7/20, but the cert's StatTimestamp week is 7/27, so the normal
    # cert_in_week auto-detection can't attribute it — the cell was showing "!".
    # Pin the 7/28 cert date. (Same out-of-window situation the 1st/8th needed.)
    (2026, 7, 24): date(2026, 7, 28),
}

# Suffix conventions per the All Clients tab key:
#   (s) SLA Client | (p) Rx Post Snap | (n) Not Delivered
CLIENT_SUFFIXES = {
    "AetnaHRP":              "(s)",
    "AetnaRCE":              "(s)",
    "AetnaRx":               "(p)(s)",
    "CareSource":            "(s)",
    "CareSourceRx":          "(n)",
    "CenteneFidelisRx":      "(p)",
    "CenteneRx":             "(p)",
    "CignaRx":               "(p)",
    "EmblemRx":              "(p)",
    "EverNorthRx":           "(p)",
    "ExcellusRx":            "(p)",
    "BCBSARRx":              "(p)",
    "HMSA_Rx":               "(n)",
    "WellCareRx":            "(p)",
    "MMOHRx":                "(n)(p)",
    "KaiserPrePayCOB":       "(s)",
    "OscarRx":               "(n)(p)",
    "PremeraMedAdvRx":       "(n)(p)",
    "NCStateRx":             "(n)(p)",
    "NCState":               "(n)",
    "Wellmark":              "(s)",
    "EmblemFacets":          "(s)",
    "HealthSpring_FWA":      "(s)",
    "Cambia":                "(n)",
    "HAPRx":                 "(p)",
    "P32-TuftsRx":           "(p)",
    # Monthly-specific suffixes per user spec 2026-05-13
    "TuftsRx":               "(p)",
    "AetnaSubro":            "(n)",
    "BCBSNC_Rx":             "(p)",
    "CareFirstRx":           "(n)(p)",
    "BCBSSCRx":              "(n)(p)",
    "EDW_ASE":               "(n)",
    "EDW_C_FAC":             "(n)",
    "EDW_C_NAS":             "(n)",
    "EDW_Empire":            "(n)",
    "EDW_WGS":               "(n)",
    "ElixirRx":              "(p)",
    "Kaiser_WARx":           "(n)",
}

WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# ---------- Client Owner tab ----------
# Per user 2026-05-20. Each owner has an "upper" list (active/current
# clients) and a "lower" list (separated by a blank row in the source).
# "*" suffix preserved verbatim from the user-provided list.
# Priority: 1 = highest, 4 = lowest.
CLIENT_OWNERS = {
    "Dave": {
        "upper": [
            ("Centene*", 1),
            ("Anthem/Wellpoint*", 2),
            ("BCBSNC*", 2),
            ("Humana*", 2),
            ("Arkansas Blue*", 3),
            ("Christus", 3),
            ("CVSPBM Rx", 3),
            ("Elixir/MCS", 3),
            ("FrontRunner", 3),
            ("HMSA*", 3),
            ("Medstar", 3),
            ("Molina", 3),
            ("NYSHIP", 3),
            ("Waystar", 3),
        ],
        "lower": [
            ("BCBSMN*", 4),
            ("BCBSLA*", 4),
            ("MVP*", 4),
            ("HealthPartners*", 4),
        ],
    },
    "Emmanuel": {
        "upper": [
            ("Aetna", 1),
            ("Evernorth Rx", 1),
            ("Excellus", 2),
            ("Medical Mutual OH*", 2),
            ("Oscar*", 2),
            ("Johns Hopkins*", 3),
            ("CareSource", 3),
            ("ESI PBM Rx", 3),
            ("HAP", 3),
            ("Ingenio", 3),
            ("Maxor", 3),
            ("Medica", 3),
            ("Work Comp", 3),
            ("United", 4),
        ],
        "lower": [
            ("BCBS_Assoc*", 4),
            ("IndepenenceHealth*", 4),
            ("BCBSND*", 4),
            ("Highmark*", 4),
        ],
    },
    "Holly": {
        "upper": [
            ("Kaiser*", 1),
            ("BSCA*", 2),
            ("Emblem*", 2),
            ("Point 32 (Tufts/Harvard Pilgrim)*", 2),
            ("Premera*", 2),
            ("Wellmark*", 2),
            ("BCBSKS", 3),
            ("BCBSSC*", 3),
            ("HealthNewEngland", 3),
            ("Medispan", 3),
            ("UPMC", 3),
            ("WebTPA", 3),
            ("NPI", 4),
        ],
        "lower": [
            ("CapitalBlueCross*", 4),
            ("BlueCrossIdaho*", 4),
            ("BCBSRI*", 4),
            ("KPS*", 4),
        ],
    },
    "Adam": {
        "upper": [
            ("Cigna*", 1),
            ("BCBSFL*", 2),
            ("CareFirst*", 2),
            ("GEHA", 2),
            ("BCBS Puerto Rico", 3),
            ("BCBSVT*", 3),
            ("HealthNow*", 3),
            ("Medimpact PBM Rx", 3),
            ("Optum", 3),
            ("Prime PBM Rx", 3),
            ("Samaritan Health", 3),
            ("Cambia", 4),
            ("Provider Solutions", 4),
        ],
        "lower": [
            ("HCSC (Cigna)*", 4),
            ("BCBSTN*", 3),
            ("BCBSMA*", 4),
        ],
    },
}


# ============================================================
#                          helpers
# ============================================================
def curl_json(url):
    r = subprocess.run(
        ["curl", "-s", "--negotiate", "-u", ":", url],
        capture_output=True, text=True, check=False,
    )
    return json.loads(r.stdout)


def curl_post_json(url, body):
    with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False, encoding="utf-8") as f:
        json.dump(body, f)
        path = f.name
    try:
        r = subprocess.run(
            ["curl", "-s", "--negotiate", "-u", ":",
             "-H", "Content-Type: application/json",
             "--data-binary", f"@{path}", url],
            capture_output=True, text=True, check=False,
        )
        return json.loads(r.stdout)
    finally:
        os.unlink(path)


def normalize(s):
    return re.sub(r"[\s_\-]+", "", s or "").lower()


def parse_dt(s):
    if not s:
        return None
    s = s.replace("Z", "").rstrip()
    try:
        return datetime.fromisoformat(s.split(".")[0])
    except Exception:
        return None


# ============================================================
#                       SQL cert fetch
# ============================================================
def fetch_dht_certs(since):
    """Query [DHTStats].[DHT].[TableList] for certifications since `since`.
    Returns a list of dicts: {DatabaseName, Name, CertTimestamp, CurrentStatus, PCN}.
    """
    q = (
        "SET NOCOUNT ON; "
        "SELECT DatabaseName, [Name], PCN, CertTimestamp, CurrentStatus, StatTimestamp "
        f"FROM [DHTStats].[DHT].[TableList] "
        f"WHERE CertTimestamp >= '{since.isoformat()}' "
        "ORDER BY CertTimestamp"
    )
    r = subprocess.run(
        ["sqlcmd", "-S", SQL_SERVER, "-d", SQL_DB, "-E", "-Q", q,
         "-W", "-s", "\t", "-h", "-1"],
        capture_output=True, text=True, check=False,
    )
    rows = []
    for line in r.stdout.splitlines():
        if not line or line.startswith("---") or "rows affected" in line:
            continue
        parts = line.split("\t")
        if len(parts) < 5:
            continue
        try:
            ts = parse_dt(parts[3])
        except Exception:
            continue
        if not ts:
            continue
        rows.append({
            "DatabaseName":   parts[0].strip(),
            "Name":           parts[1].strip(),
            "PCN":            parts[2].strip(),
            "CertTimestamp":  ts,
            "CurrentStatus":  parts[4].strip(),
            "StatTimestamp":  parse_dt(parts[5]) if len(parts) > 5 else None,
        })
    return rows


def build_cert_index(certs):
    """Return: normalized_db -> list of (datetime, status) sorted asc."""
    idx = defaultdict(list)
    for c in certs:
        key = normalize(c["DatabaseName"])
        if key:
            idx[key].append((c["CertTimestamp"], c["CurrentStatus"]))
    for k in idx:
        idx[k].sort()
    return idx


def stat_week_monday(d):
    """Monday of the delivery week a StatTimestamp (data date) belongs to.

    Per user 2026-06-24, validated against DHT: the delivery week runs
    Mon-Sat, and SUNDAY rolls forward into the next week. So a Sat data date
    stays with the week that just ended; a Sun data date belongs to the
    upcoming week.
      - Thu 6/18 -> Mon 6/15  (that week)
      - Sat 6/20 -> Mon 6/15  (week just ended; CenteneRx -> 6/19 cell)
      - Sun 6/21 -> Mon 6/22  (next week; Centene -> 6/23 cell)
    """
    if isinstance(d, datetime):
        d = d.date()
    wd = d.weekday()           # Mon=0 .. Sat=5, Sun=6
    if wd == 6:                # Sunday -> next week's Monday
        return d + timedelta(days=1)
    return d - timedelta(days=wd)   # Mon-Sat -> this week's Monday


# normalized_db -> {week_monday(date): latest Certified CertTimestamp}. Built in
# main() from DHT StatTimestamp so a weekly client's cert lands on the cell for
# the week its DATA covers (not the week it happened to certify). Used by
# cert_in_week for non-forward weekly clients.
CERT_WEEK_IDX = {}


def build_cert_week_index(certs):
    """normalized_db -> {week_monday: latest CertTimestamp} over Certified rows,
    grouping by the delivery week of each row's StatTimestamp (falls back to the
    CertTimestamp's own week when StatTimestamp is missing)."""
    idx = defaultdict(dict)
    for c in certs:
        if c.get("CurrentStatus") != "Certified":
            continue
        key = normalize(c["DatabaseName"])
        if not key:
            continue
        certdt = c["CertTimestamp"]
        wk = stat_week_monday(c.get("StatTimestamp") or certdt)
        cur = idx[key].get(wk)
        if cur is None or certdt > cur:
            idx[key][wk] = certdt
    return idx


def cert_on_day(client, day, cert_idx):
    """Return the latest CertTimestamp datetime for `client` on calendar `day`, else None.
    Friday cells additionally pick up Sat/Sun certs so weekend deliveries
    surface on the prior Friday's cell (e.g. AetnaRCE 5/9 Sat → Fri 5/8)."""
    days_to_check = {day}
    if day.weekday() == 4:  # Friday: also include Sat/Sun
        days_to_check.add(day + timedelta(days=1))
        days_to_check.add(day + timedelta(days=2))
    # CERT_CELL_REMAP: also consider cert dates explicitly reattributed TO this
    # cell (they may fall outside the windows above).
    days_to_check |= {cd for (c, cd), target in CERT_CELL_REMAP.items()
                      if c == client and target == day}
    best = None
    for key in _keys_for_client(client):
        for dt, status in cert_idx.get(key, ()):
            if status != "Certified":
                continue
            d = dt.date()
            if d not in days_to_check:
                continue
            # ...and skip cert dates reattributed AWAY from this cell, so the
            # remapped cert doesn't also stamp the day it actually ran on.
            target = CERT_CELL_REMAP.get((client, d))
            if target is not None and target != day:
                continue
            if best is None or dt > best:
                best = dt
    return best


def cert_in_week(client, scheduled_day, cert_idx):
    """Latest cert attributed to this client's delivery week (the week of
    `scheduled_day`).

    Default: attribute each cert to the week its DATA covers via the
    StatTimestamp of the certified tapes (CERT_WEEK_IDX, built by
    build_cert_week_index). Per user 2026-06-24 a single cert run can certify
    two data weeks at once — e.g. Centene's 6/22 cert had StatTimestamp 6/18
    (-> the 6/16 cell) AND 6/21 (Sun -> the 6/23 cell) — and StatTimestamp
    picks the right cell for each. Falls back to the CertTimestamp week for rows
    with no StatTimestamp (handled in build_cert_week_index).

    `CERT_DIRECTION[client] = "forward"` (Premera): keeps the explicit 7-day
    forward window on CertTimestamp.
    """
    if CERT_DIRECTION.get(client) == "forward":
        cycle_start = scheduled_day
        cycle_end   = scheduled_day + timedelta(days=6)
        best = None
        for key in _keys_for_client(client):
            for dt, status in cert_idx.get(key, ()):
                if status != "Certified":
                    continue
                if cycle_start <= dt.date() <= cycle_end:
                    if best is None or dt > best:
                        best = dt
        return best

    cell_monday = scheduled_day - timedelta(days=scheduled_day.weekday())
    best = None
    for key in _keys_for_client(client):
        dt = CERT_WEEK_IDX.get(key, {}).get(cell_monday)
        if dt and (best is None or dt > best):
            best = dt
    return best


def latest_cert(client, cert_idx, on_or_before=None):
    best = None
    for key in _keys_for_client(client):
        for dt, status in cert_idx.get(key, ()):
            if status != "Certified":
                continue
            if on_or_before and dt.date() > on_or_before:
                continue
            if best is None or dt > best:
                best = dt
    return best


def latest_cert_in_month(client, cert_idx, year, month, on_or_before=None):
    """Latest Certified datetime whose CertTimestamp falls in (year, month).
    Unlike latest_cert (which returns the single GLOBAL-latest cert), this finds
    the cert that belongs to a SPECIFIC month — so a past month tab surfaces its
    OWN cert even when the client has certified again in a later month. Without
    this, determine_monthly's `latest_cert(...).month == month` check only ever
    matched the client's most-recent month, leaving every earlier month "No Data".
    """
    # MONTHLY_CERT_MONTH_REMAP: cert dates explicitly reattributed TO this month
    # (their CertTimestamp falls outside it) count here; certs reattributed AWAY
    # from this month are skipped so one cert can't fill two month tabs.
    best = None
    for key in _keys_for_client(client):
        for dt, status in cert_idx.get(key, ()):
            if status != "Certified":
                continue
            d = dt.date()
            tgt = MONTHLY_CERT_MONTH_REMAP.get((client, d))
            if tgt is not None:
                if tgt != (year, month):
                    continue
            elif dt.year != year or dt.month != month:
                continue
            if on_or_before and d > on_or_before:
                continue
            if best is None or dt > best:
                best = dt
    return best


def _keys_for_client(client):
    # Always yield the base normalize(client) so cert/snap lookups still
    # find the natural DHT/RAMP key even when CLIENT_PRIMARY_KEY_OVERRIDE
    # has remapped the substring-target for find_matching_jobs. Per user
    # 2026-05-20: WellCare's DHT cert key is "wellcare", but the override
    # changed _keys_for_client to "wellcaremedical" only — past cert
    # dates were silently missed.
    base = normalize(client)
    seen = set()
    if base:
        seen.add(base)
        yield base
    primary = CLIENT_PRIMARY_KEY_OVERRIDE.get(client)
    if primary and primary not in seen:
        seen.add(primary)
        yield primary
    for alias in CLIENT_ALIASES.get(client, []):
        k = normalize(alias)
        if k and k not in seen:
            seen.add(k)
            yield k


# ============================================================
#                          ADO fetch
# ============================================================
def fetch_ado_tickets(min_changed_date):
    """Fetch ADO user stories tagged 'Delivery Ticket' changed since min_changed_date."""
    wiql = (
        "SELECT [System.Id] FROM WorkItems WHERE "
        "[System.TeamProject] = 'Rawlings' "
        "AND [System.Tags] CONTAINS 'Delivery Ticket' "
        f"AND [System.ChangedDate] >= '{min_changed_date.isoformat()}' "
        "AND [System.WorkItemType] = 'User Story'"
    )
    res = curl_post_json(f"{ADO_BASE}/_apis/wit/wiql?api-version=5.0", {"query": wiql})
    ids = [w["id"] for w in res.get("workItems", [])]
    if not ids:
        return []
    out = []
    fields = ",".join([
        "System.Id", "System.Title", "System.State", "System.AreaPath",
        "System.IterationPath", "System.ChangedDate", "System.CreatedDate",
        "System.AssignedTo", "System.Tags",
    ])
    for i in range(0, len(ids), 200):
        batch = ids[i:i+200]
        url = (f"{ADO_BASE}/_apis/wit/workitems?ids={','.join(map(str, batch))}"
               f"&fields={fields}&api-version=5.0")
        for w in curl_json(url).get("value", []):
            f = w["fields"]
            title = f.get("System.Title", "")
            m = TITLE_RE.match(title)
            client = m.group(2).strip() if m else ""
            kind = m.group(1) if m else ""
            assigned = f.get("System.AssignedTo", "")
            if isinstance(assigned, dict):
                assigned = assigned.get("displayName") or assigned.get("uniqueName", "")
            elif isinstance(assigned, str) and "<" in assigned:
                assigned = assigned.split("<")[0].strip()
            out.append({
                "id":       f["System.Id"],
                "title":    title,
                "state":    f.get("System.State", ""),
                "kind":     kind,
                "client":   client,
                "iter":     f.get("System.IterationPath", ""),
                "changed":  f.get("System.ChangedDate", ""),
                "created":  f.get("System.CreatedDate", ""),
                "assigned": assigned,
                "tags":     f.get("System.Tags", ""),
            })
    return out


# ============================================================
#                          RAMP fetch
# ============================================================
def fetch_ramp_jobs():
    return curl_json(f"{RAMP_BASE}/api/Ramp/Job/List").get("Data", [[]])[0]


def fetch_ramp_queue(since=None):
    """Pull RAMP queue from SQL [TRGUTIL10].RAMP.ramp.Queue directly.
    The REST endpoint /api/Ramp/Queue/List caps at 1000 items and SFTP/LogFile
    churn rotates real load entries out within hours. SQL gives full history.
    Returns dicts shaped like the REST response so downstream code is unchanged.

    `since` (a date) sets the CreateDate floor so the report can reach back to
    the first computed month and populate snap_idx with past-month snap/load
    completions (e.g. May's 'Kaiser GE 0120 Snap') — these drive the ✓ marks for
    snap-only monthly clients that the 60/89-day RAMP snap endpoint no longer
    returns. Defaults to the last 45 days. Old rows are safe: is_loading_today
    only reads Ready/Running and has_recent_failure ignores anything >3 days old.
    """
    date_floor = (f"'{since.isoformat()}'" if since
                  else "DATEADD(day, -45, GETDATE())")
    q = (
        "SET NOCOUNT ON; "
        "SELECT q.QueueId, q.JobId, q.Status, "
        "       CONVERT(varchar(23), q.StartDate, 121) AS StartDate, "
        "       CONVERT(varchar(23), q.EndDate, 121)   AS EndDate, "
        "       CONVERT(varchar(23), q.CreateDate, 121) AS CreateDate, "
        "       CAST(q.JobXml AS varchar(MAX)) AS JobXml "
        "FROM [RAMP].[ramp].[Queue] q "
        f"WHERE q.CreateDate >= {date_floor} "
        "ORDER BY q.QueueId DESC"
    )
    SEP = "\x1f"   # ASCII unit separator — unlikely to appear in any XML/data
    r = subprocess.run(
        ["sqlcmd", "-S", SQL_SERVER, "-d", "RAMP", "-E", "-Q", q,
         "-W", "-s", SEP, "-h", "-1"],
        capture_output=True, text=True, check=False,
    )
    rows = []
    for line in r.stdout.splitlines():
        if not line or line.startswith("---") or "rows affected" in line:
            continue
        parts = line.split(SEP, 6)
        if len(parts) < 7:
            continue
        try:
            qid    = int(parts[0])
            job_id = int(parts[1])
        except ValueError:
            continue
        rows.append({
            "QueueId":    qid,
            "JobId":      job_id,
            "Status":     parts[2].strip(),
            "StartDate":  parts[3].strip() if parts[3].strip() != "NULL" else "",
            "EndDate":    parts[4].strip() if parts[4].strip() != "NULL" else "",
            "CreateDate": parts[5].strip(),
            "JobXml":     parts[6],
        })
    return rows


def fetch_ramp_snaps(since=None):
    """Snap completions from the RAMP SnapQueueStatus endpoint (full payload),
    trimmed by End date. `since` (a date) sets the cutoff so the report can
    reach back far enough to fill past-month cells — the report renders every
    computed month in one run, and snap-driven ✓ marks for e.g. May are lost if
    the window only reaches ~60 days. Defaults to the last 60 days when `since`
    is not supplied. The HTTP payload is the same either way (the endpoint has no
    date filter) — a wider `since` just keeps more rows in the snap index."""
    data = curl_json(f"{RAMP_BASE}/api/Ramp/Snap/SnapQueueStatus").get("Data", [[]])
    rows = data[0] if data and isinstance(data[0], list) else data
    cutoff = (datetime.combine(since, datetime.min.time()) if since
              else datetime.now() - timedelta(days=60))
    out = []
    for s in rows:
        end = parse_dt(s.get("End"))
        if end is None or end >= cutoff:
            out.append(s)
    return out


def compute_evernorth_claims_pending(queue, jobs, since):
    r"""True if EverNorthRx has a CLAIMS file (ESI_PAID_CLAIMS_*) staged and NOT
    yet loaded — i.e. genuinely "ready for loading" right now.

    The EverNorthRx weekly claims row should flag "L" only while claims are
    staged/awaiting load — the daily Masterload also stages eligibility-only files
    (ESI_*_ELIG_*, COBC, TRR, ACUM, ABII) that must NOT light up the claims row,
    and the L must DROP the moment the claims actually load (not linger through the
    snap-awaiting-cert window). Per user 2026-06-29.

    RAMP's FileLog has no "Loaded" status (only Completed -> Staged), so "not yet
    loaded" is inferred by time: claims are pending when the latest staged
    ESI_PAID_CLAIMS file (ramp.FileLog LogDate) is NEWER than the latest successful
    'EvernorthRx Masterload 0110 Load' completion (the load that consumes them).
    Once a Masterload Load completes after the stage, the claims are loaded and
    this returns False. `[_]` escapes the literal underscores (T-SQL LIKE treats
    _ as a single-char wildcard).
    """
    q = (
        "SET NOCOUNT ON; "
        "SELECT CONVERT(varchar(19), MAX(LogDate), 121) "
        "FROM [RAMP].[ramp].[FileLog] "
        "WHERE FileName LIKE 'ESI[_]PAID[_]CLAIMS%' AND Status = 'Staged' "
        f"AND LogDate >= '{since:%Y-%m-%d}'"
    )
    r = subprocess.run(
        ["sqlcmd", "-S", SQL_SERVER, "-d", "RAMP", "-E", "-Q", q, "-W", "-h", "-1"],
        capture_output=True, text=True, check=False,
    )
    claims_stage = None
    for line in r.stdout.splitlines():
        claims_stage = parse_dt(line.strip())
        if claims_stage:
            break
    if not claims_stage:
        return False
    # Latest successful 'EvernorthRx Masterload 0110 Load' completion from queue.
    load_ids = {
        j.get("JobId") for j in jobs
        if "masterload" in (j.get("JobName") or "").lower()
        and "0110 load" in (j.get("JobName") or "").lower()
        and "evernorth" in (j.get("JobName") or "").lower()
    }
    load_end = None
    for item in queue:
        if item.get("JobId") not in load_ids:
            continue
        st = (item.get("Status") or "").lower()
        if not (st.startswith("success") or st == "resolved"):
            continue
        e = parse_dt(item.get("EndDate"))
        if e and (load_end is None or e > load_end):
            load_end = e
    return load_end is None or claims_stage > load_end


def fetch_aetna_nmsp_loads(since):
    r"""Query SQLUtilAudit.cmse_new.SourceLog for Aetna NonMSP file loads
    started since `since`. Drives "M - Aetna NMSP - MMSEA":
      - "L"  once a file at \\trgdatacap2\MMSEA\Aetna\<year>\NonMSP has an
             ImportStartDate but no ImportCompleteDate yet (loading in CMSE),
      - "✓"  once ImportCompleteDate lands (placed on the completion date).
    Returns list of dicts {"start": datetime, "done": datetime|None}.
    """
    q = (
        "SET NOCOUNT ON; "
        "SELECT CONVERT(varchar(23), ImportStartDate, 121) AS Started, "
        "ISNULL(CONVERT(varchar(23), ImportCompleteDate, 121), '') AS Done "
        "FROM [cmse_new].[dbo].[SourceLog] WITH (NOLOCK) "
        "WHERE EntryName LIKE '%MMSEA\\Aetna\\2026\\NonMSP%' "
        f"AND ImportStartDate >= '{since.isoformat()}' "
        "ORDER BY ImportStartDate"
    )
    r = subprocess.run(
        ["sqlcmd", "-S", "SQLUtilAudit", "-d", "cmse_new", "-E", "-Q", q,
         "-W", "-s", "\t", "-h", "-1"],
        capture_output=True, text=True, check=False,
    )
    out = []
    for line in r.stdout.splitlines():
        line = line.rstrip()
        if not line or line.startswith("---") or "rows affected" in line:
            continue
        parts = line.split("\t")
        start = parse_dt(parts[0].strip()) if parts else None
        done = parse_dt(parts[1].strip()) if len(parts) > 1 and parts[1].strip() else None
        if start:
            out.append({"start": start, "done": done})
    return out


def fetch_tape_loads(db, since, server="TRGETL3", name_like=None):
    """Query <server>.<db>.etl.Tape for recent successful loads (ProcessStatus=50).
    Returns list of dicts: {FileName, FileLoadDate (datetime)}.

    `server` defaults to TRGETL3 (the PBMRx tape server); JHHC Passfile lives on
    TRGINTP3.JohnsHopkins. `name_like` optionally restricts to filenames matching
    a LIKE pattern (e.g. 'PassFile') so a shared client DB only yields the rows
    for the intended feed.
    """
    where = f"ProcessStatus = 50 AND FileLoadDate >= '{since.isoformat()}'"
    if name_like:
        where += f" AND FileName LIKE '%{name_like}%'"
    q = (
        "SET NOCOUNT ON; "
        "SELECT FileName, FileLoadDate FROM [etl].[Tape] "
        f"WHERE {where} "
        "ORDER BY FileLoadDate"
    )
    r = subprocess.run(
        ["sqlcmd", "-S", server, "-d", db, "-E", "-Q", q,
         "-W", "-s", "\t", "-h", "-1"],
        capture_output=True, text=True, check=False,
    )
    rows = []
    for line in r.stdout.splitlines():
        if not line or line.startswith("---") or "rows affected" in line:
            continue
        parts = line.split("\t")
        if len(parts) < 2:
            continue
        dt = parse_dt(parts[1])
        if dt:
            rows.append({"FileName": parts[0].strip(), "FileLoadDate": dt})
    return rows


def fetch_healthnetca_claim_loads(since=None):
    """HealthNetCA backfill claims loads with the claims date range each covers.

    Reads TRGETL1.HealthNetCA dbo.tblTape (the legacy tape table — no etl.Tape /
    ProcessStatus on this client) for FileTypeID 6 (Claims). `FileDate` is when
    the file loaded; `FileLoaded` is always NULL here. The two dates in the
    filename are the claims window, the same pair the RAMP 'HealthNet 0100
    Claims Stage' job shows. See the HEALTHNETCA_* notes above.

    Returns list of dicts: {load_date: date, start: date, end: date}, one per
    distinct (load_date, range) — the per-plan files (XA/XB/XC/…) of one load all
    carry the same range and collapse to a single entry.
    """
    since = since or HEALTHNETCA_BACKFILL_FROM
    q = (
        "SET NOCOUNT ON; "
        "SELECT CONVERT(varchar(10), FileDate, 23), FileName "
        "FROM [dbo].[tblTape] "
        f"WHERE FileTypeID = {HEALTHNETCA_CLAIM_TYPE} "
        f"  AND FileDate >= '{since.isoformat()}' "
        "ORDER BY FileDate"
    )
    r = subprocess.run(
        ["sqlcmd", "-S", HEALTHNETCA_TAPE_SERVER, "-d", HEALTHNETCA_CLIENT,
         "-E", "-Q", q, "-W", "-s", "\t", "-h", "-1"],
        capture_output=True, text=True, check=False,
    )
    seen = set()
    out = []
    for line in r.stdout.splitlines():
        if not line or line.startswith("---") or "rows affected" in line:
            continue
        parts = line.split("\t")
        if len(parts) < 2:
            continue
        try:
            load_date = datetime.strptime(parts[0].strip(), "%Y-%m-%d").date()
        except ValueError:
            continue
        m = HEALTHNETCA_CLAIM_RANGE_RE.search(parts[1])
        if not m:
            continue
        try:
            start = datetime.strptime(m.group(1), "%Y%m%d").date()
            end   = datetime.strptime(m.group(2), "%Y%m%d").date()
        except ValueError:
            continue
        # A claims window whose load failed / was backed out must not label a cell.
        if (start, end) in HEALTHNETCA_FAILED_RANGES:
            continue
        key = (load_date, start, end)
        if key in seen:
            continue
        seen.add(key)
        out.append({"load_date": load_date, "start": start, "end": end})
    return out


def healthnetca_range_labels(claim_loads):
    """{cell_monday: " (M/D-M/D)"} for the HealthNetCA backfill.

    A load is attributed to the Monday of the week it ran in (the 8/13 load of
    the 3/20-3/27 claims labels the 8/10 cell). A Sunday load rolls to the NEXT
    Monday, matching closest_weekday() / the Sun→next-Monday rule the markers
    use, so the label can't land on a different cell than the marker it belongs
    to. Contiguous ranges loaded in the same week collapse into one span — the
    files arrive in consecutive weekly chunks (3/20-3/27 then 3/27-4/3), so two
    chunks in one week read as "(3/20-4/3)" rather than a two-range list.
    Non-contiguous ranges are listed.
    """
    by_week = defaultdict(list)
    for e in claim_loads or ():
        ld = e["load_date"]
        monday = (ld + timedelta(days=1) if ld.weekday() == 6
                  else ld - timedelta(days=ld.weekday()))
        by_week[monday].append((e["start"], e["end"]))
    out = {}
    for monday, ranges in by_week.items():
        merged = []
        for start, end in sorted(set(ranges)):
            if merged and start <= merged[-1][1]:
                merged[-1] = (merged[-1][0], max(merged[-1][1], end))
            else:
                merged.append((start, end))
        out[monday] = " ({})".format(
            ", ".join(f"{s.month}/{s.day}-{e.month}/{e.day}" for s, e in merged))
    out.update(HEALTHNETCA_RANGE_LABEL_OVERRIDES)
    return out


def fetch_optum_raw_instances(since):
    r"""OptumPBMRx per-RAW file instances for the broken-out report rows.

    Scans TRGETL3 OptumPBMRx etl.Tape for RAW files at ALL ProcessStatus values
    (50 = loaded; <50, e.g. 42 = staging/loading) — NOT just 50 like
    fetch_tape_loads — so a RAW shows "L" while it's still loading. Keyed by
    (raw_n, data_date) so each monthly cycle's RAW is a distinct instance.
    Returns list of dicts: {raw_n:int|str, data_date:str, loaded:bool,
    load_date:datetime|None, latest:datetime}. Snap attribution is done by the
    caller. Per user 2026-06-30. `raw_n` is an int for the normal numeric RAW
    files (1,2,3,5,6) and a STRING for alphanumeric ad-hoc RAW tokens such as
    "53YR" (RAWLINGS_RAW53YR_06282026) — 2026-07-01 the user asked that these
    ad-hoc RAW loads surface too (they were previously ignored). Numeric RAWs
    outside {1,2,3,5,6} and every alphanumeric token surface as ad-hoc rows.
    """
    rx = re.compile(r"RAW\s*0*([0-9A-Z]+)_(\d{8})", re.IGNORECASE)
    inst = {}
    SEP = "\x1f"

    def touch(raw_n, ddate, dt, loaded):
        e = inst.get((raw_n, ddate))
        if e is None:
            e = inst[(raw_n, ddate)] = {"raw_n": raw_n, "data_date": ddate,
                                        "loaded": False, "load_date": None, "latest": dt}
        if dt and (e["latest"] is None or dt > e["latest"]):
            e["latest"] = dt
        if loaded:
            e["loaded"] = True
            if dt and (e["load_date"] is None or dt > e["load_date"]):
                e["load_date"] = dt

    q = (
        "SET NOCOUNT ON; SELECT FileName, ProcessStatus, "
        "CONVERT(varchar(19), FileLoadDate, 121) FROM [etl].[Tape] "
        f"WHERE FileName LIKE '%RAW%' AND FileLoadDate >= '{since:%Y-%m-%d}' "
        "ORDER BY FileLoadDate"
    )
    r = subprocess.run(
        ["sqlcmd", "-S", "TRGETL3", "-d", "OptumPBMRx", "-E", "-Q", q,
         "-W", "-s", SEP, "-h", "-1"],
        capture_output=True, text=True, check=False,
    )
    for line in r.stdout.splitlines():
        p = line.split(SEP)
        if len(p) < 3:
            continue
        m = rx.search(p[0] or "")
        if not m:
            continue
        try:
            status = int(p[1].strip())
        except ValueError:
            continue
        # Numeric token → int (normal RAW 1/2/3/5/6); alphanumeric → uppercase
        # string (ad-hoc, e.g. "53YR"). The caller keys ad-hoc off type/value.
        tok = m.group(1)
        raw_key = int(tok) if tok.isdigit() else tok.upper()
        touch(raw_key, m.group(2), parse_dt(p[2].strip()), loaded=(status == 50))
    return list(inst.values())


# CVSPBMRx Ad Hoc detection — the giveaway is FileSize (per user 2026-07-06).
# Regular weekly CVSPBMRx eligibility files run ~0.8–12 GB; a backfill/out-of-
# cycle file is dramatically larger (e.g. Tape 340, RAW_MEMBR_ELIG_GCP_20260611,
# ~148 GB). Any CVSPBMRx tape row above this threshold is treated as an Ad Hoc
# load and broken out onto its own row, NOT attributed to the regular Monday
# weekly cells. (For reference: the FileName date normally maps to the NEXT
# Monday's weekly load; an out-of-cycle giant file is Ad Hoc.)
CVSPBMRX_ADHOC_MIN_FILESIZE = 50_000_000_000   # 50 GB (well above normal ~12 GB)


def fetch_cvspbm_adhoc(since):
    """Return CVSPBMRx Ad Hoc (backfill/out-of-cycle) tape instances.

    Scans TRGETL3 CVSPBMRx.etl.Tape for rows whose FileSize exceeds
    CVSPBMRX_ADHOC_MIN_FILESIZE — the FileSize is the giveaway that a load is a
    backfill rather than a normal weekly delivery. Returns a list of dicts:
    {load_date: datetime|None, loaded: bool, size: int, filename: str}. A row is
    "loaded" once ProcessStatus == 50 (50 = loaded; 40/42 = staging/loading).
    """
    q = (
        "SET NOCOUNT ON; SELECT FileName, FileSize, ProcessStatus, "
        "CONVERT(varchar(19), FileLoadDate, 121) "
        "FROM [etl].[Tape] "
        f"WHERE FileSize >= {CVSPBMRX_ADHOC_MIN_FILESIZE} "
        f"AND FileLoadDate >= '{since:%Y-%m-%d}' "
        "ORDER BY FileLoadDate"
    )
    SEP = "\x1f"
    r = subprocess.run(
        ["sqlcmd", "-S", "TRGETL3", "-d", "CVSPBMRx", "-E", "-Q", q,
         "-W", "-s", SEP, "-h", "-1"],
        capture_output=True, text=True, check=False,
    )
    rows = []
    for line in r.stdout.splitlines():
        p = line.split(SEP)
        if len(p) < 4:
            continue
        try:
            size = int(p[1].strip())
            status = int(p[2].strip())
        except ValueError:
            continue
        rows.append({
            "filename":  (p[0] or "").strip(),
            "size":      size,
            "loaded":    status == 50,
            "load_date": parse_dt(p[3].strip()),
        })
    return rows


# Regular weekly CVSPBMRx eligibility file: RAW_MEMBR_ELIG_YYYYMMDD.TXT. The
# 8-digit token is the file's DATA date (a Saturday); the file is expected to
# load the FOLLOWING Monday (e.g. 20260627 Sat -> the 6/29 Mon cell). GCP/backfill
# files (RAW_MEMBR_ELIG_GCP_YYYYMMDD) don't match this pattern and are handled as
# Ad Hoc by fetch_cvspbm_adhoc.
CVSPBMRX_WEEKLY_RE = re.compile(r"RAW_MEMBR_ELIG_(\d{8})\.txt", re.I)
# Ad Hoc / GCP backfill filenames carry the same trailing _YYYYMMDD.TXT data
# date (with an optional GCP_ segment before it), e.g.
# RAW_MEMBR_ELIG_GCP_20260611.TXT or RAW_MEMBR_ELIG_20260709.TXT. Capture the
# 8-digit data date so a loaded Ad Hoc ✓ can be attributed to its Monday
# delivery cell (same rule as the weekly path), not the load-completion date.
CVSPBMRX_ADHOC_DATE_RE = re.compile(r"_(\d{8})\.txt", re.I)


def _cvspbmrx_cell_monday(data_date):
    """Delivery cell (Monday) for a CVSPBMRx weekly file's data date.
    The data date is a Saturday; delivery is the next Monday (0-day offset if the
    date already lands on a Monday)."""
    return data_date + timedelta(days=(0 - data_date.weekday()) % 7)


def fetch_cvspbmrx_weekly(since):
    """Return regular (non-Ad-Hoc) CVSPBMRx weekly eligibility tape files.

    Scans TRGETL3 CVSPBMRx.etl.Tape for RAW_MEMBR_ELIG_YYYYMMDD files below the
    Ad Hoc FileSize threshold. Each file's 8-digit DATA date maps to the next
    Monday's weekly cell (see _cvspbmrx_cell_monday), so a load that arrives late
    still lands on the correct week rather than the week it happened to load.
    Per user 2026-07-21: the 20260627 file loaded 7/19 (3 weeks late) but belongs
    on the 6/29 cell — the giveaway is which claims file the '0100 Stage' job
    loaded (its FileName carries the data date, same as the tape FileName here).
    Returns list of dicts: {filename, data_date: date, cell_monday: date,
    loaded: bool, load_date: datetime|None}.
    """
    q = (
        "SET NOCOUNT ON; SELECT FileName, FileSize, ProcessStatus, "
        "CONVERT(varchar(19), FileLoadDate, 121) "
        "FROM [etl].[Tape] "
        f"WHERE FileLoadDate >= '{since:%Y-%m-%d}' "
        "ORDER BY FileLoadDate"
    )
    SEP = "\x1f"
    r = subprocess.run(
        ["sqlcmd", "-S", "TRGETL3", "-d", "CVSPBMRx", "-E", "-Q", q,
         "-W", "-s", SEP, "-h", "-1"],
        capture_output=True, text=True, check=False,
    )
    rows = []
    for line in r.stdout.splitlines():
        p = line.split(SEP)
        if len(p) < 4:
            continue
        fn = (p[0] or "").strip()
        m = CVSPBMRX_WEEKLY_RE.search(fn)
        if not m:
            continue
        try:
            size = int(p[1].strip())
            status = int(p[2].strip())
        except ValueError:
            continue
        if size >= CVSPBMRX_ADHOC_MIN_FILESIZE:
            continue   # oversized backfill -> Ad Hoc, not a weekly cell
        try:
            dd = datetime.strptime(m.group(1), "%Y%m%d").date()
        except ValueError:
            continue
        rows.append({
            "filename":    fn,
            "data_date":   dd,
            "cell_monday": _cvspbmrx_cell_monday(dd),
            "loaded":      status == 50,
            "load_date":   parse_dt(p[3].strip()),
        })
    return rows


# ------------------------------------------------------------------
# Generalized "stage-file → data-date cell" attribution (per user 2026-07-21).
#
# Same principle proven for CVSPBMRx: a client's delivery CELL is driven by the
# DATA-THROUGH date parsed from its staged CLAIMS file (the '…0100 …Stage' job's
# file — its FileName == the tape FileName), NOT by when the load/snap happened,
# so a LATE/missed load still lands on the correct week/month. Target cell = the
# client's next scheduled delivery slot ON/AFTER the data-through date (weekly:
# next scheduled weekday; monthly: the expected day of the data-through month).
#
# On that target cell the DHT cert still wins (the existing cert lookup runs
# first in resolve_marker / determine_monthly); this only fixes the ✓ FALLBACK
# and its placement for a delivery that loaded+snapped but isn't yet certified —
# per user: "add the cert date on the next scheduled day ≥ the file's data-
# through date; only use the checkmark if the client doesn't get certified."
#
# DORMANT: all of these are FORCED_INACTIVE today (no current loads), so this is
# inert until they reactivate — VALIDATE against live output on the first real
# late load. NOT included (can't be done safely off the tape yet):
#   • HealthNetCA — legacy etl.Tape (rows back to 2007, no clean ProcessStatus/
#     FileLoadDate) + each delivery fans into many sub-files (XA…XZ). Needs a
#     RAMP-based load-status source; Stage job 'HealthNet 0100 Claims Stage'.
WEEKDAY_NAME_NUM = {"Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3,
                    "Friday": 4, "Saturday": 5, "Sunday": 6}

STAGE_FILE_CELL_CLIENTS = {
    # Stage job 'Tufts RX Claims 0100 Stage'. Weekly Monday. Claims file
    # Point32Health_Rawlings_6072_YYYYMMDD_hhmmss_RXECHF70CL.txt — YYYYMMDD is
    # the Monday itself (single date).
    "TuftsRx": dict(
        server="TRGETL3", db="TuftsRx", schedule="weekly",
        claim_re=re.compile(r"Point32Health_Rawlings_\d+_(\d{8})_\d+_RXECHF70CL", re.I),
        date_fmt="%Y%m%d",
        # 2026-07-28 (reactivation): the DHT cert is a MONTHLY milestone whose
        # StatTimestamps landed across July weeks, so cert_in_week would bleed the
        # monthly cert date onto the weekly file cells. Per user, the weekly cells
        # should show a ✓ per delivered file (monthly cell keeps the cert date):
        #   require_snap=False    → a loaded claims file (PS=50) alone earns the ✓
        #                           (the catch-up batch-loaded weeks 2/16–7/27 in
        #                           one pass, so the snap-within-7-days gate can't
        #                           line up per week).
        #   checkmark_over_cert   → the weekly-cell ✓ wins over cert_in_week so the
        #                           monthly cert doesn't overwrite the weekly ✓.
        require_snap=False, checkmark_over_cert=True,
    ),
    # Stage job 'Oscar Medical 0100 Stage'. Weekly Wednesday. Claims file
    # Oscar_Weekly_Claims_MMDDYYYY_MMDDYYYY.txt — a date RANGE; use the END
    # (through) date. Skip wide (>21d) ranges: the current loads are quarter-long
    # backfills, not the weekly cadence.
    "Oscar": dict(
        server="TRGETL2", db="Oscar", schedule="weekly",
        claim_re=re.compile(r"Oscar_Weekly_Claims_(\d{8})_(\d{8})", re.I),
        date_fmt="%m%d%Y", use_end=True, max_span_days=21,
    ),
    # Stage job 'Tufts PublicPlan 0100 Stage'. Monthly. Claims file
    # THPP_Rawlings_Claim_Provider_Extract_YYYYMMDD.txt (data date the 10th).
    "Tufts_PublicPlan": dict(
        server="TRGETL3", db="Tufts_PublicPlan", schedule="monthly",
        claim_re=re.compile(r"Claim_Provider_Extract_(\d{8})", re.I),
        date_fmt="%Y%m%d",
    ),
    # Stage job 'MMOH MHS 0100 Stage'. Monthly. Claim file GB391F.MMDDYYYY.txt
    # (GB390F is eligibility; date is MM DD YYYY, not YYYYMMDD).
    "MedicalMutualMHS": dict(
        server="TRGETL4", db="MedicalMutualMHS", schedule="monthly",
        claim_re=re.compile(r"GB391F\.(\d{8})", re.I),
        date_fmt="%m%d%Y",
    ),
}


def _stage_file_data_date(filename, cfg):
    """Parse the data-through date from a staged claims FileName per `cfg`.
    Returns (data_date, span_days) or (None, None). For a date RANGE the END
    date is the data-through date (per user) and span_days is end-start."""
    m = cfg["claim_re"].search(filename or "")
    if not m:
        return None, None
    try:
        if cfg.get("use_end"):
            start = datetime.strptime(m.group(1), cfg["date_fmt"]).date()
            end = datetime.strptime(m.group(2), cfg["date_fmt"]).date()
            return end, (end - start).days
        return datetime.strptime(m.group(1), cfg["date_fmt"]).date(), 0
    except (ValueError, IndexError):
        return None, None


def _next_scheduled_weekday(client, d):
    """The client's next scheduled delivery weekday ON/AFTER date `d` (0-day
    offset if `d` already lands on a scheduled weekday). Reproduces CVSPBMRx
    (Sat 6/27 → Mon 6/29) and TuftsRx (Mon 7/13 → Mon 7/13)."""
    wanted = {WEEKDAY_NAME_NUM[n] for n in WEEKLY_CLIENTS.get(client, [])}
    if not wanted:
        return d
    for delta in range(7):
        cand = d + timedelta(days=delta)
        if cand.weekday() in wanted:
            return cand
    return d


def fetch_stage_file_loads(client, cfg, since):
    """Loaded claims files for a STAGE_FILE_CELL_CLIENTS client, from its tape.
    Returns list of dicts: {filename, data_date, load_date, loaded, cell} where
    `cell` is the weekly Monday-week target for weekly clients (next scheduled
    weekday ≥ data date) or the (year, month) tuple for monthly clients."""
    q = (
        "SET NOCOUNT ON; SELECT FileName, ProcessStatus, "
        "CONVERT(varchar(19), FileLoadDate, 121) "
        "FROM [etl].[Tape] "
        f"WHERE FileLoadDate >= '{since:%Y-%m-%d}' ORDER BY FileLoadDate"
    )
    SEP = "\x1f"
    r = subprocess.run(
        ["sqlcmd", "-S", cfg["server"], "-d", cfg["db"], "-E", "-Q", q,
         "-W", "-s", SEP, "-h", "-1"],
        capture_output=True, text=True, check=False,
    )
    rows = []
    for line in r.stdout.splitlines():
        p = line.split(SEP)
        if len(p) < 3:
            continue
        fn = (p[0] or "").strip()
        dd, span = _stage_file_data_date(fn, cfg)
        if dd is None:
            continue
        max_span = cfg.get("max_span_days")
        if max_span is not None and span is not None and span > max_span:
            continue   # backfill / catch-up range, not the normal cadence
        try:
            status = int(p[1].strip())
        except ValueError:
            continue
        cell = (_next_scheduled_weekday(client, dd) if cfg["schedule"] == "weekly"
                else (dd.year, dd.month))
        rows.append({
            "filename":  fn,
            "data_date": dd,
            "loaded":    status == 50,
            "load_date": parse_dt(p[2].strip()),
            "cell":      cell,
        })
    return rows


# JHHC Passfile: the four monthly passfiles (Trauma/Subro × Active/Closed) all
# carry TableID = 5000 in TRGETL4.JohnsHopkins.etl.Tape. FileName looks like
# …\RawlingsGroup_PassFile_TraumaClosed_SSA_20260604.txt — capture the file
# type (TraumaClosed/…) and the 8-digit data date.
JHHC_PASSFILE_RE = re.compile(r"PassFile_([A-Za-z]+)_SSA_(\d{8})", re.I)


def fetch_jhhc_passfile_loads(since):
    """JHHC Passfile monthly load instances from TRGETL4.JohnsHopkins.etl.Tape.

    The four monthly passfiles (TraumaActive / TraumaClosed / SubroActive /
    SubroClosed) all carry TableID = 5000. A cycle is COMPLETE once all four
    load (ProcessStatus = 50). Returns a list of dicts:
    {filetype, data_date, load_date: datetime|None, loaded: bool}.

    Per user 2026-07-09: after the 'JHHC Passfile Email' RAMP job finishes the
    cell shows "TBL" (to be loaded); on each refresh we check here and, once all
    four files have loaded, the cell becomes ✓ on the FileLoadDate. Replaces the
    old TRGINTP3.JohnsHopkins  FileName LIKE '%PassFile%' tape source (the load
    now lands on TRGETL4).
    """
    q = (
        "SET NOCOUNT ON; SELECT FileName, ProcessStatus, "
        "CONVERT(varchar(19), FileLoadDate, 121) FROM [etl].[Tape] "
        f"WHERE TableID = 5000 AND FileLoadDate >= '{since:%Y-%m-%d}' "
        "ORDER BY FileLoadDate"
    )
    SEP = "\x1f"
    r = subprocess.run(
        ["sqlcmd", "-S", "TRGETL4", "-d", "JohnsHopkins", "-E", "-Q", q,
         "-W", "-s", SEP, "-h", "-1"],
        capture_output=True, text=True, check=False,
    )
    rows = []
    for line in r.stdout.splitlines():
        p = line.split(SEP)
        if len(p) < 3:
            continue
        m = JHHC_PASSFILE_RE.search(p[0] or "")
        if not m:
            continue
        try:
            status = int(p[1].strip())
        except ValueError:
            continue
        rows.append({
            "filetype":  m.group(1),
            "data_date": m.group(2),
            "loaded":    status == 50,
            "load_date": parse_dt(p[2].strip()),
        })
    return rows


# Map of client → (database name, snap-index source key).
# Server defaults to TRGETL3 (see TAPE_LOAD_SERVER for overrides).
TAPE_LOAD_SOURCES = {
    "OptumPBMRx":      ("OptumPBMRx",      "optumpbmrx"),
    "ESIPBMRx":        ("ESIPBMRx",        "esipbmrx"),
    "MedImpactPBMRx":  ("MedImpactPBMRx",  "medimpactpbmrx"),
    # JHHCPassfile is NO LONGER a generic tape source — it has its own
    # TableID=5000 / TRGETL4 lookup (fetch_jhhc_passfile_loads) and a dedicated
    # determine_monthly branch that gates ✓ on all four files loading. Keeping
    # it here would feed a premature ✓ into snap_idx from a single loaded file.
}

# Per-client SQL server override for TAPE_LOAD_SOURCES (default TRGETL3).
TAPE_LOAD_SERVER = {}

# Per-client FileName LIKE filter for TAPE_LOAD_SOURCES — restricts a shared
# client DB to just the intended feed's rows.
TAPE_LOAD_NAME_FILTER = {}

# Regex for state codes inside ESIPBMRx tape filenames (e.g. Rawlings_FL_, Rawlings_GA_)
ESIPBMRX_STATE_RE = re.compile(r"Rawlings_([A-Z]{2})_", re.I)

# Multi-week load detection: client → (TRGETL3 db, regex extracting week range
# from filename). The capture groups should be (start_yyyymmdd, end_yyyymmdd)
# or a single token uniquely identifying a week's worth of data. When recent
# loads contain >1 distinct week-key, the client label gets "(N weeks)".
MULTI_WEEK_CLIENTS = {}


def find_matching_jobs(client_id, jobs):
    primary = CLIENT_PRIMARY_KEY_OVERRIDE.get(client_id) or normalize(client_id)
    if not primary:
        return []
    targets = [primary] + [normalize(a) for a in CLIENT_ALIASES.get(client_id, [])]
    targets = [t for t in targets if t]
    matches = []
    for j in jobs:
        jn = normalize(j.get("JobName", "") or "")
        # Also test a digit-code-collapsed version of jn so JobNames with a
        # numeric step code between feed and sub-feed (e.g.
        # "CareFirst 0110 Facets Load" → "carefirst0110facetsload") still match
        # aliases like "carefirstfacets" via substring.
        jn_collapsed = re.sub(r"\d+", "", jn)
        fn = normalize((j.get("Feed")   or {}).get("FeedName", "") or "")
        cn = normalize((j.get("Client") or {}).get("ClientName", "") or "")
        if any(t in jn or t in jn_collapsed or t == fn or t == cn for t in targets):
            matches.append(j)
    real = [j for j in matches if not re.search(r"(logfile|sftp)", j.get("JobName", ""), re.I)]
    return real or matches


def build_snap_index(jobs, queue, snaps, tape_loads=None):
    """Return: date -> list of (normalized_source, datetime, destination, kind, job_name).

    `kind` ∈ {"snap", "load", "tape"} — distinguishes data source so a
    snap-only client doesn't get ✓ from a load-step completion.
    `job_name` — original RAMP JobName (queue entries only; empty for tape /
    RAMP snap-endpoint entries). Used by LOAD_NAME_REQUIRED to filter
    lookups so ancillary load jobs don't trigger L for cert clients.
    """
    by_date = defaultdict(list)

    if tape_loads:
        for src_key, rows in tape_loads.items():
            for row in rows:
                dt = row.get("FileLoadDate")
                if not dt or dt.year < 2026:
                    continue
                by_date[dt.date()].append((src_key, dt, "", "tape", ""))

    for s in snaps:
        # Accept: Success, Success/ManualFix, Success/NoWork — and Resolved
        # (user: "wait for the Ramp card to be marked Resolved").
        st = str(s.get("Status", ""))
        if not (st.startswith("Success") or st == "Resolved"):
            continue
        if s.get("TaskName") not in ("DeliverFlow", "DeliverSingle", "DeliverFlowSet"):
            continue
        end = parse_dt(s.get("End"))
        if not end or end.year < 2026:
            continue
        src = s.get("Source", "") or ""
        dest = s.get("Destination", "") or ""
        src_norm = normalize(re.sub(r"_(mine|snap|stage|load|rta)$", "", src, flags=re.I))
        if src_norm:
            by_date[end.date()].append((src_norm, end, dest, "snap", ""))

    job_by_id = {j.get("JobId"): j for j in jobs}
    for q in queue:
        if q.get("Status") not in ("Successful", "Resolved"):
            continue
        j = job_by_id.get(q.get("JobId"))
        if not j:
            continue
        jn = j.get("JobName", "") or ""
        jn_lower = jn.lower()
        # Index Load, Stage AND Snap jobs (the snap step is what triggers ✓
        # for SNAP_KIND_ONLY clients like ESIPBMRx, PrimePBMRx, Kaiser_GE).
        # EXTRA_INDEXED_JOBS lets through delivery-signal jobs that lack those
        # keywords (e.g. 'JHHC Passfile Email').
        if (not any(kw in jn_lower for kw in ("load", "stage", "snap", "mine"))
                and jn_lower.strip() not in EXTRA_INDEXED_JOBS):
            continue
        # skip log/sftp noise
        if re.search(r"(logfile|sftp|upload)", jn, re.I):
            continue
        # Index by START date — a load that begins 5/13 and finishes 5/14 is
        # for 5/13's data (per user: "AetnaHRP for 5/13 finished on 5/14, so
        # the 5/13 date should get a checkmark"). Keep EndDate as the
        # sort-time for latest-match purposes.
        start_dt = parse_dt(q.get("StartDate"))
        end_dt   = parse_dt(q.get("EndDate"))
        attribution_date = (start_dt.date() if start_dt
                            else (end_dt.date() if end_dt else None))
        if attribution_date is None:
            continue
        end = end_dt or start_dt
        # Index this completion under a JobName-derived key only. The Feed
        # name is unreliable because related-but-distinct clients can share
        # one Feed (e.g. "Kaiser WA 0110 Load" and "Kaiser WARX 0110 Load"
        # both have Feed "KaiserWA" → "kaiserwa", causing KaiserWARx loads
        # to falsely match Kaiser_WA). The JobName prefix uniquely identifies
        # which client the completion belongs to.
        kind = "snap" if re.search(r"\b(snap|mine)\b", jn, re.I) else "load"
        # Take everything before the first digit sequence in the JobName,
        # then strip trailing step words.
        m = re.match(r"^([^\d]+)", jn)
        prefix = m.group(1).strip() if m else jn
        prefix = re.sub(r"\s*(load|stage|snap|etl|daily|mine|start)\s*$", "",
                        prefix, flags=re.I).strip()
        k = normalize(prefix)
        if k and len(k) >= 4:
            by_date[attribution_date].append((k, end, "", kind, jn))
        # Also emit a "feed+sub-feed" key for JobNames matching
        # "<feed> <digits> <sub-feed> <step>" — e.g. "CareFirst 0110 Facets
        # Load" → "carefirstfacets". The base key (k) above captures only the
        # part before the digit code and loses sub-feed identity.
        sub_m = re.match(
            r"^([^\d]+?)\s+\d+\s+([^\d]+?)\s+"
            r"(post\s+snap|load|stage|snap|etl|daily|mine|start)\s*$",
            jn, re.I,
        )
        if sub_m:
            combined = normalize(sub_m.group(1) + sub_m.group(2))
            if combined and combined != k and len(combined) >= 4:
                by_date[attribution_date].append((combined, end, "", kind, jn))
    return by_date


def _tape_keys_only(client):
    """For tape-only clients, return the exact tape source key (no aliases)."""
    src_keys = {sk for (db, sk) in TAPE_LOAD_SOURCES.values() if client in TAPE_LOAD_SOURCES and TAPE_LOAD_SOURCES[client][1] == sk}
    return list(src_keys) if src_keys else []


def _src_matches_client(src_norm, client_keys, min_len=4):
    """Strict equality match between snap source and client keys.

    Substring matching produced too many false positives (e.g. snap source
    "aetnarx" was being attributed to clients like "AetnaRx_LegacyDMG", and
    "bcbsfl" to "BCBSFLEligibilityLoad"). Use explicit CLIENT_ALIASES to
    register variant source names per client.
    """
    if not src_norm or len(src_norm) < min_len:
        return False
    for k in client_keys:
        if k and k == src_norm:
            return True
    return False


def _kind_allowed(client, kind):
    """Return True if a snap_idx entry of this `kind` is allowed for `client`."""
    if client in LOAD_AS_DELIVERY_CLIENTS:
        return True   # accept anything (load/tape/snap)
    if client in SNAP_KIND_ONLY_CLIENTS:
        return kind == "snap"   # only real RAMP snap completions
    return True       # daily clients accept any kind


def _load_name_allowed(client, entry_jn, kind):
    """For clients in LOAD_NAME_REQUIRED, require the snap_idx entry's
    underlying JobName to contain one of the listed keywords. Entries
    without a JobName (RAMP snap-endpoint deliveries, tape rows) are
    rejected — they can't be verified against the whitelist.

    Exception: SNAP_ONLY clients use the snap-endpoint completion AS the
    delivery signal (no JobName available). For them, snap-kind entries
    bypass the filter. Per user 2026-05-18: MMOH was snapped 5/2 and
    should show ✓ via the RAMP snap-endpoint entry.

    Clients NOT in LOAD_NAME_REQUIRED bypass the filter entirely."""
    required = LOAD_NAME_REQUIRED.get(client)
    if not required:
        return True
    if kind == "snap" and client in SNAP_ONLY_CLIENTS:
        return True
    if not entry_jn:
        return False
    jn_lower = entry_jn.lower()
    return any(kw in jn_lower for kw in required)


def snap_on_day(client, day, snap_idx, window_days=0, forward_days=0):
    """Return latest snap/load completion datetime for `client` on `day`
    (or within `window_days` calendar days BEFORE `day`, or `forward_days`
    days AFTER for late completions).
    Applies SNAP_DESTINATION_FILTER and per-client kind filter.
    """
    if client in TAPE_ONLY_CLIENTS:
        if client not in TAPE_LOAD_SOURCES:
            return None
        wanted_src = TAPE_LOAD_SOURCES[client][1]
        # For SNAP_KIND_ONLY_CLIENTS like ESIPBMRx/MedImpactPBMRx the tape
        # entries represent LOADS — only count them as ✓ if the user actually
        # treats the load as delivery (LOAD_AS_DELIVERY_CLIENTS).
        if client in SNAP_KIND_ONLY_CLIENTS and client not in LOAD_AS_DELIVERY_CLIENTS:
            wanted_src = None
        best = None
        for offset in range(-forward_days, window_days + 1):
            check_day = day - timedelta(days=offset)
            for entry in snap_idx.get(check_day, ()):
                if wanted_src and entry[0] == wanted_src:
                    dt = entry[1]
                    if best is None or dt > best:
                        best = dt
        if best:
            return best
        # fall through to regular substring matching for SNAP_KIND_ONLY tape clients

    keys = [k for k in _keys_for_client(client) if k]
    if not keys:
        return None
    dest_filter = SNAP_DESTINATION_FILTER.get(client, "").lower()
    best = None
    for offset in range(-forward_days, window_days + 1):
        check_day = day - timedelta(days=offset)
        for entry in snap_idx.get(check_day, ()):
            src_norm, dt = entry[0], entry[1]
            dest = entry[2] if len(entry) > 2 else ""
            kind = entry[3] if len(entry) > 3 else "snap"
            jn   = entry[4] if len(entry) > 4 else ""
            if dest_filter and dest_filter not in dest.lower():
                continue
            if not _kind_allowed(client, kind):
                continue
            if not _load_name_allowed(client, jn, kind):
                continue
            if _src_matches_client(src_norm, keys):
                if best is None or dt > best:
                    best = dt
    return best


def snap_in_week(client, scheduled_day, snap_idx):
    """Return latest snap/load completion in the Mon-Fri week containing
    scheduled_day. A snap or load done anywhere this week (even before the
    scheduled day) counts as "completed this week" — e.g. Wellmark Tue cell
    picks up its Mon load, WellCare Fri cell picks up its Tue load.
    Monday-scheduled clients also pick up the prior weekend (Sat-Sun).
    """
    week_start = scheduled_day - timedelta(days=scheduled_day.weekday())
    week_end   = week_start + timedelta(days=4)
    # Always look back 2 days into the prior Sat+Sun so weekend loads
    # surface on any weekday cell, not just Monday-scheduled clients.
    # Per user 2026-05-18: Wellmark 0210 Claims Load on Sun 5/17 should
    # appear as L on the Tue 5/19 cell.
    window_start = week_start - timedelta(days=2)
    best = None
    d = window_start
    while d <= week_end:
        ts = snap_on_day(client, d, snap_idx)
        if ts and (best is None or ts > best):
            best = ts
        d += timedelta(days=1)
    return best


def latest_snap_this_month(client, snap_idx, year, month, on_or_before):
    """Latest snap completion datetime for `client` in (year,month), on/before today."""
    if client in TAPE_ONLY_CLIENTS:
        if client not in TAPE_LOAD_SOURCES:
            return None
        wanted_src = TAPE_LOAD_SOURCES[client][1]
        # For SNAP_KIND_ONLY tape clients (ESIPBMRx, MedImpactPBMRx) the tape
        # entries represent LOAD activity, not snap delivery — fall through to
        # regular kind-filtered matching so only actual snap completions count.
        if client in SNAP_KIND_ONLY_CLIENTS and client not in LOAD_AS_DELIVERY_CLIENTS:
            wanted_src = None
        if wanted_src:
            best = None
            for d, entries in snap_idx.items():
                if d.year != year or d.month != month or d > on_or_before:
                    continue
                for entry in entries:
                    if entry[0] == wanted_src:
                        dt = entry[1]
                        if best is None or dt > best:
                            best = dt
            if best:
                return best
        # fall through to regular matching for SNAP_KIND_ONLY tape clients

    keys = [k for k in _keys_for_client(client) if k]
    if not keys:
        return None
    dest_filter = SNAP_DESTINATION_FILTER.get(client, "").lower()
    best = None
    for d, entries in snap_idx.items():
        if d.year != year or d.month != month or d > on_or_before:
            continue
        for entry in entries:
            src_norm, dt = entry[0], entry[1]
            dest = entry[2] if len(entry) > 2 else ""
            kind = entry[3] if len(entry) > 3 else "snap"
            jn   = entry[4] if len(entry) > 4 else ""
            if dest_filter and dest_filter not in dest.lower():
                continue
            if not _kind_allowed(client, kind):
                continue
            if not _load_name_allowed(client, jn, kind):
                continue
            if _src_matches_client(src_norm, keys):
                if best is None or dt > best:
                    best = dt
    return best


def load_this_month(client, snap_idx, year, month, on_or_before):
    """Return latest LOAD-kind completion datetime for `client` in (year,month),
    on/before today. Considers kind ∈ {"load","tape"} only — not snap completions.
    Used so monthly cert-only clients don't show L from a stray
    /Ramp/Snap completion when the load job itself hasn't run yet.
    """
    if client in TAPE_ONLY_CLIENTS:
        if client not in TAPE_LOAD_SOURCES:
            return None
        wanted_src = TAPE_LOAD_SOURCES[client][1]
        best = None
        for d, entries in snap_idx.items():
            if d.year != year or d.month != month or d > on_or_before:
                continue
            for entry in entries:
                if entry[0] == wanted_src and (len(entry) <= 3 or entry[3] in ("load", "tape")):
                    dt = entry[1]
                    if best is None or dt > best:
                        best = dt
        return best

    keys = [k for k in _keys_for_client(client) if k]
    if not keys:
        return None
    best = None
    for d, entries in snap_idx.items():
        if d.year != year or d.month != month or d > on_or_before:
            continue
        for entry in entries:
            src_norm, dt = entry[0], entry[1]
            kind = entry[3] if len(entry) > 3 else "snap"
            jn   = entry[4] if len(entry) > 4 else ""
            if kind not in ("load", "tape"):
                continue
            if not _load_name_allowed(client, jn, kind):
                continue
            if _src_matches_client(src_norm, keys):
                if best is None or dt > best:
                    best = dt
    return best


def is_loading_today(client, queue, jobs):
    """True if a matching enabled job is currently Ready/Running.

    The set of "L"-triggering job types depends on client class:
      - LOAD_AS_DELIVERY clients (AetnaRx/AetnaHRP/etc.): only LOAD jobs.
        Once their load step finishes, ✓ takes over.
      - All other clients (CenteneRx/WellCareRx/etc.): LOAD or SNAP jobs.
        These need a cert to complete the cycle, so they stay L through
        both the load and snap steps until the cert lands.

    Stage / logfile / sftp / upload jobs never count as L.
    """
    matched = find_matching_jobs(client, jobs)
    load_only = client in LOAD_AS_DELIVERY_CLIENTS or client in L_ON_LOAD_ONLY_CLIENTS
    required_kwds = LOAD_NAME_REQUIRED.get(client)
    job_ids = set()
    for j in matched:
        if j.get("Enabled") != 1:
            continue
        jn = (j.get("JobName") or "").lower()
        if any(kw in jn for kw in ("stage", "logfile", "sftp", "upload")):
            continue
        is_load = "load" in jn and "snap" not in jn and "mine" not in jn
        # EDW feeds load via 'Wellpoint 0100 EDW Pull <feed>' — treat "pull" as
        # a load verb for those clients so the running pull surfaces "L".
        if (not is_load and client in PULL_AS_LOAD_CLIENTS
                and "pull" in jn and "snap" not in jn and "mine" not in jn):
            is_load = True
        is_snap = ("snap" in jn or "mine" in jn) and "load" not in jn
        # Per-client JobName whitelist (e.g. Rx clients where only MasterLoad
        # or Claims Load jobs should signal "loading"; COBC/IHP/ABII do not).
        if required_kwds and not any(kw in jn for kw in required_kwds):
            continue
        if load_only:
            if is_load:
                job_ids.add(j.get("JobId"))
        else:
            if is_load or is_snap:
                job_ids.add(j.get("JobId"))
    for q in queue:
        if q.get("JobId") in job_ids and q.get("Status") in ("Ready", "Running"):
            return True
    return False


def scan_adhoc_loads(queue, jobs, today, since, weekend_shift=True):
    """Scan RAMP queue for ad-hoc Load jobs that have no fixed schedule on
    the report. Each completed run shows up as a one-off row in the weekly
    section on the day it landed; in-flight runs surface today with 'L'.

    Per user 2026-06-04: MSPI Load jobs and HumanaRx Load are ad-hoc — they
    only appear when they actually run. Drop the row when nothing happened.

    Completions are attributed by **Start** date (matching build_snap_index) so
    a load that begins late and finishes after midnight still counts for the day
    it ran. `weekend_shift` (default True) surfaces Sat→Fri / Sun→Mon for ad-hoc
    rows; pass False when the caller wants the true (unshifted) load weekday.

    Returns a list of dicts: {"label", "day", "marker", "alert"}.
    """
    mspi_re   = re.compile(r"\bMSPI\b.*\bLoad\b", re.IGNORECASE)
    # HumanaRx returned to Monthly classification 2026-06-04 — kept the
    # function plumbing in case the user wants to add more ad-hoc patterns
    # later; for now only MSPI surfaces here.
    # Queue rows only carry JobId, not JobName — cross-reference via jobs.
    job_by_id = {j.get("JobId"): (j.get("JobName") or "") for j in jobs}
    # Walk newest → oldest (queue is QueueId DESC). For each (label, day, marker)
    # we keep the first occurrence; we also drop any Load Failure entry for a
    # label once a strictly newer non-failure entry (Ready/Running/Success/
    # Resolved) has been seen for the same label. Per user 2026-06-04: "if a
    # new MSPI load starts for the same client that has a failure, the failure
    # can be dropped."
    seen = set()
    rows = []
    nonfailure_seen = set()  # labels whose newest run is NOT a failure
    for q in queue:
        name = (job_by_id.get(q.get("JobId")) or "").strip()
        if not name:
            continue
        if not mspi_re.search(name):
            continue
        # Skip the snap-step variant; we only key off Load completion.
        if re.search(r"\b(snap|mine|logfile|sftp|stage)\b", name, re.I):
            continue
        status = (q.get("Status") or "").strip()
        start = parse_dt(q.get("StartDate"))
        end   = parse_dt(q.get("EndDate"))
        if not start or start.date() < since:
            continue
        base = re.sub(r"\s*\d+.*$", "", name).strip()
        base = re.sub(r"\s*MSPI\s*$", "", base, flags=re.IGNORECASE).strip()
        label = f"{base} MSPI"
        s_lower = status.lower()
        if s_lower.startswith("success") or s_lower == "resolved":
            day = start.date()
            marker, alert = "✓", False
            nonfailure_seen.add(label)
        elif s_lower in ("ready", "running"):
            day, marker, alert = today, "L", False
            nonfailure_seen.add(label)
        elif s_lower == "failed":
            # Drop the failure if a strictly newer (already-seen) non-failure
            # run exists for the same label.
            if label in nonfailure_seen:
                continue
            day = start.date()
            marker, alert = "Load Failure", True
        else:
            continue
        if weekend_shift:
            if day.weekday() == 5:
                day -= timedelta(days=1)
            elif day.weekday() == 6:
                day += timedelta(days=1)
        key = (label, day, marker)
        if key in seen:
            continue
        seen.add(key)
        rows.append({"label": label, "day": day, "marker": marker, "alert": alert})
    return rows


# The two jobs whose joint completion = a finished "Kaiser Submission" (per
# user 2026-06-09). Exact lowercased names — must NOT match the separate
# 'Kaiser Pareo Audit Submission Upload'. These are Logfile/Upload jobs, which
# the normal index/L logic excludes, so they're tracked here explicitly.
KAISER_SUBMISSION_JOBS = (
    "kaiser pareo submission logfile",
    "kaiser pareo submission upload",
)


def scan_kaiser_submission(queue, jobs, today, since):
    """Per-day status for the daily 'Kaiser Submission' row.

    ✓ on a day when BOTH 'Kaiser Pareo Submission Logfile' AND
    'Kaiser Pareo Submission Upload' completed (Successful/Resolved) that day;
    'L' on today while either is still Ready/Running. Weekend completions snap
    to the closest weekday (Sat→Fri, Sun→Mon).

    Returns (done_days: set[date], running_today: bool).
    """
    job_by_id = {j.get("JobId"): (j.get("JobName") or "").strip().lower() for j in jobs}

    def snap_wd(d):
        if d.weekday() == 5:
            return d - timedelta(days=1)
        if d.weekday() == 6:
            return d + timedelta(days=1)
        return d

    done = {name: set() for name in KAISER_SUBMISSION_JOBS}
    running_today = False
    for q in queue:
        name = job_by_id.get(q.get("JobId"))
        if name not in KAISER_SUBMISSION_JOBS:
            continue
        status = (q.get("Status") or "").strip().lower()
        start = parse_dt(q.get("StartDate"))
        end   = parse_dt(q.get("EndDate"))
        dt = end or start
        if status.startswith("success") or status == "resolved":
            if dt and dt.date() >= since:
                done[name].add(dt.date())
        elif status in ("ready", "running"):
            running_today = True
    log_days = done["kaiser pareo submission logfile"]
    up_days  = done["kaiser pareo submission upload"]
    done_days = {snap_wd(d) for d in (log_days & up_days)}
    return done_days, running_today


def find_unconfigured_masterload_clients(jobs):
    """Scan RAMP jobs for `<Client> MasterLoad 0110 Load` entries whose
    derived client name isn't recognised by any existing config dict.

    Returns list of dicts:
      {"raw": str, "normalized": str, "pbmrx": bool, "enabled": bool}

    Per user 2026-06-03: keep an eye on RAMP for these and add new
    implementations to the report. PBMRx hits get SNAP_KIND_ONLY behavior.
    """
    out = []
    seen = set()
    known = (set(DAILY_CLIENTS) | set(WEEKLY_CLIENTS.keys())
             | set(MONTHLY_CLIENTS) | {KAISER_PREPAY_CLIENT})
    known_keys = set()
    for c in known:
        for k in _keys_for_client(c):
            if k:
                known_keys.add(k)
    pat = re.compile(r"^(.+?)\s*MasterLoad\s+0110\s+Load\s*$", re.IGNORECASE)
    for j in jobs:
        name = (j.get("JobName") or "").strip()
        if not name:
            continue
        m = pat.match(name)
        if not m:
            continue
        client_raw = m.group(1).strip()
        norm = normalize(client_raw)
        if not norm or norm in seen:
            continue
        # Match against known client keys with substring (matches aliases too).
        matched = any(norm == k or norm in k or k in norm for k in known_keys)
        if matched:
            continue
        seen.add(norm)
        out.append({
            "raw": client_raw,
            "normalized": norm,
            "pbmrx": "PBMRx" in name,
            "enabled": j.get("Enabled") == 1,
        })
    return out


# Ancillary sub-pipeline modifiers that should NOT count toward primary-load
# inactivity (per user 2026-06-03: the inactive-label rule targets the main
# 0100/0110 cycle only, not COBC/RTA/etc. sub-jobs).
_ANCILLARY_JOB_TOKEN_RE = re.compile(
    r"\b(cobc|rta|abii|ihp|cms\s*referral|mmsea|covid|monthly\s+claim|adjustment)\b",
    re.IGNORECASE,
)


def auto_inactive_from_ramp(jobs):
    """Set of clients (canonical names from our config) whose primary
    0100 Stage / 0110 Load jobs in RAMP are ALL Inactive (Enabled=0).
    Excludes any client whose name starts with "Kaiser" (per user
    2026-06-03 — Kaiser feeds have their own snap/inactive semantics).
    Excludes ancillary job modifiers via _ANCILLARY_JOB_TOKEN_RE.
    """
    primary_state = defaultdict(lambda: {"active": 0, "inactive": 0})
    for j in jobs:
        name = (j.get("JobName") or "").strip()
        if not name:
            continue
        if "0100 Stage" not in name and "0110 Load" not in name:
            continue
        if _ANCILLARY_JOB_TOKEN_RE.search(name):
            continue
        jn = normalize(name)
        is_active = (j.get("Enabled") == 1)
        for client in (set(DAILY_CLIENTS) | set(WEEKLY_CLIENTS.keys())
                       | set(MONTHLY_CLIENTS) | {KAISER_PREPAY_CLIENT}):
            if client.startswith("Kaiser"):
                continue
            for k in _keys_for_client(client):
                if k and (k in jn or jn.startswith(k)):
                    if is_active:
                        primary_state[client]["active"] += 1
                    else:
                        primary_state[client]["inactive"] += 1
                    break
    out = set()
    for client, state in primary_state.items():
        if state["inactive"] > 0 and state["active"] == 0:
            out.add(client)
    return out - AUTO_INACTIVE_EXCLUDE


# Clients to keep OUT of the RAMP auto-inactive sweep even though their
# 0100/0110 jobs are disabled. CareFirstRx is offboarding so its load jobs are
# disabled in RAMP, but it is still delivering + certifying (certified 7/2/26);
# forcing it Inactive made determine_monthly() step 0 short-circuit before the
# cert lookup, hiding the cert date. Per user 2026-07-02: "CareFirstRx is no
# longer Inactive and was certified today." With it excluded, the cert wins
# naturally; if a future month has no cert + no enabled jobs, has_inactive_jobs
# still surfaces "Inactive" correctly.
# Tufts_PublicPlan added 2026-07-27: reactivated per user, being certified via a
# catch-up ticket. If its load jobs are still disabled in RAMP, keep it out of the
# auto-inactive sweep so the cert wins (same rationale as CareFirstRx).
# TuftsRx added 2026-07-28: reactivated per user (catch-up cert today). Same
# rationale — keep it out of the auto-inactive sweep so the monthly cert/pin wins.
# ESIPBMRx + Tufts_Audit_CIT added 2026-07-28: reactivated per user ("only
# HealthNetCA & MedicalMutualMHS still Inactive") — keep them out of the sweep so
# a still-disabled RAMP job can't re-flag them Inactive; their real cert/snap
# history renders instead.
# HealthNetCA added 2026-08-14: reactivated per user (3/20-3/27 backfill loaded,
# certified today). Its 0100 Stage / 0110 Load jobs may still be disabled in RAMP,
# so keep it out of the sweep and let the cert/load activity win.
AUTO_INACTIVE_EXCLUDE = {"CareFirstRx", "Tufts_PublicPlan", "TuftsRx",
                        "ESIPBMRx", "Tufts_Audit_CIT", "HealthNetCA"}


def has_inactive_jobs(client, jobs, cert_idx, snap_idx, today):
    """True if RAMP has no enabled jobs for the client AND there has been
    no DHT certification or snap completion in the last 30 days.
    Clients in FORCED_INACTIVE are always inactive.
    """
    if client in FORCED_INACTIVE:
        return True
    cutoff = today - timedelta(days=30)
    for k in _keys_for_client(client):
        for dt, _ in cert_idx.get(k, ()):
            if dt.date() >= cutoff:
                return False
    for d in list(snap_idx.keys()):
        if d < cutoff:
            continue
        for entry in snap_idx[d]:
            src_norm = entry[0]
            for k in _keys_for_client(client):
                if k and (k in src_norm or src_norm in k):
                    return False
    matched = find_matching_jobs(client, jobs)
    if not matched:
        return False
    enabled = [j for j in matched if j.get("Enabled") == 1]
    return len(enabled) == 0


def has_recent_failure(client, queue, jobs, today):
    """True if a LOAD job (not stage/snap/logfile) failed in the last 3 days
    with no Successful or Resolved run after it. Stage failures and
    intermediate-step failures don't trigger "Load Failure" — per user,
    only true load-step failures count.

    Clients in LOAD_NAME_REQUIRED also restrict to those keyword patterns,
    so an ancillary "RTA Load" / "COBC Load" failure doesn't trigger
    Load Failure for the main client cycle. Per user 2026-05-19.
    """
    matched = find_matching_jobs(client, jobs)
    required_kwds = LOAD_NAME_REQUIRED.get(client)
    job_ids = set()
    for j in matched:
        jn = (j.get("JobName") or "").lower()
        # Only LOAD-named jobs count for failure detection
        if "load" not in jn:
            continue
        # COBC and ABII are ancillary file pipelines — a COBC/ABII load failure
        # must NOT mark the client's claims delivery as a failure (per user
        # 2026-06-18: CareFirstRx's "Load Failure" was the COBC load, not the
        # 'CareFirstRx 0110 Load' claims job). Excluded globally for all clients.
        if any(kw in jn for kw in ("stage", "snap", "mine", "logfile",
                                   "sftp", "upload", "cobc", "abii")):
            continue
        if required_kwds and not any(kw in jn for kw in required_kwds):
            continue
        job_ids.add(j.get("JobId"))
    if not job_ids:
        return False
    cutoff = today - timedelta(days=3)
    latest_failed = {}
    latest_success = {}
    for q in queue:
        jid = q.get("JobId")
        if jid not in job_ids:
            continue
        end = parse_dt(q.get("EndDate") or q.get("StartDate"))
        if not end or end.date() < cutoff:
            continue
        if q.get("Status") == "Failed":
            if jid not in latest_failed or end > latest_failed[jid]:
                latest_failed[jid] = end
        elif q.get("Status") in ("Successful", "Resolved"):
            # Resolved counts as success (per user: wait for the queue card
            # to be marked Resolved as a recovery from a prior failure).
            if jid not in latest_success or end > latest_success[jid]:
                latest_success[jid] = end
    for jid, fail_dt in latest_failed.items():
        succ_dt = latest_success.get(jid)
        if succ_dt is None or fail_dt > succ_dt:
            return True
    return False


# ============================================================
#                       ADO matching per client
# ============================================================
def build_ticket_index(tickets, jobs):
    """client_name -> latest ticket dict + a list of all this-month tickets (for monthly placement)."""
    latest = {}
    placements = defaultdict(list)
    for t in tickets:
        cid = t["client"]
        if not cid:
            continue
        # for monthly placement, use ticket Created date (or snap completion for Load and Snap)
        created = parse_dt(t["created"])
        placed = created.date() if created else None
        if placed:
            placements[cid].append(placed)

        existing = latest.get(cid)
        if existing is None or (parse_dt(t["changed"]) or datetime.min) > (parse_dt(existing["changed"]) or datetime.min):
            latest[cid] = t
    return latest, placements


# ============================================================
#                       calendar planning
# ============================================================
def display_name(client, monthly=False, extra_suffix=""):
    # `monthly` no longer prepends an "M - " tag — removed per user 2026-06-16.
    base = CLIENT_DISPLAY_NAME.get(client, client)
    suffix = CLIENT_SUFFIXES.get(client, "")
    return f"{base}{suffix}{extra_suffix}"


def esipbmrx_states_for_week(week_start, week_end, esipbmrx_tape_rows):
    """Return sorted list of state codes loaded between week_start and week_end."""
    states = set()
    for row in esipbmrx_tape_rows or ():
        dt = row.get("FileLoadDate")
        if not dt or not (week_start <= dt.date() <= week_end):
            continue
        m = ESIPBMRX_STATE_RE.search(row.get("FileName", "") or "")
        if m:
            states.add(m.group(1).upper())
    return sorted(states)


def count_multi_week_loads(client, week_start, week_end, multi_week_loads):
    """Return number of distinct week-key tuples captured in filenames loaded
    in the [week_start, week_end] window for this client."""
    info = MULTI_WEEK_CLIENTS.get(client)
    if not info:
        return 0
    _, pattern = info
    rows = multi_week_loads.get(client, [])
    keys = set()
    for row in rows:
        dt = row.get("FileLoadDate")
        if not dt or not (week_start <= dt.date() <= week_end):
            continue
        m = pattern.search(row.get("FileName", "") or "")
        if m:
            keys.add(m.groups())
    return len(keys)


def month_weeks(year, month):
    """Return Mon-Fri weeks rendered on this month's tab.

    A week intersecting two months is claimed by whichever month has 3+
    weekdays in it (per user 2026-06-01: "whichever month has 3 or more
    weekdays, that week should be joined with the remaining days of the
    other month"). The claiming month keeps the full 5-day week, including
    carryover dates from the adjacent month. The losing month drops that
    partial week — it will appear on the claiming month's tab instead.

    Example: 6/29 Mon and 6/30 Tue (only 2 June weekdays) drop off June and
    join July's Week 1, which has 7/1, 7/2, 7/3 (3 July weekdays).
    """
    cal = calendar.Calendar(firstweekday=0)
    out = []
    for wk in cal.monthdatescalendar(year, month):
        weekdays = wk[:5]  # Mon-Fri only
        in_month = sum(1 for d in weekdays if d.month == month)
        if in_month >= 3:
            out.append(list(weekdays))
    return out


def next_monday_if_weekend(d):
    """If d is Sat/Sun, return the following Monday; otherwise d."""
    if d.weekday() == 5:
        return d + timedelta(days=2)
    if d.weekday() == 6:
        return d + timedelta(days=1)
    return d


def closest_weekday(d):
    """Snap d to the nearest Mon-Fri. Sat → previous Fri, Sun → next Mon."""
    if d.weekday() == 5:
        return d - timedelta(days=1)
    if d.weekday() == 6:
        return d + timedelta(days=1)
    return d


def us_federal_holidays(year):
    """Return {date: name} for the 11 US federal holidays in `year`.
    Fixed-date holidays that fall on a Saturday are observed the prior
    Friday; on Sunday, the following Monday."""
    out = {}

    def nth_weekday(month, weekday, n):
        cal = calendar.monthcalendar(year, month)
        days = [w[weekday] for w in cal if w[weekday] != 0]
        return date(year, month, days[n - 1])

    def last_weekday(month, weekday):
        cal = calendar.monthcalendar(year, month)
        days = [w[weekday] for w in cal if w[weekday] != 0]
        return date(year, month, days[-1])

    def observed(d):
        if d.weekday() == 5:
            return d - timedelta(days=1)
        if d.weekday() == 6:
            return d + timedelta(days=1)
        return d

    fixed = [
        (date(year,  1,  1), "New Year's Day"),
        (date(year,  6, 19), "Juneteenth"),
        (date(year,  7,  4), "Independence Day"),
        (date(year, 11, 11), "Veterans Day"),
        (date(year, 12, 25), "Christmas Day"),
    ]
    for d, name in fixed:
        d_obs = observed(d)
        out[d_obs] = name + (" (observed)" if d_obs != d else "")

    out[nth_weekday(1,  0, 3)]  = "MLK Day"               # 3rd Mon Jan
    out[nth_weekday(2,  0, 3)]  = "Presidents' Day"       # 3rd Mon Feb
    out[last_weekday(5, 0)]     = "Memorial Day"          # last Mon May
    out[nth_weekday(9,  0, 1)]  = "Labor Day"             # 1st Mon Sep
    out[nth_weekday(10, 0, 2)]  = "Columbus Day"          # 2nd Mon Oct
    out[nth_weekday(11, 3, 4)]  = "Thanksgiving"          # 4th Thu Nov
    return out


def nmsp_mmsea_date(year, month):
    return next_monday_if_weekend(date(year, month, 15))


def average_cert_day(client, cert_idx):
    """Return the most common day-of-month this client certified on (across history)."""
    days = []
    for key in _keys_for_client(client):
        for dt, status in cert_idx.get(key, ()):
            if status == "Certified":
                days.append(dt.day)
    if not days:
        return None
    return int(round(sum(days) / len(days)))


def is_friday_or_later_in_week(today, scheduled_day):
    """True if today is on/past Friday of the same week as scheduled_day."""
    week_start = scheduled_day - timedelta(days=scheduled_day.weekday())
    week_friday = week_start + timedelta(days=4)
    return today >= week_friday


def plan_calendar(year, month, cert_idx, snap_idx, latest_tickets, monthly_placements,
                  ramp_jobs, ramp_queue, esipbmrx_tape=None, multi_week_loads=None,
                  aetna_nmsp_loads=None, optumpbmrx_tape=None, evernorth_claims_pending=False,
                  optum_raw_instances=None, cvspbm_adhoc=None, jhhc_passfile_loads=None,
                  cvspbm_weekly=None, stage_file_loads=None,
                  healthnetca_ranges=None):
    """Return ((sections, weeks)) layout.

    sections: dict {kind: dict {date: [(label, marker, alert)] }}
      kind is one of 'daily', 'weekly', 'monthly', 'kaiser'.
    weeks: list of weeks (each = list of 5 dates Mon-Fri or None).
    """
    today = date.today()
    weeks = month_weeks(year, month)
    all_days = [d for wk in weeks for d in wk if d is not None]

    # CVSPBMRx Ad Hoc (backfill) state — an oversized/out-of-cycle tape file is
    # loading through the same 'CVSPBMRx … Load' RAMP job as the regular weekly
    # feed, so is_loading_today would otherwise stamp "L" on the regular Monday
    # cell. When an Ad Hoc file is in flight, that L belongs to the backfill, not
    # the weekly cycle — suppress the regular L (see resolve_marker) and surface
    # the backfill on its own "CVSPBMRx (Ad Hoc)" row instead. Per user 2026-07-06.
    cvspbm_adhoc = cvspbm_adhoc or []
    # 2026-08-18 per user: "CVSPBMRx is still loading the Ad Hoc file, not the 8/17
    # file." The tape-only test above wasn't enough: the Ad Hoc's tape row (TapeID
    # 346, the 147 GB RAW_MEMBR_ELIG_20260709) already reads ProcessStatus 50 while
    # its 'CVS PBM RX 0110 Load' card (JobId 1942, QueueId 1413143) is STILL
    # Ready — started 7/30 11:19 and never finished. So no Ad Hoc was "in flight"
    # by the tape test, the L wasn't suppressed, and is_loading_today stamped it on
    # the 8/17 Monday cell as if this week's file were loading.
    #
    # An in-flight load card that STARTED BEFORE the current week can't be this
    # week's delivery — it's a carry-over/backfill still working an older file. Its
    # "L" belongs on the Ad Hoc row, not the weekly cell. Self-clearing: the moment
    # that card finishes (or a load starts this week) the weekly "L" resumes.
    cur_week_monday = today - timedelta(days=today.weekday())
    _cvs_load_job_ids = {
        j.get("JobId") for j in find_matching_jobs("CVSPBMRx", ramp_jobs)
        if j.get("Enabled") == 1
        and "load" in (j.get("JobName") or "").lower()
        and "stage" not in (j.get("JobName") or "").lower()
    }
    # 2026-08-21 per user: "the CVSPBMRx load/Snap is the AdHoc file." The re-run
    # of the failed backfill STARTED TODAY, so the "started before this week" test
    # no longer catches it and the L would land back on the weekly cell. The
    # data-driven giveaway: no weekly RAW_MEMBR_ELIG file has arrived for the
    # current cycle (newest weekly data date still maps to an earlier Monday
    # cell), so in-flight CVS work cannot be this week's delivery — it's the Ad Hoc.
    # The SNAP is included as well (the user's words: "the load/Snap is the AdHoc
    # file"), so the Ad Hoc row keeps the in-flight "L" through the snap step
    # instead of dropping off the moment the load card turns Successful.
    # Self-clearing: the moment a weekly file lands, the weekly "L" resumes.
    _cvs_snap_job_ids = {
        j.get("JobId") for j in find_matching_jobs("CVSPBMRx", ramp_jobs)
        if j.get("Enabled") == 1
        and "snap" in (j.get("JobName") or "").lower()
        and "load" not in (j.get("JobName") or "").lower()
    }
    _cvs_newest_weekly_cell = max(
        (f["cell_monday"] for f in (cvspbm_weekly or [])), default=None)
    _cvs_no_weekly_this_cycle = (_cvs_newest_weekly_cell is None
                                 or _cvs_newest_weekly_cell < cur_week_monday)
    cvspbm_carryover_load = None
    for q in ramp_queue:
        if (q.get("JobId") in (_cvs_load_job_ids | _cvs_snap_job_ids)
                and q.get("Status") in ("Ready", "Running", "Processing")):
            sd = parse_dt(q.get("StartDate"))
            if sd and (sd.date() < cur_week_monday or _cvs_no_weekly_this_cycle):
                if cvspbm_carryover_load is None or sd > cvspbm_carryover_load:
                    cvspbm_carryover_load = sd
    # 2026-08-20: the Ad Hoc row must be able to say "Load Failure", not just
    # "L". QueueId 1413143 ran 7/30 → 8/18 and FAILED; its tape row went back to
    # un-loaded, so the un-loaded branch below would have kept painting a stale
    # "L" on today forever. When the newest CVS load card is a recent Failed one
    # and nothing is Ready/Running, carry the failure (on its failure date)
    # instead. Self-clearing: a re-run puts a card back in Ready/Running and the
    # row reverts to "L" on today. Per user: "the Ad Hoc for CVSPBMRx should be
    # 'Load Failure'."
    # NB: the Enabled==1 filter above is deliberately dropped here — RAMP's CVS
    # PBM RX jobs were switched off (Enabled=0) after this failure, and a
    # disabled job's failed card still happened.
    _cvs_any_load_job_ids = {
        j.get("JobId") for j in find_matching_jobs("CVSPBMRx", ramp_jobs)
        if "load" in (j.get("JobName") or "").lower()
        and "stage" not in (j.get("JobName") or "").lower()
    }
    cvspbm_adhoc_failed = None
    _cvs_load_running = False
    for q in ramp_queue:
        if q.get("JobId") not in _cvs_any_load_job_ids:
            continue
        st = q.get("Status")
        if st in ("Ready", "Running", "Processing"):
            _cvs_load_running = True
        elif st == "Failed":
            ed = parse_dt(q.get("EndDate")) or parse_dt(q.get("StartDate"))
            if ed and (today - ed.date()).days <= 14:
                if cvspbm_adhoc_failed is None or ed.date() > cvspbm_adhoc_failed:
                    cvspbm_adhoc_failed = ed.date()
    # an in-flight load OR snap (cvspbm_carryover_load) means the cycle is moving
    # again — don't resurrect the old failure on the Ad Hoc row
    if _cvs_load_running or cvspbm_carryover_load is not None:
        cvspbm_adhoc_failed = None
    cvspbm_tape_adhoc_loading = any(not a.get("loaded") for a in cvspbm_adhoc)
    cvspbm_adhoc_loading = (cvspbm_tape_adhoc_loading
                            or cvspbm_carryover_load is not None)

    # CVSPBMRx weekly ✓ is attributed by the file's DATA date (from the FileName /
    # '0100 Stage' job), NOT by when the load/snap finished — so a late load lands
    # on the correct week. Build {cell_monday: snap_dt} for each weekly file that
    # has both loaded (tape ProcessStatus 50) AND snapped (a CVSPBMRx snap-kind
    # completion at/after the file's FileLoadDate). Per user 2026-07-21: the
    # 20260627 file loaded 7/19 but must ✓ on 6/29, not 7/20. cvspbm_delivered
    # overrides snap_in_week for CVSPBMRx in resolve_marker.
    cvspbm_weekly = cvspbm_weekly or []
    cvspbm_snap_dts = sorted(
        e[1]
        for d in snap_idx
        for e in snap_idx[d]
        if len(e) > 3 and e[3] == "snap" and _src_matches_client(e[0], ["cvspbmrx"])
    )
    cvspbm_delivered = {}   # cell_monday(date) -> ✓ snap datetime
    for f in cvspbm_weekly:
        if not f.get("loaded") or not f.get("load_date"):
            continue
        # This file's snap = the first CVSPBMRx snap at/after its load, within a
        # 7-day window (the snap always follows the load within a day or two; the
        # cap prevents an old file — whose real snap predates the snap-history
        # window — from matching an unrelated later snap and over-marking a stale
        # cell). Weekly cadence is 7 days, so the window can't bleed into the
        # next file's snap.
        ld = f["load_date"]
        snap_dt = next((s for s in cvspbm_snap_dts
                        if ld <= s <= ld + timedelta(days=7)), None)
        if snap_dt is None:
            continue
        cell = f["cell_monday"]
        if cell not in cvspbm_delivered or snap_dt > cvspbm_delivered[cell]:
            cvspbm_delivered[cell] = snap_dt

    # General stage-file ✓ FALLBACK for STAGE_FILE_CELL_CLIENTS (TuftsRx, Oscar,
    # Tufts_PublicPlan, MedicalMutualMHS). {client: {cell: snap_dt}} where cell is
    # the scheduled weekday date (weekly) or (year, month) (monthly). Same load+
    # snap-within-7-days rule as CVSPBMRx; cert still wins upstream. Per user
    # 2026-07-21. Dormant until these clients leave FORCED_INACTIVE.
    stage_file_loads = stage_file_loads or {}
    # {cell_monday: " (M/D-M/D)"} for the HealthNetCA claims backfill — see
    # healthnetca_range_labels() and the HEALTHNETCA_* config block.
    healthnetca_ranges = healthnetca_ranges or {}
    stage_delivered = {}
    for sf_client, rows in stage_file_loads.items():
        keys = [k for k in _keys_for_client(sf_client) if k]
        # require_snap=False (e.g. TuftsRx): a loaded claims file alone is the
        # delivery signal — the client batch-loaded many weeks at once, so a
        # per-week snap-within-7-days match can't line up. Default keeps the
        # snap gate (CVSPBMRx-style: load AND a client snap within 7 days).
        require_snap = STAGE_FILE_CELL_CLIENTS.get(sf_client, {}).get("require_snap", True)
        sf_snap_dts = sorted(
            e[1] for d in snap_idx for e in snap_idx[d]
            if len(e) > 3 and e[3] == "snap" and _src_matches_client(e[0], keys)
        )
        cells = {}
        for f in rows:
            if not f.get("loaded") or not f.get("load_date"):
                continue
            ld = f["load_date"]
            if require_snap:
                snap_dt = next((s for s in sf_snap_dts
                                if ld <= s <= ld + timedelta(days=7)), None)
                if snap_dt is None:
                    continue
            else:
                snap_dt = ld   # load itself is the delivery signal
            cell = f["cell"]
            if cell not in cells or snap_dt > cells[cell]:
                cells[cell] = snap_dt
        stage_delivered[sf_client] = cells

    daily   = defaultdict(list)
    weekly  = defaultdict(list)
    monthly = defaultdict(list)
    kaiser  = defaultdict(list)

    def is_kaiser_feed(c):
        return c.startswith("Kaiser") and c != KAISER_PREPAY_CLIENT

    def alert_state(client, day, marker):
        """Pink-fill the Date cell when client is in a problem state.
        - Load Failure / Inactive / Failed → always shade
        - No Data → only shade when today is >7 days past the expected END day
        - Kaiser feeds → shade only when past Friday without cert that week
        """
        # A real cert date is a successful delivery — NEVER pink-shade it, even
        # when a stale current-state signal (has_recent_failure /
        # has_inactive_jobs) would otherwise fire on today's cell. Per user
        # 2026-07-30: HealthNewEngland certified 7/30 (weekly Thu) but its Thursday
        # cell rendered pink because a recent load-failure / disabled-job signal
        # shaded today's cell despite the cert date already showing.
        if isinstance(marker, date):
            return False
        # Problem-state markers ALWAYS shade pink, regardless of client class.
        # Per user 2026-05-19: "Kaiser_HI does have a load failure, but is not
        # in pink." The Kaiser-feed branch below previously short-circuited
        # before this check, hiding the Load Failure shade for Kaiser feeds.
        # "Snap" added for snap-disabled clients (e.g. Kaiser_AmbM).
        # "Missing" = the scheduled delivery never arrived (added 2026-08-24 for
        # CenteneFidelis 8/19) — a problem state, always pink.
        # "Discontinued" = the client has permanently stopped delivering (added
        # 2026-08-26 for NCState's final 8/26 cell) — shaded like "Inactive".
        if marker in ("Load Failure", "Inactive", "Failed", "Deployment", "Snap",
                      "Missing", "Discontinued"):
            return True
        # "Empty" = a delivered-but-empty file (happens occasionally). Per user
        # 2026-07-08 it is NOT a problem state -> show the label, never pink.
        if marker == "Empty":
            return False
        # Implementation phase — never pink (new client; lack of jobs/data
        # would otherwise trip has_inactive_jobs).
        if marker == "Implementation":
            return False
        if is_kaiser_feed(client):
            # Snap-only / load-as-delivery / forced-inactive Kaiser feeds
            # have their own per-client semantics — fall through to the
            # generic marker rules so a valid ✓ doesn't get shaded pink
            # just because it's Friday, and Inactive still shades correctly.
            # Kaiser_WARx: per user 2026-07-17, do NOT escalate its uncertified
            # "L" to a pink alert on/after Friday — show a plain "L". Exempt it
            # from the Friday-pink rule (Load Failure / Inactive still shade via
            # the problem-state check above).
            if (client not in SNAP_KIND_ONLY_CLIENTS
                    and client not in LOAD_AS_DELIVERY_CLIENTS
                    and client not in FORCED_INACTIVE
                    and client != "Kaiser_WARx"):
                if is_friday_or_later_in_week(today, day) and not isinstance(marker, date):
                    return True
                return False
            if client == "Kaiser_WARx":
                return False
        if marker == "No Data":
            if client in FORCE_SHADE_NO_DATA:
                return True
            rng = MONTHLY_EXPECTED_DAY_RANGE.get(client)
            if rng:
                try:
                    # Shade when today is >7 days past the START of the expected
                    # window for the CELL's month (not today's). Future-month
                    # templates must not pink-shade No Data.
                    start_day = date(day.year, day.month,
                                     min(rng[0], calendar.monthrange(day.year, day.month)[1]))
                    return (today - start_day).days > 7
                except ValueError:
                    return False
            return False
        # Weekly clients — pink-shade an empty past-scheduled cell when the
        # scheduled day is 3+ days in the past with no cert / activity.
        # Monthly clients keep the 7-day threshold above. Daily Aetnas are
        # exempt (per user 2026-06-03 — they're in DAILY_CLIENTS so this
        # guard is belt-and-suspenders). IMPLEMENTATION_LOAD_ONLY clients
        # are also exempt (per user 2026-06-05 — ElevanceMMMRx showed a
        # pink "!" because we're not actively working that client yet).
        # NYShip_Rx is included even though it's not in WEEKLY_CLIENTS — it
        # uses NYSHIP_DAYS placement but is weekly-equivalent (per user
        # 2026-06-05). Threshold changed from `> 3` to `>= 3` same day so
        # a Tuesday cell pinks on Friday.
        weekly_eligible = (client in WEEKLY_CLIENTS) or client == "NYShip_Rx"
        if (not marker
                and weekly_eligible
                and client not in {"AetnaHRP", "AetnaRCE", "AetnaRx",
                                   "NCStateAetna"}
                and client not in IMPLEMENTATION_LOAD_ONLY_CLIENTS
                and (today - day).days >= 3):
            return True
        # Past-day cells reflect historical activity — don't shade them with
        # the current Inactive/Failure state of the client. Only the marker
        # text itself (handled above) determines shading for past days.
        if day < today:
            return False
        if has_inactive_jobs(client, ramp_jobs, cert_idx, snap_idx, today):
            return True
        if day == today and has_recent_failure(client, ramp_queue, ramp_jobs, today):
            # Suppress the failure shade if the client is currently loading
            # again — the retry has superseded the stale failed state.
            if not is_loading_today(client, ramp_queue, ramp_jobs):
                return True
        return False

    today_week_start = today - timedelta(days=today.weekday())
    today_week_end   = today_week_start + timedelta(days=4)

    def resolve_marker(client, day, allow_checkmark, allow_week_window):
        """Return the Date-column marker for a client placed on `day`.
          - date           : DHT certified that day (or within the 7-day cycle for weekly)
          - "Load Failure" : recent failed load (precedence over L)
          - "L"            : currently loading and scheduled day is in current week
          - "✓"            : snap/load completed (only if allow_checkmark)
          - "Inactive"     : forced inactive client
          - ""             : nothing yet
        """
        # Date-gated inactivation (FORCED_INACTIVE_FROM): any cell on/after the
        # client's cutoff is "Inactive" regardless of cert/load activity; cells
        # before the cutoff fall through and render their normal history. Checked
        # first so it outranks the cert lookup. Per user 2026-07-29
        # (TuftsMedPref & Tufts_Audit_CIT → Inactive 7/6/26 forward).
        fi_from = FORCED_INACTIVE_FROM.get(client)
        if fi_from and day >= fi_from:
            return "Inactive"
        # Forced-inactive clients show "Inactive" only for the current day
        # and future days — past-day cells keep their normal markers so a
        # newly-disabled client doesn't retroactively erase its prior cert
        # dates / ✓ from when it was active. Per user 2026-05-18: HealthNetCA
        # past Mondays should still show their cert dates. Past-day cells with
        # no activity fall back to "Inactive" at the end of this function.
        forced_inactive = client in FORCED_INACTIVE
        if forced_inactive and day >= today:
            return "Inactive"
        # checkmark_over_cert weekly stage-file clients (TuftsRx): show ✓ for a
        # delivered file BEFORE the cert lookup, so the monthly cert's per-week
        # StatTimestamps don't bleed a cert date onto the weekly file cells. The
        # monthly (determine_monthly) cell still carries the cert date. Per user
        # 2026-07-28. Only fires on the weekly path (allow_week_window) and only
        # for cells with a delivered file in stage_delivered.
        if (allow_week_window
                and client in STAGE_FILE_CELL_CLIENTS
                and STAGE_FILE_CELL_CLIENTS[client].get("checkmark_over_cert")
                and day in stage_delivered.get(client, {})):
            return "✓"
        # CELL_ACTIVITY_AFTER: this cell ignores cert/load/snap activity dated
        # on or before the cutoff (it belongs to an earlier, late delivery).
        act_after = CELL_ACTIVITY_AFTER.get((client, day))

        def _after_cutoff(ts):
            """True when `ts` (datetime or None) counts for this cell."""
            if ts is None:
                return False
            return act_after is None or ts.date() > act_after

        # cert on the exact day
        ts = cert_on_day(client, day, cert_idx)
        if _after_cutoff(ts):
            return ts.date()
        # cert anywhere in the 7-day cycle starting at this scheduled day
        if allow_week_window:
            ts = cert_in_week(client, day, cert_idx)
            if _after_cutoff(ts):
                return ts.date()
        # Forced-inactive clients never show a live "L" / "✓" / "Load Failure"
        # from resumed loading — only a genuine cert date (returned above) or
        # "Inactive". Per user 2026-07-29: HealthNetCA is loading again, but all
        # its (Monday) cells must stay Inactive until the data is reviewed for
        # activate/Snap/Certify. Without this, a current-week past cell would pick
        # up "L" from the load activity and beat the FORCED_INACTIVE fallback
        # (previously patched cell-by-cell, e.g. the 7/20 MANUAL_OVERRIDE).
        if forced_inactive:
            return "Inactive"
        in_current_week = today_week_start <= day <= today_week_end

        # Daily clients: on today's cell, L outranks Failure (active retry
        # is more useful than a stale failure). On past days, only ✓ applies.
        # Monday cells also look back Sat+Sun to catch weekend ETL loads
        # (e.g. NCStateAetna's Saturday 'Aetna RCE 310 ETL Load' → Monday ✓).
        if not allow_week_window:
            if day == today:
                if is_loading_today(client, ramp_queue, ramp_jobs):
                    return "L"
                if has_recent_failure(client, ramp_queue, ramp_jobs, today):
                    return "Load Failure"
            win_back = 2 if day.weekday() == 0 else 0     # Monday → look back Sat/Sun
            # KaiserPrePayCOB renders Sat/Sun loads on their own (Sat)/(Sun)
            # injected rows attached to Fri/Mon — don't let the Mon cell's
            # regular row pick up the Sun load too (would double-count).
            if client == KAISER_PREPAY_CLIENT:
                win_back = 0
            # Any past daily cell looks forward 1 day to catch a load
            # that crossed midnight or ran as a next-day catch-up
            # (e.g. AetnaHRP load for 5/13 finishing 5/14 → ✓ on 5/13).
            win_forward = 1 if day < today else 0
            if allow_checkmark and _after_cutoff(snap_on_day(
                    client, day, snap_idx,
                    window_days=win_back, forward_days=win_forward)):
                return "✓"
            # Past-day "L" for SNAP_KIND_ONLY daily clients (e.g. AetnaHRP):
            # if the load ran on `day` but no snap completion yet AND a job
            # is currently active for the client, the cycle is still in
            # progress — keep that prior weekday's cell at "L". Per user
            # 2026-05-19: "AetnaHRP did not Snap yet from the 5/18/26 load …
            # The 5/18/26 HRP should still be an 'L'."
            if (client in SNAP_KIND_ONLY_CLIENTS
                    and day < today
                    and today_week_start <= day <= today_week_end):
                keys = list(_keys_for_client(client))
                load_on_d = any(
                    len(entry) > 3
                    and entry[3] in ("load", "tape")
                    and _src_matches_client(entry[0], keys)
                    for entry in snap_idx.get(day, ())
                )
                if load_on_d and is_loading_today(client, ramp_queue, ramp_jobs):
                    return "L"
            # Forced-inactive clients with no prior cert/✓ on this past day
            # fall back to "Inactive" rather than leaving the cell blank.
            # Past Mondays with real cert dates already returned above.
            if forced_inactive:
                return "Inactive"
            return ""

        # Implementation-load-only clients (e.g. ElevanceMMMRx): not being
        # actively worked yet — show L only while a job is currently running
        # in RAMP; clear to blank as soon as load+snap completes. Never ✓,
        # never a cert date. Per user 2026-06-03.
        if client in IMPLEMENTATION_LOAD_ONLY_CLIENTS:
            if in_current_week and is_loading_today(client, ramp_queue, ramp_jobs):
                return "L"
            return ""

        # EverNorthRx: weekly CLAIMS row. The daily Masterload also stages
        # eligibility-only files (ELIG / COBC / TRR / ACUM / ABII), so flag "L"
        # ONLY while actual claims (ESI_PAID_CLAIMS_*) are staged and NOT yet
        # loaded (evernorth_claims_pending) — the L drops the moment the Masterload
        # Load consumes them, rather than lingering through snap-awaiting-cert.
        # Cert (handled above) still wins; a claims-load failure still surfaces.
        # Per user 2026-06-29 (tightened from the per-week claims check same day).
        if client == "EverNorthRx":
            if in_current_week:
                if has_recent_failure(client, ramp_queue, ramp_jobs, today):
                    return "Load Failure"
                if evernorth_claims_pending:
                    return "L"
            return ""

        # Weekly clients: currently-loading L outranks past failure (active
        # retry is more useful than a stale Failed entry). Cert already
        # took priority above, so cert dates aren't displaced.
        if in_current_week:
            # CVSPBMRx: while an Ad Hoc backfill is loading, the running
            # 'CVSPBMRx … Load' job is processing the backfill (not the weekly
            # cycle) — don't attribute its "L" to the regular Monday cell. The
            # backfill gets its own "CVSPBMRx (Ad Hoc)" row.
            suppress_l = client == "CVSPBMRx" and cvspbm_adhoc_loading
            if not suppress_l and is_loading_today(client, ramp_queue, ramp_jobs):
                return "L"
            if has_recent_failure(client, ramp_queue, ramp_jobs, today):
                return "Load Failure"
        # CVSPBMRx: ✓ is driven by the weekly file's DATA date (cvspbm_delivered,
        # built above from the tape FileName), NOT the snap completion date — so
        # a late load stays on its correct week (20260627 -> 6/29 cell, not the
        # 7/20 week it finally loaded). Do NOT fall back to snap_in_week here;
        # that back-attribution is exactly what put the ✓ on the wrong cell.
        # Per user 2026-07-21. Current-week "L" while loading is handled above.
        if client == "CVSPBMRx":
            cell_mon = day - timedelta(days=day.weekday())
            if allow_checkmark and cell_mon in cvspbm_delivered:
                return "✓"
        elif (client in STAGE_FILE_CELL_CLIENTS
              and STAGE_FILE_CELL_CLIENTS[client]["schedule"] == "weekly"):
            # Weekly stage-file clients (TuftsRx, Oscar): ✓ FALLBACK driven by the
            # claims file's DATA date (stage_delivered), placed on the scheduled
            # weekday cell — NOT the load/snap date. NOT gated on allow_checkmark:
            # cert clients like Oscar aren't in SNAP_ONLY_CLIENTS, but the user
            # wants a ✓ when snapped-but-not-certified. Cert already won upstream
            # (returns before this), so this only fires when uncertified. No
            # snap_in_week fallback (that would misplace a late load). Per user
            # 2026-07-21.
            if day in stage_delivered.get(client, {}):
                return "✓"
        elif allow_checkmark:
            if _after_cutoff(snap_in_week(client, day, snap_idx)):
                return "✓"
        elif in_current_week and client not in LOADING_L_ONLY_CLIENTS:
            # Weekly cert client (not snap-only) — if load/snap activity has
            # happened this week and the cert hasn't landed yet, stay L
            # (per user: "CenteneRx & WellCareRx should have an 'L' since
            # they have not been certified").
            #
            # LOADING_L_ONLY_CLIENTS (the Kaiser Pareo weekly feeds) opt OUT of
            # this fallback: their "L" comes only from is_loading_today above, so
            # the cell clears the moment the load finishes and stays blank
            # through the Wednesday PM snap until the cert lands. Per user
            # 2026-08-05.
            if _after_cutoff(snap_in_week(client, day, snap_idx)):
                return "L"
        # Forced-inactive weekly clients with no prior cert in this cycle
        # fall back to "Inactive" rather than leaving the cell blank.
        # Per user 2026-05-19: "HealthNetCA & TuftsRx for 5/18 is somehow
        # not marked 'Inactive' anymore."
        if forced_inactive:
            return "Inactive"
        return ""

    def expected_end_day(client):
        """Return the END day of the monthly client's expected delivery range.
        Falls back to historical avg, then 15th."""
        rng = MONTHLY_EXPECTED_DAY_RANGE.get(client)
        if rng:
            return rng[1]
        avg = average_cert_day(client, cert_idx)
        return avg if avg is not None else 15

    def _kaiser_amb_anchor():
        """Closest Thursday (weekday 3) to the 15th of (year, month)."""
        anchor = date(year, month, 15)
        diff = 3 - anchor.weekday()  # diff is in [-3, 3]
        return anchor + timedelta(days=diff)

    def _apply_weekday_spread(d, client):
        """If the client has a MONTHLY_PLACEMENT_WEEKDAY override, snap d to
        the matching weekday of d's Mon-Fri work-week. Clamps within month."""
        target_wd = MONTHLY_PLACEMENT_WEEKDAY.get(client)
        if target_wd is None:
            return d
        week_start = d - timedelta(days=d.weekday())
        candidate = week_start + timedelta(days=target_wd)
        if candidate.year == year and candidate.month == month:
            return candidate
        return d

    def determine_monthly(client):
        """Return (placement_date, marker) for a monthly client.
        Cert/snap dates remain on their actual date; all other markers are
        anchored to the client's expected delivery day (end of its range)."""
        # Kaiser_Amb feeds all anchor to the closest Thursday to the 15th
        # (per user 2026-05-26).
        if client.startswith("Kaiser_Amb"):
            placeholder = _kaiser_amb_anchor()
        else:
            # expected placement day (end of range; or avg if no range; fallback 15th)
            expected_d = expected_end_day(client)
            try:
                placeholder = date(year, month, min(expected_d, calendar.monthrange(year, month)[1]))
            except ValueError:
                placeholder = date(year, month, 15)
            if client in CLOSEST_WEEKDAY_CLIENTS:
                placeholder = closest_weekday(placeholder)
            else:
                placeholder = next_monday_if_weekend(placeholder)
            placeholder = _apply_weekday_spread(placeholder, client)
        # Per-month placement-day override (MONTHLY_PLACEMENT_DAY_OVERRIDES): the
        # delivery date slipped for this one month, but every marker still resolves
        # from live data. Applied to placeholder AND expected_date below.
        day_ov = MONTHLY_PLACEMENT_DAY_OVERRIDES.get((client, year, month))
        if day_ov is not None:
            placeholder = day_ov
        # 0) Forced-inactive clients always show "Inactive" on expected day
        if client in FORCED_INACTIVE:
            return placeholder, "Inactive"
        # 0-) Date-gated inactivation (FORCED_INACTIVE_FROM): once the monthly
        # placement day is on/after the client's cutoff, the monthly row reads
        # "Inactive"; earlier months keep their real cert history. Per user
        # 2026-08-19 (TuftsRx → Inactive 8/17/26 forward for the HarvardPilgrim
        # implementation), so Aug 2026 (anchored to the 10th) still renders live
        # state and Sept 2026 forward goes Inactive.
        fi_from_m = FORCED_INACTIVE_FROM.get(client)
        if fi_from_m and placeholder >= fi_from_m:
            return placeholder, "Inactive"
        # 0a) Explicit one-off placement override (per-client day + marker).
        # Highest precedence so EDW feeds can stay on 5/20 even though they
        # certified 5/21, or AetnaQNXT can show on 5/19 with L. The override
        # only applies when its date falls in the current calendar month.
        # Marker "AUTO" anchors placement to ov_day but resolves the marker
        # dynamically: cert this month → cert date; currently loading or
        # loaded this month → "L"; else "No Data".
        override = MONTHLY_PLACEMENT_OVERRIDES.get(client)
        if override:
            ov_day, ov_marker = override
            if ov_day.year == year and ov_day.month == month:
                if ov_marker == "AUTO":
                    # Per-month cert (NOT global latest_cert) so a past-month
                    # AUTO override still surfaces that month's own cert.
                    c_month = latest_cert_in_month(client, cert_idx, year, month, on_or_before=today)
                    if c_month:
                        return ov_day, c_month.date()
                    if is_loading_today(client, ramp_queue, ramp_jobs):
                        return ov_day, "L"
                    if has_recent_failure(client, ramp_queue, ramp_jobs, today):
                        return ov_day, "Load Failure"
                    if load_this_month(client, snap_idx, year, month, today):
                        return ov_day, "L"
                    return ov_day, "No Data"
                return ov_day, ov_marker
        # 0b) Snap-disabled clients (load runs but snap step is disabled in RAMP)
        # show marker "Snap" with pink shading on their expected day.
        if client in SNAP_DISABLED_CLIENTS:
            return placeholder, "Snap"
        # expected_date mirrors placeholder for Kaiser_Amb feeds and any
        # spread-adjusted client; otherwise recompute from the range end.
        if client.startswith("Kaiser_Amb"):
            expected_date = placeholder
        else:
            try:
                expected_date = date(year, month, min(expected_d, calendar.monthrange(year, month)[1]))
            except ValueError:
                expected_date = date(year, month, 15)
            if client in CLOSEST_WEEKDAY_CLIENTS:
                expected_date = closest_weekday(expected_date)
            else:
                expected_date = next_monday_if_weekend(expected_date)
            if expected_date.month != month:
                expected_date = date(year, month, calendar.monthrange(year, month)[1])
            expected_date = _apply_weekday_spread(expected_date, client)
        if day_ov is not None:
            expected_date = day_ov

        # 1) Certified in THIS month → place on that month's actual cert date.
        # Per-month lookup (NOT the global latest_cert) so a PAST month surfaces
        # its OWN cert even after the client certified again in a later month —
        # otherwise every month but the client's most-recent one showed "No Data".
        c_month = latest_cert_in_month(client, cert_idx, year, month, on_or_before=today)
        if c_month:
            d = c_month.date()
            if d.year != year or d.month != month:
                # MONTHLY_CERT_MONTH_REMAP cert from another calendar month (a
                # late delivery certified after the month rolled over): the cert
                # date can't be a placement day on this tab (rows whose date
                # falls outside the month are dropped), so anchor it to this
                # month's expected day and render the cert date there.
                d = expected_date
            else:
                d = next_monday_if_weekend(d) if d.weekday() >= 5 else d
            return d, c_month.date()

        # 1-stage) Monthly stage-file clients (Tufts_PublicPlan, MedicalMutualMHS):
        # ✓ FALLBACK on the expected day of the data-through MONTH, driven by the
        # claims file's data date (stage_delivered) rather than the load/snap
        # month — so a late load stays in its correct month. Cert (step 1) already
        # won; this only shows when snapped-but-not-certified. Per user 2026-07-21.
        if (client in STAGE_FILE_CELL_CLIENTS
                and STAGE_FILE_CELL_CLIENTS[client]["schedule"] == "monthly"
                and (year, month) in stage_delivered.get(client, {})):
            return expected_date, "✓"

        # 1a) Per-month marker override (e.g. force "L" until cert) — placed on
        # the client's expected day. A cert this month (handled above) always
        # wins, so this only shows while the month is uncertified.
        mm_ov = MONTHLY_MONTH_MARKER_OVERRIDES.get((client, year, month))
        if mm_ov is not None:
            return expected_date, mm_ov

        # "Today" rules only fire when today actually falls within the target
        # month — future-month templates must not pull placements back into the
        # current week (per user 2026-05-26: BCBSFL Elig was missing from June
        # because is_loading_today returned today=5/26 and the row got dropped).
        today_in_month = (today.year == year and today.month == month)

        # 1b) Monthly stage-file clients (Tufts_PublicPlan, MedicalMutualMHS) in
        # the CURRENT month: delivery evidence is the client's own etl.Tape
        # (step 1-stage) plus the DHT cert (step 1) — never the RAMP queue. A
        # load that finishes Successful/Resolved but whose data is then BACKED
        # OUT leaves no tape row, yet latest_snap_this_month / load_this_month
        # still saw the queue row and painted a false ✓ on the load day. Per user
        # 2026-08-04: "the Tufts_PublicPlan load failed and was backed out —
        # remove from update and keep on the 10th until the next load."
        # So stay anchored to the expected day: "L" only while a load is
        # genuinely in flight, "Load Failure" on a real failure, otherwise
        # "No Data" until the next load actually reaches the tape. Self-
        # correcting — step 1-stage flips it to ✓ the moment the tape confirms a
        # loaded claims file, and step 1 to the cert date after that. Scoped to
        # today_in_month so already-published past-month cells don't change.
        if (today_in_month and client in STAGE_FILE_CELL_CLIENTS
                and STAGE_FILE_CELL_CLIENTS[client]["schedule"] == "monthly"):
            if is_loading_today(client, ramp_queue, ramp_jobs):
                return expected_date, "L"
            if has_recent_failure(client, ramp_queue, ramp_jobs, today):
                return expected_date, "Load Failure"
            return expected_date, "No Data"

        # JHHC Passfile (per user 2026-07-09): the 'JHHC Passfile Email' RAMP
        # job delivers four files (Trauma/Subro × Active/Closed) that then load
        # into TRGETL4.JohnsHopkins.etl.Tape (TableID = 5000). Lifecycle:
        #   email job finished           → "TBL" (to be loaded)
        #   all four files loaded (PS=50) → ✓ on the latest FileLoadDate
        # The load is checked live on every refresh via fetch_jhhc_passfile_loads.
        if client == "JHHCPassfile":
            month_loads = [r for r in (jhhc_passfile_loads or [])
                           if r.get("loaded") and r.get("load_date")
                           and r["load_date"].year == year
                           and r["load_date"].month == month]
            if month_loads:
                # Group by data date; a cycle is complete once all four
                # distinct file types have loaded.
                types_by_dd = defaultdict(set)
                latest_by_dd = {}
                for r in month_loads:
                    dd = r["data_date"]
                    types_by_dd[dd].add(r["filetype"].lower())
                    if dd not in latest_by_dd or r["load_date"] > latest_by_dd[dd]:
                        latest_by_dd[dd] = r["load_date"]
                complete = [dd for dd, t in types_by_dd.items() if len(t) >= 4]
                if complete:
                    d = max(latest_by_dd[dd] for dd in complete).date()
                    if d.weekday() >= 5:
                        d = next_monday_if_weekend(d)
                    return d, "✓"
                # Some files loaded but not all four yet → still to-be-loaded.
                d = max(latest_by_dd.values()).date()
                if d.weekday() >= 5:
                    d = next_monday_if_weekend(d)
                return d, "TBL"
            # No load rows yet this month. If the 'JHHC Passfile Email' job has
            # finished (Successful/Resolved) this month, the files are delivered
            # and awaiting load → "TBL".
            email_dt = None
            for d_, entries in snap_idx.items():
                if d_.year != year or d_.month != month:
                    continue
                for e in entries:
                    if e and e[0] == "jhhcpassfileemail":
                        if email_dt is None or e[1] > email_dt:
                            email_dt = e[1]
            if email_dt:
                d = email_dt.date()
                if d.weekday() >= 5:
                    d = next_monday_if_weekend(d)
                return d, "TBL"
            # Email still running → L; recent failure → Load Failure; else No Data.
            if today_in_month and is_loading_today(client, ramp_queue, ramp_jobs):
                return today, "L"
            if today_in_month and has_recent_failure(client, ramp_queue, ramp_jobs, today):
                return today, "Load Failure"
            return expected_date, "No Data"

        # Cert-only clients (BCBSKS/BCBSKSMedAdv/BCBSSCRx) stay on expected
        # day until DHT cert lands.
        if client in MONTHLY_CERT_ONLY_CLIENTS:
            # Blank-until-cert clients (BCBSKSMedAdv) ignore mid-process
            # activity entirely — they stay empty until a real cert arrives.
            if client in MONTHLY_BLANK_UNTIL_CERT:
                return expected_date, ""
            if today_in_month and is_loading_today(client, ramp_queue, ramp_jobs):
                return expected_date, "L"
            if has_recent_failure(client, ramp_queue, ramp_jobs, today):
                return expected_date, "Load Failure"
            # Show L only when the actual LOAD job has run this month (not
            # just a stray /Ramp/Snap completion). Per user 2026-05-15:
            # "BCBSSC RX 0110 Load has not started loading yet this month —
            # don't show as loading."
            ln = load_this_month(client, snap_idx, year, month, today)
            if ln:
                return expected_date, "L"
            return expected_date, "No Data"

        # 2) Currently loading right now → today + L (outranks past completion).
        if today_in_month and is_loading_today(client, ramp_queue, ramp_jobs):
            return today, "L"

        # 3) Recent failure today → today + Load Failure
        if today_in_month and has_recent_failure(client, ramp_queue, ramp_jobs, today):
            return today, "Load Failure"

        # 4) Successfully snapped this month → ✓ on the snap date.
        # Catches monthly clients whose snap completed but cert is pending.
        sn = latest_snap_this_month(client, snap_idx, year, month, today)
        if sn:
            d = sn.date()
            if d.weekday() >= 5:
                d = next_monday_if_weekend(d)
            return d, "✓"

        # 4b) No snap yet, but the LOAD job has run this month → keep at "L"
        # on the expected day. Per user 2026-05-15: Kaiser_AmbN has loaded
        # but not snapped yet — should stay "L" instead of "No Data".
        ln = load_this_month(client, snap_idx, year, month, today)
        if ln:
            return expected_date, "L"

        # 5) ADO delivery ticket changed this week → move up to that day with L.
        # Only valid in the current month's rendering.
        ticket = latest_tickets.get(client) if today_in_month else None
        if ticket:
            ch = parse_dt(ticket.get("changed", "")) or parse_dt(ticket.get("created", ""))
            if ch and today_week_start <= ch.date() <= today_week_end \
               and ch.date() <= today:
                state = (ticket.get("state") or "").lower()
                if state in ("active", "in progress", "new", "committed"):
                    d = ch.date()
                    if d.weekday() >= 5:
                        d = next_monday_if_weekend(d)
                    return d, "L"

        # 5) Inactive / No Data on expected day
        if has_inactive_jobs(client, ramp_jobs, cert_idx, snap_idx, today):
            return expected_date, "Inactive"
        return expected_date, "No Data"

    def place(bucket, kind, client, day, marker_override=None, label_prefix=""):
        # `from_manual` is True only when the marker comes from the
        # MANUAL_OVERRIDES dict — caller-passed marker_override values (e.g.
        # determine_monthly results) are still eligible for sticky-cert restore.
        from_manual = False
        if marker_override is None:
            mov = MANUAL_OVERRIDES.get((client, day))
            if mov is not None:
                marker_override = mov
                from_manual = True
        # File-gated override: pin the marker only while the named file has not
        # landed in the client's "Loaded" folder yet. Self-clearing — no manual
        # cleanup once the load completes. MANUAL_OVERRIDES still outranks it.
        if marker_override is None:
            fgo = FILE_GATED_OVERRIDES.get((client, day))
            if fgo is not None and not file_gate_satisfied(fgo[1], fgo[2]):
                marker_override = fgo[0]
                from_manual = True
        if marker_override is not None:
            marker = marker_override
        elif kind == "daily":
            marker = resolve_marker(client, day, allow_checkmark=True, allow_week_window=False)
        elif kind == "weekly":
            allow_check = client in SNAP_ONLY_CLIENTS
            marker = resolve_marker(client, day, allow_checkmark=allow_check, allow_week_window=True)
        elif kind == "kaiser":
            marker = resolve_marker(client, day, allow_checkmark=True, allow_week_window=False)
        else:  # monthly
            marker = resolve_marker(client, day, allow_checkmark=True, allow_week_window=False)
        # Soft, self-clearing label (SOFT_OVERRIDES): applies only while nothing
        # has happened yet, and is treated as manual so the sticky-cert cache
        # can't restore a stale ✓ over it. Real activity ("L" / "✓" / cert date)
        # always wins. Per user 2026-08-13 (AetnaRCE / NCStateAetna delayed).
        if not from_manual and marker in ("", "No Data"):
            soft = SOFT_OVERRIDES.get((client, day))
            if soft is not None:
                marker = soft
                from_manual = True
        # A manual override of "!" means "flag this cell as a miss": render an
        # EMPTY pink-shaded cell (the renderers write the "!" glyph into any
        # alerted empty cell). Needed when the auto marker would show a real
        # signal that belongs to a different row — e.g. CVSPBMRx 8/17/26, whose
        # "Load Failure" is the Ad Hoc backfill's, not the weekly cycle's. A
        # plain "" override can't do this: it deliberately suppresses the shade.
        force_alert = False
        if marker == "!":
            marker = ""
            force_alert = True
        alert  = alert_state(client, day, marker)
        # An explicit manual blank ("") means the team has intentionally emptied
        # this cell — never pink-shade it (otherwise a live has_recent_failure on
        # today's cell would repaint it as a stray "!" even though the failure is
        # already shown on its correct day). Per user 2026-07-29 (AetnaRCE /
        # NCStateAetna 7/29 re-send cell).
        # A manual "L" override is the same intent for an in-progress reload: the
        # team has deliberately marked the cell "loading", so a stale
        # has_recent_failure on today's cell must NOT repaint it pink (auto "L"
        # already avoids pink via the is_loading_today suppress, but a manual "L"
        # can be pinned before the resuming job shows Ready/Running in RAMP). Per
        # user 2026-07-31 (AetnaRCE / NCStateAetna resuming load covers 7/29-7/31).
        if from_manual and marker in ("", "L"):
            alert = False
        if force_alert:
            alert = True
        marker, alert = apply_sticky_cert(client, day, marker, alert, from_manual)
        wk_start = day - timedelta(days=day.weekday())
        wk_end   = wk_start + timedelta(days=4)
        # ESIPBMRx state-list: per-state placement by round (load → snap → ✓).
        # Each tape row is a state-file load. A state is "covered" by the
        # next snap completion that lands AFTER its FileLoadDate — those
        # states get ✓ on the snap date. States loaded after the most recent
        # snap (or never snapped) show "L" on their load date.
        # Scan the WHOLE month — bucket[cell_day] routes each row to the
        # week-block where cell_day falls during rendering.
        if client == "ESIPBMRx" and esipbmrx_tape:
            tape_in_month = []
            for row in esipbmrx_tape:
                fdt = row.get("FileLoadDate")
                if not fdt or fdt.year != year or fdt.month != month:
                    continue
                sm = ESIPBMRX_STATE_RE.search(row.get("FileName", "") or "")
                if sm:
                    tape_in_month.append((sm.group(1).upper(), fdt))
            if tape_in_month:
                snap_completions = []
                for d_, entries in snap_idx.items():
                    if d_.year != year or d_.month != month:
                        continue
                    for entry in entries:
                        kind_e = entry[3] if len(entry) > 3 else "snap"
                        if entry[0] == "esipbmrx" and kind_e == "snap":
                            snap_completions.append(entry[1])
                snap_completions.sort()

                def covering_snap(load_dt):
                    for sdt in snap_completions:
                        if sdt >= load_dt:
                            return sdt
                    return None

                # Group state codes by (cell_day, marker). Dedupe states
                # within a group so a state with multiple files in one round
                # only appears once per cell.
                # Per user 2026-05-18: if a load/snap finishes over the
                # weekend, it should still appear on the closest weekday
                # (Sat → previous Friday, Sun → next Monday) instead of
                # disappearing because the calendar only renders Mon-Fri.
                def _closest_weekday(d):
                    if d.weekday() == 5:
                        return d - timedelta(days=1)   # Sat → Fri
                    if d.weekday() == 6:
                        return d + timedelta(days=1)   # Sun → Mon
                    return d
                state_groups = defaultdict(set)
                for state, load_dt in tape_in_month:
                    cov = covering_snap(load_dt)
                    if cov:
                        cell_day = _closest_weekday(cov.date())
                        marker_s = "✓"
                    else:
                        cell_day = _closest_weekday(load_dt.date())
                        marker_s = "L"
                    state_groups[(cell_day, marker_s)].add(state)

                highlight = "yellow" if client in YELLOW_HIGHLIGHT else None
                for (cell_day, marker_s) in sorted(state_groups,
                                                   key=lambda k: (k[0], 0 if k[1] == "✓" else 1)):
                    states_sorted = sorted(state_groups[(cell_day, marker_s)])
                    chunks = [states_sorted[i:i+4] for i in range(0, len(states_sorted), 4)]
                    alert_s = alert_state(client, cell_day, marker_s)
                    for chunk in chunks:
                        extra = f" ({', '.join(chunk)})"
                        label = f"{label_prefix}{display_name(client, monthly=(kind=='monthly'), extra_suffix=extra)}"
                        bucket[cell_day].append((label, marker_s, alert_s, highlight))
                return
        # Multi-week load flag for CenteneRx and similar
        extra = ""
        if client in MULTI_WEEK_CLIENTS and multi_week_loads:
            n = count_multi_week_loads(client, wk_start, wk_end, multi_week_loads)
            if n > 1:
                extra = f" ({n} weeks)"
        # HealthNetCA backfill: label the cell with the claims date range that
        # loaded that week, e.g. `HealthNetCA (3/20-3/27)` on 8/10. Keyed by the
        # cell's own Monday so it only tags the week the load actually ran in.
        if client == HEALTHNETCA_CLIENT and healthnetca_ranges:
            extra = healthnetca_ranges.get(day, extra)
        label = f"{label_prefix}{display_name(client, monthly=(kind=='monthly'), extra_suffix=extra)}"
        if client in BOLD_LABEL:
            highlight = "bold"
        elif client in YELLOW_HIGHLIGHT:
            highlight = "yellow"
        else:
            highlight = None
        # If this is a Load Failure with a registered ADO ticket, attach
        # the work-item URL so the marker cell becomes a clickable link.
        link = None
        if marker == "Load Failure":
            ticket = LOAD_FAILURE_ADO_LINKS.get((client, day))
            if ticket:
                link = ADO_LINK.format(ticket)
        bucket[day].append((label, marker, alert, highlight, link))

    # daily clients on every weekday (alphabetical)
    for d in all_days:
        for c in sorted(DAILY_CLIENTS):
            # Blank-before clients: skip cells before their go-live cutoff
            # (e.g. ElevanceMMMRx shows June 2026 forward only).
            bb = BLANK_BEFORE.get(c)
            if bb and d < bb:
                continue
            # Blank-from clients: drop the standing row on/after the cutoff
            # (e.g. ElevanceMMMRx removed 7/13/26 forward — AdHoc only).
            bf = BLANK_FROM.get(c)
            if bf and d >= bf:
                continue
            place(daily, "daily", c, d)

    # KaiserPrePayCOB — placed in the DAILY section (sorted alphabetically with
    # the other daily clients) per user 2026-06-16. (Was previously its own
    # bottom-of-week row in the `kaiser` bucket.)
    for d in all_days:
        place(daily, "daily", KAISER_PREPAY_CLIENT, d)

    # weekly clients on assigned weekday (alphabetical within column)
    for c in sorted(WEEKLY_CLIENTS):
        days = WEEKLY_CLIENTS[c]
        impl_start = IMPLEMENTATION_CLIENTS.get(c)
        blank_before = BLANK_BEFORE.get(c)
        blank_from = BLANK_FROM.get(c)
        for d in all_days:
            if d.strftime("%A") not in days:
                continue
            # Blank-before clients: render empty (skip the cell) for any day
            # before the cutoff — no marker, no pink "!".
            if blank_before and d < blank_before:
                continue
            # Blank-from clients: drop the standing row on/after the cutoff.
            if blank_from and d >= blank_from:
                continue
            # Implementation clients are suppressed before their start date.
            if impl_start and d < impl_start:
                continue
            # During implementation phase (before first cert), render
            # "Implementation" in the date cell as the manual marker.
            if impl_start:
                latest = latest_cert(c, cert_idx, on_or_before=d)
                if latest is None or latest.date() < impl_start:
                    place(weekly, "weekly", c, d, marker_override="Implementation")
                    continue
            place(weekly, "weekly", c, d)

    # NYShip_Rx rotates 4x/month
    for daynum in NYSHIP_DAYS:
        try:
            tgt = next_monday_if_weekend(date(year, month, daynum))
        except ValueError:
            continue
        if tgt.month != month:
            continue
        label = f"NYShip_Rx ({NYSHIP_LABEL[daynum]})"
        ov = NYSHIP_OVERRIDES.get((year, month, daynum))
        if ov is not None:
            # One-off override (e.g. 1st & 8th combined cycle): force the
            # marker. date(...) renders as the cert date; strings ("L") render
            # as-is. A date is never alerted; a string defers to alert_state.
            marker = ov
            alert  = False if isinstance(ov, date) else alert_state("NYShip_Rx", tgt, ov)
            marker, alert = apply_sticky_cert("NYShip_Rx", tgt, marker, alert, True)
        else:
            # Cert-style client — stay L when loaded this week until cert lands.
            marker = resolve_marker("NYShip_Rx", tgt, allow_checkmark=False, allow_week_window=True)
            alert  = alert_state("NYShip_Rx", tgt, marker)
            marker, alert = apply_sticky_cert("NYShip_Rx", tgt, marker, alert, False)
        weekly[tgt].append((label, marker, alert, None))

    # Monthly clients: state-driven placement (cert→cert date, loading→today,
    # snap→snap date, failure→today, else→expected day).
    for c in sorted(MONTHLY_CLIENTS):
        # OptumPBMRx is special — placed twice/month (early-month + end-month sets).
        if c == "OptumPBMRx":
            continue
        # Clients retired mid-year drop off starting their retirement month.
        retired = MONTHLY_RETIRED_FROM.get(c)
        if retired and (year, month) >= retired:
            continue
        # New clients don't appear before their start month (ModaRx: Sept 2026).
        starts = MONTHLY_STARTS_FROM.get(c)
        if starts and (year, month) < starts:
            continue
        d, marker = determine_monthly(c)
        # Lock-in (per user 2026-07-28): remember a certified/snapped monthly cell
        # and restore it if a later run regresses to No Data after the source data
        # ages out of the fetch window. "Told otherwise" = a manual override or
        # FORCED_INACTIVE governs this cell → don't restore over it.
        told_otherwise = (
            c in FORCED_INACTIVE
            or (c, year, month) in MONTHLY_MONTH_MARKER_OVERRIDES
            or (c in MONTHLY_PLACEMENT_OVERRIDES
                and MONTHLY_PLACEMENT_OVERRIDES[c][0].year == year
                and MONTHLY_PLACEMENT_OVERRIDES[c][0].month == month)
            # A per-month placement-day override is an explicit instruction about
            # WHICH DAY this month's row sits on — don't let the sticky cache drag
            # it back to the auto-computed day.
            or (c, year, month) in MONTHLY_PLACEMENT_DAY_OVERRIDES
        )
        d, marker = apply_sticky_monthly(c, d, marker, year, month, today, told_otherwise)
        if d.month != month:
            continue
        # Only-when-active clients (HumanaRx): no placeholder row on a fixed
        # expected day — drop the month entirely unless something happened.
        if c in MONTHLY_ONLY_WHEN_ACTIVE and marker in ("", "No Data", "Inactive"):
            continue
        place(monthly, "monthly", c, d, marker_override=marker)

    # OptumPBMRx: per-RAW breakout — one row per RAW file ("OptumPBMRx - Raw N").
    # Per user 2026-06-30: show each RAW loading individually instead of the two
    # grouped (RAW 1/2/3) / (RAW 5/6) rows. Lifecycle per file: "L" while the
    # specific file is staging/loading (Optum 0100 PBM Stage → 0110 Load; tape
    # ProcessStatus 42), "✓" once its Snap (Optum 0200 PBM Start Snap) completes.
    # Only show a RAW once it's PRESENT (no placeholder/No-Data rows). RAW numbers
    # outside the expected set surface as "(ad hoc)" rows (pink) so unexpected
    # loads are visible. Each instance is placed on its own load-date week and
    # emitted only on the month-tab that renders that week (so the RAW5/6 end-of-
    # month set lands correctly on the next month's tab when the week straddles
    # the boundary — no special-casing needed).
    _OPTUM_SNAP_RE = re.compile(r"optum.*0200.*pbm.*start.*snap", re.IGNORECASE)
    _optum_job_by_id = {j.get("JobId"): (j.get("JobName") or "") for j in ramp_jobs}
    OPTUM_EXPECTED_RAWS = {1, 2, 3, 5, 6}

    # Snap completion datetimes. The Optum pipeline is sequential (Stage→Load→Snap
    # per file, one file fully processed before the next — confirmed by the user),
    # so a loaded RAW is "snapped" once a snap completes AFTER its load: the next
    # snap can only be that file's (the prior file's snap ran before this load).
    _optum_snap_dts = []
    for q in ramp_queue:
        if not _OPTUM_SNAP_RE.search(_optum_job_by_id.get(q.get("JobId"), "")):
            continue
        stt = (q.get("Status") or "").lower()
        if not (stt.startswith("success") or stt == "resolved"):
            continue
        ed = parse_dt(q.get("EndDate"))
        if ed:
            _optum_snap_dts.append(ed)

    _rendered_days = set(all_days)
    # raw_n is an int (normal 1/2/3/5/6) or a str (ad-hoc token like "53YR").
    # Sort numeric RAWs first (by value), then alphanumeric ad-hoc (by string);
    # never compare int vs str directly (TypeError in py3).
    def _optum_sort_key(x):
        rn = x.get("raw_n")
        is_str = isinstance(rn, str)
        return (1 if is_str else 0, str(rn) if is_str else rn, x.get("data_date", ""))
    for ri in sorted((optum_raw_instances or ()), key=_optum_sort_key):
        load_dt = ri.get("load_date")
        activity = load_dt or ri.get("latest")
        if not activity:
            continue
        placement = next_monday_if_weekend(activity.date())
        if placement not in _rendered_days:
            continue  # this RAW's load week is rendered on another month's tab
        loaded = bool(ri.get("loaded"))
        snapped = loaded and load_dt is not None and any(s > load_dt for s in _optum_snap_dts)
        marker = "✓" if snapped else "L"
        raw_n = ri.get("raw_n")
        # Alphanumeric tokens (e.g. "53YR") are legitimate ad-hoc RAW loads — show
        # them as a plain row (no "(ad hoc)" suffix, no pink; per user 2026-07-01
        # the 0110 Load did not fail). A NUMERIC RAW outside the expected set still
        # flags as ad-hoc pink so a genuinely unexpected numeric load is visible.
        numeric_adhoc = (not isinstance(raw_n, str)) and (raw_n not in OPTUM_EXPECTED_RAWS)
        label = f"OptumPBMRx - Raw {raw_n}" + (" (ad hoc)" if numeric_adhoc else "")
        monthly[placement].append((label, marker, numeric_adhoc, None))

    # Aetna NMSP - MMSEA: ✓ once SourceLog shows a NonMSP file fully imported
    # (ImportCompleteDate) this month, placed on the completion date. While a
    # file has only started loading in CMSE (ImportStartDate, no complete date)
    # the cell shows "L" on the start date. Otherwise the 15th rule (or next
    # Monday) with No Data.
    nmsp_day = nmsp_mmsea_date(year, month)
    if nmsp_day.month == month:
        done_dt = None   # latest completed load this month
        start_dt = None  # latest in-progress (started, not yet complete) load
        for rec in (aetna_nmsp_loads or ()):
            d, s = rec.get("done"), rec.get("start")
            if d and d.year == year and d.month == month and d.date() <= today:
                if done_dt is None or d > done_dt:
                    done_dt = d
            if s and not d and s.year == year and s.month == month and s.date() <= today:
                if start_dt is None or s > start_dt:
                    start_dt = s
        if done_dt:
            placement = done_dt.date()
            if placement.weekday() >= 5:
                placement = next_monday_if_weekend(placement)
            if placement.month != month:
                placement = nmsp_day
            marker = "✓"
            alert  = False
        elif start_dt:
            placement = start_dt.date()
            if placement.weekday() >= 5:
                placement = next_monday_if_weekend(placement)
            if placement.month != month:
                placement = nmsp_day
            marker = "L"
            alert  = False
        else:
            placement = nmsp_day
            marker = "No Data"
            alert = alert_state("AetnaMMSEA", nmsp_day, marker)
        monthly[placement].append(("Aetna NMSP - MMSEA", marker, alert, "bold"))

    # (Removed) loading-today extras pass — L is now surfaced on each client's
    # scheduled weekday cell via resolve_marker / determine_monthly directly.

    # One-off injected entries (catch-up loads, etc.)
    section_map = {"daily": daily, "weekly": weekly, "monthly": monthly, "kaiser": kaiser}
    for section, day, label, marker, alert, highlight in ADDITIONAL_ENTRIES:
        bucket = section_map.get(section)
        if bucket is None or day.month != month or day.year != year:
            continue
        bucket[day].append((label, marker, alert, highlight))

    # (KaiserPrePayCOB Sat/Sun weekend tracking removed 2026-06-16 per user —
    # no need to surface Saturday/Sunday loads anymore.)

    # Ad-hoc MSPI loads -> "Monthly Ad Hoc" (per user 2026-06-25): each MSPI
    # client (Aetna MSPI, BCBSNC MSPI, Aetna QNXT MSPI) appears ONCE per month,
    # on the date of its LATEST ad-hoc load that month, in the MONTHLY section.
    # No persistent daily row, no per-run weekly entries, never flagged missing
    # (matches HumanaRx, which is already monthly). Weekend-start loads shift to
    # an adjacent weekday for display.
    since_adhoc = date(year, month, 1) - timedelta(days=14)
    mspi_latest = {}   # label -> (orig_day, marker, alert)
    for ah in scan_adhoc_loads(ramp_queue, ramp_jobs, today, since_adhoc,
                               weekend_shift=False):
        d = ah["day"]
        if d.year != year or d.month != month:
            continue
        prev = mspi_latest.get(ah["label"])
        if prev is None or d >= prev[0]:
            mspi_latest[ah["label"]] = (d, ah["marker"], ah["alert"])
    for label, (d, mk, al) in mspi_latest.items():
        disp = d
        if disp.weekday() == 5:
            disp -= timedelta(days=1)
        elif disp.weekday() == 6:
            disp += timedelta(days=1)
        monthly[disp].append((label, mk, al, None, None))

    # Ad-hoc MONTHLY snap-driven clients (per user 2026-06-25): surface ONCE per
    # month in the MONTHLY section, dated to when the named RAMP job FINISHES
    # (LatestJobRun EndDate, Successful/Resolved) this month; 'L' on today while
    # it's Ready/Running; no row if it hasn't run this month. e.g. UnitedRx(P)
    # <- 'United 0130 RX Post Snap' (UnitedRx can't certify).
    for label, jobname in ADHOC_MONTHLY_SNAP_CLIENTS.items():
        jn_l = jobname.strip().lower()
        for j in ramp_jobs:
            if (j.get("JobName") or "").strip().lower() != jn_l:
                continue
            lr = j.get("LatestJobRun") or {}
            status = (lr.get("Status") or "").strip().lower()
            if status.startswith("success") or status == "resolved":
                fin = parse_dt(lr.get("EndDate")) or parse_dt(lr.get("StartDate"))
                if fin and fin.year == year and fin.month == month:
                    disp = fin.date()
                    if disp.weekday() == 5:
                        disp -= timedelta(days=1)
                    elif disp.weekday() == 6:
                        disp += timedelta(days=1)
                    monthly[disp].append((label, fin.date(), False, None, None))
            elif status in ("ready", "running"):
                monthly[today].append((label, "L", False, None, None))
            break

    # CignaRx EOM/SOM injection — second CignaRx cycle closing out prior month.
    # Surfaces on the first Tuesday of each month (per user 2026-06-03), EXCEPT
    # while it is actively loading in the current month: then it moves to TODAY
    # with an "L" so the live cycle shows where the activity is (per user
    # 2026-07-01 — the current 'Cigna RX 0110 Load' IS the EOM/SOM). Once the
    # cert lands the row returns to the first Tuesday with the cert date.
    cigna_target = None
    for d in all_days:
        if d.month == month and d.weekday() == 1:  # Tuesday
            cigna_target = d
            break
    if cigna_target is not None:
        cig_label = "CignaRx (EOM/SOM)(p)"
        placement_override = None
        override = CIGNARX_EOM_SOM_OVERRIDES.get((year, month))
        if override is not None:
            if isinstance(override, tuple):
                placement_override, marker = override
            else:
                marker = override
        else:
            # Window: from the 1st of the month through 14 days in. Restricted
            # to the current month (was -7 days) so the prior month's regular
            # weekly CignaRx certs (e.g. 6/24, 6/30) aren't pulled onto this
            # cell — the EOM/SOM cycle is the NEXT CignaRx load in the new
            # month, which stays blank until it loads. Per user 2026-07-01.
            win_start = date(year, month, 1)
            win_end   = date(year, month, 1) + timedelta(days=14)
            # Cert preference: earliest cert in the window.
            cig_cert = None
            for key in _keys_for_client("CignaRx"):
                for dt, status in cert_idx.get(key, ()):
                    if status != "Certified":
                        continue
                    if win_start <= dt.date() <= win_end:
                        if cig_cert is None or dt < cig_cert:
                            cig_cert = dt
            if cig_cert and cig_cert.date() <= today:
                marker = cig_cert.date()
            else:
                cig_keys = list(_keys_for_client("CignaRx"))
                loaded = False
                for d_scan, entries in snap_idx.items():
                    if not (win_start <= d_scan <= win_end):
                        continue
                    for entry in entries:
                        if _src_matches_client(entry[0], cig_keys):
                            jn = entry[4] if len(entry) > 4 else ""
                            if _load_name_allowed("CignaRx", jn, entry[3] if len(entry) > 3 else "load"):
                                loaded = True
                                break
                    if loaded:
                        break
                # Also show "L" while the EOM/SOM Cigna RX 0110 Load is
                # Ready/Running right now (before it lands in snap_idx) — only
                # for the CURRENT month tab so past months aren't disturbed.
                is_current = (year == today.year and month == today.month)
                if loaded or (is_current and is_loading_today("CignaRx", ramp_queue, ramp_jobs)):
                    marker = "L"
                else:
                    marker = ""   # blank until the next load
        # A tuple override forces the placement day; otherwise, while loading in
        # the current month place the row on TODAY; else (blank / landed cert)
        # keep it on the first Tuesday.
        placement = cigna_target
        if placement_override is not None and placement_override in all_days:
            placement = placement_override
        elif marker == "L" and year == today.year and month == today.month and today in all_days:
            placement = today
        alert = alert_state("CignaRx", placement, marker)
        weekly[placement].append((cig_label, marker, alert, None))

    # CVSPBMRx Ad Hoc (backfill) row — one per oversized/out-of-cycle tape file
    # (FileSize giveaway; see fetch_cvspbm_adhoc). While loading → "L" on today
    # (per user 2026-07-29: "continue to show the Ad Hoc on the current day
    # during the load"). Once loaded → "✓" attributed to the file's DATA date
    # Monday delivery cell (parsed from the filename), NOT the load-completion
    # date — same rule as the regular weekly path (per user 2026-07-29: the
    # in-flight RAW_MEMBR_ELIG_20260709 Ad Hoc is "the 7/13 load"; 7/9 data →
    # 7/13 Mon cell). Falls back to the weekend-shifted load date if the filename
    # has no parseable date. Kept separate from the regular weekly Monday cells
    # (per user 2026-07-06). Emitted only on the month tab whose rendered days
    # contain the placement day.
    _cvspbm_ah_seen = set()
    for a in cvspbm_adhoc:
        if a.get("loaded"):
            cell = None
            m_ah = CVSPBMRX_ADHOC_DATE_RE.search(a.get("filename") or "")
            if m_ah:
                try:
                    cell = _cvspbmrx_cell_monday(
                        datetime.strptime(m_ah.group(1), "%Y%m%d").date())
                except ValueError:
                    cell = None
            if cell is None:
                ld = a.get("load_date")
                if not ld:
                    continue
                cell = ld.date()
                if cell.weekday() == 5:
                    cell -= timedelta(days=1)
                elif cell.weekday() == 6:
                    cell += timedelta(days=1)
            mk = "✓"
        elif cvspbm_adhoc_failed is not None:
            cell = cvspbm_adhoc_failed
            mk = "Load Failure"
        else:
            cell = today
            mk = "L"
        if cell not in all_days:
            continue
        key = (cell, mk)
        if key in _cvspbm_ah_seen:
            continue
        _cvspbm_ah_seen.add(key)
        # Highlight the in-flight "L" (per user 2026-07-15) so the Ad Hoc
        # backfill stands out while it loads; the highlight clears once the
        # load job finishes and snaps (mk flips to ✓).
        ah_highlight = "yellow" if mk == "L" else None
        weekly[cell].append(("CVSPBMRx (Ad Hoc)", mk,
                             mk == "Load Failure", ah_highlight))
    # Carry-over in-flight load with no un-loaded tape row to hang the row off
    # (the Ad Hoc file's tape row already reads ProcessStatus 50 but its RAMP load
    # card is still Ready/Running — see cvspbm_carryover_load above). The regular
    # Monday "L" is suppressed for it, so surface it here instead of losing the
    # in-flight signal entirely. Clears itself when the card finishes.
    if (cvspbm_carryover_load is not None and not cvspbm_tape_adhoc_loading
            and today in all_days and (today, "L") not in _cvspbm_ah_seen):
        _cvspbm_ah_seen.add((today, "L"))
        weekly[today].append(("CVSPBMRx (Ad Hoc)", "L", False, "yellow"))

    # One-off (per user 2026-06-11): Kaiser_AmbM runs a SECOND June cycle — it
    # certified 6/11, and a new monthly load lands ~6/18. determine_monthly only
    # places one row/month and the 6/11 cert takes precedence, so the new cycle
    # wouldn't surface until it re-certifies. Inject a 6/18 row that tracks the
    # new load (No Data → L) using only activity on/after the 6/12 cutoff (so the
    # prior 6/11 cert/load is ignored). Once the new cert lands, determine_monthly
    # shows it on the cert date and this injection is skipped (no duplicate).
    # Remove after the June 2026 cycle.
    if (year, month) == (2026, 6):
        ka_place  = date(2026, 6, 18)
        ka_cutoff = date(2026, 6, 12)        # ignore the prior 6/11 cert/load
        ka_keys   = list(_keys_for_client("Kaiser_AmbM"))
        ka_new_cert = any(
            status == "Certified" and ka_cutoff <= dt.date() <= today
            for k in ka_keys for dt, status in cert_idx.get(k, ())
        )
        if not ka_new_cert:
            ka_loaded = any(
                _src_matches_client(e[0], ka_keys)
                for d_scan, entries in snap_idx.items() if ka_cutoff <= d_scan <= today
                for e in entries
            )
            ka_marker = "L" if (is_loading_today("Kaiser_AmbM", ramp_queue, ramp_jobs)
                                or ka_loaded) else "No Data"
            monthly[ka_place].append(("Kaiser_AmbM", ka_marker, False, None))

    # Kaiser Submission daily row — ✓ when both 'Kaiser Pareo Submission Logfile'
    # AND '...Upload' finished that day; 'L' on today while in progress; blank
    # otherwise. Placed before the sort so DAILY_ORDER positions it (between
    # KaiserPrePayCOB and NCStateAetna).
    ks_done, ks_running = scan_kaiser_submission(
        ramp_queue, ramp_jobs, today, date(year, month, 1) - timedelta(days=14))
    for d in all_days:
        if d in ks_done:
            mk = "✓"
        elif d == today and ks_running:
            mk = "L"
        else:
            mk = ""
        daily[d].append(("Kaiser Submission", mk, False, None, None))

    # Sort: Daily by the fixed DAILY_ORDER; Weekly/Monthly/Kaiser alphabetical.
    for d in daily:
        daily[d].sort(key=lambda r: _daily_order_key(r[0]))
    for bucket in (weekly, monthly, kaiser):
        for d in bucket:
            bucket[d].sort(key=lambda r: r[0].lower())

    return {"daily": daily, "weekly": weekly, "monthly": monthly, "kaiser": kaiser}, weeks


# ============================================================
#                           rendering (no colour fills)
# ============================================================
HEADER_FILL = PatternFill("solid", fgColor="2C5F8A")
HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
DAY_FILL    = PatternFill("solid", fgColor="E3EBF4")
DAY_FONT    = Font(name="Segoe UI", bold=True, size=10, color="1F3D5C")
TODAY_FILL  = PatternFill("solid", fgColor="FFD180")
ALERT_FILL  = PatternFill("solid", fgColor="FFC7CE")
ALERT_FONT  = Font(name="Segoe UI", size=9, bold=True, color="9C0006")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2A8")
YELLOW_FONT = Font(name="Segoe UI", size=9, bold=True, color="7F6000")
CELL_FONT   = Font(name="Segoe UI", size=9)
THIN        = Side(style="thin", color="C8C8C8")
BORDER      = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def fmt_marker(m):
    """Pass through date objects for native Excel date formatting; stringify
    anything else (✓, L, No Data, Inactive, Load Failure, blank…)."""
    if isinstance(m, date):
        return m
    return str(m or "")


def _write_section_rows(ws, cur_row, wk, plan_section):
    """Write one section (daily/weekly/monthly) of stacked rows for a week.
    Returns the next free row after the section.
    Each row spans the 5 day-columns; cell content + date marker per column.
    """
    max_clients = max((len(plan_section.get(d, [])) for d in wk if d), default=0)
    for ci in range(max_clients):
        for i, d in enumerate(wk):
            col_day = i * 2 + 1
            col_dat = i * 2 + 2
            cell      = ws.cell(row=cur_row, column=col_day, value=None)
            date_cell = ws.cell(row=cur_row, column=col_dat, value=None)
            for c in (cell, date_cell):
                c.font = CELL_FONT
                c.border = BORDER
                c.alignment = Alignment(vertical="top")
            date_cell.alignment = Alignment(horizontal="center", vertical="top")
            if d is None:
                cell.fill = PatternFill("solid", fgColor="F5F5F5")
                date_cell.fill = PatternFill("solid", fgColor="F5F5F5")
                continue
            clients = plan_section.get(d, [])
            if ci < len(clients):
                row = clients[ci]
                name, marker, alert = row[0], row[1], row[2]
                highlight = row[3] if len(row) > 3 else None
                link      = row[4] if len(row) > 4 else None
                cell.value = name
                v = fmt_marker(marker)
                date_cell.value = v
                if isinstance(v, date):
                    date_cell.number_format = "mm/dd/yy"
                if highlight == "yellow":
                    cell.fill = YELLOW_FILL
                    cell.font = YELLOW_FONT
                    date_cell.fill = YELLOW_FILL
                elif highlight == "bold":
                    cell.font = Font(name="Segoe UI", size=9, bold=True)
                if alert:
                    date_cell.fill = ALERT_FILL
                    if not date_cell.value:
                        date_cell.value = "!"
                    date_cell.font = ALERT_FONT
                if link:
                    # Underline the alert font so the cell visibly reads
                    # as a clickable link (cursor changes on hover too).
                    date_cell.hyperlink = link
                    date_cell.font = Font(name="Segoe UI", size=9,
                                          bold=True, color="9C0006",
                                          underline="single")
        cur_row += 1
    return cur_row


def _blank_separator_row(ws, cur_row):
    for i in range(10):
        c = ws.cell(row=cur_row, column=i + 1, value=None)
        c.fill = PatternFill("solid", fgColor="FFFFFF")
        c.border = Border()
    ws.row_dimensions[cur_row].height = 6
    return cur_row + 1


# Section-header band (Daily / Weekly / Monthly). Doubles as the visual
# separator between sections, so no blank row is needed alongside it.
SECTION_LABEL_FILL = PatternFill("solid", fgColor="E8EDF3")
SECTION_LABEL_FONT = Font(name="Segoe UI", bold=True, size=9, color="2C5F8A")


def _section_label_row(ws, cur_row, text):
    """Write a thin labeled band spanning all 10 columns as a section header."""
    for i in range(10):
        c = ws.cell(row=cur_row, column=i + 1, value=None)
        c.fill = SECTION_LABEL_FILL
        c.border = Border()
    c0 = ws.cell(row=cur_row, column=1, value=text)
    c0.font = SECTION_LABEL_FONT
    c0.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=10)
    ws.row_dimensions[cur_row].height = 14
    return cur_row + 1


def write_weekly_stacked(ws, year, month, sections, weeks, today):
    holidays = us_federal_holidays(year)
    cur_row = 1
    week_no = 0
    for wk in weeks:
        week_no += 1
        first_d = next((d for d in wk if d), None)
        last_d  = next((d for d in reversed(wk) if d), None)
        label = f"Week {week_no}: {first_d:%m/%d} – {last_d:%m/%d}" if first_d and last_d else f"Week {week_no}"
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=10)
        c = ws.cell(row=cur_row, column=1, value=label)
        c.font = Font(name="Segoe UI", bold=True, size=12, color="1F3D5C")
        cur_row += 1

        for i, day in enumerate(WEEKDAYS):
            col_day = i * 2 + 1
            col_dat = i * 2 + 2
            for col, val in ((col_day, day), (col_dat, "Date")):
                cell = ws.cell(row=cur_row, column=col, value=val)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
                cell.border = BORDER
        cur_row += 1

        # date strip row (real Excel dates, formatted mm/dd/yy).
        # Federal-holiday labels are written into the day-name sub-column
        # of this same row so the label sits directly under "Monday" and
        # right next to the date — not as a separate banner row.
        holiday_fill = PatternFill("solid", fgColor="FFE4B5")
        holiday_font = Font(name="Segoe UI", italic=True, size=9,
                            bold=True, color="7C4A00")
        for i, d in enumerate(wk):
            col_day = i * 2 + 1
            col_dat = i * 2 + 2
            if d:
                cell = ws.cell(row=cur_row, column=col_dat, value=d)
                cell.number_format = "mm/dd/yy"
                cell.fill = DAY_FILL if d != today else TODAY_FILL
                cell.font = DAY_FONT
                cell.alignment = Alignment(horizontal="center")
                cell.border = BORDER
                hname = holidays.get(d)
                if hname:
                    hc = ws.cell(row=cur_row, column=col_day, value=hname)
                    hc.fill = holiday_fill
                    hc.font = holiday_font
                    hc.alignment = Alignment(horizontal="center", vertical="center")
                    hc.border = BORDER
        cur_row += 1

        # Each section is introduced by its header band (which also separates
        # it from the section above — no blank row needed). KaiserPrePayCOB now
        # lives in the Daily section; the `kaiser` bucket is only populated for
        # historical snapshot months and is rendered after Monthly when present.
        cur_row = _section_label_row(ws, cur_row, "Daily")
        cur_row = _write_section_rows(ws, cur_row, wk, sections["daily"])
        cur_row = _section_label_row(ws, cur_row, "Weekly")
        cur_row = _write_section_rows(ws, cur_row, wk, sections["weekly"])
        cur_row = _section_label_row(ws, cur_row, "Monthly")
        cur_row = _write_section_rows(ws, cur_row, wk, sections["monthly"])
        if any(sections["kaiser"].get(d) for d in wk if d):
            cur_row = _write_section_rows(ws, cur_row, wk, sections["kaiser"])

        # per-week key block
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=10)
        kc = ws.cell(row=cur_row, column=1,
                     value="Key:  Date = Certified  |  ✓ = Loaded/Snapped  |  L = Loading"
                           "  |  TBL = To Be Loaded"
                           "  |  pink = Failure/Inactive  |  (s) SLA  |  (p) Rx Post Snap"
                           "  |  (n) Not Delivered  |  -  = No load that day")
        kc.font = Font(name="Segoe UI", italic=True, size=9, color="555555")
        kc.alignment = Alignment(horizontal="left")
        cur_row += 2

    # Client-name columns set to ≈190 px (width 27.07) per user 2026-05-20.
    # Excel pixels ≈ 7 * width + 0.5.
    # Column D (Tuesday date) was widened to ≈95 px for BCBSAR's "Implementation"
    # text; reverted to the standard 11 once BCBSAR certified (2026-06-12).
    client_w = (190 - 0.5) / 7
    for i in range(10):
        w = client_w if i % 2 == 0 else 11
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.sheet_view.showGridLines = False


def _write_month_section(ws, cur_row, wk, plan_section):
    """Single-column-per-day version for the Month sheet."""
    max_clients = max((len(plan_section.get(d, [])) for d in wk if d), default=0)
    for ci in range(max_clients):
        for col, d in enumerate(wk, start=1):
            cell = ws.cell(row=cur_row, column=col, value=None)
            cell.font = CELL_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if d is None:
                cell.fill = PatternFill("solid", fgColor="F5F5F5")
                continue
            clients = plan_section.get(d, [])
            if ci < len(clients):
                row = clients[ci]
                name, marker, alert = row[0], row[1], row[2]
                highlight = row[3] if len(row) > 3 else None
                m_str = fmt_marker(marker)
                cell.value = f"{name}  [{m_str}]" if m_str else name
                if highlight == "yellow":
                    cell.fill = YELLOW_FILL
                    cell.font = YELLOW_FONT
                if alert:
                    cell.fill = ALERT_FILL
                    cell.font = ALERT_FONT
        cur_row += 1
    return cur_row


def write_month_sheet(ws, year, month, sections, weeks, today):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    tc = ws.cell(row=1, column=1, value=f"{calendar.month_name[month]} {year} — Delivery Calendar")
    tc.font = Font(name="Segoe UI", bold=True, size=14, color="1F3D5C")
    tc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    for i, d in enumerate(WEEKDAYS, start=1):
        c = ws.cell(row=2, column=i, value=d)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = 32

    cur_row = 3
    for wk in weeks:
        for col, d in enumerate(wk, start=1):
            label = d.strftime("%a %m/%d") if d else ""
            cell = ws.cell(row=cur_row, column=col, value=label)
            cell.fill = TODAY_FILL if d == today else DAY_FILL
            cell.font = DAY_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
        cur_row += 1

        cur_row = _write_month_section(ws, cur_row, wk, sections["daily"])
        cur_row += 1  # blank
        cur_row = _write_month_section(ws, cur_row, wk, sections["weekly"])
        cur_row += 1
        cur_row = _write_month_section(ws, cur_row, wk, sections["monthly"])
        cur_row += 1
        cur_row = _write_month_section(ws, cur_row, wk, sections["kaiser"])

        # per-week key
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
        kc = ws.cell(row=cur_row, column=1,
                     value="Key:  Date = Certified  |  ✓ = Loaded/Snapped  |  L = Loading"
                           "  |  TBL = To Be Loaded"
                           "  |  pink Date = Failure/Inactive  |  (s) SLA  |  (p) Rx Post Snap"
                           "  |  (n) Not Delivered  |  -  = No load that day")
        kc.font = Font(name="Segoe UI", italic=True, size=9, color="555555")
        cur_row += 2

    ws.sheet_view.showGridLines = False


def write_tickets_sheet(ws, latest_tickets, ramp_jobs):
    headers = ["Client", "Title Kind", "ADO Ticket", "ADO State", "Tags",
               "RAMP Jobs (enabled)", "Assigned To", "Changed"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left")
        cell.border = BORDER

    rows = sorted(latest_tickets.values(), key=lambda t: (t["state"], t["client"].lower()))
    for r_i, t in enumerate(rows, start=2):
        matched = find_matching_jobs(t["client"], ramp_jobs)
        enabled = sum(1 for j in matched if j.get("Enabled") == 1)
        vals = [
            t["client"], t["kind"], f"#{t['id']}", t["state"], t["tags"],
            enabled, t["assigned"], (t["changed"][:10] if t["changed"] else ""),
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r_i, column=c, value=v)
            cell.font = CELL_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER

    ws.freeze_panes = "A2"
    for i, w in enumerate([22, 18, 12, 14, 28, 14, 24, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_key_sheet(ws):
    rows = [
        ("Marker / Suffix",  "Meaning"),
        ("[Date]",           "Client was certified that day in DHT.TableList (CertTimestamp)."),
        ("Snap",             "Snap completed that day in RAMP /api/Ramp/Snap/SnapQueueStatus."),
        ("L",                "Load or snap currently in progress for the client today."),
        ("(s)",              "SLA Client — tight delivery window."),
        ("(p)",              "Rx Client Post Snap."),
        ("(n)",              "Special handling — historically Not Delivered."),
        ("M -",              "Monthly client — placed on the day its delivery ticket fired."),
    ]
    for c, h in enumerate(rows[0], start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
    for r, row in enumerate(rows[1:], start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    notes_row = len(rows) + 3
    notes = [
        "Data sources",
        f"  • SQL  TRGUTIL10.DHTStats [DHT].[TableList]  (CertTimestamp + CurrentStatus)",
        f"  • RAMP {RAMP_BASE}/api/Ramp/Snap/SnapQueueStatus  (snap completions)",
        f"  • RAMP {RAMP_BASE}/api/Ramp/Queue/List           (load-job completions)",
        f"  • RAMP {RAMP_BASE}/api/Ramp/Job/List             (enabled-job detection)",
        f"  • ADO  {ADO_BASE}  (User Stories tagged 'Delivery Ticket')",
        "",
        "Title formats recognised on ADO tickets:",
        "  • 'Snap and Mine - <Client> - ...'",
        "  • 'Load and Snap - <Client> - ...'",
        "  • 'Kaiser - SNAP/MINE - <Client> - ...'",
    ]
    for i, n in enumerate(notes):
        cell = ws.cell(row=notes_row + i, column=1, value=n)
        cell.font = Font(name="Segoe UI", bold=(i == 0 or n.startswith("Title")), size=10)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def write_client_owner_sheet(ws):
    """Render the Client Owner tab — four owner groups side-by-side, each
    a (Owner, Client, Priority) sub-block. The Owner sub-column is merged
    vertically across the whole block with the name centered both axes.
    Each owner's 3-column header is a distinct color. A blank row
    separates each owner's upper list from their lower list, matching the
    source layout.
    """
    # Distinct header colors per owner. White HEADER_FONT remains readable
    # on each. Per user 2026-05-20.
    owner_header_colors = {
        "Dave":     "2C5F8A",   # blue (matches main report header)
        "Emmanuel": "4A7C59",   # green
        "Holly":    "7C4A6E",   # mauve
        "Adam":     "B86F2E",   # orange
    }

    owners = list(CLIENT_OWNERS.items())
    upper_max = max(len(d["upper"]) for _, d in owners)
    lower_max = max(len(d["lower"]) for _, d in owners)

    blank_row = 2 + upper_max               # row index where the gap sits
    bottom_end_row = blank_row + lower_max  # last data row

    owner_font   = Font(name="Segoe UI", bold=True, size=14, color="1F3D5C")
    owner_align  = Alignment(horizontal="center", vertical="center")
    center       = Alignment(horizontal="center", vertical="center")
    plain        = Alignment(vertical="top", wrap_text=False)

    for col_idx, (owner, data) in enumerate(owners):
        col_o = col_idx * 3 + 1
        col_c = col_idx * 3 + 2
        col_p = col_idx * 3 + 3

        # ---- per-owner header row (row 1) ----
        fill = PatternFill("solid", fgColor=owner_header_colors.get(owner, "2C5F8A"))
        for col, label in ((col_o, "Owner"), (col_c, "Client"), (col_p, "Priority")):
            cell = ws.cell(row=1, column=col, value=label)
            cell.fill = fill
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

        # ---- pre-border the Client + Priority columns ----
        for row in range(2, bottom_end_row + 1):
            for col in (col_c, col_p):
                ws.cell(row=row, column=col).border = BORDER

        # ---- merge Owner column vertically and center the name ----
        ws.merge_cells(start_row=2, start_column=col_o,
                       end_row=bottom_end_row, end_column=col_o)
        oc = ws.cell(row=2, column=col_o, value=owner)
        oc.font = owner_font
        oc.alignment = owner_align
        # Apply BORDER to every cell within the merged Owner range so the
        # bottom edge actually renders in Excel — without this, a merged
        # cell only borders the top-left occurrence and the bottom edge of
        # the merged block disappears. Per user 2026-05-20.
        for r in range(2, bottom_end_row + 1):
            ws.cell(row=r, column=col_o).border = BORDER

        # ---- upper entries ----
        for r_off, entry in enumerate(data["upper"]):
            row = 2 + r_off
            client, priority = entry
            cc = ws.cell(row=row, column=col_c, value=client)
            cc.font = CELL_FONT
            cc.alignment = plain
            pc = ws.cell(row=row, column=col_p, value=priority)
            pc.font = CELL_FONT
            pc.alignment = center

        # ---- lower entries (after the blank separator row) ----
        for r_off, entry in enumerate(data["lower"]):
            row = blank_row + 1 + r_off
            client, priority = entry
            cc = ws.cell(row=row, column=col_c, value=client)
            cc.font = CELL_FONT
            cc.alignment = plain
            pc = ws.cell(row=row, column=col_p, value=priority)
            pc.font = CELL_FONT
            pc.alignment = center

    # Column widths — user-requested pixel targets (105 px Owner, 185 px Client).
    # Excel width-to-pixels: pixels ≈ 7 * width + 0.5 (Calibri 11 baseline).
    # Owner widened 2026-05-20 from 91→105 px to give the centered name more room.
    OWNER_WIDTH    = (105 - 0.5) / 7   # ≈ 14.93
    CLIENT_WIDTH   = (185 - 0.5) / 7   # ≈ 26.36
    PRIORITY_WIDTH = 9
    for col_idx in range(len(owners)):
        ws.column_dimensions[get_column_letter(col_idx * 3 + 1)].width = OWNER_WIDTH
        ws.column_dimensions[get_column_letter(col_idx * 3 + 2)].width = CLIENT_WIDTH
        ws.column_dimensions[get_column_letter(col_idx * 3 + 3)].width = PRIORITY_WIDTH

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


# ============================================================
#                            main
# ============================================================
def parse_all_clients_xlsx(file_path, year):
    """Read an ExpectedClientDates_<MMM>.xlsx 'All Clients' tab into the
    same `(sections, weeks)` shape that `plan_calendar` returns.

    Two header styles are accepted:
      Jan: `All Clients - Week N - M/D-M/D` (Monday date encoded in header)
      Feb-Apr (and later): `<MonthName> - Week N - All Clients` (no date in
        header — the Monday date is inferred by scanning the first real
        date cell in the block).

    Daily rows: same client name across all populated weekday cells. In the
    Feb+ format, non-cert weekday cells contain the boolean `True` (a load-
    verification checkbox); these become blank markers, not strings.
    """
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    if "All Clients" not in wb.sheetnames:
        return None, None
    ws = wb["All Clients"]

    sections = {"daily": defaultdict(list), "weekly": defaultdict(list),
                "monthly": defaultdict(list), "kaiser": defaultdict(list)}
    weeks = []

    header_re_jan = re.compile(
        r"All Clients\s*-\s*Week\s*(\d+)\s*-\s*(\d+)/(\d+)-(\d+)/(\d+)",
        re.IGNORECASE,
    )
    header_re_feb = re.compile(
        r"([A-Za-z]+)\s*-\s*Week\s*(\d+)\s*-\s*All\s*Clients",
        re.IGNORECASE,
    )

    # Pending block state when the Monday date hasn't been resolved yet.
    pending_week = None  # dict: {'block_rows': [(row_idx, cells)], ...}
    current_week_dates = None
    in_data_block = False

    def to_marker(raw):
        # Feb+ files put boolean True/False in non-cert date cells as a
        # load-verification flag; treat these as blank, not "True"/"False".
        if isinstance(raw, bool):
            return ""
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date):
            return raw
        if raw is None:
            return ""
        return str(raw).strip()

    def _name_matches_fill(display):
        """True when `display` is one of the ALL_CLIENTS_FILL_CHECKMARK
        clients (allowing common suffix conventions like (s), (p), (n),
        " - ", and the leading `M - ` monthly tag)."""
        base = display.strip()
        base = re.sub(r"^M\s*-\s*", "", base).strip()
        base = base.lstrip("* ").strip()
        for target in ALL_CLIENTS_FILL_CHECKMARK:
            if (base == target
                    or base.startswith(target + "(")
                    or base.startswith(target + " ")
                    or base.startswith(target + "-")):
                return True
        return False

    def flush_block(block_rows, week_days):
        """Commit a block of pending rows once week_days has been resolved."""
        for cells in block_rows:
            populated = []
            for i in range(5):
                n = cells[i * 2]
                if n is None:
                    continue
                ns = str(n).strip()
                if not ns:
                    continue
                populated.append((i, ns))
            is_daily_row = False
            daily_name = None
            if populated and len(populated) >= 2:
                base = populated[0][1].lstrip("* ").strip()
                if all(p[1].lstrip("* ").strip() == base for p in populated):
                    is_daily_row = True
                    daily_name = base

            for i in range(5):
                name = cells[i * 2]
                mark = cells[i * 2 + 1]
                if name is None and (mark is None or isinstance(mark, bool)):
                    continue
                name_s = "" if name is None else str(name).strip()
                if not name_s:
                    continue
                display = name_s
                if display.startswith("*"):
                    display = display.lstrip("* ").strip()

                kind = "weekly"
                if is_daily_row:
                    kind = "daily"
                    display = daily_name
                elif display.startswith("M -") or display.startswith("M-"):
                    kind = "monthly"
                    # Strip the "M - " tag from historical snapshot labels too,
                    # so it's gone everywhere (per user 2026-06-16).
                    display = re.sub(r"^M\s*-\s*", "", display).strip()
                elif "KaiserPrePayCOB" in display:
                    kind = "kaiser"

                marker = to_marker(mark)
                # Empty / True placeholder cells for the daily-Aetna and
                # PBMRx group: render as ✓ (loaded/snapped) instead of blank.
                if (isinstance(marker, str) and marker == ""
                        and _name_matches_fill(display)):
                    marker = "✓"
                alert = (isinstance(marker, str) and marker
                         and marker.strip().lower() in ALL_CLIENTS_ALERT_MARKERS)

                cell_day = week_days[i]
                sections[kind][cell_day].append((display, marker, alert))

    def commit_pending():
        nonlocal pending_week
        if pending_week is None:
            return
        # Anchor the Mon-Fri cycle Monday from a date cell whose weekday
        # matches its column position (typical case — most cert dates equal
        # the scheduled cycle day). Falls back to any date's own-week Monday.
        anchor_mon = None
        for i in range(5):
            for cells in pending_week["block_rows"]:
                v = cells[i * 2 + 1]
                d_val = None
                if isinstance(v, datetime):
                    d_val = v.date()
                elif isinstance(v, date) and not isinstance(v, bool):
                    d_val = v
                if d_val is None or d_val.weekday() != i:
                    continue
                anchor_mon = d_val - timedelta(days=i)
                break
            if anchor_mon:
                break
        if anchor_mon is None:
            # Fallback: take the first real date and snap to its own Monday.
            for cells in pending_week["block_rows"]:
                if anchor_mon:
                    break
                for i in range(5):
                    v = cells[i * 2 + 1]
                    d_val = None
                    if isinstance(v, datetime):
                        d_val = v.date()
                    elif isinstance(v, date) and not isinstance(v, bool):
                        d_val = v
                    if d_val is None:
                        continue
                    anchor_mon = d_val - timedelta(days=d_val.weekday())
                    break
        if anchor_mon is None and pending_week.get("month_hint"):
            # Fallback: compute Mon by week number within the month.
            yr = pending_week.get("year_hint", year)
            mn = pending_week["month_hint"]
            wn = pending_week["week_no"]
            cal_obj = calendar.Calendar(firstweekday=0)
            month_mons = [d for d in cal_obj.itermonthdates(yr, mn)
                          if d.weekday() == 0 and (d.month == mn
                                                   or (d - timedelta(days=4)).month == mn)]
            if 1 <= wn <= len(month_mons):
                anchor_mon = month_mons[wn - 1]
        if anchor_mon is None:
            pending_week = None
            return
        week_days = [anchor_mon + timedelta(days=i) for i in range(5)]
        weeks.append(week_days)
        flush_block(pending_week["block_rows"], week_days)
        pending_week = None

    month_name_to_num = {
        m.lower(): i for i, m in enumerate(
            ["", "January", "February", "March", "April", "May", "June",
             "July", "August", "September", "October", "November", "December"]
        ) if m
    }

    for row_idx in range(1, ws.max_row + 1):
        cells = [ws.cell(row=row_idx, column=c).value for c in range(1, 11)]
        first = cells[0]
        first_s = "" if first is None else str(first).strip()

        m1 = header_re_jan.match(first_s)
        m2 = header_re_feb.match(first_s) if not m1 else None
        if m1:
            commit_pending()
            mon_m, mon_d = int(m1.group(2)), int(m1.group(3))
            try:
                mon_date = date(year, mon_m, mon_d)
            except ValueError:
                continue
            week_days = [mon_date + timedelta(days=i) for i in range(5)]
            current_week_dates = week_days
            weeks.append(week_days)
            in_data_block = False
            pending_week = None
            continue
        if m2:
            commit_pending()
            mname = m2.group(1).lower()
            wnum = int(m2.group(2))
            month_hint = month_name_to_num.get(mname)
            pending_week = {
                "block_rows": [],
                "month_hint": month_hint,
                "year_hint": year,
                "week_no": wnum,
            }
            current_week_dates = None
            in_data_block = False
            continue

        if first_s == "Monday" and cells[2] in ("Tuesday", " Tuesday"):
            in_data_block = True
            continue

        if (first_s.startswith("Verify SNAP") or first_s.startswith("Verify Load")
                or first_s.startswith("SNAP pattern") or first_s.startswith("The three way")
                or (in_data_block and not any(c is not None and not isinstance(c, bool)
                                              and str(c).strip() for c in cells))):
            in_data_block = False
            continue

        if not in_data_block:
            continue

        if pending_week is not None:
            pending_week["block_rows"].append(cells)
            continue
        if current_week_dates is None:
            continue
        flush_block([cells], current_week_dates)

    # End-of-file: flush any open pending block.
    commit_pending()
    return sections, weeks


def _html_escape(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def _fmt_marker_html(m):
    """Marker rendering for HTML cells."""
    if isinstance(m, date):
        return m.strftime("%m/%d/%y")
    return _html_escape(m or "")


def _render_section_rows_html(wk, plan_section, today):
    """Return one HTML <tbody> for a section. Each row has 10 cells.
    Per user 2026-06-03: today highlighting is restricted to the date-strip
    row only — data-row cells do NOT get a today-column accent, so the table
    body reads cleanly without a vertical line down every row."""
    max_clients = max((len(plan_section.get(d, [])) for d in wk if d), default=0)
    if max_clients == 0:
        return ""
    rows_html = []
    for ci in range(max_clients):
        cells = []
        for i, d in enumerate(wk):
            if d is None:
                cells.append(f'<td class="name dim-month"></td>')
                cells.append(f'<td class="marker dim-month"></td>')
                continue
            entries = plan_section.get(d, [])
            if ci >= len(entries):
                cells.append(f'<td class="name"></td>')
                cells.append(f'<td class="marker"></td>')
                continue
            row = entries[ci]
            name, marker, alert = row[0], row[1], row[2]
            highlight = row[3] if len(row) > 3 else None
            link      = row[4] if len(row) > 4 else None
            name_classes = ["name", "client-cell"]
            marker_classes = ["marker"]
            if highlight == "yellow":
                name_classes.append("hl-yellow")
                marker_classes.append("hl-yellow")
            elif highlight == "bold":
                name_classes.append("hl-bold")
            if alert:
                marker_classes.append("alert")
            v = _fmt_marker_html(marker)
            if alert and not v:
                v = "!"
            data_client = _html_escape(name or "")
            name_html = data_client
            marker_html = v
            if link:
                marker_classes.append("link")
                marker_html = (f'<a href="{_html_escape(link)}" target="_blank" '
                               f'rel="noopener">{v}</a>')
            cells.append(f'<td class="{" ".join(name_classes)}" '
                         f'data-client="{data_client}">{name_html}</td>')
            cells.append(f'<td class="{" ".join(marker_classes)}">{marker_html}</td>')
        rows_html.append(f'<tr class="data-row">{"".join(cells)}</tr>')
    return "\n".join(rows_html)


def _render_week_card_html(wk, week_no, sections, today, holidays):
    first_d = next((d for d in wk if d), None)
    last_d  = next((d for d in reversed(wk) if d), None)
    if first_d and last_d:
        label = (f"Week {week_no}: {first_d.strftime('%m/%d')} – "
                 f"{last_d.strftime('%m/%d')}")
    else:
        label = f"Week {week_no}"

    # Header + date-strip rows.
    # Per user 2026-06-03: only the actual date cell is highlighted as
    # "today" — no vertical column accent. Header row stays uniform.
    header_cells = []
    date_cells = []
    for i, d in enumerate(wk):
        day_name = WEEKDAYS[i]
        header_cells.append(
            f'<th class="hdr-day">{day_name}</th>'
            f'<th class="hdr-date">Date</th>')
        if d is None:
            date_cells.append(
                '<td class="strip-name dim-month"></td>'
                '<td class="strip-date dim-month"></td>')
            continue
        hname = holidays.get(d, "")
        hname_cls = " holiday" if hname else ""
        is_today = " is-today" if d == today else ""
        date_str = d.strftime("%m/%d/%y")
        date_cells.append(
            f'<td class="strip-name{hname_cls}">{_html_escape(hname)}</td>'
            f'<td class="strip-date{is_today}">{date_str}</td>')

    sec_labels = {"daily": "Daily", "weekly": "Weekly", "monthly": "Monthly"}
    sec_html_parts = []
    for sec_key in ("daily", "weekly", "monthly", "kaiser"):
        body = _render_section_rows_html(wk, sections[sec_key], today)
        if body:
            lbl = sec_labels.get(sec_key)
            label_row = (f'<tr class="sec-label"><td colspan="10">{lbl}</td></tr>'
                         if lbl else "")
            sec_html_parts.append(
                f'<tbody class="sec sec-{sec_key}">{label_row}{body}</tbody>'
            )

    key_line = ("Key:  Date = Certified  |  ✓ = Loaded/Snapped  |  L = Loading"
                "  |  TBL = To Be Loaded"
                "  |  pink = Failure/Inactive  |  (s) SLA  |  (p) Rx Post Snap"
                "  |  (n) Not Delivered  |  -  = No load that day")

    # Explicit column widths so the table sizes to content, not the page.
    # Client-name columns ~165 px, date columns ~70 px, Tuesday date 95 px.
    colgroup = (
        '<colgroup>'
        '<col class="cn"><col class="dt">'
        '<col class="cn"><col class="dt-wide">'
        '<col class="cn"><col class="dt">'
        '<col class="cn"><col class="dt">'
        '<col class="cn"><col class="dt">'
        '</colgroup>'
    )

    return (
        f'<section class="week-card">'
        f'  <div class="week-label">{_html_escape(label)}</div>'
        f'  <table class="grid">'
        f'    {colgroup}'
        f'    <thead class="grid-head"><tr>{"".join(header_cells)}</tr></thead>'
        f'    <tbody class="strip"><tr>{"".join(date_cells)}</tr></tbody>'
        f'    {"".join(sec_html_parts)}'
        f'  </table>'
        f'  <div class="key">{_html_escape(key_line)}</div>'
        f'</section>'
    )


def build_dashboard_html(month_packs, today, current_month_name):
    """Render the dashboard HTML.
    month_packs: list of dicts {name, year, month, sections, weeks, holidays}
    """
    import calendar as _cal_mod

    tabs_html = []
    panels_html = []
    for mp in month_packs:
        is_current = (mp["name"] == current_month_name)
        active = " active" if is_current else ""
        tab_id = f'tab-{mp["name"].replace(" ", "-")}'
        tabs_html.append(
            f'<button class="tab{active}" data-target="{tab_id}">'
            f'{_html_escape(mp["name"])}</button>')

        week_blocks = []
        for week_no, wk in enumerate(mp["weeks"], start=1):
            week_blocks.append(_render_week_card_html(
                wk, week_no, mp["sections"], today, mp["holidays"]))
        panels_html.append(
            f'<section class="month-panel{active}" id="{tab_id}">'
            f'  {"".join(week_blocks)}'
            f'</section>')

    generated = datetime.now().strftime("%Y-%m-%d %H:%M")
    title = "Client Delivery Status"

    css = """
:root {
  --bg: #f4f6f9; --card: #ffffff; --border: #c8c8c8; --text: #1f2a37;
  --muted: #5b6776; --accent: #2C5F8A; --accent-dark: #1F3D5C;
  --day-fill: #E3EBF4; --today-fill: #FFD180; --today-strong: #f08c00;
  --alert: #FFC7CE; --alert-dark: #9C0006;
  --yellow: #FFF2A8; --yellow-dark: #7F6000;
  --holiday: #FFE4B5; --holiday-dark: #7C4A00;
}
* { box-sizing: border-box; }
body {
  margin: 0;
  font-family: "Segoe UI", -apple-system, BlinkMacSystemFont, sans-serif;
  background: var(--bg); color: var(--text); font-size: 13px;
}
header.bar {
  background: var(--accent-dark); color: #fff;
  padding: 10px 18px; position: sticky; top: 0; z-index: 5;
  display: flex; align-items: center; gap: 16px; flex-wrap: wrap;
}
header.bar h1 { margin: 0; font-size: 16px; font-weight: 600; }
header.bar .meta { font-size: 11px; opacity: 0.85; }
header.bar .grow { flex: 1; }
header.bar input[type=search] {
  border: 0; border-radius: 4px; padding: 6px 10px; font-size: 13px; min-width: 220px;
}
header.bar button {
  background: var(--accent); color: #fff; border: 0; border-radius: 4px;
  padding: 6px 12px; cursor: pointer; font-size: 12px;
}
header.bar .legend {
  font-size: 11px; opacity: 0.92; display: flex; align-items: center; gap: 6px;
}
header.bar .legend b {
  display: inline-block; background: var(--alert); color: var(--alert-dark);
  font-weight: 700; border-radius: 3px; padding: 0 6px; line-height: 16px;
}
nav.tabs {
  background: #fff; border-bottom: 1px solid var(--border);
  padding: 4px 12px; position: sticky; top: 44px; z-index: 4;
  display: flex; gap: 4px; flex-wrap: wrap;
}
nav.tabs button.tab {
  background: transparent; color: var(--accent-dark);
  border: 1px solid transparent; border-bottom: 0;
  border-radius: 4px 4px 0 0;
  padding: 6px 12px; cursor: pointer; font-size: 12px; font-weight: 600;
}
nav.tabs button.tab.active {
  background: var(--accent); color: #fff;
}
main { padding: 12px 18px 40px; }
.month-panel { display: none; }
.month-panel.active { display: block; }
.week-card {
  margin: 0 0 14px;
}
.week-label {
  font-weight: 700; font-size: 14px; color: var(--accent-dark);
  margin: 12px 0 4px;
}
/* Tightened 2026-06-03: table sizes to content (width: auto), not the page.
   Explicit column widths via <colgroup>: name ≈ 165 px, date ≈ 70 px,
   Tuesday date 95 px while BCBSAR is in Implementation. */
table.grid {
  border-collapse: separate; border-spacing: 0; table-layout: fixed;
  background: var(--card); width: auto;
}
table.grid col.cn       { width: 165px; }
table.grid col.dt       { width:  70px; }
table.grid col.dt-wide  { width:  95px; }
table.grid th, table.grid td {
  border: 1px solid var(--border); padding: 3px 6px;
  font-size: 12px; vertical-align: top;
  overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
}
table.grid th {
  background: var(--accent); color: #fff;
  font-weight: 600; font-size: 11px; text-align: center; letter-spacing: 0.3px;
}
.strip td { background: var(--day-fill); text-align: center; font-weight: 600; color: var(--accent-dark); }
.strip td.is-today {
  background: var(--today-fill); color: var(--accent-dark);
  box-shadow: inset 0 0 0 2px var(--today-strong);
}
.strip td.holiday { background: var(--holiday); color: var(--holiday-dark); font-style: italic; font-weight: 700; }
.strip td.dim-month, td.dim-month { background: #F5F5F5 !important; color: #aaa; }
td.name { white-space: nowrap; }
td.marker { text-align: center; font-variant-numeric: tabular-nums; }
td.alert { background: var(--alert); color: var(--alert-dark); font-weight: 700; }
td.hl-yellow { background: var(--yellow); color: var(--yellow-dark); font-weight: 700; }
.hl-bold { font-weight: 700; }
td.marker.link a { color: var(--alert-dark); font-weight: 700; text-decoration: underline; }
.sec-gap td { background: #fff; border: 0; height: 4px; padding: 0; }
.sec-label td { background: #E8EDF3; color: #2C5F8A; font-weight: 700; font-size: 10px; text-transform: uppercase; letter-spacing: 0.5px; padding: 2px 6px; border: 0; }
.key {
  font-style: italic; font-size: 11px; color: #555;
  margin: 4px 0 12px;
}
/* Search: dim all client-name cells, then re-emphasize the matched ones
   only. Per user 2026-06-03: highlight just the client name, not the
   whole row. */
body.search-active td.client-cell { opacity: 0.25; }
body.search-active td.client-cell.match-cell { opacity: 1; background: #fffbe8; font-weight: 600; }
"""

    js = """
(function() {
  function show(tabBtn) {
    document.querySelectorAll('nav.tabs button.tab').forEach(b => b.classList.remove('active'));
    document.querySelectorAll('.month-panel').forEach(p => p.classList.remove('active'));
    tabBtn.classList.add('active');
    const target = document.getElementById(tabBtn.dataset.target);
    if (target) target.classList.add('active');
    window.scrollTo({ top: 0 });
  }
  document.querySelectorAll('nav.tabs button.tab').forEach(btn => {
    btn.addEventListener('click', () => show(btn));
  });
  // Search
  const search = document.getElementById('search');
  function applySearch() {
    const q = search.value.trim().toLowerCase();
    document.body.classList.toggle('search-active', q.length > 0);
    // Per-cell highlight: only the matching client-name cell pops; the rest
    // of the row stays at normal/dimmed weight.
    document.querySelectorAll('td.client-cell').forEach(td => {
      const c = (td.dataset.client || '').toLowerCase();
      const isMatch = q && c && c.indexOf(q) !== -1;
      td.classList.toggle('match-cell', isMatch);
    });
  }
  search.addEventListener('input', applySearch);
  // Today jump. The today cell can live in a NON-active month panel — a
  // month-end week (e.g. 6/29-7/3) is rendered on the next month's tab, and a
  // hidden (display:none) panel can't be scrolled to. So switch to the panel
  // that contains today first, then scroll.
  const todayBtn = document.getElementById('today-jump');
  if (todayBtn) {
    todayBtn.addEventListener('click', () => {
      const el = document.querySelector('.strip td.is-today');
      if (!el) return;
      const panel = el.closest('.month-panel');
      if (panel && !panel.classList.contains('active')) {
        const tabBtn = document.querySelector(
          'nav.tabs button.tab[data-target="' + panel.id + '"]');
        if (tabBtn) show(tabBtn);
      }
      el.scrollIntoView({ behavior: 'smooth', block: 'center' });
    });
  }
})();
"""

    head = (
        "<!doctype html><html lang='en'><head>"
        "<meta charset='utf-8'>"
        f"<title>{_html_escape(title)}</title>"
        f"<style>{css}</style>"
        "</head>"
    )
    body = (
        "<body>"
        "<header class='bar'>"
        f"  <h1>{_html_escape(title)}</h1>"
        f"  <div class='meta'>Generated {generated}</div>"
        "  <div class='legend'><b>!</b> = flagged / needs attention "
        "(past due or problem, no delivery recorded yet)</div>"
        "  <div class='grow'></div>"
        "  <input type='search' id='search' placeholder='Filter by client name…' autocomplete='off'>"
        "  <button id='today-jump'>Jump to today</button>"
        "</header>"
        f"<nav class='tabs'>{''.join(tabs_html)}</nav>"
        f"<main>{''.join(panels_html)}</main>"
        f"<script>{js}</script>"
        "</body></html>"
    )
    return head + body


def main():
    today = date.today()
    year, month = today.year, today.month
    month_start = date(year, month, 1)

    print(f"[info] Today: {today}  Month: {year}-{month:02d}")

    print("[info] Querying DHT cert table…")
    # pull 3 months back for monthly clients' avg-day calc
    certs = fetch_dht_certs(since=date(year, month, 1) - timedelta(days=90))
    cert_idx = build_cert_index(certs)
    global CERT_WEEK_IDX
    CERT_WEEK_IDX = build_cert_week_index(certs)
    print(f"[info]   {len(certs)} cert rows / {len(cert_idx)} distinct DatabaseNames")

    print("[info] Fetching ADO tickets…")
    tickets = fetch_ado_tickets(min_changed_date=month_start - timedelta(days=14))
    print(f"[info]   {len(tickets)} delivery tickets in window")

    print("[info] Fetching RAMP jobs…")
    jobs = fetch_ramp_jobs()
    enabled_n = sum(1 for j in jobs if j.get("Enabled") == 1)
    print(f"[info]   {len(jobs)} total ({enabled_n} enabled)")

    # ----- Auto-discover new MasterLoad 0110 Load implementations (per user
    # 2026-06-03, updated 2026-06-08). New clients default to Weekly/Monday
    # placement as CERT-STYLE clients: they stay "L" until the cert date lands
    # and never show a ✓ (no SNAP_ONLY membership). Per user 2026-06-08: "All
    # new 'MasterLoad 0110 Load' will also be in this format." The old
    # PBMRx→SNAP_KIND_ONLY (✓ after snap) and non-PBMRx→IMPLEMENTATION_LOAD_ONLY
    # (blank after snap) branches were removed. (ElevanceMMMRx was promoted to a
    # DAILY_CLIENTS client 2026-06-16 and is now "known", so it's no longer
    # rediscovered here.)
    new_impls = find_unconfigured_masterload_clients(jobs)
    if new_impls:
        for entry in new_impls:
            client = entry["raw"]
            print(f"[info]   NEW MasterLoad implementation: {client} "
                  f"(pbmrx={entry['pbmrx']}, enabled={entry['enabled']})")
            if client not in WEEKLY_CLIENTS:
                WEEKLY_CLIENTS[client] = ["Monday"]
            # Add alias so RAMP job lookups find the new client.
            aliases = CLIENT_ALIASES.setdefault(client, [])
            if entry["normalized"] not in aliases:
                aliases.append(entry["normalized"])

    # ----- Auto-Inactive clients whose primary 0100/0110 jobs are all
    # Inactive in RAMP (Enabled=0). Kaiser feeds excluded per user.
    auto_inactive = auto_inactive_from_ramp(jobs)
    if auto_inactive:
        print(f"[info]   Auto-Inactive (0100/0110 disabled, non-Kaiser): "
              f"{sorted(auto_inactive)}")
        FORCED_INACTIVE.update(auto_inactive)

    # Reach back to the first COMPUTED month (months 1..N are frozen xlsx
    # snapshots that never consult live data) so past-month tabs get their
    # snap/load-driven markers. Both the queue (snap/load completions, esp. for
    # snap-only monthly clients like Kaiser GE / MMOH whose May snaps have rolled
    # off the RAMP snap endpoint) and the snap endpoint honor this floor.
    first_computed_month = (max(EXPECTED_DATES_FILES) + 1) if year == 2026 else 1
    snap_since = (date(year, first_computed_month, 1)
                  if 1 <= first_computed_month <= 12 else None)
    print("[info] Fetching RAMP queue + snap history…")
    queue = fetch_ramp_queue(since=snap_since)
    snaps = fetch_ramp_snaps(since=snap_since)
    print(f"[info]   queue={len(queue)}, snaps={len(snaps)}")

    print("[info] Fetching tape loads…")
    since_dt = date(year, month, 1) - timedelta(days=14)
    tape_loads = {}
    for client, (db, src_key) in TAPE_LOAD_SOURCES.items():
        server = TAPE_LOAD_SERVER.get(client, "TRGETL3")
        name_like = TAPE_LOAD_NAME_FILTER.get(client)
        rows = fetch_tape_loads(db, since_dt, server=server, name_like=name_like)
        tape_loads[src_key] = rows
        print(f"[info]   {server}.{db}: {len(rows)} rows")

    print("[info] Fetching multi-week client tape loads…")
    multi_week_loads = {}
    for client, (db, _pattern) in MULTI_WEEK_CLIENTS.items():
        rows = fetch_tape_loads(db, since_dt)
        multi_week_loads[client] = rows
        print(f"[info]   {client}@{db}: {len(rows)} rows")

    snap_idx = build_snap_index(jobs, queue, snaps, tape_loads=tape_loads)
    print(f"[info] snap index dates: {len(snap_idx)}")

    print("[info] Querying SQLUtilAudit for Aetna NMSP NonMSP loads…")
    # Reach back to the first computed month (not just 30 days) so past-month
    # tabs get their MMSEA ✓ — e.g. the 5/28 NonMSP import for May.
    aetna_nmsp_loads = fetch_aetna_nmsp_loads(
        since=(snap_since or date(year, month, 1) - timedelta(days=30)))
    print(f"[info]   {len(aetna_nmsp_loads)} Aetna NonMSP entries")

    print("[info] Checking EverNorthRx claims (ESI_PAID_CLAIMS_*) staged-not-loaded…")
    evernorth_claims_pending = compute_evernorth_claims_pending(
        queue, jobs, since=date(year, month, 1) - timedelta(days=14))
    print(f"[info]   EverNorthRx claims pending load: {evernorth_claims_pending}")

    print("[info] Fetching OptumPBMRx per-RAW instances (TRGETL3 tape, all statuses)…")
    optum_raw_instances = fetch_optum_raw_instances(since=date(year, month, 1) - timedelta(days=14))
    print(f"[info]   OptumPBMRx RAW instances: {len(optum_raw_instances)}")

    print("[info] Fetching CVSPBMRx Ad Hoc (oversized/backfill) tape files…")
    cvspbm_adhoc = fetch_cvspbm_adhoc(since=date(year, month, 1) - timedelta(days=14))
    print(f"[info]   CVSPBMRx Ad Hoc instances: {len(cvspbm_adhoc)}")

    print("[info] Fetching CVSPBMRx weekly eligibility files (data-date cells)…")
    # Wide window: the report renders every month of the year, and a late load
    # (e.g. the 6/27 file loaded 7/19) must still ✓ on its data-date week.
    cvspbm_weekly = fetch_cvspbmrx_weekly(since=date(year, 1, 1))
    print(f"[info]   CVSPBMRx weekly files: {len(cvspbm_weekly)} "
          f"(loaded={sum(1 for r in cvspbm_weekly if r['loaded'])})")

    print("[info] Fetching stage-file data-date loads (TuftsRx/Oscar/Tufts_PublicPlan/MedicalMutualMHS)…")
    # Data-date-driven delivery cells for late/missed loads (see
    # STAGE_FILE_CELL_CLIENTS). Whole-year window like CVSPBMRx.
    stage_file_loads = {}
    for sf_client, sf_cfg in STAGE_FILE_CELL_CLIENTS.items():
        try:
            stage_file_loads[sf_client] = fetch_stage_file_loads(
                sf_client, sf_cfg, since=date(year, 1, 1))
        except Exception as e:
            print(f"[warn]   {sf_client} stage-file fetch failed: {e}")
            stage_file_loads[sf_client] = []
    print("[info]   stage-file loads: "
          + ", ".join(f"{c}={len(r)}" for c, r in stage_file_loads.items()))

    print("[info] Fetching HealthNetCA backfill claims ranges (TRGETL1 tblTape)…")
    try:
        healthnetca_claim_loads = fetch_healthnetca_claim_loads()
    except Exception as e:
        print(f"[warn]   HealthNetCA claims fetch failed: {e}")
        healthnetca_claim_loads = []
    healthnetca_ranges = healthnetca_range_labels(healthnetca_claim_loads)
    if healthnetca_ranges:
        print("[info]   HealthNetCA backfill: "
              + ", ".join(f"{d.month}/{d.day} cell{lbl}"
                          for d, lbl in sorted(healthnetca_ranges.items())))
    else:
        print("[info]   HealthNetCA backfill: no claims loads since "
              f"{HEALTHNETCA_BACKFILL_FROM:%Y-%m-%d}")

    print("[info] Fetching JHHC Passfile loads (TRGETL4 TableID=5000)…")
    # Widest window: the report renders every month of the year in one run, and
    # the JHHCPassfile ✓ for a past month (e.g. June) is anchored to that
    # month's FileLoadDate — so fetch the whole year, not just a 2-week window.
    jhhc_passfile_loads = fetch_jhhc_passfile_loads(since=date(year, 1, 1))
    print(f"[info]   JHHC Passfile tape rows: {len(jhhc_passfile_loads)} "
          f"(loaded={sum(1 for r in jhhc_passfile_loads if r['loaded'])})")

    latest_tickets, monthly_placements = build_ticket_index(tickets, jobs)
    print(f"[info] latest tickets indexed for {len(latest_tickets)} clients")

    wb = Workbook()
    # Drop the auto-created default sheet so we control ordering.
    del wb[wb.sheetnames[0]]

    # Client Owner sits first.
    ws_owner = wb.create_sheet("Client Owner")
    write_client_owner_sheet(ws_owner)

    import calendar as _cal_mod
    current_tab_name = f"{_cal_mod.month_name[month]} {year}"

    # Generate every month of the year so prior months don't drop off as they
    # conclude (per user 2026-06-03). Compute current month last in the loop
    # iteration but build tabs in Jan→Dec order for natural tab ordering.
    load_sticky_certs()
    print(f"[info] sticky cert cache: {len(STICKY_CERTS)} remembered cells")

    month_packs = []
    for m in range(1, 13):
        sec_m = wk_m = None
        # For closed months Jan–Apr 2026 the manually-maintained
        # ExpectedClientDates_*.xlsx files are the source of truth. Live DHT
        # / RAMP history rolls off too quickly to reconstruct.
        snapshot_file = EXPECTED_DATES_FILES.get(m) if year == 2026 else None
        if snapshot_file:
            try:
                snap_path = os.path.join(EXPECTED_DATES_DIR, snapshot_file)
                if os.path.exists(snap_path):
                    sec_m, wk_m = parse_all_clients_xlsx(snap_path, year)
                    if sec_m is not None and wk_m:
                        print(f"[info]   {_cal_mod.month_name[m]}: loaded from {snapshot_file}")
                    else:
                        sec_m = wk_m = None
            except Exception as e:
                print(f"[warn] failed to parse {snapshot_file}: {e}")
                sec_m = wk_m = None

        if sec_m is None or wk_m is None:
            sec_m, wk_m = plan_calendar(year, m, cert_idx, snap_idx,
                                        latest_tickets, monthly_placements,
                                        jobs, queue,
                                        esipbmrx_tape=tape_loads.get("esipbmrx"),
                                        multi_week_loads=multi_week_loads,
                                        aetna_nmsp_loads=aetna_nmsp_loads,
                                        optumpbmrx_tape=tape_loads.get("optumpbmrx"),
                                        evernorth_claims_pending=evernorth_claims_pending,
                                        optum_raw_instances=optum_raw_instances,
                                        cvspbm_adhoc=cvspbm_adhoc,
                                        jhhc_passfile_loads=jhhc_passfile_loads,
                                        cvspbm_weekly=cvspbm_weekly,
                                        stage_file_loads=stage_file_loads,
                                        healthnetca_ranges=healthnetca_ranges)

        tab_name = f"{_cal_mod.month_name[m]} {year}"
        ws_m = wb.create_sheet(tab_name)
        write_weekly_stacked(ws_m, year, m, sec_m, wk_m, today)
        month_packs.append({
            "name": tab_name, "year": year, "month": m,
            "sections": sec_m, "weeks": wk_m,
            "holidays": us_federal_holidays(year),
        })
        if m == month:
            sections, weeks = sec_m, wk_m

    # Persist any newly-seen cert dates so future runs can't regress them to "!".
    save_sticky_certs()
    print(f"[info] sticky cert cache saved: {len(STICKY_CERTS)} cells")

    # Open the workbook to the current-month tab by default.
    wb.active = wb.sheetnames.index(current_tab_name)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, f"ClientDeliveryStatus_{today:%Y-%m-%d}.xlsx")
    try:
        wb.save(out_path)
    except PermissionError:
        # file is open in Excel — fall back to a timestamped name so we can still verify
        fallback = os.path.join(OUTPUT_DIR,
                                f"ClientDeliveryStatus_{today:%Y-%m-%d}_{datetime.now():%H%M%S}.xlsx")
        wb.save(fallback)
        out_path = fallback
        print(f"[warn] primary file locked; wrote {fallback} instead")
    print(f"[done] Wrote {out_path}")

    # Project-folder copy (dated filename for archival inspection).
    try:
        os.makedirs(LOCAL_COPY_DIR, exist_ok=True)
        local_path = os.path.join(LOCAL_COPY_DIR,
                                  f"ClientDeliveryStatus_{today:%Y-%m-%d}.xlsx")
        import shutil
        shutil.copyfile(out_path, local_path)
        print(f"[done] Local copy: {local_path}")
    except Exception as e:
        print(f"[warn] Local-copy failed: {e}")

    # OneDrive copy with a FIXED filename so a single Notion link stays valid
    # across runs (OneDrive auto-syncs to SharePoint; same URL serves latest).
    try:
        os.makedirs(os.path.dirname(ONEDRIVE_COPY_PATH), exist_ok=True)
        import shutil
        shutil.copyfile(out_path, ONEDRIVE_COPY_PATH)
        print(f"[done] OneDrive copy: {ONEDRIVE_COPY_PATH}")
    except Exception as e:
        print(f"[warn] OneDrive-copy failed: {e}")

    # Dashboard HTML — self-contained file with search + today emphasis +
    # sticky toolbar/tabs. Same color codes and weekly-stacked layout as the
    # .xlsx, just rendered in the browser. Sits alongside the .xlsx.
    try:
        html_str = build_dashboard_html(month_packs, today, current_tab_name)
        html_paths = [
            os.path.join(OUTPUT_DIR, "ClientDeliveryStatus.html"),
            os.path.join(LOCAL_COPY_DIR, "ClientDeliveryStatus.html"),
            os.path.join(os.path.dirname(ONEDRIVE_COPY_PATH),
                         "ClientDeliveryStatus.html"),
        ]
        wrote_one = False
        for p in html_paths:
            try:
                os.makedirs(os.path.dirname(p), exist_ok=True)
                with open(p, "w", encoding="utf-8") as f:
                    f.write(html_str)
                print(f"[done] HTML dashboard: {p}")
                wrote_one = True
            except Exception as e:
                print(f"[warn] HTML write failed for {p}: {e}")
        if not wrote_one:
            print("[warn] HTML dashboard not written to any path")
    except Exception as e:
        print(f"[warn] HTML dashboard build failed: {e}")

    return out_path


if __name__ == "__main__":
    main()
